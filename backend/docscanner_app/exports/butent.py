import os
import logging
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import date, datetime
from typing import List, Dict, Any

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from ..utils.extra_fields import get_extra_for_export

logger = logging.getLogger(__name__)

# =========================
# Конфиг путей
# =========================
env_value = os.getenv("BUTENT_TEMPLATES_DIR")

if not env_value:
    raise ValueError("BUTENT_TEMPLATES_DIR not set in .env")

TEMPLATES_DIR = Path(env_value)

# Единый шаблон для обоих режимов (suminis и kiekinis)
BUTENT_TEMPLATE_FILE = "Butent_Import_Template.xlsx"

# =========================
# Helpers
# =========================

def _safe_D(x):
    """Безопасное преобразование в Decimal."""
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")


def _s(v):
    """Безопасная строка с strip()."""
    return str(v).strip() if v is not None else ""


def _use_matched_catalog(item) -> bool:
    """Есть ли валидный каталог-матч на строке (и юзер его не отключил)."""
    if getattr(item, "catalog_match_user_override", False):
        return False

    matched_code = _s(getattr(item, "matched_prekes_kodas", ""))
    return bool(matched_code) and matched_code.upper() != "UKN0"


def _resolved_field(item, field_name: str) -> str:
    """
    Если есть валидный каталог-матч — берём matched_ поле.
    Если matched_ поле пустое — fallback на оригинальное.
    """
    if _use_matched_catalog(item):
        matched_value = _s(
            getattr(item, f"matched_{field_name}", "")
        )
        if matched_value:
            return matched_value

    return _s(getattr(item, field_name, ""))


def _is_zero(v) -> bool:
    """Нулевая ставка НДС? None/'' считаем как 0."""
    try:
        return Decimal(str(v)) == 0
    except Exception:
        return True


EU_ISO2 = {
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR", "HU", "IE",
    "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK", "SI", "ES", "SE"
}


def _is_eu_country(iso: object) -> bool:
    """True только для явных ISO2 из списка ЕС. Пустое значение -> False."""
    if not iso:
        return False
    return str(iso).strip().upper() in EU_ISO2


def _pick_isaf_for_purchase(doc):
    """
    Возвращает:
      - '12' -> Neformuoti (НЕ включать в i.SAF)
      - None -> не ставить тег вовсе (включать по умолчанию)

    Правило:
      если (seller_country_iso пусто ИЛИ не-ЕС) И ВСЕ ставки vat_percent по строкам == 0
      -> '12', иначе None.
    """
    country = getattr(doc, "seller_country_iso", "") or ""
    is_eu = _is_eu_country(country)
    non_eu_or_empty = not is_eu

    line_items = getattr(doc, "line_items", None)
    if line_items and hasattr(line_items, "all") and line_items.exists():
        vat_zero_all = all(_is_zero(getattr(it, "vat_percent", None)) for it in line_items.all())
    else:
        vat_zero_all = _is_zero(getattr(doc, "vat_percent", None))

    if non_eu_or_empty and vat_zero_all:
        return "12"

    return None


def _get_butent_isaf_flag(doc) -> int:
    """
    Для Būtent возвращает 1 или 0:
      - 0 -> НЕ включать в i.SAF
      - 1 -> включать в i.SAF (по умолчанию)
    """
    _isaf = getattr(doc, "report_to_isaf", None)
    if _isaf is False:
        return 0
    rivile_code = _pick_isaf_for_purchase(doc)
    return 0 if rivile_code == "12" else 1


def get_party_code(
    doc,
    *,
    role: str,
    id_field: str,
    vat_field: str,
    id_programoje_field: str,
) -> str:
    """
    Код стороны (seller/buyer) по приоритету:
      1) *_id
      2) *_vat_code
      3) *_id_programoje
    Если все пусто - вернётся "".
    """
    sid = _s(getattr(doc, id_field, None))
    if sid:
        logger.info("[BUTENT:PARTY] %s: %s -> %s", role, id_field, sid)
        return sid

    svat = _s(getattr(doc, vat_field, None))
    if svat:
        logger.info("[BUTENT:PARTY] %s: %s -> %s", role, vat_field, svat)
        return svat

    sidp = _s(getattr(doc, id_programoje_field, None))
    if sidp:
        logger.info("[BUTENT:PARTY] %s: %s -> %s", role, id_programoje_field, sidp)
        return sidp

    logger.info("[BUTENT:PARTY] %s: empty id/vat/id_programoje -> ''", role)
    return ""


# =========================
# Per-company extra fields helper
# =========================

def _parse_cp_key(cp_key):
    if not cp_key:
        return ""

    cp = str(cp_key).strip()
    if cp.lower().startswith("id:"):
        return cp.split(":", 1)[1].strip()
    return cp


def _get_own_company_code_from_doc(doc):
    """
    Определяет код своей фирмы из документа.

    - pirkimas -> своя фирма buyer
    - pardavimas -> своя фирма seller
    """
    doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()

    if doc_type == "pirkimas":
        candidates = [
            getattr(doc, "buyer_id", ""),
            getattr(doc, "buyer_vat_code", ""),
            getattr(doc, "buyer_id_programoje", ""),
        ]
    else:
        candidates = [
            getattr(doc, "seller_id", ""),
            getattr(doc, "seller_vat_code", ""),
            getattr(doc, "seller_id_programoje", ""),
        ]

    for value in candidates:
        code = _s(value)
        if code:
            return code
    return ""


def _get_butent_extra_for_doc(user, doc, own_company_code=None) -> Dict[str, Any]:
    """
    Получает extra fields для конкретного документа.

    Приоритет:
    1. Профиль конкретной фирмы по own_company_code
    2. Профиль фирмы, определённой из документа
    3. Глобальный профиль (__all__)
    4. Пустой dict
    """
    if not user:
        return {}

    requested_code = _parse_cp_key(own_company_code)
    doc_company_code = _get_own_company_code_from_doc(doc)

    extra = {}
    resolved_by = ""

    if requested_code:
        extra = get_extra_for_export(user, "butent", requested_code)
        if extra:
            resolved_by = requested_code

    if not extra and doc_company_code and doc_company_code != requested_code:
        extra = get_extra_for_export(user, "butent", doc_company_code)
        if extra:
            resolved_by = doc_company_code

    if not extra:
        extra = get_extra_for_export(user, "butent", None)
        if extra:
            resolved_by = "__all__/legacy"

    logger.info(
        "[BUTENT:EXTRA] doc=%s own_company_code=%r requested_code=%r doc_company_code=%r resolved_by=%r fields=%s",
        getattr(doc, "pk", None),
        own_company_code,
        requested_code,
        doc_company_code,
        resolved_by,
        {k: v for k, v in extra.items() if v} if extra else {},
    )

    return extra or {}


def _get_client_data_for_butent(doc) -> Dict[str, Any]:
    """
    Возвращает словарь с данными клиента для колонок K-Q.
    Логика:
      - pirkimas -> seller (продавец)
      - pardavimas -> buyer (покупатель)
    """
    doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()

    if doc_type == "pirkimas":
        code = get_party_code(
            doc,
            role="seller",
            id_field="seller_id",
            vat_field="seller_vat_code",
            id_programoje_field="seller_id_programoje",
        )
        return {
            "code": code,
            "fizinis": 1 if getattr(doc, "seller_is_person", False) else 0,
            "vat": _s(getattr(doc, "seller_vat_code", "")),
            "name": _s(getattr(doc, "seller_name", "")),
            "address": _s(getattr(doc, "seller_address", "")),
            "country": _s(getattr(doc, "seller_country_iso", "")),
            "iban": _s(getattr(doc, "seller_iban", "")),
        }
    else:
        code = get_party_code(
            doc,
            role="buyer",
            id_field="buyer_id",
            vat_field="buyer_vat_code",
            id_programoje_field="buyer_id_programoje",
        )
        return {
            "code": code,
            "fizinis": 1 if getattr(doc, "buyer_is_person", False) else 0,
            "vat": _s(getattr(doc, "buyer_vat_code", "")),
            "name": _s(getattr(doc, "buyer_name", "")),
            "address": _s(getattr(doc, "buyer_address", "")),
            "country": _s(getattr(doc, "buyer_country_iso", "")),
            "iban": _s(getattr(doc, "buyer_iban", "")),
        }


def _format_date_for_butent(dt) -> str:
    """
    Форматирует дату в строку 'YYYY.MM.DD' для Būtent.
    """
    if not dt:
        return ""
    if isinstance(dt, str):
        try:
            dt = datetime.strptime(dt, "%Y-%m-%d").date()
        except ValueError:
            return ""
    if isinstance(dt, (date, datetime)):
        return dt.strftime("%Y.%m.%d")
    return ""


def _get_operacija(doc, extra_fields=None) -> str:
    """
    Определяет операцию для колонки H:
      - pirkimas -> "Pajamavimas" (или из extra_fields['pirkimas_operacija'])
      - pardavimas -> "Pardavimas" (или из extra_fields['pardavimas_operacija'])
    """
    doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()
    extra_fields = extra_fields or {}

    if doc_type == "pirkimas":
        custom_op = _s(extra_fields.get("pirkimas_operacija", ""))
        if custom_op:
            logger.info(
                "[BUTENT:OPERACIJA] doc=%s using custom pirkimas_operacija=%r",
                getattr(doc, "pk", None), custom_op
            )
            return custom_op
        return "Pajamavimas"

    if doc_type == "pardavimas":
        custom_op = _s(extra_fields.get("pardavimas_operacija", ""))
        if custom_op:
            logger.info(
                "[BUTENT:OPERACIJA] doc=%s using custom pardavimas_operacija=%r",
                getattr(doc, "pk", None), custom_op
            )
            return custom_op
        return "Pardavimas"

    logger.warning("[BUTENT] Unknown pirkimas_pardavimas=%r, defaulting to Pajamavimas", doc_type)
    return "Pajamavimas"


def _get_sandelis(doc, extra_fields=None) -> str:
    """
    Определяет склад для колонки I:
      - doc.sandelio_kodas если есть
      - иначе из extra_fields (pirkimas_sandelis/pardavimas_sandelis)
      - иначе "S1"
    """
    sandelis = _s(getattr(doc, "sandelio_kodas", ""))
    if sandelis:
        return sandelis

    extra_fields = extra_fields or {}
    doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()

    if doc_type == "pirkimas":
        custom_sandelis = _s(extra_fields.get("pirkimas_sandelis", ""))
        if custom_sandelis:
            logger.info(
                "[BUTENT:SANDELIS] doc=%s using custom pirkimas_sandelis=%r",
                getattr(doc, "pk", None), custom_sandelis
            )
            return custom_sandelis
    elif doc_type == "pardavimas":
        custom_sandelis = _s(extra_fields.get("pardavimas_sandelis", ""))
        if custom_sandelis:
            logger.info(
                "[BUTENT:SANDELIS] doc=%s using custom pardavimas_sandelis=%r",
                getattr(doc, "pk", None), custom_sandelis
            )
            return custom_sandelis

    return "S1"


def _format_decimal(value, decimals=2) -> float:
    """
    Преобразует значение в float с округлением.
    Возвращает ЧИСЛО (float), не строку.
    """
    try:
        d = Decimal(str(value))
        rounded = d.quantize(Decimal(10) ** -decimals, rounding=ROUND_HALF_UP)
        return float(rounded)
    except Exception:
        return 0.0


def _distribute_discount_to_butent_lines(doc, items_list: list) -> None:
    """
    Распределяет скидку документа (invoice_discount_wo_vat) на строки товаров.

    ВАЖНО: Būtent не имеет поля для скидки документа, поэтому мы:
      1. ВЫЧИТАЕМ долю скидки из subtotal каждой строки
      2. ПЕРЕСЧИТЫВАЕМ price = new_subtotal / quantity
      3. ПЕРЕСЧИТЫВАЕМ vat = new_subtotal × vat_percent / 100

    Args:
        doc: документ с полем invoice_discount_wo_vat
        items_list: список объектов LineItem (модифицируется in-place)

    Модифицирует:
        Устанавливает атрибуты _butent_price_after_discount и _butent_vat_after_discount
    """
    if not items_list:
        return

    discount_raw = getattr(doc, "invoice_discount_wo_vat", None)
    if discount_raw in (None, "", 0, "0"):
        return

    try:
        discount_wo = Decimal(str(discount_raw))
    except (ValueError, InvalidOperation):
        logger.warning(
            "[BUTENT:DISCOUNT] doc=%s invalid discount value: %r",
            getattr(doc, "pk", None), discount_raw
        )
        return

    if discount_wo <= 0:
        return

    logger.info(
        "[BUTENT:DISCOUNT] doc=%s distributing discount=%.2f across %d lines",
        getattr(doc, "pk", None), discount_wo, len(items_list)
    )

    sum_subtotal_before = Decimal("0")
    for item in items_list:
        price = Decimal(str(getattr(item, "price", 0) or 0))
        qty = Decimal(str(getattr(item, "quantity", 1) or 1))
        sum_subtotal_before += price * qty

    if sum_subtotal_before <= 0:
        logger.warning(
            "[BUTENT:DISCOUNT] doc=%s sum_subtotal=0, cannot distribute",
            getattr(doc, "pk", None)
        )
        return

    discount_distributed = Decimal("0")

    for i, item in enumerate(items_list):
        qty = Decimal(str(getattr(item, "quantity", 1) or 1))
        price_before = Decimal(str(getattr(item, "price", 0) or 0))
        vat_percent = Decimal(str(getattr(item, "vat_percent", 0) or 0))

        subtotal_before = price_before * qty

        if i == len(items_list) - 1:
            line_discount = discount_wo - discount_distributed
        else:
            share = subtotal_before / sum_subtotal_before
            line_discount = (discount_wo * share).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            discount_distributed += line_discount

        subtotal_after = subtotal_before - line_discount

        if qty > 0:
            price_after = (subtotal_after / qty).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
        else:
            price_after = Decimal("0")

        if vat_percent > 0 and subtotal_after > 0:
            vat_after = (subtotal_after * vat_percent / Decimal("100")).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
        else:
            vat_after = Decimal("0")

        setattr(item, "_butent_price_after_discount", float(price_after))
        setattr(item, "_butent_vat_after_discount", float(vat_after))

        logger.debug(
            "[BUTENT:DISCOUNT] line=%d qty=%.2f price: %.2f->%.2f vat: %.2f->%.2f (discount=%.2f)",
            i,
            float(qty),
            float(price_before),
            float(price_after),
            float(Decimal(str(getattr(item, "vat", 0) or 0))),
            float(vat_after),
            float(line_discount)
        )


# =========================
# PVM Kodas helpers
# =========================

def _get_pvm_kodas_for_item(doc, item, line_map=None, default="") -> str:
    """
    Получает PVM kodas для строки с учётом резолвера и separate_vat.

    Приоритет:
    1. line_map[item.id] - от резолвера (если есть)
    2. item.pvm_kodas - из БД
    3. default - fallback

    ВАЖНО: "Keli skirtingi PVM" - это маркер, не реальный код -> возвращаем default
    """
    item_id = getattr(item, "id", None)

    if line_map is not None and item_id is not None and item_id in line_map:
        pvm = _s(line_map.get(item_id, ""))
        if pvm and pvm != "Keli skirtingi PVM":
            return pvm

    pvm = _s(getattr(item, "pvm_kodas", ""))
    if pvm and pvm != "Keli skirtingi PVM":
        return pvm

    return default


def _get_pvm_kodas_for_doc(doc, default="") -> str:
    """
    Получает PVM kodas для документа (sumiskai режим).

    ВАЖНО:
    - При separate_vat=True -> пустой (смешанные ставки)
    - "Keli skirtingi PVM" -> пустой (это маркер, не код)
    """
    separate_vat = bool(getattr(doc, "separate_vat", False))
    scan_type = _s(getattr(doc, "scan_type", "")).lower()

    if separate_vat and scan_type in ("sumiskai", "summary", "suminis"):
        return default

    pvm = _s(getattr(doc, "pvm_kodas", ""))

    if pvm == "Keli skirtingi PVM":
        return default

    return pvm or default


# =========================
# Основная функция экспорта
# =========================

def export_to_butent(
    documents: List,
    mode: str = "auto",
    user=None,
    own_company_code=None,
) -> Dict[str, bytes]:
    """
    Экспортирует документы в формат Būtent Excel.

    Args:
        documents: список документов для экспорта
        mode: 'auto' | 'suminis' | 'kiekinis'
              'auto' - автоматически разделяет документы на два файла
              'suminis' - принудительно все в один файл (suminis режим)
              'kiekinis' - принудительно все в один файл (kiekinis режим)
        user: пользователь
        own_company_code: код своей фирмы для поиска профиля extra_fields

    Returns:
        Dict[str, bytes]: словарь вида {"suminis": bytes, "kiekinis": bytes}
                         где ключи присутствуют только если есть соответствующие документы
    """
    logger.info(
        "[BUTENT:EXPORT] Starting export, docs=%d mode=%s own_company_code=%r",
        len(documents), mode, own_company_code
    )

    if not documents:
        logger.warning("[BUTENT:EXPORT] No documents to export")
        raise ValueError("No documents provided for export")

    docs_suminis = []
    docs_kiekinis = []

    if mode == "auto":
        for doc in documents:
            line_items = getattr(doc, "line_items", None)
            has_items = False
            if line_items and hasattr(line_items, "all"):
                has_items = line_items.exists()

            if has_items:
                docs_kiekinis.append(doc)
            else:
                docs_suminis.append(doc)

        logger.info(
            "[BUTENT:EXPORT] Auto mode: suminis=%d kiekinis=%d",
            len(docs_suminis),
            len(docs_kiekinis)
        )
    elif mode == "suminis":
        docs_suminis = documents
        logger.info("[BUTENT:EXPORT] Force suminis mode: %d docs", len(docs_suminis))
    elif mode == "kiekinis":
        docs_kiekinis = documents
        logger.info("[BUTENT:EXPORT] Force kiekinis mode: %d docs", len(docs_kiekinis))
    else:
        logger.error("[BUTENT:EXPORT] Unknown mode: %s", mode)
        raise ValueError(f"Unknown mode: {mode}")

    result = {}

    if docs_suminis:
        logger.info("[BUTENT:EXPORT] Generating suminis file...")
        result["suminis"] = _generate_butent_file(docs_suminis, "suminis", user, own_company_code)

    if docs_kiekinis:
        logger.info("[BUTENT:EXPORT] Generating kiekinis file...")
        result["kiekinis"] = _generate_butent_file(docs_kiekinis, "kiekinis", user, own_company_code)

    if not result:
        logger.warning("[BUTENT:EXPORT] No files generated")
        raise ValueError("No documents to export")

    logger.info("[BUTENT:EXPORT] Export completed, files=%s", list(result.keys()))
    return result


def _generate_butent_file(documents: List, mode: str, user=None, own_company_code=None) -> bytes:
    """
    Генерирует один Excel-файл для Būtent.

    Args:
        documents: список документов для экспорта
        mode: 'suminis' | 'kiekinis'
        user: пользователь
        own_company_code: код своей фирмы для поиска профиля extra_fields

    Returns:
        bytes: содержимое Excel-файла
    """
    logger.info(
        "[BUTENT:FILE] Generating %s file for %d docs own_company_code=%r",
        mode, len(documents), own_company_code
    )

    template_path = TEMPLATES_DIR / BUTENT_TEMPLATE_FILE
    if not template_path.exists():
        logger.error("[BUTENT:FILE] Template not found: %s", template_path)
        raise FileNotFoundError(f"Būtent template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    rows = []

    for doc in documents:
        line_map = getattr(doc, "_pvm_line_map", None)
        extra_fields = _get_butent_extra_for_doc(user, doc, own_company_code)

        operacija = _get_operacija(doc, extra_fields)
        sandelis = _get_sandelis(doc, extra_fields)
        isaf = _get_butent_isaf_flag(doc)
        client = _get_client_data_for_butent(doc)

        doc_common = [
            _format_date_for_butent(getattr(doc, "invoice_date", None)),
            _format_date_for_butent(getattr(doc, "operation_date", None)),
            _format_date_for_butent(getattr(doc, "due_date", None)),
            isaf,
            _s(getattr(doc, "document_series", "")),
            _s(getattr(doc, "document_number", "")),
            _s(getattr(doc, "order_number", "")),
            operacija,
            sandelis,
            _s(getattr(doc, "preview_url", "")),
            client["code"],
            client["fizinis"],
            client["vat"],
            client["name"],
            client["address"],
            client["country"],
            client["iban"],
        ]

        if mode == "suminis":
            preke_kodas = (
                _s(getattr(doc, "prekes_kodas", ""))
                or _s(getattr(doc, "prekes_barkodas", ""))
                or "PREKE001"
            )

            pvm_kodas = _get_pvm_kodas_for_doc(doc, default="")

            row = doc_common + [
                preke_kodas,
                1,
                _format_decimal(getattr(doc, "amount_wo_vat", 0)),
                _s(getattr(doc, "currency", "EUR") or "EUR"),
                _format_decimal(getattr(doc, "vat_amount", 0)),
                0,
                pvm_kodas,
                _s(getattr(doc, "prekes_barkodas", "")),
                _s(getattr(doc, "prekes_pavadinimas", "")),
            ]
            rows.append(row)
            logger.info("[BUTENT:SUMINIS] doc=%s row added", getattr(doc, "pk", None))

        else:
            line_items = getattr(doc, "line_items", None)
            if not line_items or not hasattr(line_items, "all"):
                logger.warning(
                    "[BUTENT:KIEKINIS] doc=%s has no line_items, skipping",
                    getattr(doc, "pk", None)
                )
                continue

            items_list = list(line_items.all())
            _distribute_discount_to_butent_lines(doc, items_list)

            items_added = 0
            for item in items_list:
                preke_kodas = (
                    _resolved_field(item, "prekes_kodas")
                    or _resolved_field(item, "prekes_barkodas")
                    or _s(getattr(doc, "prekes_kodas", ""))
                    or "PREKE001"
                )

                price_to_use = getattr(item, "_butent_price_after_discount", None)
                if price_to_use is None:
                    price_to_use = getattr(item, "price", 0)

                vat_to_use = getattr(item, "_butent_vat_after_discount", None)
                if vat_to_use is None:
                    vat_to_use = getattr(item, "vat", 0)

                pvm_kodas = _get_pvm_kodas_for_item(doc, item, line_map, default="")

                row = doc_common + [
                    preke_kodas,
                    getattr(item, "quantity", 1),
                    _format_decimal(price_to_use),
                    _s(getattr(doc, "currency", "EUR") or "EUR"),
                    _format_decimal(vat_to_use),
                    0,
                    pvm_kodas,
                    _resolved_field(item, "prekes_barkodas"),
                    _resolved_field(item, "prekes_pavadinimas"),
                ]
                rows.append(row)
                items_added += 1

            discount_val = getattr(doc, "invoice_discount_wo_vat", None)
            if discount_val is not None and discount_val > 0:
                sum_price_qty = sum(
                    Decimal(str(getattr(item, "_butent_price_after_discount", None)
                                or getattr(item, "price", 0))) *
                    Decimal(str(getattr(item, "quantity", 1)))
                    for item in items_list
                )
                sum_vat = sum(
                    Decimal(str(getattr(item, "_butent_vat_after_discount", None)
                                or getattr(item, "vat", 0)))
                    for item in items_list
                )
                logger.info(
                    "[BUTENT:KIEKINIS] doc=%s after discount: Σ(price×qty)=%.2f Σ(vat)=%.2f",
                    getattr(doc, "pk", None),
                    float(sum_price_qty),
                    float(sum_vat)
                )

            logger.info(
                "[BUTENT:KIEKINIS] doc=%s items=%d",
                getattr(doc, "pk", None),
                items_added,
            )

    start_row = 2
    for idx, row_data in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_idx)

            if col_idx in [4, 12, 19, 20, 22, 23]:
                if isinstance(value, (int, float)):
                    cell.data_type = "n"
                    cell.value = value

                    if idx == start_row and col_idx in [19, 20, 22]:
                        logger.debug(
                            "[BUTENT:CELL] row=%d col=%d value=%r type=%s data_type=%s",
                            idx, col_idx, value, type(value).__name__, cell.data_type
                        )
                elif isinstance(value, str):
                    try:
                        num_value = float(value)
                        cell.data_type = "n"
                        cell.value = num_value
                        logger.warning(
                            "[BUTENT:CELL] Converted string to float: row=%d col=%d '%s'->%f",
                            idx, col_idx, value, num_value
                        )
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = value

                if col_idx in [19, 20, 22]:
                    cell.number_format = "0.00"
            else:
                cell.value = value

    logger.info("[BUTENT:FILE] Written %d rows to Excel", len(rows))

    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info("[BUTENT:FILE] File generation completed")
    return output.read()


# =========================
# Вспомогательная функция для создания шаблона
# =========================

def create_butent_template():
    """
    Создаёт пустой шаблон Excel для импорта в Būtent.
    Используется для первоначальной настройки.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"

    headers = [
        "Data",
        "Kita data",
        "Terminas",
        "iSAF požymis",
        "Serija",
        "Kiti dok. Nr.",
        "Kiti dok. Nr.2",
        "Operacija",
        "Sandėlis",
        "Pastabos",
        "Kliento kodas",
        "Požymis jei fizinis",
        "PVM mokėtojo kodas",
        "Pavadinimas",
        "Adresas",
        "Šalis",
        "Atsiskaitomoji sąskaita",
        "Prekės kodas",
        "Kiekis",
        "Kaina",
        "Valiuta",
        "PVM suma",
        "Atv. PVM taikymas",
        "PVM kodas",
        "Prekės barkodas",
        "Prekės pavadinimas",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 18
    ws.column_dimensions["N"].width = 25
    ws.column_dimensions["O"].width = 30
    ws.column_dimensions["P"].width = 8
    ws.column_dimensions["Q"].width = 25
    ws.column_dimensions["R"].width = 15
    ws.column_dimensions["S"].width = 10
    ws.column_dimensions["T"].width = 12
    ws.column_dimensions["U"].width = 10
    ws.column_dimensions["V"].width = 12
    ws.column_dimensions["W"].width = 12
    ws.column_dimensions["X"].width = 12
    ws.column_dimensions["Y"].width = 15
    ws.column_dimensions["Z"].width = 25

    template_path = TEMPLATES_DIR / BUTENT_TEMPLATE_FILE
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(template_path)

    logger.info("[BUTENT:TEMPLATE] Created template: %s", template_path)
    return template_path






# import os
# import logging
# from pathlib import Path
# from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
# from datetime import date, datetime
# from typing import List, Dict, Any

# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment

# from ..utils.extra_fields import get_extra_for_export

# logger = logging.getLogger(__name__)

# # =========================
# # Конфиг путей
# # =========================
# env_value = os.getenv("BUTENT_TEMPLATES_DIR")

# if not env_value:
#     raise ValueError("BUTENT_TEMPLATES_DIR not set in .env")

# TEMPLATES_DIR = Path(env_value)

# # Единый шаблон для обоих режимов (suminis и kiekinis)
# BUTENT_TEMPLATE_FILE = "Butent_Import_Template.xlsx"

# # =========================
# # Helpers
# # =========================

# def _safe_D(x):
#     """Безопасное преобразование в Decimal."""
#     try:
#         return Decimal(str(x))
#     except Exception:
#         return Decimal("0")


# def _s(v):
#     """Безопасная строка с strip()."""
#     return str(v).strip() if v is not None else ""


# def _is_zero(v) -> bool:
#     """Нулевая ставка НДС? None/'' считаем как 0."""
#     try:
#         return Decimal(str(v)) == 0
#     except Exception:
#         return True


# EU_ISO2 = {
#     "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR", "HU", "IE",
#     "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK", "SI", "ES", "SE"
# }


# def _is_eu_country(iso: object) -> bool:
#     """True только для явных ISO2 из списка ЕС. Пустое значение -> False."""
#     if not iso:
#         return False
#     return str(iso).strip().upper() in EU_ISO2


# def _pick_isaf_for_purchase(doc):
#     """
#     Возвращает:
#       - '12' -> Neformuoti (НЕ включать в i.SAF)
#       - None -> не ставить тег вовсе (включать по умолчанию)

#     Правило:
#       если (seller_country_iso пусто ИЛИ не-ЕС) И ВСЕ ставки vat_percent по строкам == 0
#       -> '12', иначе None.
#     """
#     country = getattr(doc, "seller_country_iso", "") or ""
#     is_eu = _is_eu_country(country)
#     non_eu_or_empty = not is_eu

#     line_items = getattr(doc, "line_items", None)
#     if line_items and hasattr(line_items, "all") and line_items.exists():
#         vat_zero_all = all(_is_zero(getattr(it, "vat_percent", None)) for it in line_items.all())
#     else:
#         vat_zero_all = _is_zero(getattr(doc, "vat_percent", None))

#     if non_eu_or_empty and vat_zero_all:
#         return "12"

#     return None


# def _get_butent_isaf_flag(doc) -> int:
#     """
#     Для Būtent возвращает 1 или 0:
#       - 0 -> НЕ включать в i.SAF
#       - 1 -> включать в i.SAF (по умолчанию)
#     """
#     _isaf = getattr(doc, "report_to_isaf", None)
#     if _isaf is False:
#         return 0
#     rivile_code = _pick_isaf_for_purchase(doc)
#     return 0 if rivile_code == "12" else 1


# def get_party_code(
#     doc,
#     *,
#     role: str,
#     id_field: str,
#     vat_field: str,
#     id_programoje_field: str,
# ) -> str:
#     """
#     Код стороны (seller/buyer) по приоритету:
#       1) *_id
#       2) *_vat_code
#       3) *_id_programoje
#     Если все пусто - вернётся "".
#     """
#     sid = _s(getattr(doc, id_field, None))
#     if sid:
#         logger.info("[BUTENT:PARTY] %s: %s -> %s", role, id_field, sid)
#         return sid

#     svat = _s(getattr(doc, vat_field, None))
#     if svat:
#         logger.info("[BUTENT:PARTY] %s: %s -> %s", role, vat_field, svat)
#         return svat

#     sidp = _s(getattr(doc, id_programoje_field, None))
#     if sidp:
#         logger.info("[BUTENT:PARTY] %s: %s -> %s", role, id_programoje_field, sidp)
#         return sidp

#     logger.info("[BUTENT:PARTY] %s: empty id/vat/id_programoje -> ''", role)
#     return ""


# # =========================
# # Per-company extra fields helper
# # =========================

# def _parse_cp_key(cp_key):
#     if not cp_key:
#         return ""

#     cp = str(cp_key).strip()
#     if cp.lower().startswith("id:"):
#         return cp.split(":", 1)[1].strip()
#     return cp


# def _get_own_company_code_from_doc(doc):
#     """
#     Определяет код своей фирмы из документа.

#     - pirkimas -> своя фирма buyer
#     - pardavimas -> своя фирма seller
#     """
#     doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()

#     if doc_type == "pirkimas":
#         candidates = [
#             getattr(doc, "buyer_id", ""),
#             getattr(doc, "buyer_vat_code", ""),
#             getattr(doc, "buyer_id_programoje", ""),
#         ]
#     else:
#         candidates = [
#             getattr(doc, "seller_id", ""),
#             getattr(doc, "seller_vat_code", ""),
#             getattr(doc, "seller_id_programoje", ""),
#         ]

#     for value in candidates:
#         code = _s(value)
#         if code:
#             return code
#     return ""


# def _get_butent_extra_for_doc(user, doc, own_company_code=None) -> Dict[str, Any]:
#     """
#     Получает extra fields для конкретного документа.

#     Приоритет:
#     1. Профиль конкретной фирмы по own_company_code
#     2. Профиль фирмы, определённой из документа
#     3. Глобальный профиль (__all__)
#     4. Пустой dict
#     """
#     if not user:
#         return {}

#     requested_code = _parse_cp_key(own_company_code)
#     doc_company_code = _get_own_company_code_from_doc(doc)

#     extra = {}
#     resolved_by = ""

#     if requested_code:
#         extra = get_extra_for_export(user, "butent", requested_code)
#         if extra:
#             resolved_by = requested_code

#     if not extra and doc_company_code and doc_company_code != requested_code:
#         extra = get_extra_for_export(user, "butent", doc_company_code)
#         if extra:
#             resolved_by = doc_company_code

#     if not extra:
#         extra = get_extra_for_export(user, "butent", None)
#         if extra:
#             resolved_by = "__all__/legacy"

#     logger.info(
#         "[BUTENT:EXTRA] doc=%s own_company_code=%r requested_code=%r doc_company_code=%r resolved_by=%r fields=%s",
#         getattr(doc, "pk", None),
#         own_company_code,
#         requested_code,
#         doc_company_code,
#         resolved_by,
#         {k: v for k, v in extra.items() if v} if extra else {},
#     )

#     return extra or {}


# def _get_client_data_for_butent(doc) -> Dict[str, Any]:
#     """
#     Возвращает словарь с данными клиента для колонок K-Q.
#     Логика:
#       - pirkimas -> seller (продавец)
#       - pardavimas -> buyer (покупатель)
#     """
#     doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()

#     if doc_type == "pirkimas":
#         code = get_party_code(
#             doc,
#             role="seller",
#             id_field="seller_id",
#             vat_field="seller_vat_code",
#             id_programoje_field="seller_id_programoje",
#         )
#         return {
#             "code": code,
#             "fizinis": 1 if getattr(doc, "seller_is_person", False) else 0,
#             "vat": _s(getattr(doc, "seller_vat_code", "")),
#             "name": _s(getattr(doc, "seller_name", "")),
#             "address": _s(getattr(doc, "seller_address", "")),
#             "country": _s(getattr(doc, "seller_country_iso", "")),
#             "iban": _s(getattr(doc, "seller_iban", "")),
#         }
#     else:
#         code = get_party_code(
#             doc,
#             role="buyer",
#             id_field="buyer_id",
#             vat_field="buyer_vat_code",
#             id_programoje_field="buyer_id_programoje",
#         )
#         return {
#             "code": code,
#             "fizinis": 1 if getattr(doc, "buyer_is_person", False) else 0,
#             "vat": _s(getattr(doc, "buyer_vat_code", "")),
#             "name": _s(getattr(doc, "buyer_name", "")),
#             "address": _s(getattr(doc, "buyer_address", "")),
#             "country": _s(getattr(doc, "buyer_country_iso", "")),
#             "iban": _s(getattr(doc, "buyer_iban", "")),
#         }


# def _format_date_for_butent(dt) -> str:
#     """
#     Форматирует дату в строку 'YYYY.MM.DD' для Būtent.
#     """
#     if not dt:
#         return ""
#     if isinstance(dt, str):
#         try:
#             dt = datetime.strptime(dt, "%Y-%m-%d").date()
#         except ValueError:
#             return ""
#     if isinstance(dt, (date, datetime)):
#         return dt.strftime("%Y.%m.%d")
#     return ""


# def _get_operacija(doc, extra_fields=None) -> str:
#     """
#     Определяет операцию для колонки H:
#       - pirkimas -> "Pajamavimas" (или из extra_fields['pirkimas_operacija'])
#       - pardavimas -> "Pardavimas" (или из extra_fields['pardavimas_operacija'])
#     """
#     doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()
#     extra_fields = extra_fields or {}

#     if doc_type == "pirkimas":
#         custom_op = _s(extra_fields.get("pirkimas_operacija", ""))
#         if custom_op:
#             logger.info(
#                 "[BUTENT:OPERACIJA] doc=%s using custom pirkimas_operacija=%r",
#                 getattr(doc, "pk", None), custom_op
#             )
#             return custom_op
#         return "Pajamavimas"

#     if doc_type == "pardavimas":
#         custom_op = _s(extra_fields.get("pardavimas_operacija", ""))
#         if custom_op:
#             logger.info(
#                 "[BUTENT:OPERACIJA] doc=%s using custom pardavimas_operacija=%r",
#                 getattr(doc, "pk", None), custom_op
#             )
#             return custom_op
#         return "Pardavimas"

#     logger.warning("[BUTENT] Unknown pirkimas_pardavimas=%r, defaulting to Pajamavimas", doc_type)
#     return "Pajamavimas"


# def _get_sandelis(doc, extra_fields=None) -> str:
#     """
#     Определяет склад для колонки I:
#       - doc.sandelio_kodas если есть
#       - иначе из extra_fields (pirkimas_sandelis/pardavimas_sandelis)
#       - иначе "S1"
#     """
#     sandelis = _s(getattr(doc, "sandelio_kodas", ""))
#     if sandelis:
#         return sandelis

#     extra_fields = extra_fields or {}
#     doc_type = _s(getattr(doc, "pirkimas_pardavimas", "")).lower()

#     if doc_type == "pirkimas":
#         custom_sandelis = _s(extra_fields.get("pirkimas_sandelis", ""))
#         if custom_sandelis:
#             logger.info(
#                 "[BUTENT:SANDELIS] doc=%s using custom pirkimas_sandelis=%r",
#                 getattr(doc, "pk", None), custom_sandelis
#             )
#             return custom_sandelis
#     elif doc_type == "pardavimas":
#         custom_sandelis = _s(extra_fields.get("pardavimas_sandelis", ""))
#         if custom_sandelis:
#             logger.info(
#                 "[BUTENT:SANDELIS] doc=%s using custom pardavimas_sandelis=%r",
#                 getattr(doc, "pk", None), custom_sandelis
#             )
#             return custom_sandelis

#     return "S1"


# def _format_decimal(value, decimals=2) -> float:
#     """
#     Преобразует значение в float с округлением.
#     Возвращает ЧИСЛО (float), не строку.
#     """
#     try:
#         d = Decimal(str(value))
#         rounded = d.quantize(Decimal(10) ** -decimals, rounding=ROUND_HALF_UP)
#         return float(rounded)
#     except Exception:
#         return 0.0


# def _distribute_discount_to_butent_lines(doc, items_list: list) -> None:
#     """
#     Распределяет скидку документа (invoice_discount_wo_vat) на строки товаров.

#     ВАЖНО: Būtent не имеет поля для скидки документа, поэтому мы:
#       1. ВЫЧИТАЕМ долю скидки из subtotal каждой строки
#       2. ПЕРЕСЧИТЫВАЕМ price = new_subtotal / quantity
#       3. ПЕРЕСЧИТЫВАЕМ vat = new_subtotal × vat_percent / 100

#     Args:
#         doc: документ с полем invoice_discount_wo_vat
#         items_list: список объектов LineItem (модифицируется in-place)

#     Модифицирует:
#         Устанавливает атрибуты _butent_price_after_discount и _butent_vat_after_discount
#     """
#     if not items_list:
#         return

#     discount_raw = getattr(doc, "invoice_discount_wo_vat", None)
#     if discount_raw in (None, "", 0, "0"):
#         return

#     try:
#         discount_wo = Decimal(str(discount_raw))
#     except (ValueError, InvalidOperation):
#         logger.warning(
#             "[BUTENT:DISCOUNT] doc=%s invalid discount value: %r",
#             getattr(doc, "pk", None), discount_raw
#         )
#         return

#     if discount_wo <= 0:
#         return

#     logger.info(
#         "[BUTENT:DISCOUNT] doc=%s distributing discount=%.2f across %d lines",
#         getattr(doc, "pk", None), discount_wo, len(items_list)
#     )

#     sum_subtotal_before = Decimal("0")
#     for item in items_list:
#         price = Decimal(str(getattr(item, "price", 0) or 0))
#         qty = Decimal(str(getattr(item, "quantity", 1) or 1))
#         sum_subtotal_before += price * qty

#     if sum_subtotal_before <= 0:
#         logger.warning(
#             "[BUTENT:DISCOUNT] doc=%s sum_subtotal=0, cannot distribute",
#             getattr(doc, "pk", None)
#         )
#         return

#     discount_distributed = Decimal("0")

#     for i, item in enumerate(items_list):
#         qty = Decimal(str(getattr(item, "quantity", 1) or 1))
#         price_before = Decimal(str(getattr(item, "price", 0) or 0))
#         vat_percent = Decimal(str(getattr(item, "vat_percent", 0) or 0))

#         subtotal_before = price_before * qty

#         if i == len(items_list) - 1:
#             line_discount = discount_wo - discount_distributed
#         else:
#             share = subtotal_before / sum_subtotal_before
#             line_discount = (discount_wo * share).quantize(
#                 Decimal("0.01"), rounding=ROUND_HALF_UP
#             )
#             discount_distributed += line_discount

#         subtotal_after = subtotal_before - line_discount

#         if qty > 0:
#             price_after = (subtotal_after / qty).quantize(
#                 Decimal("0.01"), rounding=ROUND_HALF_UP
#             )
#         else:
#             price_after = Decimal("0")

#         if vat_percent > 0 and subtotal_after > 0:
#             vat_after = (subtotal_after * vat_percent / Decimal("100")).quantize(
#                 Decimal("0.01"), rounding=ROUND_HALF_UP
#             )
#         else:
#             vat_after = Decimal("0")

#         setattr(item, "_butent_price_after_discount", float(price_after))
#         setattr(item, "_butent_vat_after_discount", float(vat_after))

#         logger.debug(
#             "[BUTENT:DISCOUNT] line=%d qty=%.2f price: %.2f->%.2f vat: %.2f->%.2f (discount=%.2f)",
#             i,
#             float(qty),
#             float(price_before),
#             float(price_after),
#             float(Decimal(str(getattr(item, "vat", 0) or 0))),
#             float(vat_after),
#             float(line_discount)
#         )


# # =========================
# # PVM Kodas helpers
# # =========================

# def _get_pvm_kodas_for_item(doc, item, line_map=None, default="") -> str:
#     """
#     Получает PVM kodas для строки с учётом резолвера и separate_vat.

#     Приоритет:
#     1. line_map[item.id] - от резолвера (если есть)
#     2. item.pvm_kodas - из БД
#     3. default - fallback

#     ВАЖНО: "Keli skirtingi PVM" - это маркер, не реальный код -> возвращаем default
#     """
#     item_id = getattr(item, "id", None)

#     if line_map is not None and item_id is not None and item_id in line_map:
#         pvm = _s(line_map.get(item_id, ""))
#         if pvm and pvm != "Keli skirtingi PVM":
#             return pvm

#     pvm = _s(getattr(item, "pvm_kodas", ""))
#     if pvm and pvm != "Keli skirtingi PVM":
#         return pvm

#     return default


# def _get_pvm_kodas_for_doc(doc, default="") -> str:
#     """
#     Получает PVM kodas для документа (sumiskai режим).

#     ВАЖНО:
#     - При separate_vat=True -> пустой (смешанные ставки)
#     - "Keli skirtingi PVM" -> пустой (это маркер, не код)
#     """
#     separate_vat = bool(getattr(doc, "separate_vat", False))
#     scan_type = _s(getattr(doc, "scan_type", "")).lower()

#     if separate_vat and scan_type in ("sumiskai", "summary", "suminis"):
#         return default

#     pvm = _s(getattr(doc, "pvm_kodas", ""))

#     if pvm == "Keli skirtingi PVM":
#         return default

#     return pvm or default


# # =========================
# # Основная функция экспорта
# # =========================

# def export_to_butent(
#     documents: List,
#     mode: str = "auto",
#     user=None,
#     own_company_code=None,
# ) -> Dict[str, bytes]:
#     """
#     Экспортирует документы в формат Būtent Excel.

#     Args:
#         documents: список документов для экспорта
#         mode: 'auto' | 'suminis' | 'kiekinis'
#               'auto' - автоматически разделяет документы на два файла
#               'suminis' - принудительно все в один файл (suminis режим)
#               'kiekinis' - принудительно все в один файл (kiekinis режим)
#         user: пользователь
#         own_company_code: код своей фирмы для поиска профиля extra_fields

#     Returns:
#         Dict[str, bytes]: словарь вида {"suminis": bytes, "kiekinis": bytes}
#                          где ключи присутствуют только если есть соответствующие документы
#     """
#     logger.info(
#         "[BUTENT:EXPORT] Starting export, docs=%d mode=%s own_company_code=%r",
#         len(documents), mode, own_company_code
#     )

#     if not documents:
#         logger.warning("[BUTENT:EXPORT] No documents to export")
#         raise ValueError("No documents provided for export")

#     docs_suminis = []
#     docs_kiekinis = []

#     if mode == "auto":
#         for doc in documents:
#             line_items = getattr(doc, "line_items", None)
#             has_items = False
#             if line_items and hasattr(line_items, "all"):
#                 has_items = line_items.exists()

#             if has_items:
#                 docs_kiekinis.append(doc)
#             else:
#                 docs_suminis.append(doc)

#         logger.info(
#             "[BUTENT:EXPORT] Auto mode: suminis=%d kiekinis=%d",
#             len(docs_suminis),
#             len(docs_kiekinis)
#         )
#     elif mode == "suminis":
#         docs_suminis = documents
#         logger.info("[BUTENT:EXPORT] Force suminis mode: %d docs", len(docs_suminis))
#     elif mode == "kiekinis":
#         docs_kiekinis = documents
#         logger.info("[BUTENT:EXPORT] Force kiekinis mode: %d docs", len(docs_kiekinis))
#     else:
#         logger.error("[BUTENT:EXPORT] Unknown mode: %s", mode)
#         raise ValueError(f"Unknown mode: {mode}")

#     result = {}

#     if docs_suminis:
#         logger.info("[BUTENT:EXPORT] Generating suminis file...")
#         result["suminis"] = _generate_butent_file(docs_suminis, "suminis", user, own_company_code)

#     if docs_kiekinis:
#         logger.info("[BUTENT:EXPORT] Generating kiekinis file...")
#         result["kiekinis"] = _generate_butent_file(docs_kiekinis, "kiekinis", user, own_company_code)

#     if not result:
#         logger.warning("[BUTENT:EXPORT] No files generated")
#         raise ValueError("No documents to export")

#     logger.info("[BUTENT:EXPORT] Export completed, files=%s", list(result.keys()))
#     return result


# def _generate_butent_file(documents: List, mode: str, user=None, own_company_code=None) -> bytes:
#     """
#     Генерирует один Excel-файл для Būtent.

#     Args:
#         documents: список документов для экспорта
#         mode: 'suminis' | 'kiekinis'
#         user: пользователь
#         own_company_code: код своей фирмы для поиска профиля extra_fields

#     Returns:
#         bytes: содержимое Excel-файла
#     """
#     logger.info(
#         "[BUTENT:FILE] Generating %s file for %d docs own_company_code=%r",
#         mode, len(documents), own_company_code
#     )

#     template_path = TEMPLATES_DIR / BUTENT_TEMPLATE_FILE
#     if not template_path.exists():
#         logger.error("[BUTENT:FILE] Template not found: %s", template_path)
#         raise FileNotFoundError(f"Būtent template not found: {template_path}")

#     wb = load_workbook(template_path)
#     ws = wb.active

#     rows = []

#     for doc in documents:
#         line_map = getattr(doc, "_pvm_line_map", None)
#         extra_fields = _get_butent_extra_for_doc(user, doc, own_company_code)

#         operacija = _get_operacija(doc, extra_fields)
#         sandelis = _get_sandelis(doc, extra_fields)
#         isaf = _get_butent_isaf_flag(doc)
#         client = _get_client_data_for_butent(doc)

#         doc_common = [
#             _format_date_for_butent(getattr(doc, "invoice_date", None)),
#             _format_date_for_butent(getattr(doc, "operation_date", None)),
#             _format_date_for_butent(getattr(doc, "due_date", None)),
#             isaf,
#             _s(getattr(doc, "document_series", "")),
#             _s(getattr(doc, "document_number", "")),
#             _s(getattr(doc, "order_number", "")),
#             operacija,
#             sandelis,
#             _s(getattr(doc, "preview_url", "")),
#             client["code"],
#             client["fizinis"],
#             client["vat"],
#             client["name"],
#             client["address"],
#             client["country"],
#             client["iban"],
#         ]

#         if mode == "suminis":
#             preke_kodas = (
#                 _s(getattr(doc, "prekes_kodas", ""))
#                 or _s(getattr(doc, "prekes_barkodas", ""))
#                 or "PREKE001"
#             )

#             pvm_kodas = _get_pvm_kodas_for_doc(doc, default="")

#             row = doc_common + [
#                 preke_kodas,
#                 1,
#                 _format_decimal(getattr(doc, "amount_wo_vat", 0)),
#                 _s(getattr(doc, "currency", "EUR") or "EUR"),
#                 _format_decimal(getattr(doc, "vat_amount", 0)),
#                 0,
#                 pvm_kodas,
#                 _s(getattr(doc, "prekes_barkodas", "")),
#                 _s(getattr(doc, "prekes_pavadinimas", "")),
#             ]
#             rows.append(row)
#             logger.info("[BUTENT:SUMINIS] doc=%s row added", getattr(doc, "pk", None))

#         else:
#             line_items = getattr(doc, "line_items", None)
#             if not line_items or not hasattr(line_items, "all"):
#                 logger.warning(
#                     "[BUTENT:KIEKINIS] doc=%s has no line_items, skipping",
#                     getattr(doc, "pk", None)
#                 )
#                 continue

#             items_list = list(line_items.all())
#             _distribute_discount_to_butent_lines(doc, items_list)

#             items_added = 0
#             for item in items_list:
#                 preke_kodas = (
#                     _s(getattr(item, "prekes_kodas", ""))
#                     or _s(getattr(item, "prekes_barkodas", ""))
#                     or _s(getattr(doc, "prekes_kodas", ""))
#                     or "PREKE001"
#                 )

#                 price_to_use = getattr(item, "_butent_price_after_discount", None)
#                 if price_to_use is None:
#                     price_to_use = getattr(item, "price", 0)

#                 vat_to_use = getattr(item, "_butent_vat_after_discount", None)
#                 if vat_to_use is None:
#                     vat_to_use = getattr(item, "vat", 0)

#                 pvm_kodas = _get_pvm_kodas_for_item(doc, item, line_map, default="")

#                 row = doc_common + [
#                     preke_kodas,
#                     getattr(item, "quantity", 1),
#                     _format_decimal(price_to_use),
#                     _s(getattr(doc, "currency", "EUR") or "EUR"),
#                     _format_decimal(vat_to_use),
#                     0,
#                     pvm_kodas,
#                     _s(getattr(item, "prekes_barkodas", "")),
#                     _s(getattr(item, "prekes_pavadinimas", "")),
#                 ]
#                 rows.append(row)
#                 items_added += 1

#             discount_val = getattr(doc, "invoice_discount_wo_vat", None)
#             if discount_val is not None and discount_val > 0:
#                 sum_price_qty = sum(
#                     Decimal(str(getattr(item, "_butent_price_after_discount", None)
#                                 or getattr(item, "price", 0))) *
#                     Decimal(str(getattr(item, "quantity", 1)))
#                     for item in items_list
#                 )
#                 sum_vat = sum(
#                     Decimal(str(getattr(item, "_butent_vat_after_discount", None)
#                                 or getattr(item, "vat", 0)))
#                     for item in items_list
#                 )
#                 logger.info(
#                     "[BUTENT:KIEKINIS] doc=%s after discount: Σ(price×qty)=%.2f Σ(vat)=%.2f",
#                     getattr(doc, "pk", None),
#                     float(sum_price_qty),
#                     float(sum_vat)
#                 )

#             logger.info(
#                 "[BUTENT:KIEKINIS] doc=%s items=%d",
#                 getattr(doc, "pk", None),
#                 items_added,
#             )

#     start_row = 2
#     for idx, row_data in enumerate(rows, start=start_row):
#         for col_idx, value in enumerate(row_data, start=1):
#             cell = ws.cell(row=idx, column=col_idx)

#             if col_idx in [4, 12, 19, 20, 22, 23]:
#                 if isinstance(value, (int, float)):
#                     cell.data_type = "n"
#                     cell.value = value

#                     if idx == start_row and col_idx in [19, 20, 22]:
#                         logger.debug(
#                             "[BUTENT:CELL] row=%d col=%d value=%r type=%s data_type=%s",
#                             idx, col_idx, value, type(value).__name__, cell.data_type
#                         )
#                 elif isinstance(value, str):
#                     try:
#                         num_value = float(value)
#                         cell.data_type = "n"
#                         cell.value = num_value
#                         logger.warning(
#                             "[BUTENT:CELL] Converted string to float: row=%d col=%d '%s'->%f",
#                             idx, col_idx, value, num_value
#                         )
#                     except (ValueError, TypeError):
#                         cell.value = value
#                 else:
#                     cell.value = value

#                 if col_idx in [19, 20, 22]:
#                     cell.number_format = "0.00"
#             else:
#                 cell.value = value

#     logger.info("[BUTENT:FILE] Written %d rows to Excel", len(rows))

#     from io import BytesIO

#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     logger.info("[BUTENT:FILE] File generation completed")
#     return output.read()


# # =========================
# # Вспомогательная функция для создания шаблона
# # =========================

# def create_butent_template():
#     """
#     Создаёт пустой шаблон Excel для импорта в Būtent.
#     Используется для первоначальной настройки.
#     """
#     from openpyxl import Workbook

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Import"

#     headers = [
#         "Data",
#         "Kita data",
#         "Terminas",
#         "iSAF požymis",
#         "Serija",
#         "Kiti dok. Nr.",
#         "Kiti dok. Nr.2",
#         "Operacija",
#         "Sandėlis",
#         "Pastabos",
#         "Kliento kodas",
#         "Požymis jei fizinis",
#         "PVM mokėtojo kodas",
#         "Pavadinimas",
#         "Adresas",
#         "Šalis",
#         "Atsiskaitomoji sąskaita",
#         "Prekės kodas",
#         "Kiekis",
#         "Kaina",
#         "Valiuta",
#         "PVM suma",
#         "Atv. PVM taikymas",
#         "PVM kodas",
#         "Prekės barkodas",
#         "Prekės pavadinimas",
#     ]

#     for col_idx, header in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col_idx)
#         cell.value = header
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal="center")

#     ws.column_dimensions["A"].width = 12
#     ws.column_dimensions["B"].width = 12
#     ws.column_dimensions["C"].width = 12
#     ws.column_dimensions["D"].width = 12
#     ws.column_dimensions["E"].width = 10
#     ws.column_dimensions["F"].width = 15
#     ws.column_dimensions["G"].width = 15
#     ws.column_dimensions["H"].width = 15
#     ws.column_dimensions["I"].width = 12
#     ws.column_dimensions["J"].width = 30
#     ws.column_dimensions["K"].width = 15
#     ws.column_dimensions["L"].width = 12
#     ws.column_dimensions["M"].width = 18
#     ws.column_dimensions["N"].width = 25
#     ws.column_dimensions["O"].width = 30
#     ws.column_dimensions["P"].width = 8
#     ws.column_dimensions["Q"].width = 25
#     ws.column_dimensions["R"].width = 15
#     ws.column_dimensions["S"].width = 10
#     ws.column_dimensions["T"].width = 12
#     ws.column_dimensions["U"].width = 10
#     ws.column_dimensions["V"].width = 12
#     ws.column_dimensions["W"].width = 12
#     ws.column_dimensions["X"].width = 12
#     ws.column_dimensions["Y"].width = 15
#     ws.column_dimensions["Z"].width = 25

#     template_path = TEMPLATES_DIR / BUTENT_TEMPLATE_FILE
#     TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
#     wb.save(template_path)

#     logger.info("[BUTENT:TEMPLATE] Created template: %s", template_path)
#     return template_path



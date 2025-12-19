import os
import logging
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import date, datetime

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# =========================
# Конфиг путей / шаблонов
# =========================

def _get_templates_dir() -> Path:
    """
    Шаблоны лежат там же, где и для других программ.
    Приоритет:
      1) SITE_PRO_TEMPLATES_DIR
    """
    env_value = os.getenv("SITE_PRO_TEMPLATES_DIR")
    if not env_value:
        raise ValueError("SITE_PRO_TEMPLATES_DIR not set in .env")
    return Path(env_value)

TEMPLATES_DIR = _get_templates_dir()

# Файлы шаблонов (твои текущие)
SITE_PRO_TEMPLATE_CLIENTS   = "site_pro_import_klientai.xlsx"
SITE_PRO_TEMPLATE_ITEMS     = "site_pro_import_prekes_paslaugos.xlsx"
SITE_PRO_TEMPLATE_PURCHASES = "site_pro_import_pirkimai.xlsx"
SITE_PRO_TEMPLATE_SALES     = "site_pro_import_pardavimai.xlsx"


# =========================
# Helpers
# =========================

EU_ISO2 = {
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR", "HU", "IE",
    "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK", "SI", "ES", "SE"
}

def _s(v) -> str:
    return str(v).strip() if v is not None else ""

# =========================
# Units normalization
# =========================

_UNITS_FORCE_DOT = {"vnt", "kg", "l", "m", "val"}

def _normalize_measure_unit(unit: str) -> str:
    u = _s(unit)
    if not u:
        return ""

    # убираем пробелы по краям, но сохраняем исходный регистр
    stripped = u.strip()

    # для сравнения: lowercase и без завершающей точки
    base = stripped.lower().rstrip(".").strip()

    # только эти 5 -> гарантируем точку на конце
    if base in _UNITS_FORCE_DOT:
        return base + "."

    # остальные оставляем как есть (как в БД)
    return stripped

def _safe_D(x) -> Decimal:
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")

def _format_date_yyyy_mm_dd(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(dt, fmt).date().strftime("%Y-%m-%d")
            except ValueError:
                pass
        return ""
    if isinstance(dt, datetime):
        return dt.date().strftime("%Y-%m-%d")
    if isinstance(dt, date):
        return dt.strftime("%Y-%m-%d")
    return ""

def _quantize_2(x: Decimal) -> Decimal:
    return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def _is_eu_country(iso) -> bool:
    if not iso:
        return False
    return str(iso).strip().upper() in EU_ISO2

def _location_from_country_iso(iso) -> str:
    """
    Правило:
      LT -> "lt"
      EU -> "eu"
      иначе -> "rest" (включая пустое)
    """
    iso = _s(iso).upper()
    if iso == "LT":
        return "lt"
    if _is_eu_country(iso):
        return "eu"
    return "rest"

def _attribute_name_from_preke_paslauga(code) -> str:
    """
    1 - Prekės
    2 - Paslaugos
    3 - Prekės
    4 - Paslaugos
    """
    try:
        v = int(str(code).strip())
    except Exception:
        v = 1
    return "Paslaugos" if v in (2, 4) else "Prekės"

def _get_party_code(doc, *, role: str, id_field: str, vat_field: str, id_programoje_field: str) -> str:
    """
    Код стороны (seller/buyer) по приоритету:
      1) *_id
      2) *_vat_code
      3) *_id_programoje
    """
    sid = _s(getattr(doc, id_field, None))
    if sid:
        return sid
    svat = _s(getattr(doc, vat_field, None))
    if svat:
        return svat
    sidp = _s(getattr(doc, id_programoje_field, None))
    if sidp:
        return sidp
    return ""

def _doc_type(doc) -> str:
    return _s(getattr(doc, "pirkimas_pardavimas", "")).lower()

def _get_extra(user, key: str, default=""):
    """
    user.site_pro_extra_fields
    """
    if not user or not hasattr(user, "site_pro_extra_fields") or not user.site_pro_extra_fields:
        return default
    v = user.site_pro_extra_fields.get(key, default)
    return v if v is not None else default

def _get_warehouse_name(user, doc_type_str: str) -> str:
    if doc_type_str == "pirkimas":
        v = _s(_get_extra(user, "pirkimas_sandelis", ""))
    else:
        v = _s(_get_extra(user, "pardavimas_sandelis", ""))
    return v or "Pagrindinis"

def _get_employee_name(user) -> str:
    v = _s(_get_extra(user, "pardavimas_darbuotojas", ""))
    return v or "Vardenis Pavardenis"

def _get_purchase_employee_name(user) -> str:
    return _s(_get_extra(user, "pirkimas_darbuotojas", ""))

def _get_cost_center(user, doc_type_str: str) -> str:
    if doc_type_str == "pirkimas":
        return _s(_get_extra(user, "pirkimas_kastu_centras", ""))
    return _s(_get_extra(user, "pardavimas_kastu_centras", ""))

def _get_group_name(user, doc_type_str: str) -> str:
    if doc_type_str == "pirkimas":
        return _s(_get_extra(user, "pirkimas_prekes_grupe", ""))
    return _s(_get_extra(user, "pardavimas_prekes_grupe", ""))

def _get_measure_unit(item) -> str:
    raw = _s(getattr(item, "unit", ""))
    unit = _normalize_measure_unit(raw)
    return unit or "vnt."

def _get_currency(doc) -> str:
    return _s(getattr(doc, "currency", "")) or "EUR"

def _get_doc_number(doc) -> str:
    return _s(getattr(doc, "document_number", "")) or _s(getattr(doc, "number", "")) or _s(getattr(doc, "pk", ""))

def _get_doc_series_for_sales(doc) -> str:
    return _s(getattr(doc, "document_series", "")) or "SF"

def _get_doc_date(doc) -> str:
    return _format_date_yyyy_mm_dd(getattr(doc, "invoice_date", None)) or _format_date_yyyy_mm_dd(getattr(doc, "operation_date", None))

def _get_seller_fields(doc):
    return {
        "name": _s(getattr(doc, "seller_name", "")),
        "address": _s(getattr(doc, "seller_address", "")),
        "country": _s(getattr(doc, "seller_country_iso", "")),
        "vat": _s(getattr(doc, "seller_vat_code", "")),
        "code": _get_party_code(
            doc,
            role="seller",
            id_field="seller_id",
            vat_field="seller_vat_code",
            id_programoje_field="seller_id_programoje",
        ),
        "is_person": bool(getattr(doc, "seller_is_person", False)),
    }

def _get_buyer_fields(doc):
    return {
        "name": _s(getattr(doc, "buyer_name", "")),
        "address": _s(getattr(doc, "buyer_address", "")),
        "country": _s(getattr(doc, "buyer_country_iso", "")),
        "vat": _s(getattr(doc, "buyer_vat_code", "")),
        "code": _get_party_code(
            doc,
            role="buyer",
            id_field="buyer_id",
            vat_field="buyer_vat_code",
            id_programoje_field="buyer_id_programoje",
        ),
        "is_person": bool(getattr(doc, "buyer_is_person", False)),
    }

def _get_item_identity(doc, it=None):
    """
    Возвращает (name, code, barcode) с приоритетом item -> doc
    """
    if it is not None:
        name = _s(getattr(it, "prekes_pavadinimas", "")) or _s(getattr(doc, "prekes_pavadinimas", ""))
        code = _s(getattr(it, "prekes_kodas", "")) or _s(getattr(doc, "prekes_kodas", ""))
        barcode = _s(getattr(it, "prekes_barkodas", "")) or _s(getattr(doc, "prekes_barkodas", ""))
        return name, code, barcode

    name = _s(getattr(doc, "prekes_pavadinimas", ""))
    code = _s(getattr(doc, "prekes_kodas", ""))
    barcode = _s(getattr(doc, "prekes_barkodas", ""))
    return name, code, barcode

def _get_operation_type_name(doc_type_str: str, attr_name: str) -> str:
    """
    Правило:
      - Pirkimas + Prekės   -> "Pirkimas"
      - Pirkimas + Paslaugos -> "Pirkimas (paslaugos)"
      - Pardavimas + Prekės -> "Pardavimai"
      - Pardavimas + Paslaugos -> "Pardavimas (paslaugos)"
    """
    if doc_type_str == "pirkimas":
        return "Pirkimas (paslaugos)" if attr_name == "Paslaugos" else "Pirkimas"
    return "Pardavimas (paslaugos)" if attr_name == "Paslaugos" else "Pardavimai"

def _get_vat_classifier(doc, it=None) -> str:
    """
    Приоритет:
      1) item.pvm_kodas (если есть)
      2) doc._pvm_line_map[item.id] (если resolver положил)
      3) doc.pvm_kodas
      4) пусто
    """
    if it is not None:
        v = _s(getattr(it, "pvm_kodas", ""))
        if v:
            return v

        line_map = getattr(doc, "_pvm_line_map", None) or {}
        it_id = getattr(it, "id", None)
        if it_id is not None:
            v = _s(line_map.get(it_id, ""))
            if v:
                return v

    return _s(getattr(doc, "pvm_kodas", ""))


# =========================
# Работа с Excel-шаблоном (по строке ключей)
# =========================

def _load_template(template_filename: str):
    path = TEMPLATES_DIR / template_filename
    if not path.exists():
        raise FileNotFoundError(f"site_pro template not found: {path}")
    wb = load_workbook(path)
    ws = wb.active
    return wb, ws

def _build_field_map(ws, key_row: int = 2):
    """
    Во 2-й строке лежат ключи полей (включая '*').
    Вернём base_name -> column_index, где base_name = key без '*'.
    """
    base_to_col = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(key_row, c).value
        if not raw:
            continue
        key = str(raw).strip()
        base = key.rstrip("*").strip()
        if base:
            base_to_col[base] = c
    return base_to_col

def _clear_data_area(ws, start_row: int, base_to_col: dict):
    if ws.max_row < start_row:
        return
    cols = sorted(set(base_to_col.values()))
    for r in range(start_row, ws.max_row + 1):
        for c in cols:
            ws.cell(r, c).value = None

def _write_rows(ws, base_to_col: dict, rows: list[dict], start_row: int = 3):
    for i, row in enumerate(rows):
        r = start_row + i
        for base_field, value in row.items():
            c = base_to_col.get(base_field)
            if not c:
                continue
            cell = ws.cell(r, c)
            if isinstance(value, (int, float, Decimal)):
                cell.value = float(value) if isinstance(value, Decimal) else value
            else:
                cell.value = value

def _get_doc_series_optional(doc) -> str:
    # без fallback
    return _s(getattr(doc, "document_series", ""))

def _normalize_number_remove_series_prefix(series: str, number: str) -> str:
    """
    Если number уже начинается с series (с разделителями или без) — удаляем series с начала.
    Примеры:
      series="SF", number="SF123"    -> "123"
      series="SF", number="SF-123"   -> "123"
      series="SF", number="SF/123"   -> "123"
      series="SF", number="SF 123"   -> "123"
      series="SF", number="sf  123"  -> "123"
    """
    series = _s(series)
    number = _s(number)
    if not series or not number:
        return number

    s_up = series.upper()
    n = number.strip()
    n_up = n.upper()

    if not n_up.startswith(s_up):
        return n

    rest = n[len(series):].lstrip()
    # убрать типичные разделители
    while rest and rest[0] in ("-", "/", "\\", "_", " ", ".", ":", "#"):
        rest = rest[1:].lstrip()

    # если после удаления стало пусто — оставим исходное число (чтоб не сделать number="")
    return rest or n

# =========================
# Скидка: только для экспорта (без мутаций item и без БД)
# =========================

def _calc_discounted_price_map(doc, items_list: list) -> dict:
    """
    Возвращает dict: {id(item) -> new_price_excl_vat_float}
    Скидка invoice_discount_wo_vat распределяется пропорционально subtotal = qty * price.
    Последняя строка получает остаток скидки.
    """
    if not items_list:
        return {}

    discount_raw = getattr(doc, "invoice_discount_wo_vat", None)
    if discount_raw in (None, "", 0, "0"):
        return {}

    try:
        discount_wo = Decimal(str(discount_raw))
    except (ValueError, InvalidOperation):
        logger.warning("[SITE_PRO:DISCOUNT] invalid invoice_discount_wo_vat=%r doc=%s",
                       discount_raw, getattr(doc, "pk", None))
        return {}

    if discount_wo <= 0:
        return {}

    subtotals = []
    sum_subtotal = Decimal("0")

    for it in items_list:
        qty = _safe_D(getattr(it, "quantity", 1) or 1)
        price = _safe_D(getattr(it, "price", 0) or 0)
        subtotal = qty * price
        subtotals.append((qty, price, subtotal))
        sum_subtotal += subtotal

    if sum_subtotal <= 0:
        return {}

    distributed = Decimal("0")
    price_map = {}

    for idx, it in enumerate(items_list):
        qty, price_before, subtotal_before = subtotals[idx]

        if qty <= 0 or subtotal_before <= 0:
            price_map[id(it)] = float(_quantize_2(price_before))
            continue

        if idx == len(items_list) - 1:
            line_discount = discount_wo - distributed
        else:
            share = subtotal_before / sum_subtotal
            line_discount = _quantize_2(discount_wo * share)
            distributed += line_discount

        subtotal_after = subtotal_before - line_discount
        if subtotal_after < 0:
            subtotal_after = Decimal("0")

        price_after = _quantize_2(subtotal_after / qty)
        price_map[id(it)] = float(price_after)

    return price_map

# =========================
# Экспорт (4 файла, но пустые НЕ возвращаем)
# =========================

def export_to_site_pro(documents: list, user=None) -> dict:
    """
    Возвращает только НЕпустые файлы:
      {
        "clients": bytes,    (если есть клиенты)
        "items": bytes,      (если есть товары/услуги)
        "purchases": bytes,  (если есть pirkimai)
        "sales": bytes,      (если есть pardavimai)
      }
    """
    if not documents:
        raise ValueError("No documents provided for export")

    out = {}

    clients_bytes = _export_clients(documents)
    if clients_bytes:
        out["clients"] = clients_bytes

    items_bytes = _export_items(documents, user=user)
    if items_bytes:
        out["items"] = items_bytes

    purchases_bytes = _export_purchases(documents, user=user)
    if purchases_bytes:
        out["purchases"] = purchases_bytes

    sales_bytes = _export_sales(documents, user=user)
    if sales_bytes:
        out["sales"] = sales_bytes

    return out

# ---------- CLIENTS ----------

def _export_clients(documents: list) -> bytes:
    wb, ws = _load_template(SITE_PRO_TEMPLATE_CLIENTS)
    base_to_col = _build_field_map(ws, key_row=2)

    start_row = 3
    _clear_data_area(ws, start_row, base_to_col)

    seen = set()
    rows = []

    for doc in documents:
        t = _doc_type(doc)
        p = _get_seller_fields(doc) if t == "pirkimas" else _get_buyer_fields(doc)

        name = p["name"]
        if not name:
            continue

        is_juridical = 0 if p["is_person"] else 1
        location = _location_from_country_iso(p["country"])

        key = (name.lower(), is_juridical, location)
        if key in seen:
            continue
        seen.add(key)

        row = {"name": name, "isJuridical": is_juridical, "location": location}
        if p["vat"]:
            row["vatCode"] = p["vat"]
        if p["code"]:
            row["code"] = p["code"]
        rows.append(row)

    if not rows:
        return b""

    _write_rows(ws, base_to_col, rows, start_row=start_row)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# ---------- ITEMS ----------

def _export_items(documents: list, user=None) -> bytes:
    wb, ws = _load_template(SITE_PRO_TEMPLATE_ITEMS)
    base_to_col = _build_field_map(ws, key_row=2)

    start_row = 3
    _clear_data_area(ws, start_row, base_to_col)

    # key = (name.lower(), attributeName, measurementUnitName)
    seen = {}
    rows = []

    def add_item(name: str, attribute_name: str, unit: str, group_name: str | None, code: str = "", barcode: str = ""):
        nonlocal rows, seen
        name = _s(name)
        if not name:
            return
        attribute_name = attribute_name or "Prekės"
        unit = unit or _normalize_measure_unit("vnt")
        k = (name.lower(), attribute_name, unit)

        existing = seen.get(k)
        if existing is None:
            row = {"name": name, "attributeName": attribute_name, "measurementUnitName": unit}
            if group_name:
                row["groupName"] = group_name
            if code:
                row["code"] = code
            if barcode:
                row["barcode"] = barcode
            seen[k] = row
            rows.append(row)
            return

        if group_name and not existing.get("groupName"):
            existing["groupName"] = group_name
        if code and not existing.get("code"):
            existing["code"] = code
        if barcode and not existing.get("barcode"):
            existing["barcode"] = barcode

    for doc in documents:
        t = _doc_type(doc)
        group_name = _get_group_name(user, t) or None

        line_items = getattr(doc, "line_items", None)
        if line_items and hasattr(line_items, "all") and line_items.exists():
            for it in list(line_items.all()):
                attr = _attribute_name_from_preke_paslauga(
                    getattr(it, "preke_paslauga", None) or getattr(doc, "preke_paslauga", None)
                )
                unit = _get_measure_unit(it)
                name, code, barcode = _get_item_identity(doc, it)
                add_item(name, attr, unit, group_name, code=code, barcode=barcode)
        else:
            doc_attr = _attribute_name_from_preke_paslauga(getattr(doc, "preke_paslauga", None))
            name, code, barcode = _get_item_identity(doc, None)
            if not name:
                name = "Paslauga" if doc_attr == "Paslaugos" else "Preke"
            unit = _normalize_measure_unit("vnt")  # станет "vnt."
            add_item(name, doc_attr, unit, group_name, code=code, barcode=barcode)

    if not rows:
        return b""

    _write_rows(ws, base_to_col, rows, start_row=start_row)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# ---------- PURCHASES ----------

def _export_purchases(documents: list, user=None) -> bytes:
    wb, ws = _load_template(SITE_PRO_TEMPLATE_PURCHASES)
    base_to_col = _build_field_map(ws, key_row=2)

    start_row = 3
    _clear_data_area(ws, start_row, base_to_col)

    rows = []

    for doc in documents:
        if _doc_type(doc) != "pirkimas":
            continue

        supplier = _get_seller_fields(doc)
        purchase_date = _get_doc_date(doc)
        series = _get_doc_series_optional(doc)   # без fallback
        number_raw = _get_doc_number(doc)
        number = _normalize_number_remove_series_prefix(series, number_raw)
        currency = _get_currency(doc)

        warehouse = _get_warehouse_name(user, "pirkimas")
        cost_center = _get_cost_center(user, "pirkimas") or ""
        employee = _get_purchase_employee_name(user)

        line_items = getattr(doc, "line_items", None)
        if line_items and hasattr(line_items, "all") and line_items.exists():
            items_list = list(line_items.all())
            price_map = _calc_discounted_price_map(doc, items_list)

            for it in items_list:
                # тип строки (Prekės/Paslaugos) -> operationTypeName*
                attr = _attribute_name_from_preke_paslauga(
                    getattr(it, "preke_paslauga", None) or getattr(doc, "preke_paslauga", None)
                )
                op_type = _get_operation_type_name("pirkimas", attr)

                base = {
                    "purchaseDate": purchase_date,
                    "number": number,
                    "operationTypeName": op_type,
                    "currencyId": currency,
                    "supplierName": supplier["name"],
                    "warehouseName": warehouse,
                }
                if supplier["code"]:
                    base["supplierCode"] = supplier["code"]
                if cost_center:
                    base["costCenter"] = cost_center
                if employee:
                    base["employee"] = employee
                if series:
                    base["series"] = series

                item_name, code, barcode = _get_item_identity(doc, it)
                qty = _safe_D(getattr(it, "quantity", 1) or 1)

                if price_map:
                    price = _safe_D(price_map.get(id(it), getattr(it, "price", 0) or 0))
                else:
                    price = _safe_D(getattr(it, "price", 0) or 0)

                row = dict(base)
                row.update({
                    "items": item_name,            # pavadinimas
                    "quantity": float(qty),
                    "priceExclVat": float(_quantize_2(price)),
                })

                # 2) обязательно пробуем проставить и код, и баркод (если есть)
                if code:
                    row["code"] = code
                if barcode:
                    row["barcode"] = barcode

                # 1) vatRate: item.vat_percent если есть lineitems
                vat_rate = getattr(it, "vat_percent", None)
                if vat_rate is not None and _s(vat_rate) != "":
                    try:
                        row["vatRate"] = float(_safe_D(vat_rate))
                    except Exception:
                        pass

                vat_classifier = _get_vat_classifier(doc, it)
                if vat_classifier:
                    row["vatClassifier"] = vat_classifier

                rows.append(row)

        else:
            # суммарный документ: vatRate берём из doc.vat_percent (если есть)
            doc_attr = _attribute_name_from_preke_paslauga(getattr(doc, "preke_paslauga", None))
            op_type = _get_operation_type_name("pirkimas", doc_attr)

            base = {
                "purchaseDate": purchase_date,
                "number": number,
                "operationTypeName": op_type,
                "currencyId": currency,
                "supplierName": supplier["name"],
                "warehouseName": warehouse,
            }
            if supplier["code"]:
                base["supplierCode"] = supplier["code"]
            if cost_center:
                base["costCenter"] = cost_center
            if employee:
                base["employee"] = employee
            if series:
                base["series"] = series

            amount_wo = _safe_D(getattr(doc, "amount_wo_vat", 0) or 0)
            discount_wo = _safe_D(getattr(doc, "invoice_discount_wo_vat", 0) or 0)
            if discount_wo > 0:
                amount_wo = amount_wo - discount_wo
                if amount_wo < 0:
                    amount_wo = Decimal("0")

            item_name, code, barcode = _get_item_identity(doc, None)
            if not item_name:
                item_name = "Preke"

            row = dict(base)
            row.update({
                "items": item_name,
                "quantity": 1,
                "priceExclVat": float(_quantize_2(amount_wo)),
            })

            if code:
                row["code"] = code
            if barcode:
                row["barcode"] = barcode

            vat_rate = getattr(doc, "vat_percent", None)
            if vat_rate is not None and _s(vat_rate) != "":
                try:
                    row["vatRate"] = float(_safe_D(vat_rate))
                except Exception:
                    pass
            
            vat_classifier = _get_vat_classifier(doc, None)
            if vat_classifier:
                row["vatClassifier"] = vat_classifier

            rows.append(row)

    if not rows:
        return b""

    _write_rows(ws, base_to_col, rows, start_row=start_row)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# ---------- SALES ----------

def _export_sales(documents: list, user=None) -> bytes:
    wb, ws = _load_template(SITE_PRO_TEMPLATE_SALES)
    base_to_col = _build_field_map(ws, key_row=2)

    start_row = 3
    _clear_data_area(ws, start_row, base_to_col)

    rows = []

    for doc in documents:
        if _doc_type(doc) != "pardavimas":
            continue

        buyer = _get_buyer_fields(doc)
        sale_date = _get_doc_date(doc)
        series = _get_doc_series_for_sales(doc)  # fallback SF остаётся
        number_raw = _get_doc_number(doc)
        number = _normalize_number_remove_series_prefix(series, number_raw)
        currency = _get_currency(doc)

        warehouse = _get_warehouse_name(user, "pardavimas")
        employee = _get_employee_name(user)
        cost_center = _get_cost_center(user, "pardavimas") or ""

        line_items = getattr(doc, "line_items", None)
        if line_items and hasattr(line_items, "all") and line_items.exists():
            items_list = list(line_items.all())
            price_map = _calc_discounted_price_map(doc, items_list)

            for it in items_list:
                attr = _attribute_name_from_preke_paslauga(
                    getattr(it, "preke_paslauga", None) or getattr(doc, "preke_paslauga", None)
                )
                op_type = _get_operation_type_name("pardavimas", attr)

                base = {
                    "saleDate": sale_date,
                    "series": series,
                    "number": number,
                    "operationTypeName": op_type,
                    "currencyId": currency,
                    "employee": employee,
                    "clientName": buyer["name"],
                    "warehouseName": warehouse,
                }
                if buyer["code"]:
                    base["clientCode"] = buyer["code"]
                if cost_center:
                    base["costCenter"] = cost_center

                item_name, code, barcode = _get_item_identity(doc, it)
                qty = _safe_D(getattr(it, "quantity", 1) or 1)

                if price_map:
                    price = _safe_D(price_map.get(id(it), getattr(it, "price", 0) or 0))
                else:
                    price = _safe_D(getattr(it, "price", 0) or 0)

                row = dict(base)
                row.update({
                    "items": item_name,
                    "quantity": float(qty),
                    "priceExclVat": float(_quantize_2(price)),
                })

                if code:
                    row["code"] = code
                if barcode:
                    row["barcode"] = barcode

                # 1) vatRate: item.vat_percent если есть lineitems
                vat_rate = getattr(it, "vat_percent", None)
                if vat_rate is not None and _s(vat_rate) != "":
                    try:
                        row["vatRate"] = float(_safe_D(vat_rate))
                    except Exception:
                        pass

                vat_classifier = _get_vat_classifier(doc, it)
                if vat_classifier:
                    row["vatClassifier"] = vat_classifier

                rows.append(row)

        else:
            doc_attr = _attribute_name_from_preke_paslauga(getattr(doc, "preke_paslauga", None))
            op_type = _get_operation_type_name("pardavimas", doc_attr)

            base = {
                "saleDate": sale_date,
                "series": series,
                "number": number,
                "operationTypeName": op_type,
                "currencyId": currency,
                "employee": employee,
                "clientName": buyer["name"],
                "warehouseName": warehouse,
            }
            if buyer["code"]:
                base["clientCode"] = buyer["code"]
            if cost_center:
                base["costCenter"] = cost_center

            amount_wo = _safe_D(getattr(doc, "amount_wo_vat", 0) or 0)
            discount_wo = _safe_D(getattr(doc, "invoice_discount_wo_vat", 0) or 0)
            if discount_wo > 0:
                amount_wo = amount_wo - discount_wo
                if amount_wo < 0:
                    amount_wo = Decimal("0")

            item_name, code, barcode = _get_item_identity(doc, None)
            if not item_name:
                item_name = "Preke"

            row = dict(base)
            row.update({
                "items": item_name,
                "quantity": 1,
                "priceExclVat": float(_quantize_2(amount_wo)),
            })

            if code:
                row["code"] = code
            if barcode:
                row["barcode"] = barcode

            vat_rate = getattr(doc, "vat_percent", None)
            if vat_rate is not None and _s(vat_rate) != "":
                try:
                    row["vatRate"] = float(_safe_D(vat_rate))
                except Exception:
                    pass

            vat_classifier = _get_vat_classifier(doc, None)
            if vat_classifier:
                row["vatClassifier"] = vat_classifier

            rows.append(row)

    if not rows:
        return b""

    _write_rows(ws, base_to_col, rows, start_row=start_row)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

from __future__ import annotations

import os
import random
import logging
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)

# =========================
# Конфиг путей
# =========================
env_value = os.getenv("RIVILE_ERP_TEMPLATES_DIR")

TEMPLATES_DIR = Path(env_value)

PREKES_TEMPLATE_FILE = "Prekės, paslaugos.xlsx"
CLIENTS_TEMPLATE_FILE = "Klientai.xlsx"
PIRK_TEMPLATE_FILE = "Pirkimai.xlsx"
PARD_TEMPLATE_FILE = "Pardavimai.xlsx"


# =========================
# Константы/форматы Excel
# =========================
DATE_FMT = "yyyy-mm-dd"
MONEY_FMT = "0.00"
PRICE_FMT = "0.00##"
QTY_FMT = "0.#####"

EXCEL_BAD_PREFIXES = ("=", "+", "-", "@", "\t")
DEFAULT_UNIT = "VNT"
DEFAULT_CURRENCY = "EUR"
DEFAULT_DEPT = "01"


# =========================
# Колонки (1-based)
# =========================
class PrekesCols:
    REF_ID = 1
    TYPE = 2
    CODE = 3
    NAME = 4
    BASE_UOM = 5


class ClientCols:
    REF_ID = 1
    NAME = 2
    CODE = 3
    TYPE_ID = 4
    REG_CODE = 7
    VAT = 8
    ADDRESS = 9
    IS_CUSTOMER = 21
    IS_SUPPLIER = 27


class HeaderCols:
    REF_ID = 1
    CLIENT_CODE = 2
    OP_DATE = 3
    INV_DATE = 4
    DOC_NO = 5
    DOC_TYPE = 6
    JOURNAL = 7
    CURRENCY = 9
    DEPARTMENT = 20
    OBJECT = 21


class LineCols:
    REF_ID = 1
    ITEM_CODE = 2
    UOM = 3
    BARCODE = 4
    DEPT = 5
    QTY = 6
    PRICE = 7
    DISCOUNT_PCT = 8
    VAT_CODE = 9
    VAT_AMOUNT = 10
    NAME = 11
    OBJECT_CODE = 12


# =========================
# Хелперы нормализации/форматирования
# =========================
def _s(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _safe_D(x: Any) -> Decimal:
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")


def safe_excel_text(value: Optional[str]) -> str:
    s = _s(value)
    if not s:
        return ""
    if s.startswith(EXCEL_BAD_PREFIXES):
        return "'" + s
    return s


def normalize_code(code: Optional[str]) -> str:
    return safe_excel_text(_s(code).upper())


def to_decimal(x: Any, q: str = "0.01") -> Decimal:
    if x is None or x == "":
        d = Decimal("0")
    elif isinstance(x, Decimal):
        d = x
    else:
        d = Decimal(str(x).replace(",", "."))
    return d.quantize(Decimal(q), rounding=ROUND_HALF_UP)


def set_cell_money(ws: Worksheet, row: int, col: int, amount: Any):
    c = ws.cell(row=row, column=col, value=float(to_decimal(amount, q="0.01")))
    c.number_format = MONEY_FMT


def set_cell_price(ws: Worksheet, row: int, col: int, price: Any):
    c = ws.cell(row=row, column=col, value=float(to_decimal(price, q="0.0001")))
    c.number_format = PRICE_FMT


def set_cell_qty(ws: Worksheet, row: int, col: int, qty: Any):
    s = _s(qty)
    if s:
        s = s.strip()
        if s.endswith((",", ".")):
            s = s[:-1]
        s = s.replace(",", ".")
    try:
        d = Decimal(s) if s else Decimal("0")
    except Exception:
        d = Decimal("0")

    d = d.quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)

    if d == d.to_integral_value():
        fmt = "0"
    else:
        fmt = QTY_FMT

    c = ws.cell(row=row, column=col, value=float(d))
    c.number_format = fmt


def to_excel_date(d: Any) -> Optional[date]:
    if not d:
        return None
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    try:
        return datetime.fromisoformat(str(d)).date()
    except Exception:
        return None


def set_cell_date(ws: Worksheet, row: int, col: int, d: Any):
    dt = to_excel_date(d)
    if dt:
        c = ws.cell(row=row, column=col, value=dt)
        c.number_format = DATE_FMT
    else:
        ws.cell(row=row, column=col, value="")


# =========================
# Расчёт процента скидки документа
# =========================
def compute_global_invoice_discount_pct(doc: Any) -> Optional[Decimal]:
    disc = _safe_D(getattr(doc, "invoice_discount_wo_vat", 0) or 0)
    if disc <= 0:
        return None

    line_items = getattr(doc, "line_items", None)
    if line_items and hasattr(line_items, "all") and line_items.exists():
        base_total = Decimal("0")
        for it in line_items.all():
            qty = _safe_D(getattr(it, "quantity", 1) or 1)
            price = _safe_D(getattr(it, "price", 0) or 0)
            base_total += (price * qty)
    else:
        base_total = _safe_D(getattr(doc, "amount_wo_vat", 0) or 0)

    if base_total <= 0:
        return None

    pct = (disc / base_total) * Decimal("100")
    if pct < 0:
        pct = Decimal("0")
    if pct > Decimal("99.99"):
        pct = Decimal("99.99")
    return pct.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def compute_global_invoice_discount_pct_for_merge_vat(doc: Any) -> Optional[Decimal]:
    disc_wo = _safe_D(getattr(doc, "invoice_discount_wo_vat", 0) or 0)
    if disc_wo <= 0:
        return None

    line_items = getattr(doc, "line_items", None)
    if line_items and hasattr(line_items, "all") and line_items.exists():
        gross_total = Decimal("0")
        for it in line_items.all():
            qty = _safe_D(getattr(it, "quantity", 1) or 1)
            price_wo = _safe_D(getattr(it, "price", 0) or 0)
            vat_line = _safe_D(getattr(it, "vat", 0) or 0)  # PVM na stroku
            gross_total += (price_wo * qty) + vat_line
    else:
        gross_total = _safe_D(getattr(doc, "amount_wo_vat", 0) or 0) + _safe_D(getattr(doc, "vat_amount", 0) or 0)

    if gross_total <= 0:
        return None

    pct = (disc_wo / gross_total) * Decimal("100")
    if pct < 0:
        pct = Decimal("0")
    if pct > Decimal("99.99"):
        pct = Decimal("99.99")
    return pct.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# =========================
# Документ: номер/рефы
# =========================
def build_dok_nr(series: str, number: str) -> str:
    s = (series or "").strip()
    n = (number or "").strip()

    if not s:
        return n
    if not n:
        return s

    if n.startswith(s):
        tail = n[len(s):]
        tail = tail.lstrip("-/ .")
        return f"{s}{tail}"

    return f"{s}{n}"


def _fallback_doc_num(series: str, number: str) -> str:
    s = _s(series)
    n = _s(number)
    if not s and not n:
        return f"NERANUMERIO{random.randint(0, 99999):05d}"
    if s and not n:
        return s
    if n and not s:
        return n
    return n if n.startswith(s) else f"{s}{n}"


def build_ref_id(series: str, number: str) -> str:
    ref = build_dok_nr(series, number)
    if not _s(ref):
        ref = _fallback_doc_num(series, number)
    return ref


# =========================
# Уникальный REF_ID с дедупликацией
# =========================
def build_unique_ref_id(
    series: str,
    number: str,
    seen_refs: set[str],
    doc_pk: Any = None
) -> str:
    """
    Строит уникальный ref_id.
    
    Алгоритм:
    1. Строим базовый ref_id (series + number)
    2. Если НЕ в seen_refs → возвращаем как есть
    3. Если УЖЕ в seen_refs (дубликат):
       a) Пробуем "{base}_{doc_pk}" (например SF001_12345)
       b) Если и это занято → "{base}_2", "{base}_3"...
    
    Добавляет результат в seen_refs.
    """
    base_ref = build_ref_id(series, number)
    
    # Если уникальный — сразу возвращаем
    if base_ref not in seen_refs:
        seen_refs.add(base_ref)
        return base_ref
    
    # Дубликат! Пробуем с doc_pk
    if doc_pk is not None:
        ref_with_pk = f"{base_ref}_{doc_pk}"
        if ref_with_pk not in seen_refs:
            seen_refs.add(ref_with_pk)
            logger.debug(
                "[RIVILE_ERP:REF_ID] Дубликат '%s' -> '%s' (добавлен pk)",
                base_ref, ref_with_pk
            )
            return ref_with_pk
    
    # doc_pk нет или тоже занят — добавляем порядковый суффикс
    counter = 2
    while True:
        candidate = f"{base_ref}_{counter}"
        if candidate not in seen_refs:
            seen_refs.add(candidate)
            logger.debug(
                "[RIVILE_ERP:REF_ID] Дубликат '%s' -> '%s' (добавлен суффикс)",
                base_ref, candidate
            )
            return candidate
        counter += 1
        # Защита от бесконечного цикла (маловероятно, но на всякий случай)
        if counter > 9999:
            candidate = f"{base_ref}_{random.randint(10000, 99999)}"
            seen_refs.add(candidate)
            return candidate


# =========================
# Код контрагента
# =========================
def get_party_code(doc: Any, *, id_field: str, vat_field: str, id_programoje_field: str) -> str:
    sid = _s(getattr(doc, id_field, None))
    if sid:
        return sid
    svat = _s(getattr(doc, vat_field, None))
    if svat:
        return svat
    sidp = _s(getattr(doc, id_programoje_field, None))
    if sidp:
        return sidp
    return ""


# =========================
# Нормализация типа позиции
# =========================
def normalize_tip_lineitem(value: Any) -> str:
    s = str(value).strip() if value is not None else ""
    if not s:
        return "1"
    
    try:
        n = int(float(s.replace(",", ".")))
        if n == 1:
            return "1"
        elif n == 2:
            return "2"
        elif n == 3:
            return "1"
        elif n == 4:
            return "2"
        return "1"
    except Exception:
        pass
    
    low = s.lower()
    if low in ("preke", "prekė", "prekes", "prekės"):
        return "1"
    if low in ("paslauga", "paslaugos"):
        return "2"
    
    return "1"


def normalize_tip_doc(value: Any) -> str:
    s = _s(value)
    if not s:
        return "1"
    try:
        n = int(float(s.replace(",", ".")))
    except Exception:
        low = s.lower()
        if low in ("preke", "prekė", "prekes", "prekės"):
            return "1"
        if low in ("paslauga", "paslaugos"):
            return "2"
        return "1"
    if n == 1:
        return "1"
    if n == 2:
        return "2"
    if n == 3:
        return "1"
    if n == 4:
        return "2"
    return "1"


# =========================
# Загрузка шаблона
# =========================
def _load_template(filename: str):
    path = TEMPLATES_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"Не найден шаблон: {path}")
    wb = load_workbook(path)
    return wb


# =========================
# PVM Kodas helpers
# =========================
def _get_pvm_kodas_for_export(doc, item=None, line_map=None) -> str:
    """
    Получает PVM kodas для экспорта с учётом separate_vat.
    
    ПРАВИЛО:
    - Если separate_vat=True и нет line items (sumiskai) -> пустой
    - Если есть line_map и item -> берём из line_map
    - Иначе -> из item.pvm_kodas или doc.pvm_kodas
    
    ВАЖНО: "Keli skirtingi PVM" — это маркер, не реальный код -> возвращаем пустой
    """
    separate_vat = bool(getattr(doc, "separate_vat", False))
    scan_type = _s(getattr(doc, "scan_type", "")).lower()
    
    # Случай: sumiskai + separate_vat=True -> пустой PVM код
    if separate_vat and scan_type == "sumiskai":
        logger.debug(
            "[RIVILE_ERP:PVM] doc=%s sumiskai+separate_vat=True -> empty pvm_kodas",
            getattr(doc, "pk", None)
        )
        return ""
    
    # Случай: есть item и line_map (detaliai режим)
    if item is not None and line_map is not None:
        item_id = getattr(item, "id", None)
        if item_id is not None and item_id in line_map:
            pvm = _s(line_map.get(item_id, ""))
            # Фильтруем маркер
            if pvm == "Keli skirtingi PVM":
                return ""
            return pvm
    
    # Случай: есть item, но нет line_map -> берём из item
    if item is not None:
        pvm = _s(getattr(item, "pvm_kodas", ""))
        if pvm == "Keli skirtingi PVM":
            return ""
        return pvm
    
    # Случай: нет item (sumiskai) -> берём из документа
    pvm = _s(getattr(doc, "pvm_kodas", ""))
    if pvm == "Keli skirtingi PVM":
        return ""
    return pvm


def _get_vat_percent_for_export(doc, item=None) -> Any:
    """
    Получает vat_percent для экспорта.
    
    ПРАВИЛО:
    - Если separate_vat=True и нет line items (sumiskai) -> None (пустой)
    - Иначе -> из item.vat_percent или doc.vat_percent
    """
    separate_vat = bool(getattr(doc, "separate_vat", False))
    scan_type = _s(getattr(doc, "scan_type", "")).lower()
    
    # Случай: sumiskai + separate_vat=True -> пустой
    if separate_vat and scan_type == "sumiskai":
        logger.debug(
            "[RIVILE_ERP:VAT] doc=%s sumiskai+separate_vat=True -> empty vat_percent",
            getattr(doc, "pk", None)
        )
        return None
    
    # Иначе возвращаем реальное значение
    if item is not None:
        return getattr(item, "vat_percent", None)
    return getattr(doc, "vat_percent", None)


# =========================================================
# 1) PREKĖS / PASLAUGOS
# =========================================================
def export_prekes_and_paslaugos_to_rivile_erp_xlsx(documents: Iterable[Any], output_path: str | Path) -> Path:
    wb = _load_template(PREKES_TEMPLATE_FILE)
    ws = wb.active

    prekes_rows: list[list[str]] = []
    seen: set[str] = set()

    for doc in documents or []:
        line_items = getattr(doc, "line_items", None)
        has_items = bool(line_items and hasattr(line_items, "all") and line_items.exists())

        if has_items:
            for item in line_items.all():
                kodas_raw = getattr(item, "prekes_kodas", None) or getattr(item, "prekes_barkodas", None)
                kodas = normalize_code(kodas_raw)
                if not kodas or kodas in seen:
                    continue

                tipas = normalize_tip_lineitem(getattr(item, "preke_paslauga", None))
                unit = normalize_code(getattr(item, "unit", None) or DEFAULT_UNIT)
                pavadinimas = safe_excel_text(getattr(item, "prekes_pavadinimas", None) or "Prekė")

                prekes_rows.append([kodas, tipas, kodas, pavadinimas, unit])
                seen.add(kodas)
        else:
            kodas_raw = getattr(doc, "prekes_kodas", None) or getattr(doc, "prekes_barkodas", None)
            kodas = normalize_code(kodas_raw)
            if not kodas or kodas in seen:
                continue

            tipas = normalize_tip_doc(getattr(doc, "preke_paslauga", None))
            unit = normalize_code(getattr(doc, "unit", None) or DEFAULT_UNIT)
            pavadinimas = safe_excel_text(getattr(doc, "prekes_pavadinimas", None) or "Prekė")

            prekes_rows.append([kodas, tipas, kodas, pavadinimas, unit])
            seen.add(kodas)

    start_row = 6
    for i, row_data in enumerate(prekes_rows):
        r = start_row + i
        ws.cell(row=r, column=PrekesCols.REF_ID, value=row_data[0])
        ws.cell(row=r, column=PrekesCols.TYPE, value=row_data[1])
        ws.cell(row=r, column=PrekesCols.CODE, value=row_data[2])
        ws.cell(row=r, column=PrekesCols.NAME, value=row_data[3])
        ws.cell(row=r, column=PrekesCols.BASE_UOM, value=row_data[4])

    wb.save(output_path)
    return Path(output_path)


# =========================================================
# 2) KLIENTAI (с дедупликацией)
# =========================================================
def export_clients_to_rivile_erp_xlsx(clients: Iterable[dict], output_path: str | Path) -> Path:
    wb = _load_template(CLIENTS_TEMPLATE_FILE)
    ws = wb.active

    # Проверяем наличие дополнительных листов
    ws_addresses = wb["ClientAddresses"] if "ClientAddresses" in wb.sheetnames else None
    ws_banks = wb["ClientBankAccounts"] if "ClientBankAccounts" in wb.sheetnames else None

    start_row = 6
    row_idx = 0
    addr_row = 3   # стартовая строка для ClientAddresses
    bank_row = 3   # стартовая строка для ClientBankAccounts
    seen_clients: set[str] = set()

    for client in clients or []:
        cid = normalize_code(client.get("id"))
        if not cid or cid in seen_clients:
            continue
        seen_clients.add(cid)

        row = start_row + row_idx

        doc_type = (_s(client.get("type")) or "pirkimas").lower()
        is_person = bool(client.get("is_person", False))

        code = cid
        name = safe_excel_text(client.get("name"))
        vat = normalize_code(client.get("vat"))
        adr = safe_excel_text(client.get("address"))
        country_iso = safe_excel_text(client.get("country_iso"))
        iban = safe_excel_text(client.get("iban"))

        # === Первый таб (основной) ===
        ws.cell(row=row, column=ClientCols.REF_ID, value=cid)
        ws.cell(row=row, column=ClientCols.NAME, value=name)
        ws.cell(row=row, column=ClientCols.CODE, value=code)
        ws.cell(row=row, column=ClientCols.TYPE_ID, value=1 if is_person else 0)
        ws.cell(row=row, column=ClientCols.REG_CODE, value=cid)
        ws.cell(row=row, column=ClientCols.VAT, value=vat)
        ws.cell(row=row, column=ClientCols.ADDRESS, value=adr)

        ws.cell(row=row, column=ClientCols.IS_CUSTOMER, value=1 if doc_type == "pardavimas" else "")
        ws.cell(row=row, column=ClientCols.IS_SUPPLIER, value=1 if doc_type == "pirkimas" else "")

        # === ClientAddresses (если есть address или country_iso) ===
        if ws_addresses is not None and (adr or country_iso):
            ws_addresses.cell(row=addr_row, column=1, value=cid)          # A: RefID
            ws_addresses.cell(row=addr_row, column=2, value=0)            # B: всегда 0
            ws_addresses.cell(row=addr_row, column=3, value=name)         # C: Name
            ws_addresses.cell(row=addr_row, column=4, value=adr)          # D: Address
            ws_addresses.cell(row=addr_row, column=7, value=country_iso)  # G: Country ISO
            addr_row += 1

        # === ClientBankAccounts (если есть iban) ===
        if ws_banks is not None and iban:
            ws_banks.cell(row=bank_row, column=1, value=cid)   # A: RefID
            ws_banks.cell(row=bank_row, column=2, value=name)  # B: Name
            ws_banks.cell(row=bank_row, column=3, value=iban)  # C: IBAN
            bank_row += 1

        row_idx += 1

    wb.save(output_path)
    return Path(output_path)





# def export_clients_to_rivile_erp_xlsx(clients: Iterable[dict], output_path: str | Path) -> Path:
#     wb = _load_template(CLIENTS_TEMPLATE_FILE)
#     ws = wb.active

#     start_row = 6
#     row_idx = 0
#     seen_clients: set[str] = set()  # Для дедупликации

#     for client in clients or []:
#         cid = normalize_code(client.get("id"))
        
#         # Пропускаем пустые или дублирующиеся id
#         if not cid or cid in seen_clients:
#             continue
#         seen_clients.add(cid)

#         row = start_row + row_idx

#         doc_type = (_s(client.get("type")) or "pirkimas").lower()
#         is_person = bool(client.get("is_person", False))

#         code = cid
#         name = safe_excel_text(client.get("name"))
#         vat = normalize_code(client.get("vat"))
#         adr = safe_excel_text(client.get("address"))

#         ws.cell(row=row, column=ClientCols.REF_ID, value=cid)
#         ws.cell(row=row, column=ClientCols.NAME, value=name)
#         ws.cell(row=row, column=ClientCols.CODE, value=code)
#         ws.cell(row=row, column=ClientCols.TYPE_ID, value=1 if is_person else 0)
#         ws.cell(row=row, column=ClientCols.REG_CODE, value=cid)
#         ws.cell(row=row, column=ClientCols.VAT, value=vat)
#         ws.cell(row=row, column=ClientCols.ADDRESS, value=adr)

#         ws.cell(row=row, column=ClientCols.IS_CUSTOMER, value=1 if doc_type == "pardavimas" else "")
#         ws.cell(row=row, column=ClientCols.IS_SUPPLIER, value=1 if doc_type == "pirkimas" else "")

#         row_idx += 1

#     wb.save(output_path)
#     return Path(output_path)


# =========================================================
# 3) PIRKIMAI/PARDAVIMAI (Headers/Lines) с уникальными ref_id
# =========================================================
def export_documents_to_rivile_erp_xlsx(
    documents: Iterable[Any],
    output_path: str | Path,
    doc_type: str = "pirkimai",
    rivile_erp_extra_fields: Optional[dict] = None,
) -> Path:
    """
    Экспорт документов в XLSX-шаблоны Rivile ERP.
    doc_type: 'pirkimai' или 'pardavimai'
    """
    if doc_type == "pirkimai":
        wb = _load_template(PIRK_TEMPLATE_FILE)
        client_id_field = "seller_id"
        client_vat_field = "seller_vat_code"
        client_id_programoje_field = "seller_id_programoje"
        default_journal = "0201"
        prefix = "pirkimas"
    elif doc_type == "pardavimai":
        wb = _load_template(PARD_TEMPLATE_FILE)
        client_id_field = "buyer_id"
        client_vat_field = "buyer_vat_code"
        client_id_programoje_field = "buyer_id_programoje"
        default_journal = "0101"
        prefix = "pardavimas"
    else:
        raise ValueError("doc_type должен быть 'pirkimai' или 'pardavimai'")

    if "Headers" not in wb.sheetnames or "Lines" not in wb.sheetnames:
        raise ValueError("Шаблон должен содержать листы 'Headers' и 'Lines'")

    extra = rivile_erp_extra_fields or {}
    user = extra.get("user") if isinstance(extra.get("user"), dict) else {}
    extra_settings = user.get("extra_settings") if isinstance(user.get("extra_settings"), dict) else {}
    merge_vat = str(extra_settings.get("merge_vat", "0")).strip() == "1"
    journal_key = f"{prefix}_zurnalo_kodas"
    dept_key    = f"{prefix}_padalinio_kodas"
    obj_key     = f"{prefix}_objekto_kodas"

    user_journal = _s(extra.get(journal_key) or "")
    user_dept    = _s(extra.get(dept_key) or "")
    user_obj     = _s(extra.get(obj_key) or "")

    ws_headers = wb["Headers"]
    ws_lines = wb["Lines"]

    header_idx = 6
    line_idx = 3

    # ====== НОВОЕ: отслеживание уникальности ref_id ======
    seen_refs: set[str] = set()

    for doc in documents or []:
        dok_nr = _s(getattr(doc, "document_number", "") or "")
        series = _s(getattr(doc, "document_series", "") or "")
        doc_pk = getattr(doc, "pk", None) or getattr(doc, "id", None)

        # ====== ИЗМЕНЕНИЕ: используем build_unique_ref_id ======
        ref_id = build_unique_ref_id(series, dok_nr, seen_refs, doc_pk)

        client_code = get_party_code(
            doc,
            id_field=client_id_field,
            vat_field=client_vat_field,
            id_programoje_field=client_id_programoje_field,
        )

        if merge_vat:
            discount_pct = compute_global_invoice_discount_pct_for_merge_vat(doc)
        else:
            discount_pct = compute_global_invoice_discount_pct(doc)

        # === Headers ===
        header_row = header_idx

        ws_headers.cell(row=header_row, column=HeaderCols.REF_ID, value=safe_excel_text(ref_id))
        ws_headers.cell(row=header_row, column=HeaderCols.CLIENT_CODE, value=safe_excel_text(client_code))
        set_cell_date(
            ws_headers,
            header_row,
            HeaderCols.OP_DATE,
            getattr(doc, "operation_date", None) or getattr(doc, "invoice_date", None),
        )
        set_cell_date(ws_headers, header_row, HeaderCols.INV_DATE, getattr(doc, "invoice_date", None))
        ws_headers.cell(row=header_row, column=HeaderCols.DOC_NO, value=safe_excel_text(ref_id))
        ws_headers.cell(row=header_row, column=HeaderCols.DOC_TYPE, value=0)

        if user_journal:
            zurnalo_kodas = user_journal
        else:
            zurnalo_kodas = _s(getattr(doc, "zurnalo_kodas", "")) or default_journal

        ws_headers.cell(row=header_row, column=HeaderCols.JOURNAL, value=safe_excel_text(zurnalo_kodas))

        currency = _s(getattr(doc, "currency", "") or DEFAULT_CURRENCY) or DEFAULT_CURRENCY
        ws_headers.cell(row=header_row, column=HeaderCols.CURRENCY, value=currency)

        header_idx += 1

        # === Lines ===
        line_map = getattr(doc, "_pvm_line_map", None)
        line_items = getattr(doc, "line_items", None)
        has_items = bool(line_items and hasattr(line_items, "all") and line_items.exists())

        if has_items:
            # ========== DETALIAI режим ==========
            for item in line_items.all():
                ws_lines.cell(row=line_idx, column=LineCols.REF_ID, value=safe_excel_text(ref_id))
                ws_lines.cell(
                    row=line_idx,
                    column=LineCols.ITEM_CODE,
                    value=normalize_code(getattr(item, "prekes_kodas", None) or ""),
                )
                ws_lines.cell(
                    row=line_idx,
                    column=LineCols.UOM,
                    value=normalize_code(getattr(item, "unit", None) or DEFAULT_UNIT),
                )
                ws_lines.cell(
                    row=line_idx,
                    column=LineCols.BARCODE,
                    value=normalize_code(getattr(item, "prekes_barkodas", None) or ""),
                )

                item_dept = _s(getattr(item, "padalinio_kodas", None) or "")
                padalinio_kodas = item_dept or user_dept or DEFAULT_DEPT
                ws_lines.cell(
                    row=line_idx,
                    column=LineCols.DEPT,
                    value=safe_excel_text(padalinio_kodas),
                )

                # qty nuzhen dlia rascheta unit_vat
                qty_val = getattr(item, "quantity", None) or 1
                set_cell_qty(ws_lines, line_idx, LineCols.QTY, qty_val)

                if merge_vat:
                    qty_dec = _safe_D(qty_val)
                    price_wo = _safe_D(getattr(item, "price", 0) or 0)
                    vat_line = _safe_D(getattr(item, "vat", 0) or 0)  # PVM na stroku

                    unit_vat = (vat_line / qty_dec) if qty_dec != 0 else Decimal("0")
                    price_gross = price_wo + unit_vat

                    set_cell_price(ws_lines, line_idx, LineCols.PRICE, price_gross)

                    if discount_pct is not None:
                        ws_lines.cell(row=line_idx, column=LineCols.DISCOUNT_PCT, value=float(discount_pct))

                    # Ne pildom I i J
                    ws_lines.cell(row=line_idx, column=LineCols.VAT_CODE, value="")
                    set_cell_money(ws_lines, line_idx, LineCols.VAT_AMOUNT, 0)
                else:
                    set_cell_price(ws_lines, line_idx, LineCols.PRICE, getattr(item, "price", None) or 0)

                    if discount_pct is not None:
                        ws_lines.cell(row=line_idx, column=LineCols.DISCOUNT_PCT, value=float(discount_pct))
                    else:
                        set_cell_money(ws_lines, line_idx, LineCols.VAT_AMOUNT, getattr(item, "vat", None) or 0)

                    pvm_code = _get_pvm_kodas_for_export(doc, item=item, line_map=line_map)
                    ws_lines.cell(row=line_idx, column=LineCols.VAT_CODE, value=safe_excel_text(pvm_code))

                name = safe_excel_text(getattr(item, "prekes_pavadinimas", None) or "")
                ws_lines.cell(row=line_idx, column=LineCols.NAME, value=name)

                item_obj = _s(getattr(item, "objekto_kodas", None) or "")
                objekto_kodas = item_obj or user_obj
                if objekto_kodas:
                    ws_lines.cell(
                        row=line_idx,
                        column=LineCols.OBJECT_CODE,
                        value=safe_excel_text(objekto_kodas),
                    )

                line_idx += 1
        else:
            # ========== SUMISKAI режим ==========
            ws_lines.cell(row=line_idx, column=LineCols.REF_ID, value=safe_excel_text(ref_id))
            ws_lines.cell(
                row=line_idx,
                column=LineCols.ITEM_CODE,
                value=normalize_code(getattr(doc, "prekes_kodas", None) or ""),
            )
            ws_lines.cell(
                row=line_idx,
                column=LineCols.UOM,
                value=normalize_code(getattr(doc, "unit", None) or DEFAULT_UNIT),
            )
            ws_lines.cell(
                row=line_idx,
                column=LineCols.BARCODE,
                value=normalize_code(getattr(doc, "prekes_barkodas", None) or ""),
            )

            doc_dept = _s(getattr(doc, "padalinio_kodas", None) or "")
            padalinio_kodas = doc_dept or user_dept or DEFAULT_DEPT
            ws_lines.cell(
                row=line_idx,
                column=LineCols.DEPT,
                value=safe_excel_text(padalinio_kodas),
            )

            set_cell_qty(ws_lines, line_idx, LineCols.QTY, getattr(doc, "quantity", None) or 1)

            amount_wo = _safe_D(getattr(doc, "amount_wo_vat", None) or 0)
            vat_amount = _safe_D(getattr(doc, "vat_amount", None) or 0)

            if merge_vat:
                amount_gross = amount_wo + vat_amount
                set_cell_money(ws_lines, line_idx, LineCols.PRICE, amount_gross)

                if discount_pct is not None:
                    ws_lines.cell(row=line_idx, column=LineCols.DISCOUNT_PCT, value=float(discount_pct))

                # Ne pildom I i J
                ws_lines.cell(row=line_idx, column=LineCols.VAT_CODE, value="")
                set_cell_money(ws_lines, line_idx, LineCols.VAT_AMOUNT, 0)
            else:
                set_cell_money(ws_lines, line_idx, LineCols.PRICE, amount_wo)

                if discount_pct is not None:
                    ws_lines.cell(row=line_idx, column=LineCols.DISCOUNT_PCT, value=float(discount_pct))
                else:
                    set_cell_money(ws_lines, line_idx, LineCols.VAT_AMOUNT, vat_amount)

                pvm_code = _get_pvm_kodas_for_export(doc, item=None, line_map=None)
                ws_lines.cell(row=line_idx, column=LineCols.VAT_CODE, value=safe_excel_text(pvm_code))

            name = safe_excel_text(getattr(doc, "prekes_pavadinimas", None) or "")
            ws_lines.cell(row=line_idx, column=LineCols.NAME, value=name)

            doc_obj = _s(getattr(doc, "objekto_kodas", None) or "")
            objekto_kodas = doc_obj or user_obj
            if objekto_kodas:
                ws_lines.cell(
                    row=line_idx,
                    column=LineCols.OBJECT_CODE,
                    value=safe_excel_text(objekto_kodas),
                )

            ws_headers.cell(
                row=header_row,
                column=HeaderCols.DEPARTMENT,
                value=safe_excel_text(padalinio_kodas),
            )
            if objekto_kodas:
                ws_headers.cell(
                    row=header_row,
                    column=HeaderCols.OBJECT,
                    value=safe_excel_text(objekto_kodas),
                )

            line_idx += 1

    wb.save(output_path)
    return Path(output_path)







# from __future__ import annotations

# import os
# import random
# import logging
# from pathlib import Path
# from decimal import Decimal, ROUND_HALF_UP
# from datetime import date, datetime
# from typing import Any, Iterable, Optional

# from openpyxl import load_workbook
# from openpyxl.worksheet.worksheet import Worksheet


# logger = logging.getLogger(__name__)

# # =========================
# # Конфиг путей
# # =========================
# env_value = os.getenv("RIVILE_ERP_TEMPLATES_DIR")

# TEMPLATES_DIR = Path(env_value)

# PREKES_TEMPLATE_FILE = "Prekės, paslaugos.xlsx"
# CLIENTS_TEMPLATE_FILE = "Klientai.xlsx"
# PIRK_TEMPLATE_FILE = "Pirkimai.xlsx"
# PARD_TEMPLATE_FILE = "Pardavimai.xlsx"


# # =========================
# # Константы/форматы Excel
# # =========================
# DATE_FMT = "yyyy-mm-dd"
# MONEY_FMT = "0.00"
# PRICE_FMT = "0.00##"
# QTY_FMT = "0.#####"

# EXCEL_BAD_PREFIXES = ("=", "+", "-", "@", "\t")
# DEFAULT_UNIT = "VNT"
# DEFAULT_CURRENCY = "EUR"
# DEFAULT_DEPT = "01"


# # =========================
# # Колонки (1-based)
# # =========================
# class PrekesCols:
#     REF_ID = 1
#     TYPE = 2
#     CODE = 3
#     NAME = 4
#     BASE_UOM = 5


# class ClientCols:
#     REF_ID = 1
#     NAME = 2
#     CODE = 3
#     TYPE_ID = 4
#     REG_CODE = 7
#     VAT = 8
#     ADDRESS = 9
#     IS_CUSTOMER = 21
#     IS_SUPPLIER = 27


# class HeaderCols:
#     REF_ID = 1
#     CLIENT_CODE = 2
#     OP_DATE = 3
#     INV_DATE = 4
#     DOC_NO = 5
#     DOC_TYPE = 6
#     JOURNAL = 7
#     CURRENCY = 9
#     DEPARTMENT = 20
#     OBJECT = 21


# class LineCols:
#     REF_ID = 1
#     ITEM_CODE = 2
#     UOM = 3
#     BARCODE = 4
#     DEPT = 5
#     QTY = 6
#     PRICE = 7
#     DISCOUNT_PCT = 8
#     VAT_CODE = 9
#     VAT_AMOUNT = 10
#     NAME = 11
#     OBJECT_CODE = 12


# # =========================
# # Хелперы нормализации/форматирования
# # =========================
# def _s(v: Any) -> str:
#     return str(v).strip() if v is not None else ""


# def _safe_D(x: Any) -> Decimal:
#     try:
#         return Decimal(str(x))
#     except Exception:
#         return Decimal("0")


# def safe_excel_text(value: Optional[str]) -> str:
#     s = _s(value)
#     if not s:
#         return ""
#     if s.startswith(EXCEL_BAD_PREFIXES):
#         return "'" + s
#     return s


# def normalize_code(code: Optional[str]) -> str:
#     return safe_excel_text(_s(code).upper())


# def to_decimal(x: Any, q: str = "0.01") -> Decimal:
#     if x is None or x == "":
#         d = Decimal("0")
#     elif isinstance(x, Decimal):
#         d = x
#     else:
#         d = Decimal(str(x).replace(",", "."))
#     return d.quantize(Decimal(q), rounding=ROUND_HALF_UP)


# def set_cell_money(ws: Worksheet, row: int, col: int, amount: Any):
#     c = ws.cell(row=row, column=col, value=float(to_decimal(amount, q="0.01")))
#     c.number_format = MONEY_FMT


# def set_cell_price(ws: Worksheet, row: int, col: int, price: Any):
#     c = ws.cell(row=row, column=col, value=float(to_decimal(price, q="0.0001")))
#     c.number_format = PRICE_FMT


# def set_cell_qty(ws: Worksheet, row: int, col: int, qty: Any):
#     s = _s(qty)
#     if s:
#         s = s.strip()
#         if s.endswith((",", ".")):
#             s = s[:-1]
#         s = s.replace(",", ".")
#     try:
#         d = Decimal(s) if s else Decimal("0")
#     except Exception:
#         d = Decimal("0")

#     d = d.quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)

#     if d == d.to_integral_value():
#         fmt = "0"
#     else:
#         fmt = QTY_FMT

#     c = ws.cell(row=row, column=col, value=float(d))
#     c.number_format = fmt


# def to_excel_date(d: Any) -> Optional[date]:
#     if not d:
#         return None
#     if isinstance(d, datetime):
#         return d.date()
#     if isinstance(d, date):
#         return d
#     try:
#         return datetime.fromisoformat(str(d)).date()
#     except Exception:
#         return None


# def set_cell_date(ws: Worksheet, row: int, col: int, d: Any):
#     dt = to_excel_date(d)
#     if dt:
#         c = ws.cell(row=row, column=col, value=dt)
#         c.number_format = DATE_FMT
#     else:
#         ws.cell(row=row, column=col, value="")


# # =========================
# # Расчёт процента скидки документа
# # =========================
# def compute_global_invoice_discount_pct(doc: Any) -> Optional[Decimal]:
#     disc = _safe_D(getattr(doc, "invoice_discount_wo_vat", 0) or 0)
#     if disc <= 0:
#         return None

#     line_items = getattr(doc, "line_items", None)
#     if line_items and hasattr(line_items, "all") and line_items.exists():
#         base_total = Decimal("0")
#         for it in line_items.all():
#             qty = _safe_D(getattr(it, "quantity", 1) or 1)
#             price = _safe_D(getattr(it, "price", 0) or 0)
#             base_total += (price * qty)
#     else:
#         base_total = _safe_D(getattr(doc, "amount_wo_vat", 0) or 0)

#     if base_total <= 0:
#         return None

#     pct = (disc / base_total) * Decimal("100")
#     if pct < 0:
#         pct = Decimal("0")
#     if pct > Decimal("99.99"):
#         pct = Decimal("99.99")
#     return pct.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# # =========================
# # Документ: номер/рефы
# # =========================
# def build_dok_nr(series: str, number: str) -> str:
#     s = (series or "").strip()
#     n = (number or "").strip()

#     if not s:
#         return n
#     if not n:
#         return s

#     if n.startswith(s):
#         tail = n[len(s):]
#         tail = tail.lstrip("-/ .")
#         return f"{s}{tail}"

#     return f"{s}{n}"


# def _fallback_doc_num(series: str, number: str) -> str:
#     s = _s(series)
#     n = _s(number)
#     if not s and not n:
#         return f"NERANUMERIO{random.randint(0, 99999):05d}"
#     if s and not n:
#         return s
#     if n and not s:
#         return n
#     return n if n.startswith(s) else f"{s}{n}"


# def build_ref_id(series: str, number: str) -> str:
#     ref = build_dok_nr(series, number)
#     if not _s(ref):
#         ref = _fallback_doc_num(series, number)
#     return ref


# # =========================
# # Уникальный REF_ID с дедупликацией
# # =========================
# def build_unique_ref_id(
#     series: str,
#     number: str,
#     seen_refs: set[str],
#     doc_pk: Any = None
# ) -> str:
#     """
#     Строит уникальный ref_id.
    
#     Алгоритм:
#     1. Строим базовый ref_id (series + number)
#     2. Если НЕ в seen_refs → возвращаем как есть
#     3. Если УЖЕ в seen_refs (дубликат):
#        a) Пробуем "{base}_{doc_pk}" (например SF001_12345)
#        b) Если и это занято → "{base}_2", "{base}_3"...
    
#     Добавляет результат в seen_refs.
#     """
#     base_ref = build_ref_id(series, number)
    
#     # Если уникальный — сразу возвращаем
#     if base_ref not in seen_refs:
#         seen_refs.add(base_ref)
#         return base_ref
    
#     # Дубликат! Пробуем с doc_pk
#     if doc_pk is not None:
#         ref_with_pk = f"{base_ref}_{doc_pk}"
#         if ref_with_pk not in seen_refs:
#             seen_refs.add(ref_with_pk)
#             logger.debug(
#                 "[RIVILE_ERP:REF_ID] Дубликат '%s' -> '%s' (добавлен pk)",
#                 base_ref, ref_with_pk
#             )
#             return ref_with_pk
    
#     # doc_pk нет или тоже занят — добавляем порядковый суффикс
#     counter = 2
#     while True:
#         candidate = f"{base_ref}_{counter}"
#         if candidate not in seen_refs:
#             seen_refs.add(candidate)
#             logger.debug(
#                 "[RIVILE_ERP:REF_ID] Дубликат '%s' -> '%s' (добавлен суффикс)",
#                 base_ref, candidate
#             )
#             return candidate
#         counter += 1
#         # Защита от бесконечного цикла (маловероятно, но на всякий случай)
#         if counter > 9999:
#             candidate = f"{base_ref}_{random.randint(10000, 99999)}"
#             seen_refs.add(candidate)
#             return candidate


# # =========================
# # Код контрагента
# # =========================
# def get_party_code(doc: Any, *, id_field: str, vat_field: str, id_programoje_field: str) -> str:
#     sid = _s(getattr(doc, id_field, None))
#     if sid:
#         return sid
#     svat = _s(getattr(doc, vat_field, None))
#     if svat:
#         return svat
#     sidp = _s(getattr(doc, id_programoje_field, None))
#     if sidp:
#         return sidp
#     return ""


# # =========================
# # Нормализация типа позиции
# # =========================
# def normalize_tip_lineitem(value: Any) -> str:
#     s = str(value).strip() if value is not None else ""
#     if not s:
#         return "1"
    
#     try:
#         n = int(float(s.replace(",", ".")))
#         if n == 1:
#             return "1"
#         elif n == 2:
#             return "2"
#         elif n == 3:
#             return "1"
#         elif n == 4:
#             return "2"
#         return "1"
#     except Exception:
#         pass
    
#     low = s.lower()
#     if low in ("preke", "prekė", "prekes", "prekės"):
#         return "1"
#     if low in ("paslauga", "paslaugos"):
#         return "2"
    
#     return "1"


# def normalize_tip_doc(value: Any) -> str:
#     s = _s(value)
#     if not s:
#         return "1"
#     try:
#         n = int(float(s.replace(",", ".")))
#     except Exception:
#         low = s.lower()
#         if low in ("preke", "prekė", "prekes", "prekės"):
#             return "1"
#         if low in ("paslauga", "paslaugos"):
#             return "2"
#         return "1"
#     if n == 1:
#         return "1"
#     if n == 2:
#         return "2"
#     if n == 3:
#         return "1"
#     if n == 4:
#         return "2"
#     return "1"


# # =========================
# # Загрузка шаблона
# # =========================
# def _load_template(filename: str):
#     path = TEMPLATES_DIR / filename
#     if not path.exists():
#         raise FileNotFoundError(f"Не найден шаблон: {path}")
#     wb = load_workbook(path)
#     return wb


# # =========================
# # PVM Kodas helpers
# # =========================
# def _get_pvm_kodas_for_export(doc, item=None, line_map=None) -> str:
#     """
#     Получает PVM kodas для экспорта с учётом separate_vat.
    
#     ПРАВИЛО:
#     - Если separate_vat=True и нет line items (sumiskai) -> пустой
#     - Если есть line_map и item -> берём из line_map
#     - Иначе -> из item.pvm_kodas или doc.pvm_kodas
    
#     ВАЖНО: "Keli skirtingi PVM" — это маркер, не реальный код -> возвращаем пустой
#     """
#     separate_vat = bool(getattr(doc, "separate_vat", False))
#     scan_type = _s(getattr(doc, "scan_type", "")).lower()
    
#     # Случай: sumiskai + separate_vat=True -> пустой PVM код
#     if separate_vat and scan_type == "sumiskai":
#         logger.debug(
#             "[RIVILE_ERP:PVM] doc=%s sumiskai+separate_vat=True -> empty pvm_kodas",
#             getattr(doc, "pk", None)
#         )
#         return ""
    
#     # Случай: есть item и line_map (detaliai режим)
#     if item is not None and line_map is not None:
#         item_id = getattr(item, "id", None)
#         if item_id is not None and item_id in line_map:
#             pvm = _s(line_map.get(item_id, ""))
#             # Фильтруем маркер
#             if pvm == "Keli skirtingi PVM":
#                 return ""
#             return pvm
    
#     # Случай: есть item, но нет line_map -> берём из item
#     if item is not None:
#         pvm = _s(getattr(item, "pvm_kodas", ""))
#         if pvm == "Keli skirtingi PVM":
#             return ""
#         return pvm
    
#     # Случай: нет item (sumiskai) -> берём из документа
#     pvm = _s(getattr(doc, "pvm_kodas", ""))
#     if pvm == "Keli skirtingi PVM":
#         return ""
#     return pvm


# def _get_vat_percent_for_export(doc, item=None) -> Any:
#     """
#     Получает vat_percent для экспорта.
    
#     ПРАВИЛО:
#     - Если separate_vat=True и нет line items (sumiskai) -> None (пустой)
#     - Иначе -> из item.vat_percent или doc.vat_percent
#     """
#     separate_vat = bool(getattr(doc, "separate_vat", False))
#     scan_type = _s(getattr(doc, "scan_type", "")).lower()
    
#     # Случай: sumiskai + separate_vat=True -> пустой
#     if separate_vat and scan_type == "sumiskai":
#         logger.debug(
#             "[RIVILE_ERP:VAT] doc=%s sumiskai+separate_vat=True -> empty vat_percent",
#             getattr(doc, "pk", None)
#         )
#         return None
    
#     # Иначе возвращаем реальное значение
#     if item is not None:
#         return getattr(item, "vat_percent", None)
#     return getattr(doc, "vat_percent", None)


# # =========================================================
# # 1) PREKĖS / PASLAUGOS
# # =========================================================
# def export_prekes_and_paslaugos_to_rivile_erp_xlsx(documents: Iterable[Any], output_path: str | Path) -> Path:
#     wb = _load_template(PREKES_TEMPLATE_FILE)
#     ws = wb.active

#     prekes_rows: list[list[str]] = []
#     seen: set[str] = set()

#     for doc in documents or []:
#         line_items = getattr(doc, "line_items", None)
#         has_items = bool(line_items and hasattr(line_items, "all") and line_items.exists())

#         if has_items:
#             for item in line_items.all():
#                 kodas_raw = getattr(item, "prekes_kodas", None) or getattr(item, "prekes_barkodas", None)
#                 kodas = normalize_code(kodas_raw)
#                 if not kodas or kodas in seen:
#                     continue

#                 tipas = normalize_tip_lineitem(getattr(item, "preke_paslauga", None))
#                 unit = normalize_code(getattr(item, "unit", None) or DEFAULT_UNIT)
#                 pavadinimas = safe_excel_text(getattr(item, "prekes_pavadinimas", None) or "Prekė")

#                 prekes_rows.append([kodas, tipas, kodas, pavadinimas, unit])
#                 seen.add(kodas)
#         else:
#             kodas_raw = getattr(doc, "prekes_kodas", None) or getattr(doc, "prekes_barkodas", None)
#             kodas = normalize_code(kodas_raw)
#             if not kodas or kodas in seen:
#                 continue

#             tipas = normalize_tip_doc(getattr(doc, "preke_paslauga", None))
#             unit = normalize_code(getattr(doc, "unit", None) or DEFAULT_UNIT)
#             pavadinimas = safe_excel_text(getattr(doc, "prekes_pavadinimas", None) or "Prekė")

#             prekes_rows.append([kodas, tipas, kodas, pavadinimas, unit])
#             seen.add(kodas)

#     start_row = 6
#     for i, row_data in enumerate(prekes_rows):
#         r = start_row + i
#         ws.cell(row=r, column=PrekesCols.REF_ID, value=row_data[0])
#         ws.cell(row=r, column=PrekesCols.TYPE, value=row_data[1])
#         ws.cell(row=r, column=PrekesCols.CODE, value=row_data[2])
#         ws.cell(row=r, column=PrekesCols.NAME, value=row_data[3])
#         ws.cell(row=r, column=PrekesCols.BASE_UOM, value=row_data[4])

#     wb.save(output_path)
#     return Path(output_path)


# # =========================================================
# # 2) KLIENTAI (с дедупликацией)
# # =========================================================
# def export_clients_to_rivile_erp_xlsx(clients: Iterable[dict], output_path: str | Path) -> Path:
#     wb = _load_template(CLIENTS_TEMPLATE_FILE)
#     ws = wb.active

#     start_row = 6
#     row_idx = 0
#     seen_clients: set[str] = set()  # Для дедупликации

#     for client in clients or []:
#         cid = normalize_code(client.get("id"))
        
#         # Пропускаем пустые или дублирующиеся id
#         if not cid or cid in seen_clients:
#             continue
#         seen_clients.add(cid)

#         row = start_row + row_idx

#         doc_type = (_s(client.get("type")) or "pirkimas").lower()
#         is_person = bool(client.get("is_person", False))

#         code = cid
#         name = safe_excel_text(client.get("name"))
#         vat = normalize_code(client.get("vat"))
#         adr = safe_excel_text(client.get("address"))

#         ws.cell(row=row, column=ClientCols.REF_ID, value=cid)
#         ws.cell(row=row, column=ClientCols.NAME, value=name)
#         ws.cell(row=row, column=ClientCols.CODE, value=code)
#         ws.cell(row=row, column=ClientCols.TYPE_ID, value=1 if is_person else 0)
#         ws.cell(row=row, column=ClientCols.REG_CODE, value=cid)
#         ws.cell(row=row, column=ClientCols.VAT, value=vat)
#         ws.cell(row=row, column=ClientCols.ADDRESS, value=adr)

#         ws.cell(row=row, column=ClientCols.IS_CUSTOMER, value=1 if doc_type == "pardavimas" else "")
#         ws.cell(row=row, column=ClientCols.IS_SUPPLIER, value=1 if doc_type == "pirkimas" else "")

#         row_idx += 1

#     wb.save(output_path)
#     return Path(output_path)


# # =========================================================
# # 3) PIRKIMAI/PARDAVIMAI (Headers/Lines) с уникальными ref_id
# # =========================================================
# def export_documents_to_rivile_erp_xlsx(
#     documents: Iterable[Any],
#     output_path: str | Path,
#     doc_type: str = "pirkimai",
#     rivile_erp_extra_fields: Optional[dict] = None,
# ) -> Path:
#     """
#     Экспорт документов в XLSX-шаблоны Rivile ERP.
#     doc_type: 'pirkimai' или 'pardavimai'
#     """
#     if doc_type == "pirkimai":
#         wb = _load_template(PIRK_TEMPLATE_FILE)
#         client_id_field = "seller_id"
#         client_vat_field = "seller_vat_code"
#         client_id_programoje_field = "seller_id_programoje"
#         default_journal = "0201"
#         prefix = "pirkimas"
#     elif doc_type == "pardavimai":
#         wb = _load_template(PARD_TEMPLATE_FILE)
#         client_id_field = "buyer_id"
#         client_vat_field = "buyer_vat_code"
#         client_id_programoje_field = "buyer_id_programoje"
#         default_journal = "0101"
#         prefix = "pardavimas"
#     else:
#         raise ValueError("doc_type должен быть 'pirkimai' или 'pardavimai'")

#     if "Headers" not in wb.sheetnames or "Lines" not in wb.sheetnames:
#         raise ValueError("Шаблон должен содержать листы 'Headers' и 'Lines'")

#     extra = rivile_erp_extra_fields or {}
#     journal_key = f"{prefix}_zurnalo_kodas"
#     dept_key    = f"{prefix}_padalinio_kodas"
#     obj_key     = f"{prefix}_objekto_kodas"

#     user_journal = _s(extra.get(journal_key) or "")
#     user_dept    = _s(extra.get(dept_key) or "")
#     user_obj     = _s(extra.get(obj_key) or "")

#     ws_headers = wb["Headers"]
#     ws_lines = wb["Lines"]

#     header_idx = 6
#     line_idx = 3

#     # ====== НОВОЕ: отслеживание уникальности ref_id ======
#     seen_refs: set[str] = set()

#     for doc in documents or []:
#         dok_nr = _s(getattr(doc, "document_number", "") or "")
#         series = _s(getattr(doc, "document_series", "") or "")
#         doc_pk = getattr(doc, "pk", None) or getattr(doc, "id", None)

#         # ====== ИЗМЕНЕНИЕ: используем build_unique_ref_id ======
#         ref_id = build_unique_ref_id(series, dok_nr, seen_refs, doc_pk)

#         client_code = get_party_code(
#             doc,
#             id_field=client_id_field,
#             vat_field=client_vat_field,
#             id_programoje_field=client_id_programoje_field,
#         )

#         discount_pct = compute_global_invoice_discount_pct(doc)

#         # === Headers ===
#         header_row = header_idx

#         ws_headers.cell(row=header_row, column=HeaderCols.REF_ID, value=safe_excel_text(ref_id))
#         ws_headers.cell(row=header_row, column=HeaderCols.CLIENT_CODE, value=safe_excel_text(client_code))
#         set_cell_date(
#             ws_headers,
#             header_row,
#             HeaderCols.OP_DATE,
#             getattr(doc, "operation_date", None) or getattr(doc, "invoice_date", None),
#         )
#         set_cell_date(ws_headers, header_row, HeaderCols.INV_DATE, getattr(doc, "invoice_date", None))
#         ws_headers.cell(row=header_row, column=HeaderCols.DOC_NO, value=safe_excel_text(ref_id))
#         ws_headers.cell(row=header_row, column=HeaderCols.DOC_TYPE, value=0)

#         if user_journal:
#             zurnalo_kodas = user_journal
#         else:
#             zurnalo_kodas = _s(getattr(doc, "zurnalo_kodas", "")) or default_journal

#         ws_headers.cell(row=header_row, column=HeaderCols.JOURNAL, value=safe_excel_text(zurnalo_kodas))

#         currency = _s(getattr(doc, "currency", "") or DEFAULT_CURRENCY) or DEFAULT_CURRENCY
#         ws_headers.cell(row=header_row, column=HeaderCols.CURRENCY, value=currency)

#         header_idx += 1

#         # === Lines ===
#         line_map = getattr(doc, "_pvm_line_map", None)
#         line_items = getattr(doc, "line_items", None)
#         has_items = bool(line_items and hasattr(line_items, "all") and line_items.exists())

#         if has_items:
#             # ========== DETALIAI режим ==========
#             for item in line_items.all():
#                 ws_lines.cell(row=line_idx, column=LineCols.REF_ID, value=safe_excel_text(ref_id))
#                 ws_lines.cell(
#                     row=line_idx,
#                     column=LineCols.ITEM_CODE,
#                     value=normalize_code(getattr(item, "prekes_kodas", None) or ""),
#                 )
#                 ws_lines.cell(
#                     row=line_idx,
#                     column=LineCols.UOM,
#                     value=normalize_code(getattr(item, "unit", None) or DEFAULT_UNIT),
#                 )
#                 ws_lines.cell(
#                     row=line_idx,
#                     column=LineCols.BARCODE,
#                     value=normalize_code(getattr(item, "prekes_barkodas", None) or ""),
#                 )

#                 item_dept = _s(getattr(item, "padalinio_kodas", None) or "")
#                 padalinio_kodas = item_dept or user_dept or DEFAULT_DEPT
#                 ws_lines.cell(
#                     row=line_idx,
#                     column=LineCols.DEPT,
#                     value=safe_excel_text(padalinio_kodas),
#                 )

#                 set_cell_qty(ws_lines, line_idx, LineCols.QTY, getattr(item, "quantity", None) or 1)
#                 set_cell_price(ws_lines, line_idx, LineCols.PRICE, getattr(item, "price", None) or 0)

#                 if discount_pct is not None:
#                     ws_lines.cell(row=line_idx, column=LineCols.DISCOUNT_PCT, value=float(discount_pct))
#                 else:
#                     set_cell_money(ws_lines, line_idx, LineCols.VAT_AMOUNT, getattr(item, "vat", None) or 0)

#                 # ====== ИСПРАВЛЕНИЕ: Используем helper для PVM кода ======
#                 pvm_code = _get_pvm_kodas_for_export(doc, item=item, line_map=line_map)
#                 ws_lines.cell(
#                     row=line_idx,
#                     column=LineCols.VAT_CODE,
#                     value=safe_excel_text(pvm_code),
#                 )

#                 name = safe_excel_text(getattr(item, "prekes_pavadinimas", None) or "")
#                 ws_lines.cell(row=line_idx, column=LineCols.NAME, value=name)

#                 item_obj = _s(getattr(item, "objekto_kodas", None) or "")
#                 objekto_kodas = item_obj or user_obj
#                 if objekto_kodas:
#                     ws_lines.cell(
#                         row=line_idx,
#                         column=LineCols.OBJECT_CODE,
#                         value=safe_excel_text(objekto_kodas),
#                     )

#                 line_idx += 1
#         else:
#             # ========== SUMISKAI режим ==========
#             ws_lines.cell(row=line_idx, column=LineCols.REF_ID, value=safe_excel_text(ref_id))
#             ws_lines.cell(
#                 row=line_idx,
#                 column=LineCols.ITEM_CODE,
#                 value=normalize_code(getattr(doc, "prekes_kodas", None) or ""),
#             )
#             ws_lines.cell(
#                 row=line_idx,
#                 column=LineCols.UOM,
#                 value=normalize_code(getattr(doc, "unit", None) or DEFAULT_UNIT),
#             )
#             ws_lines.cell(
#                 row=line_idx,
#                 column=LineCols.BARCODE,
#                 value=normalize_code(getattr(doc, "prekes_barkodas", None) or ""),
#             )

#             doc_dept = _s(getattr(doc, "padalinio_kodas", None) or "")
#             padalinio_kodas = doc_dept or user_dept or DEFAULT_DEPT
#             ws_lines.cell(
#                 row=line_idx,
#                 column=LineCols.DEPT,
#                 value=safe_excel_text(padalinio_kodas),
#             )

#             set_cell_qty(ws_lines, line_idx, LineCols.QTY, getattr(doc, "quantity", None) or 1)

#             price = getattr(doc, "amount_wo_vat", None)
#             set_cell_money(ws_lines, line_idx, LineCols.PRICE, price if price is not None else 0)

#             if discount_pct is not None:
#                 ws_lines.cell(row=line_idx, column=LineCols.DISCOUNT_PCT, value=float(discount_pct))
#             else:
#                 vat_amount = getattr(doc, "vat_amount", None)
#                 set_cell_money(ws_lines, line_idx, LineCols.VAT_AMOUNT, vat_amount if vat_amount is not None else 0)

#             # ====== ИСПРАВЛЕНИЕ: Используем helper для PVM кода ======
#             # При sumiskai + separate_vat=True -> пустой
#             pvm_code = _get_pvm_kodas_for_export(doc, item=None, line_map=None)
#             ws_lines.cell(
#                 row=line_idx,
#                 column=LineCols.VAT_CODE,
#                 value=safe_excel_text(pvm_code),
#             )

#             name = safe_excel_text(getattr(doc, "prekes_pavadinimas", None) or "")
#             ws_lines.cell(row=line_idx, column=LineCols.NAME, value=name)

#             doc_obj = _s(getattr(doc, "objekto_kodas", None) or "")
#             objekto_kodas = doc_obj or user_obj
#             if objekto_kodas:
#                 ws_lines.cell(
#                     row=line_idx,
#                     column=LineCols.OBJECT_CODE,
#                     value=safe_excel_text(objekto_kodas),
#                 )

#             ws_headers.cell(
#                 row=header_row,
#                 column=HeaderCols.DEPARTMENT,
#                 value=safe_excel_text(padalinio_kodas),
#             )
#             if objekto_kodas:
#                 ws_headers.cell(
#                     row=header_row,
#                     column=HeaderCols.OBJECT,
#                     value=safe_excel_text(objekto_kodas),
#                 )

#             line_idx += 1

#     wb.save(output_path)
#     return Path(output_path)



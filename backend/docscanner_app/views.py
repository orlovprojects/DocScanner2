# --- Standard library ---
import hashlib
import io
import logging
import logging.config
import os
import uuid
import re
import tempfile
import zipfile, tarfile
import json
from datetime import date, datetime, timedelta, time as dt_time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import unicodedata
from django.http import HttpRequest
from django.contrib.auth import get_user_model
import openpyxl
from django.conf import settings as django_settings
from django.http import FileResponse
from django.db import models
from django.core.mail import EmailMultiAlternatives
from email.utils import formataddr
from openpyxl import Workbook
from rest_framework.exceptions import ValidationError

from django.core.files.base import ContentFile
from .tasks import process_uploaded_file_task 

from .tasks import start_session_processing, export_to_optimum_task, export_to_dineta_task, export_to_rivile_gama_api_task


from rest_framework.parsers import MultiPartParser, FormParser
from django.utils.dateparse import parse_date
from .serializers import ScannedDocumentListSerializer
from .pagination import DocumentsCursorPagination, UsersCursorPagination, MobileInboxCursorPagination, LineItemPagination


import hmac
from .utils.file_converter import SUPPORTED_EXTS
from django.views.decorators.http import require_POST

from .utils.invoice_pdf import save_invoice_pdf


# --- Django ---
from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.http import FileResponse, Http404, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.core.files.base import File



# --- Django REST Framework ---
from rest_framework import generics, permissions, status
from rest_framework.decorators import api_view, permission_classes, authentication_classes, parser_classes
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.views.decorators.csrf import csrf_exempt

# --- DRF SimpleJWT ---
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework_simplejwt.tokens import AccessToken

# --- Local (project) imports ---
from .data_import.data_import_from_buh import import_products_from_xlsx, import_clients_from_xlsx
from .exports.apskaita5 import export_documents_group_to_apskaita5_files
from .exports.centas import export_documents_group_to_centras_xml, generate_prekes_paslaugos_csv, generate_pradiniai_likuciai_csv
from .exports.finvalda import (
    export_pirkimai_group_to_finvalda,
    export_pardavimai_group_to_finvalda,
)
from .exports.agnum import (
    export_pirkimai_group_to_agnum,
    export_pardavimai_group_to_agnum,
)
from .exports.rivile import (
    export_clients_group_to_rivile,
    export_pirkimai_group_to_rivile,
    export_pardavimai_group_to_rivile,
    export_prekes_paslaugos_kodai_group_to_rivile,
)
from .exports.rivile_erp import (
    export_clients_to_rivile_erp_xlsx,
    export_prekes_and_paslaugos_to_rivile_erp_xlsx,
    export_documents_to_rivile_erp_xlsx,
    classify_isaf_for_erp,
)
from .exports.stekas import export_documents_group_to_stekas_files
from .exports.apsa import export_to_apsa
from .exports.dineta import dineta_hello, DinetaError
from .exports.optimum import optimum_hello, OptimumError
from .utils.password_encryption import decrypt_password
from .utils.password_encryption import encrypt_password
from .exports.rivile_gama_api import verify_api_key


from .exports.pragma4 import export_to_pragma40_xml
from .exports.pragma3 import export_to_pragma_full, save_pragma_export_to_files
from .exports.butent import export_to_butent
from .exports.site_pro import export_to_site_pro
from .exports.debetas import export_to_debetas
from .validators.required_fields_checker import check_required_fields_for_export
from .validators.math_validator_for_export import validate_document_math_for_export
from .exports.formatters import COUNTRY_NAME_LT


from .models import (
    CustomUser,
    ScannedDocument,
    ProductAutocomplete,
    ClientAutocomplete,
    LineItem,
    AdClick,
    MobileAccessKey,
    MobileInboxDocument,
    Payments,
    UploadSession,
    ChunkedUpload,
    MeasurementUnit,
    InvoiceSeries,
    Product,
    RecurringInvoice,
    Invoice,
    InvoiceEmail,
    InvoiceSettings,
    CreditUsageLog,
    InvSubscription,
    RivileGamaAPIKey,
    PaymentAllocation,
    CompanyProfile,
    Purchase,
    PurchaseLine,
    JournalEntry,
    JournalEntryLine,
    Company,
)

from .serializers import (
    CustomUserSerializer,
    ViewModeSerializer,
    ScannedDocumentSerializer,
    ScannedDocumentListSerializer,
    ScannedDocumentDetailSerializer,
    ScannedDocumentAdminDetailSerializer,
    AdClickSerializer,
    LineItemSerializer,
    CustomUserAdminListSerializer,
    DinetaSettingsSerializer,
    OptimumSettingsSerializer,
    MobileAccessKeySerializer,
    MobileInboxDocumentSerializer,
    PaymentSerializer,
    CounterpartySerializer,
    InvoiceSeriesSerializer,
    MeasurementUnitSerializer,
    ProductListSerializer,
    ProductSerializer,
    RecurringInvoiceListSerializer,
    RecurringInvoiceDetailSerializer,
    RecurringInvoiceWriteSerializer,
    RivileGamaAPIKeySerializer,
    RivileGamaAPIKeyCreateSerializer,
    RivileGamaAPIKeyUpdateSerializer,
    InvoiceAdminListSerializer,
    RecurringInvoiceAdminListSerializer,
    NewsletterCampaignCreateSerializer,
    NewsletterRecipientSerializer,
    NewsletterCampaignSerializer,
    CompanyProfileSerializer,
    PurchaseSerializer,
    PurchaseLineSerializer,
    JournalEntrySerializer,
    JournalEntryLineSerializer
)
from django.db.models import Prefetch
from django.db.models import Count, Sum

from typing import Any, Optional

from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

# <-- поправь пути импорта под свой проект
from .models import ScannedDocument
from .serializers import LineItemSerializer
from .pagination import LineItemPagination

from .utils.data_resolver import (
    ResolveContext,
    resolve_direction,
    _pvm_label,
    _nz,
    _normalize_vat_percent,
    _normalize_ps,
    _ps_to_bin,
    _need_geo,
)

from .tasks import process_uploaded_file_task
from .utils.data_resolver import build_preview
from .utils.pirkimas_pardavimas import determine_pirkimas_pardavimas
from .utils.prekes_kodas import assign_random_prekes_kodai
from .utils.save_document import _apply_sumiskai_defaults_from_user
from .utils.update_currency_rates import update_currency_rates
from .validators.vat_klas import auto_select_pvm_code

#dlia superuser dashboard
from django.db.models import Count
from .permissions import IsSuperUser, IsOwner

#wagtail imports
from rest_framework import viewsets, mixins
from .models import GuidePage, GuideCategoryPage
from rest_framework.decorators import action
from .serializers import (
    GuideCategoryListSerializer,
    GuideCategoryDetailSerializer,
    GuideArticleListSerializer,
    GuideArticleDetailSerializer,
)


#emails
from .emails import siusti_sveikinimo_laiska, siusti_kontakto_laiska
from .emails import siusti_masini_laiska_visiems
from .emails import siusti_mobilios_apps_kvietima
from .utils.play_store_link_gen import build_mobile_play_store_link

from time import perf_counter


# --- Logging setup ---
logging.config.dictConfig(settings.LOGGING)
logger = logging.getLogger('docscanner_app')
site_url = settings.SITE_URL_FRONTEND  # берём из settings.py


#admin dashboard
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
import json

# from .models import ScannedDocument, CustomUser
# from .permissions import IsSuperUser
# from .views import summarize_doc_issues  # если в том же файле — не нужно

def _today_dates():
    today = timezone.localdate()
    yesterday = today - timedelta(days=1)
    return today, yesterday

def _qs_by_date(model, date_field: str, target_date, exclude_archives=True):
    start = timezone.make_aware(datetime.combine(target_date, dt_time.min))
    end = timezone.make_aware(datetime.combine(target_date, dt_time.max))
    qs = model.objects.filter(**{f"{date_field}__range": (start, end)})
    if exclude_archives:
        qs = qs.filter(is_archive_container=False)
    return qs

def _qs_last_n_days(model, date_field: str, days: int, exclude_archives=True):
    since = timezone.now() - timedelta(days=days)
    qs = model.objects.filter(**{f"{date_field}__gte": since})
    if exclude_archives:
        qs = qs.filter(is_archive_container=False)
    return qs

def _qs_all_docs(exclude_archives=True):
    qs = ScannedDocument.objects.all()
    if exclude_archives:
        qs = qs.filter(is_archive_container=False)
    return qs

def _count_errors_in_qs(qs):
    """Ошибка = math_validation_passed=False ИЛИ ready_for_export=False"""
    return qs.filter(
        Q(math_validation_passed=False) | Q(ready_for_export=False)
    ).count()

def _count_rejected_in_qs(qs):
    """Количество rejected документов"""
    return qs.filter(status="rejected").count()

def _pct(part, whole):
    return round((part / whole * 100.0), 2) if whole else 0.0

def _rate(ok_count, total_count):
    """% успешных (без ошибок) от общего количества."""
    return _pct(ok_count, total_count)

def _rejected_stats(rejected, total):
    """Статистика rejected: count, total, percent"""
    return {
        "rejected": rejected,
        "total": total,
        "pct": _pct(rejected, total),
    }

def _payments_agg(qs):
    r = qs.aggregate(
        total_eur=Sum('amount_total'),
        net_eur=Sum('net_amount'),
        count=Count('id'),
    )
    return {
        "total_eur": round((r['total_eur'] or 0) / 100, 2),
        "net_eur":   round((r['net_eur'] or 0) / 100, 2),
        "count":     r['count'] or 0,
    }

def _invoice_count_by_date(date_val):
    """Считает выставленные счета за дату (status != draft)"""
    start = timezone.make_aware(datetime.combine(date_val, dt_time.min))
    end = timezone.make_aware(datetime.combine(date_val, dt_time.max))
    return Invoice.objects.filter(
        created_at__range=(start, end)
    ).exclude(status="draft").count()

def _invoice_count_last_n_days(n):
    since = timezone.now() - timedelta(days=n)
    return Invoice.objects.filter(created_at__gte=since).exclude(status="draft").count()

def _invoice_count_total():
    return Invoice.objects.exclude(status="draft").count()

def _email_stats_by_date(date_val):
    """Возвращает {'sent': X, 'failed': Y} за дату"""
    start = timezone.make_aware(datetime.combine(date_val, dt_time.min))
    end = timezone.make_aware(datetime.combine(date_val, dt_time.max))
    qs = InvoiceEmail.objects.filter(sent_at__range=(start, end))
    return {
        "sent": qs.filter(status="sent").count(),
        "failed": qs.filter(status__in=["failed", "bounced"]).count(),
    }

def _email_stats_last_n_days(n):
    since = timezone.now() - timedelta(days=n)
    qs = InvoiceEmail.objects.filter(sent_at__gte=since)
    return {
        "sent": qs.filter(status="sent").count(),
        "failed": qs.filter(status__in=["failed", "bounced"]).count(),
    }

def _email_stats_total():
    qs = InvoiceEmail.objects.all()
    return {
        "sent": qs.filter(status="sent").count(),
        "failed": qs.filter(status__in=["failed", "bounced"]).count(),
    }

def _inv_subscription_stats():
    """Статистика подписок Išrašymas"""
    now = timezone.now()
    
    # Активные триалы (status=trial И trial_end ещё не прошёл)
    trial_active = InvSubscription.objects.filter(
        status="trial",
        trial_end__gte=now
    ).count()
    
    # Завершённые триалы (trial_used=True И (status != trial ИЛИ trial_end < now))
    trial_expired = InvSubscription.objects.filter(
        trial_used=True
    ).exclude(
        status="trial",
        trial_end__gte=now
    ).count()
    
    # Платные подписки
    paid_monthly = InvSubscription.objects.filter(
        status="active",
        plan__icontains="monthly"
    ).count()
    
    paid_yearly = InvSubscription.objects.filter(
        status="active",
        plan__icontains="yearly"
    ).count()
    
    return {
        "trial_active": trial_active,
        "trial_expired": trial_expired,
        "paid_monthly": paid_monthly,
        "paid_yearly": paid_yearly,
    }

@api_view(["GET"])
@permission_classes([IsSuperUser])
def superuser_dashboard_stats(request):
    doc_date_field = "uploaded_at"
    user_date_field = "date_joined"

    today, yesterday = _today_dates()

    # Все QuerySet'ы исключают is_archive_container=True
    qs_all       = _qs_all_docs()
    qs_today     = _qs_by_date(ScannedDocument, doc_date_field, today)
    qs_yesterday = _qs_by_date(ScannedDocument, doc_date_field, yesterday)
    qs_7d        = _qs_last_n_days(ScannedDocument, doc_date_field, 7)
    qs_30d       = _qs_last_n_days(ScannedDocument, doc_date_field, 30)

    docs_today     = qs_today.count()
    docs_yesterday = qs_yesterday.count()
    docs_7d        = qs_7d.count()
    docs_30d       = qs_30d.count()
    total_docs     = qs_all.count()

    # Ошибки (math_validation_passed=False OR ready_for_export=False)
    err_today     = _count_errors_in_qs(qs_today)
    err_yesterday = _count_errors_in_qs(qs_yesterday)
    err_7d        = _count_errors_in_qs(qs_7d)
    err_30d       = _count_errors_in_qs(qs_30d)
    err_total     = _count_errors_in_qs(qs_all)

    ok_today     = max(docs_today - err_today, 0)
    ok_yesterday = max(docs_yesterday - err_yesterday, 0)
    ok_7d        = max(docs_7d - err_7d, 0)
    ok_30d       = max(docs_30d - err_30d, 0)
    ok_total     = max(total_docs - err_total, 0)

    # Rejected документы (status="rejected")
    rej_today     = _count_rejected_in_qs(qs_today)
    rej_yesterday = _count_rejected_in_qs(qs_yesterday)
    rej_7d        = _count_rejected_in_qs(qs_7d)
    rej_30d       = _count_rejected_in_qs(qs_30d)
    rej_total     = _count_rejected_in_qs(qs_all)

    # уникальные пользователи
    start_today = timezone.make_aware(datetime.combine(today, dt_time.min))
    end_today   = timezone.make_aware(datetime.combine(today, dt_time.max))
    unique_users_excl_1_2_today = (
        ScannedDocument.objects
        .filter(is_archive_container=False)
        .exclude(user_id__in=[1, 2])
        .filter(**{f"{doc_date_field}__range": (start_today, end_today)})
        .values("user_id").distinct().count()
    )

    # пользователи/регистрации
    new_users_today     = CustomUser.objects.filter(**{f"{user_date_field}__date": today}).count()
    new_users_yesterday = CustomUser.objects.filter(**{f"{user_date_field}__date": yesterday}).count()
    new_users_7d        = _qs_last_n_days(CustomUser, user_date_field, 7, exclude_archives=False).count()
    new_users_30d       = _qs_last_n_days(CustomUser, user_date_field, 30, exclude_archives=False).count()
    total_users         = CustomUser.objects.count()

    # разбивка по типам
    st_sumiskai = qs_all.filter(scan_type="sumiskai").count()
    st_detaliai = qs_all.filter(scan_type="detaliai").count()

    # Payments статистика
    pay_base = Payments.objects.filter(payment_status='paid')

    start_today = timezone.make_aware(datetime.combine(today, dt_time.min))
    end_today   = timezone.make_aware(datetime.combine(today, dt_time.max))
    start_yest  = timezone.make_aware(datetime.combine(yesterday, dt_time.min))
    end_yest    = timezone.make_aware(datetime.combine(yesterday, dt_time.max))

    week_start  = today - timedelta(days=today.weekday())  # Понедельник текущей недели
    month_start = today.replace(day=1)

    pay_today      = _payments_agg(pay_base.filter(paid_at__range=(start_today, end_today)))
    pay_yesterday  = _payments_agg(pay_base.filter(paid_at__range=(start_yest, end_yest)))
    pay_this_week  = _payments_agg(pay_base.filter(paid_at__gte=timezone.make_aware(datetime.combine(week_start, dt_time.min))))
    pay_this_month = _payments_agg(pay_base.filter(paid_at__gte=timezone.make_aware(datetime.combine(month_start, dt_time.min))))
    pay_30d        = _payments_agg(pay_base.filter(paid_at__gte=timezone.now() - timedelta(days=30)))
    pay_total      = _payments_agg(pay_base)

    # ========== Išrašymas stats ==========
    inv_today = _invoice_count_by_date(today)
    inv_yesterday = _invoice_count_by_date(yesterday)
    inv_7d = _invoice_count_last_n_days(7)
    inv_30d = _invoice_count_last_n_days(30)
    inv_total = _invoice_count_total()

    inv_subs = _inv_subscription_stats()

    email_today = _email_stats_by_date(today)
    email_yesterday = _email_stats_by_date(yesterday)
    email_7d = _email_stats_last_n_days(7)
    email_30d = _email_stats_last_n_days(30)
    email_total = _email_stats_total()

    # ========== Važtaraščiai stats ==========
    wb_qs_all = ScannedWaybill.objects.filter(
        is_archive_container=False,
        is_multi_doc_container=False,
    ).exclude(user_id__in=[1, 2, 31, 105])
    wb_qs_today = wb_qs_all.filter(uploaded_at__date=today)
    wb_qs_yesterday = wb_qs_all.filter(uploaded_at__date=yesterday)
    wb_qs_7d = wb_qs_all.filter(uploaded_at__gte=timezone.now() - timedelta(days=7))
    wb_qs_30d = wb_qs_all.filter(uploaded_at__gte=timezone.now() - timedelta(days=30))

    def _wb_stats(qs):
        total = qs.count()
        ok = qs.filter(status__in=("completed", "exported")).count()
        rej = qs.filter(status="rejected").count()
        return {"total": total, "ok": ok, "rejected": rej}

    data = {
        "documents": {
            "today":       {"count": docs_today,     "errors": err_today},
            "yesterday":   {"count": docs_yesterday, "errors": err_yesterday},
            "last_7_days": {"count": docs_7d,        "errors": err_7d},
            "last_30_days":{"count": docs_30d,       "errors": err_30d},
            "total":       {"count": total_docs,     "errors": err_total},

            "success_rate": {
                "today":       _rate(ok_today,     docs_today),
                "yesterday":   _rate(ok_yesterday, docs_yesterday),
                "last_7_days": _rate(ok_7d,        docs_7d),
                "last_30_days":_rate(ok_30d,       docs_30d),
                "total":       _rate(ok_total,     total_docs),
            },

            # ✅ Новый блок — Rejected статистика
            "rejected": {
                "today":       _rejected_stats(rej_today,     docs_today),
                "yesterday":   _rejected_stats(rej_yesterday, docs_yesterday),
                "last_7_days": _rejected_stats(rej_7d,        docs_7d),
                "last_30_days":_rejected_stats(rej_30d,       docs_30d),
                "total":       _rejected_stats(rej_total,     total_docs),
            },

            "unique_users_excluding_1_2_today": unique_users_excl_1_2_today,
            "scan_types": {
                "sumiskai": {"count": st_sumiskai, "pct": _pct(st_sumiskai, total_docs)},
                "detaliai": {"count": st_detaliai, "pct": _pct(st_detaliai, total_docs)},
            },
        },
        "vaztarasciai": {
            "today": _wb_stats(wb_qs_today),
            "yesterday": _wb_stats(wb_qs_yesterday),
            "last_7_days": _wb_stats(wb_qs_7d),
            "last_30_days": _wb_stats(wb_qs_30d),
            "total": _wb_stats(wb_qs_all),
        },
        "users": {
            "new_today":        new_users_today,
            "new_yesterday":    new_users_yesterday,
            "new_last_7_days":  new_users_7d,
            "new_last_30_days": new_users_30d,
            "total":            total_users,
        },
        "payments": {
            "today":       pay_today,
            "yesterday":   pay_yesterday,
            "this_week":   pay_this_week,
            "this_month":  pay_this_month,
            "last_30_days":pay_30d,
            "total":       pay_total,
        },
        "israsymas": {
            "invoices": {
                "today": inv_today,
                "yesterday": inv_yesterday,
                "last_7_days": inv_7d,
                "last_30_days": inv_30d,
                "total": inv_total,
            },
            "subscriptions": inv_subs,
            "emails": {
                "today": email_today,
                "yesterday": email_yesterday,
                "last_7_days": email_7d,
                "last_30_days": email_30d,
                "total": email_total,
            },
        },
        "meta": {
            "timezone": str(timezone.get_current_timezone()),
            "generated_at": timezone.now().isoformat(),
        },
    }
    return Response(data)








def strip_diacritics(text):
    """
    Pakeičia visas lietuviškas ir kitas lotyniškas raides su diakritika
    į paprastas: š->s, ą->a, Ž->Z ir t.t.
    """
    if not isinstance(text, str):
        return text
    normalized = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_documents(request):
    from datetime import date
    import io
    import zipfile
    import tempfile
    import logging

    logger = logging.getLogger(__name__)
    log_ctx = {"user": getattr(request.user, "id", None)}

    # ---- входные параметры
    ids = request.data.get('ids', [])
    export_type = request.data.get('export_type') or getattr(request.user, 'default_accounting_program', 'centas')
    raw_overrides = request.data.get('overrides', {}) or {}
    mode_raw = (request.data.get('mode') or "").strip().lower()  # <<< NEW

    scope = (request.data.get("scope") or "").strip().lower()
    filters = request.data.get("filters") or {}
    cp_key = (request.data.get("cp_key") or "").strip()

    excluded_ids = request.data.get("excluded_ids") or []
    excluded_ids = [int(x) for x in excluded_ids if str(x).isdigit()]

    source = (request.data.get("source") or "scanned").strip().lower()
    user = request.user                          
    export_type = str(export_type).lower() 
    today_str = date.today().strftime('%Y-%m-%d')  
    inv_sub = None
    inv_usage = None    

    logger.info("[EXP] start user=%s export_type_raw=%r ids=%s raw_overrides=%r mode_raw=%r",
                log_ctx["user"], export_type, ids, raw_overrides, mode_raw)
    
    if source == "invoice":
        from docscanner_app.models import Invoice
        from docscanner_app.utils.invoice_export_adapter import adapt_invoices_for_export

        if not ids:
            return Response({"error": "No invoice ids provided"}, status=400)

        invoices = Invoice.objects.filter(
            pk__in=ids, user=user
        ).prefetch_related("line_items")

        if not invoices.exists():
            return Response({"error": "No invoices found"}, status=404)

        # --- Block isankstine ---
        blocked = invoices.filter(invoice_type="isankstine")
        if blocked.exists():
            return Response({
                "error": "Išankstinės sąskaitos negali būti eksportuojamos. "
                         "Konvertuokite į SF arba PVM SF."
            }, status=400)

        # --- Only issued/sent/paid ---
        not_ready = invoices.exclude(status__in=["issued", "sent", "paid"])
        if not_ready.exists():
            return Response(
                {"error": "Galima eksportuoti tik israsytas saskaitas."},
                status=400,
            )

        # --- Inv subscription: export limit check ---
        inv_sub = None
        inv_usage = None
        try:
            from .models import InvSubscription, InvMonthlyUsage
            inv_sub = InvSubscription.objects.filter(user=user).first()
            if inv_sub:
                inv_sub.check_and_expire()
                if inv_sub.status == "free":
                    inv_usage = InvMonthlyUsage.get_current(user)
                    new_ids = [inv.pk for inv in invoices if inv.pk not in inv_usage.exported_invoice_ids]
                    slots_left = 10 - inv_usage.exports_used
                    if len(new_ids) > slots_left:
                        return Response({
                            "error": "limit_reached",
                            "feature": "export",
                            "message": (
                                f"Mėnesio eksporto limitas: {inv_usage.exports_used}/10 panaudota. "
                                f"Bandote eksportuoti {len(new_ids)} naujų sąskaitų, "
                                f"bet liko tik {max(0, slots_left)} eksportų."
                            ),
                            "exports_used": inv_usage.exports_used,
                            "exports_max": 10,
                        }, status=403)
        except Exception as e:
            logger.warning("[EXP] inv subscription check failed: %s", e)

        # --- Adapt for exporters ---
        adapted = adapt_invoices_for_export(invoices, user=user)

        pirkimai_docs = []
        pardavimai_docs = adapted
        unknown_docs = []

        documents = pardavimai_docs

        cp_key = "__israsymas__"

        logger.info(
            "[EXP] INVOICE source: %d invoices for export_type=%s",
            len(adapted), export_type,
        )

    else:
    
        if scope == "filtered":
            # multi требует выбранного контрагента
            if mode_raw == "multi" and not cp_key:
                return Response({"error": "Choose counterparty (cp_key) for multi export"}, status=400)

            q = filters or {}
            status_param = (q.get("status") or "").strip()
            date_from = (q.get("from") or "").strip()
            date_to = (q.get("to") or "").strip()
            search = (q.get("search") or "").strip()

            qs = ScannedDocument.objects.filter(user=request.user)

            # --- те же фильтры, что в get_user_documents ---
            if status_param:
                qs = qs.filter(status=status_param)

            tz = timezone.get_current_timezone()

            if date_from:
                d = parse_date(date_from)
                if d:
                    dt_from = timezone.make_aware(datetime.combine(d, dt_time.min), tz)
                    qs = qs.filter(uploaded_at__gte=dt_from)

            if date_to:
                d = parse_date(date_to)
                if d:
                    dt_to = timezone.make_aware(datetime.combine(d, dt_time.min), tz) + timedelta(days=1)
                    qs = qs.filter(uploaded_at__lt=dt_to)

            if search:
                qs = qs.filter(document_number__icontains=search)

            # --- фильтр по контрагенту (как у тебя в /documents/) ---
            if cp_key:
                cp = cp_key.strip().lower()
                if cp.startswith("id:"):
                    cp_id = cp.split("id:", 1)[1].strip()
                    if cp_id:
                        qs = qs.filter(Q(seller_id=cp_id) | Q(buyer_id=cp_id))
                else:
                    qs = qs.filter(
                        Q(seller_vat_code__iexact=cp) |
                        Q(buyer_vat_code__iexact=cp) |
                        Q(seller_name__icontains=cp) |
                        Q(buyer_name__icontains=cp)
                    )

            # --- ВАЖНО: экспортируем только "не серые" как в таблице ---
            qs = qs.filter(
                status__in=["completed", "exported"],
                ready_for_export=True,
                math_validation_passed=True,
            )

            ids = list(qs.values_list("id", flat=True))

            if excluded_ids:
                ids = [i for i in ids if i not in set(excluded_ids)]


            if not ids:
                logger.warning("[EXP] no ids provided")
                return Response({"error": "No document ids provided"}, status=400)

        # user = request.user
        # export_type = str(export_type).lower()

        # Rivilė: ar reikia nuimti lietuviškas raides (š->s ir t.t.)
        extra_settings = getattr(user, "extra_settings", {}) or {}
        rivile_strip_lt = bool(extra_settings.get("rivile_strip_lt_letters"))
        logger.info("[EXP] user extra_settings: rivile_strip_lt_letters=%s", rivile_strip_lt)

        # --- нормализация overrides (id -> 'pirkimas'|'pardavimas')
        overrides = {}
        for k, v in raw_overrides.items():
            key = str(k)
            val = str(v).lower()
            if val in ('pirkimas', 'pardavimas'):
                overrides[key] = val
            else:
                logger.warning("[EXP] skip override key=%r val=%r (invalid)", key, v)

        # --- определить mode: берём из клиента, иначе как раньше (по overrides)
        if mode_raw in ("multi", "single"):                       # <<< NEW
            mode = mode_raw
            logger.info("[EXP] view mode taken from request: %s", mode)
        else:
            mode = 'multi' if overrides else 'single'
            logger.info("[EXP] view mode inferred for backward-compat: %s", mode)

        # Доп. диагностика: если пришёл multi, но overrides пустой
        if mode == "multi" and not overrides:
            logger.info("[EXP] mode is 'multi' but overrides are EMPTY (will rely on resolver/doc DB fields)")

        logger.info("[EXP] export_type=%s overrides_norm=%r", export_type, overrides)

        # today_str = date.today().strftime('%Y-%m-%d')

        documents = ScannedDocument.objects.filter(pk__in=ids, user=user).prefetch_related('line_items')
        # documents = ScannedDocument.objects.filter(pk__in=ids, user=user)
        if not documents:
            logger.warning("[EXP] no documents found by ids=%s user=%s", ids, log_ctx["user"])
            return Response({"error": "No documents found"}, status=404)

        # === резолвер ===
        from .utils.data_resolver import prepare_export_groups
        logger.info("[EXP] resolver_mode=%s", mode)

        try:
            prepared = prepare_export_groups(
                documents,
                user=user,
                overrides=overrides if mode == "multi" else {},
                view_mode=mode,
                cp_key=cp_key if mode == "multi" else None,   
            )
        except Exception as e:
            logger.exception("[EXP] prepare_export_groups failed: %s", e)
            return Response({"error": "Resolver failed", "detail": str(e)}, status=500)

        # быстрый дамп того, что пришло из резолвера
        def _debug_dump(prepared_obj, where):
            for bucket in ("pirkimai", "pardavimai", "unknown"):
                packs = prepared_obj.get(bucket) or []
                logger.info("[EXPDBG:%s] bucket=%s count=%d", where, bucket, len(packs))
                for p in packs:
                    d = p.get("doc")
                    dpk = getattr(d, "pk", None)
                    li = p.get("line_items") or []
                    logger.info(
                        "[EXPDBG:%s] bucket=%s doc=%s dir=%r pack_keys=%s pvm=%r lines=%d",
                        where, bucket, dpk, p.get("direction"), list(p.keys()),
                        p.get("pvm_kodas", None), len(li)
                    )
                    if li:
                        preview = [(x.get("id"), x.get("pvm_kodas")) for x in li[:3]]
                        logger.info("[EXPDBG:%s] doc=%s sample_line_items=%s", where, dpk, preview)

        _debug_dump(prepared, "after_resolver")

        # применяем «в память» (без сохранения в БД)
        def _apply_resolved(pack_list, tag):
            out_docs = []
            for pack in pack_list:
                d = pack["doc"]
                setattr(d, "pirkimas_pardavimas", pack.get("direction"))
                setattr(d, "pvm_kodas", pack.get("pvm_kodas", None))  # явное перетирание

                line_map = {}
                for li in (pack.get("line_items") or []):
                    li_id = li.get("id")
                    if li_id is not None:
                        line_map[li_id] = li.get("pvm_kodas")
                setattr(d, "_pvm_line_map", line_map)

                logger.info("[EXPDBG:apply] tag=%s doc=%s dir=%r pvm_kodas=%r line_map_size=%d",
                            tag, getattr(d, "pk", None), getattr(d, "pirkimas_pardavimas", None),
                            getattr(d, "pvm_kodas", None), len(line_map))
                out_docs.append(d)
            return out_docs

        pirkimai_docs   = _apply_resolved(prepared.get("pirkimai", []), "pirkimai")
        pardavimai_docs = _apply_resolved(prepared.get("pardavimai", []), "pardavimai")
        unknown_docs    = _apply_resolved(prepared.get("unknown", []), "unknown")

        logger.info("[EXP] ready_for_export counts: pirkimai=%d pardavimai=%d unknown=%d",
                    len(pirkimai_docs), len(pardavimai_docs), len(unknown_docs))

    if source == "invoice":
        prepared = {
            "pirkimai": [],
            "pardavimai": [
                {"doc": d, "direction": "pardavimas"}
                for d in pardavimai_docs
            ],
            "unknown": [],
        }

    # --- переменные для универсального финализатора
    response = None
    export_success = False
    exported_ids = [d.pk for d in (pirkimai_docs + pardavimai_docs)]

    # общий контейнер (внутри веток можно переопределять/очищать)
    files_to_export = []

    # ========================= CENTAS =========================
    if export_type == 'centas':
        logger.info("[EXP] CENTAS export started")
        assign_random_prekes_kodai(documents)

        if pirkimai_docs:
            logger.info("[EXP] CENTAS exporting pirkimai: %d docs", len(pirkimai_docs))
            xml_bytes = export_documents_group_to_centras_xml(
                pirkimai_docs,
                direction="pirkimas",
                user=request.user,
                own_company_code=cp_key,
            )
            files_to_export.append((f"{today_str}_pirkimai.xml", xml_bytes))

        if pardavimai_docs:
            logger.info("[EXP] CENTAS exporting pardavimai: %d docs", len(pardavimai_docs))
            xml_bytes = export_documents_group_to_centras_xml(
                pardavimai_docs,
                direction="pardavimas",
                user=request.user,
                own_company_code=cp_key,
            )
            files_to_export.append((f"{today_str}_pardavimai.xml", xml_bytes))

        # ── prekės/paslaugos CSV ──
        csv_bytes = generate_prekes_paslaugos_csv(
            pirkimai_docs=pirkimai_docs or [],
            pardavimai_docs=pardavimai_docs or [],
            user=request.user,
            own_company_code=cp_key,
        )
        if csv_bytes:
            files_to_export.append(("prekes_paslaugos.csv", csv_bytes))

        # ── pradiniai likučiai CSV (только для pardavimai prekės) ──
        if pardavimai_docs:
            likuciai_bytes = generate_pradiniai_likuciai_csv(
                pardavimai_docs=pardavimai_docs,
                user=request.user,
                own_company_code=cp_key,
            )
            if likuciai_bytes:
                files_to_export.append(("pradiniai_likuciai.csv", likuciai_bytes))

        logger.info("[EXP] CENTAS files_to_export=%s", [n for n, _ in files_to_export])

        if len(files_to_export) > 1:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for filename, content in files_to_export:
                    zf.writestr(filename, content)
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename={today_str}_importui.zip'
            export_success = True
        elif len(files_to_export) == 1:
            filename, content = files_to_export[0]
            ct = 'text/csv; charset=windows-1257' if filename.endswith('.csv') else 'application/xml; charset=utf-8'
            response = HttpResponse(content, content_type=ct)
            response['Content-Disposition'] = f'attachment; filename={filename}'
            export_success = True
        else:
            logger.warning("[EXP] CENTAS nothing to export")
            response = Response({"error": "No documents to export"}, status=400)

    # ========================= RIVILĖ (EIP) =========================

    elif export_type == 'rivile':
        logger.info("[EXP] RIVILE export started")
        assign_random_prekes_kodai(documents)

        files_to_export = []

        # 1) Клиенты (N08+N33): собираем ИЗ ДОКУМЕНТОВ; кэш больше не нужен
        docs_for_clients = (pirkimai_docs or []) + (pardavimai_docs or [])
        if docs_for_clients:
            klientai_xml = export_clients_group_to_rivile(
                clients=None,
                documents=docs_for_clients,
            )
            if klientai_xml and klientai_xml.strip():
                files_to_export.append(('klientai.eip', klientai_xml))
                logger.info("[EXP] RIVILE clients exported")

        # 2) ПИРКИМАИ (I06/I07)
        if pirkimai_docs:
            logger.info("[EXP] RIVILE exporting pirkimai: %d docs", len(pirkimai_docs))
            pirkimai_xml = export_pirkimai_group_to_rivile(pirkimai_docs, request.user, own_company_code=cp_key)
            files_to_export.append(('pirkimai.eip', pirkimai_xml))

        # 3) ПАРДАВИМАИ (I06/I07)
        if pardavimai_docs:
            logger.info("[EXP] RIVILE exporting pardavimai: %d docs", len(pardavimai_docs))
            pardavimai_xml = export_pardavimai_group_to_rivile(pardavimai_docs, request.user, own_company_code=cp_key)
            files_to_export.append(('pardavimai.eip', pardavimai_xml))

        # 4) N17/N25 - ИЗМЕНЕНО: передаём request.user
        prekes_xml, paslaugos_xml, kodai_xml = export_prekes_paslaugos_kodai_group_to_rivile(
            documents, 
            request.user,
            own_company_code=cp_key, 
        )
        if prekes_xml and prekes_xml.strip():
            files_to_export.append(('prekes.eip', prekes_xml))
        if paslaugos_xml and paslaugos_xml.strip():
            files_to_export.append(('paslaugos.eip', paslaugos_xml))
        if kodai_xml and kodai_xml.strip():
            files_to_export.append(('kodai.eip', kodai_xml))

        logger.info("[EXP] RIVILE files_to_export=%s", [n for n, _ in files_to_export])

        if files_to_export:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for filename, xml_content in files_to_export:
                    zf.writestr(filename, xml_content)
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename={today_str}_rivile_eip.zip'
            export_success = True
        else:
            logger.warning("[EXP] RIVILE nothing to export")
            response = Response({"error": "No documents to export"}, status=400)

    # ========================= RIVILE GAMA API =========================

    elif export_type == "rivile_gama_api":
        from .models import ExportSession
        from .utils.api_key_resolver import resolve_api_key

        assign_random_prekes_kodai(documents)

        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])
        if not all_docs:
            logger.warning("[EXP] RIVILE_GAMA_API no documents to export")
            return Response({"error": "No documents to export"}, status=400)

        # --- Определяем company_code для поиска API ключа ---
        if source == "invoice":
            own_company_code = "__israsymas__"
        elif cp_key:
            cp = cp_key.strip()
            own_company_code = cp.split(":", 1)[1].strip() if cp.lower().startswith("id:") else cp
        else:
            own_company_code = str(getattr(user, "company_code", "") or "").strip()

        logger.info(
            "[EXP] RIVILE_GAMA_API source=%s cp_key=%r own_company=%s",
            source, cp_key, own_company_code,
        )

        # --- Ищем API ключ ---
        if source == "invoice":
            api_key_obj = resolve_api_key(user, "rivile_gama_api", own_company_code, strict=True)
        else:
            api_key_obj = resolve_api_key(user, "rivile_gama_api", own_company_code)

        if not api_key_obj:
            return JsonResponse(
                {"error": f"Rivile GAMA API raktas nerastas įmonei {own_company_code}. "
                          "Pridėkite raktą Nustatymuose."},
                status=400,
            )

        doc_ids = [d.pk for d in all_docs]

        # --- ExportSession + Celery ---
        session_obj = ExportSession.objects.create(
            user=user,
            program="rivile_gama_api",
            stage=ExportSession.Stage.QUEUED,
            total_documents=len(doc_ids),
        )
        if source == "invoice":
            session_obj.invoice_documents.set(doc_ids)
        else:
            session_obj.documents.set(doc_ids)

        task = export_to_rivile_gama_api_task.delay(
            session_obj.id,
            api_key_obj.pk,
            own_company_code,
        )
        session_obj.task_id = task.id
        session_obj.save(update_fields=["task_id"])

        logger.info(
            "[EXP] RIVILE_GAMA_API session=%s task=%s docs=%d company=%s key=%s",
            session_obj.pk, task.id, len(doc_ids), own_company_code, api_key_obj.company_code,
        )

        return Response({
            "status": "ok",
            "session_id": session_obj.pk,
            "total_documents": len(doc_ids),
            "message": "Export started",
        }, status=202)

    # ========================= FINVALDA =========================
    elif export_type == 'finvalda':
        logger.info("[EXP] FINVALDA export started")
        assign_random_prekes_kodai(documents)

        files_to_export = []

        if pirkimai_docs:
            logger.info("[EXP] FINVALDA exporting pirkimai: %d docs", len(pirkimai_docs))
            pirk_files = export_pirkimai_group_to_finvalda(pirkimai_docs, user=request.user, own_company_code=cp_key)
            if pirk_files.get("pirkimai"):
                files_to_export.append((f"{today_str}_pirkimai_finvalda.xml", pirk_files["pirkimai"]))
            if pirk_files.get("pirkimu_grazinimai"):
                files_to_export.append((f"{today_str}_pirkimu_grazinimai_finvalda.xml", pirk_files["pirkimu_grazinimai"]))
        if pardavimai_docs:
            logger.info("[EXP] FINVALDA exporting pardavimai: %d docs", len(pardavimai_docs))
            pard_files = export_pardavimai_group_to_finvalda(pardavimai_docs, user=request.user, own_company_code=cp_key)
            if pard_files.get("pardavimai"):
                files_to_export.append((f"{today_str}_pardavimai_finvalda.xml", pard_files["pardavimai"]))
            if pard_files.get("pardavimo_grazinimai"):
                files_to_export.append((f"{today_str}_pardavimo_grazinimai_finvalda.xml", pard_files["pardavimo_grazinimai"]))

        logger.info("[EXP] FINVALDA files_to_export=%s", [n for n, _ in files_to_export])

        if len(files_to_export) > 1:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for filename, xml_content in files_to_export:
                    zf.writestr(filename, xml_content)
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename={today_str}_finvalda.zip'
            export_success = True
        elif len(files_to_export) == 1:
            filename, xml_content = files_to_export[0]
            response = HttpResponse(xml_content, content_type='application/xml')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            export_success = True
        else:
            logger.warning("[EXP] FINVALDA nothing to export")
            response = Response({"error": "No documents to export"}, status=400)



    # ========================= PRAGMA 3.2 =========================
    elif export_type == 'pragma3':
        logger.info("[EXP] PRAGMA32 export started")
        assign_random_prekes_kodai(documents)

        # Используем уже подготовленные документы с атрибутами от _apply_resolved
        # (pirkimas_pardavimas, pvm_kodas, _pvm_line_map)
        # ВАЖНО: добавь .prefetch_related('line_items') в начале функции где documents = ...
        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])

        files_to_export = []

        try:
            # Полный экспорт (4 или 6 файлов)
            export_data = export_to_pragma_full(
                documents=all_docs,
                user=request.user,
                include_reference_data=True,
                own_company_code=cp_key, 
            )
            
            logger.info("[EXP] PRAGMA32 export_data keys: %s", list(export_data.keys()))

            # Pirkimai (если есть)
            if export_data.get('pirkimai'):
                files_to_export.append((
                    'Pirkimai.txt',
                    export_data['pirkimai']
                ))

            if export_data.get('pirkimai_det'):
                files_to_export.append((
                    'Pirkimai_Prekes.txt',
                    export_data['pirkimai_det']
                ))

            # Pardavimai (если есть)
            if export_data.get('pardavimai'):
                files_to_export.append((
                    'PARDAVIMAI.txt',
                    export_data['pardavimai']
                ))

            if export_data.get('pardavimai_det'):
                files_to_export.append((
                    'PARDAVIMAI_PREKES.txt',
                    export_data['pardavimai_det']
                ))

            # Справочники (общие)
            if export_data.get('companies'):
                files_to_export.append((
                    'Imones.txt',
                    export_data['companies']
                ))

            if export_data.get('products'):
                files_to_export.append((
                    'Prekes.txt',
                    export_data['products']
                ))

            logger.info("[EXP] PRAGMA32 files_to_export=%s", [n for n, _ in files_to_export])

        except Exception as e:
            logger.exception("[EXP] PRAGMA32 export failed: %s", e)
            return Response({"error": "Pragma 3.2 export failed", "detail": str(e)}, status=500)

        # Формирование ответа
        if len(files_to_export) > 1:
            # Несколько файлов -> ZIP
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for filename, txt_content in files_to_export:
                    zf.writestr(filename, txt_content)
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename={today_str}_pragma32.zip'
            export_success = True
            
        elif len(files_to_export) == 1:
            # Один файл -> прямая отдача
            filename, txt_content = files_to_export[0]
            response = HttpResponse(
                txt_content,
                content_type='text/plain; charset=windows-1257'
            )
            response['Content-Disposition'] = f'attachment; filename={filename}'
            export_success = True
            
        else:
            logger.warning("[EXP] PRAGMA32 nothing to export")
            response = Response({"error": "No documents to export"}, status=400)


    # ========================= PRAGMA 4.0 =========================
    elif export_type == 'pragma4':
        logger.info("[EXP] PRAGMA40 export started")
        assign_random_prekes_kodai(documents)

        from types import SimpleNamespace

        if not cp_key:
            return Response(
                {"error": "Counterparty (CP) is required for Pragma 4.0 export. Select a counterparty."},
                status=400,
            )

        # --- Парсим cp_key и ищем CP в документах ---
        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])
        if not all_docs:
            all_docs = list(documents)

        cp_raw = cp_key.strip()
        counterparty = None

        def _build_cp(doc, prefix):
            """Собираем SimpleNamespace из полей документа с указанным prefix."""
            return SimpleNamespace(
                company_code=getattr(doc, f'{prefix}id', '') or '',
                name=getattr(doc, f'{prefix}name', '') or '',
                vat_code=getattr(doc, f'{prefix}vat_code', '') or '',
                email=getattr(doc, f'{prefix}email', '') or '',
                address=getattr(doc, f'{prefix}address', '') or '',
                city=getattr(doc, f'{prefix}city', '') or '',
                country=getattr(doc, f'{prefix}country', '') or '',
                country_iso=getattr(doc, f'{prefix}country_iso', '') or '',
                post_code=getattr(doc, f'{prefix}post_code', '') or '',
                iban=getattr(doc, f'{prefix}iban', '') or '',
            )

        if cp_raw.lower().startswith("id:"):
            # Формат 1: "id:304401940" → ищем по seller_id / buyer_id
            code = cp_raw.split(":", 1)[1].strip()
            for doc in all_docs:
                if str(getattr(doc, 'seller_id', '') or '') == code:
                    counterparty = _build_cp(doc, 'seller_')
                    break
                if str(getattr(doc, 'buyer_id', '') or '') == code:
                    counterparty = _build_cp(doc, 'buyer_')
                    break
        else:
            # Формат 2 или 3: VAT code или имя (lowercase)
            cp_lower = cp_raw.lower()
            for doc in all_docs:
                for prefix in ('seller_', 'buyer_'):
                    vat = (str(getattr(doc, f'{prefix}vat_code', '') or '')).strip().lower()
                    name = (str(getattr(doc, f'{prefix}name', '') or '')).strip().lower()
                    if vat and vat == cp_lower:
                        counterparty = _build_cp(doc, prefix)
                        break
                    if name and name == cp_lower:
                        counterparty = _build_cp(doc, prefix)
                        break
                if counterparty:
                    break

        if counterparty is None:
            # Последний fallback — минимальный CP
            if cp_raw.lower().startswith("id:"):
                counterparty = SimpleNamespace(
                    company_code=cp_raw.split(":", 1)[1].strip(),
                    name='', vat_code='', email='', address='',
                    city='', country='', country_iso='', post_code='', iban='',
                )
            else:
                counterparty = SimpleNamespace(
                    company_code='', name=cp_raw, vat_code=cp_raw if cp_raw[:2].isalpha() else '',
                    email='', address='', city='', country='', country_iso='',
                    post_code='', iban='',
                )
            logger.warning("[EXP] PRAGMA40 CP not found in docs, fallback cp_key=%s", cp_key)

        logger.info("[EXP] PRAGMA40 CP: code=%s name=%s vat=%s",
                     counterparty.company_code, counterparty.name, counterparty.vat_code)

        try:
            result = export_to_pragma40_xml(
                documents=all_docs,
                counterparty=counterparty,
                user=request.user,
                own_company_code=cp_key,
            )
        except Exception as e:
            logger.exception("[EXP] PRAGMA40 export failed: %s", e)
            return Response({"error": "Pragma 4.0 export failed", "detail": str(e)}, status=500)

        if not result:
            logger.warning("[EXP] PRAGMA40 nothing to export")
            return Response({"error": "No documents to export"}, status=400)

        if len(result) == 1:
            doc_type_key, xml_bytes = list(result.items())[0]
            filename = f"{today_str}_pragma40_{doc_type_key}.xml"
            response = HttpResponse(xml_bytes, content_type="application/xml; charset=utf-8")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            export_success = True
        else:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for doc_type_key, xml_bytes in result.items():
                    fname = f"{today_str}_pragma40_{doc_type_key}.xml"
                    zip_file.writestr(fname, xml_bytes)
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{today_str}_pragma40.zip"'
            export_success = True



    # ========================= Butent =========================
    elif export_type == 'butent':
        logger.info("[EXP] BUTENT export started")
        assign_random_prekes_kodai(documents)

        # Объединяем все документы для экспорта (Būtent поддерживает смешивание)
        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])
        
        if not all_docs:
            logger.warning("[EXP] BUTENT no documents to export")
            return Response({"error": "No documents to export"}, status=400)

        try:
            # Экспортируем в Excel (mode='auto' возвращает Dict[str, bytes])
            result = export_to_butent(
                documents=all_docs,
                mode='auto',
                user=request.user,
                own_company_code=cp_key,
            )
            
            logger.info("[EXP] BUTENT export completed, files=%s", list(result.keys()))
            
            # Если один файл - отдаем его напрямую
            if len(result) == 1:
                mode, excel_bytes = list(result.items())[0]
                filename = f'{today_str}_butent_{mode}_import.xlsx'
                
                logger.info("[EXP] BUTENT single file: %s, size=%d bytes", filename, len(excel_bytes))
                
                response = HttpResponse(
                    excel_bytes,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                export_success = True
            
            # Если два файла - создаем ZIP архив
            else:
                import zipfile
                from io import BytesIO
                
                zip_buffer = BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for mode, excel_bytes in result.items():
                        filename = f'{today_str}_butent_{mode}_import.xlsx'
                        zip_file.writestr(filename, excel_bytes)
                        logger.info("[EXP] BUTENT added to ZIP: %s, size=%d bytes", filename, len(excel_bytes))
                
                zip_buffer.seek(0)
                zip_bytes = zip_buffer.read()
                
                logger.info("[EXP] BUTENT ZIP created, size=%d bytes", len(zip_bytes))
                
                response = HttpResponse(zip_bytes, content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="{today_str}_butent_import.zip"'
                export_success = True

        except FileNotFoundError as e:
            logger.error("[EXP] BUTENT template not found: %s", e)
            return Response({
                "error": "Būtent template not found",
                "detail": "Please create template using create_butent_template()"
            }, status=500)
        
        except Exception as e:
            logger.exception("[EXP] BUTENT export failed: %s", e)
            return Response({
                "error": "Būtent export failed",
                "detail": str(e)
            }, status=500)
        

    # ========================= SITE.PRO (B1) =========================
    elif export_type == 'site_pro':
        logger.info("[EXP] SITE.PRO(B1) export started")
        assign_random_prekes_kodai(documents)

        # Экспортируем только уже классифицированные документы (pirkimai + pardavimai)
        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])

        if not all_docs:
            logger.warning("[EXP] SITE.PRO(B1) no documents to export (no pirkimai/pardavimai)")
            return Response({"error": "No documents to export"}, status=400)

        try:
            # result: {"clients": bytes, "items": bytes, "purchases": bytes, "sales": bytes}
            result = export_to_site_pro(all_docs, user=request.user, own_company_code=cp_key)
            logger.info("[EXP] SITE.PRO(B1) export completed, keys=%s", list(result.keys()))

            files_to_export = []

            if result.get("clients"):
                files_to_export.append((f"{today_str}_site_pro_klientai.xlsx", result["clients"]))
            if result.get("items"):
                files_to_export.append((f"{today_str}_site_pro_prekes_paslaugos.xlsx", result["items"]))
            if result.get("purchases"):
                files_to_export.append((f"{today_str}_site_pro_pirkimai.xlsx", result["purchases"]))
            if result.get("sales"):
                files_to_export.append((f"{today_str}_site_pro_pardavimai.xlsx", result["sales"]))

            if not files_to_export:
                logger.warning("[EXP] SITE.PRO(B1) nothing to export (empty bytes)")
                return Response({"error": "No documents to export"}, status=400)

            # B1 обычно = 4 файла -> ZIP
            if len(files_to_export) > 1:
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    for filename, content in files_to_export:
                        zf.writestr(filename, content)
                        logger.info("[EXP] SITE.PRO(B1) added to ZIP: %s size=%d", filename, len(content))
                zip_buffer.seek(0)

                response = HttpResponse(zip_buffer.read(), content_type="application/zip")
                response["Content-Disposition"] = f'attachment; filename="{today_str}_site_pro_importas.zip"'
                export_success = True

            else:
                filename, content = files_to_export[0]
                response = HttpResponse(
                    content,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                export_success = True

        except FileNotFoundError as e:
            logger.exception("[EXP] SITE.PRO(B1) template not found: %s", e)
            return Response(
                {"error": "B1 template not found", "detail": str(e)},
                status=500
            )
        except Exception as e:
            logger.exception("[EXP] SITE.PRO(B1) export failed: %s", e)
            return Response(
                {"error": "B1 export failed", "detail": str(e)},
                status=500
            )



    # ========================= APSKAITA5 =========================

    elif export_type == 'apskaita5':
        logger.info("[EXP] APSKAITA5 export started")
        assign_random_prekes_kodai(documents)

        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])

        if not all_docs:
            logger.warning("[EXP] APSKAITA5 no documents to export")
            return Response({"error": "No documents to export"}, status=400)

        extra_fields = {
            "user": {
                "extra_settings": getattr(user, "extra_settings", {}) or {},
            }
        }

        content, filename, content_type = export_documents_group_to_apskaita5_files(
            documents=all_docs,
            site_url="",
            company_code=None,
            direction=None,
            apskaita5_extra_fields=extra_fields,
        )
        logger.info("[EXP] APSKAITA5 produced file=%s content_type=%s size=%d",
                     filename, content_type, len(content))
        response = HttpResponse(content, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['X-Content-Type-Options'] = 'nosniff'
        export_success = True

    # elif export_type == 'apskaita5':
    #     logger.info("[EXP] APSKAITA5 export started")
    #     assign_random_prekes_kodai(documents)

    #     content, filename, content_type = export_documents_group_to_apskaita5_files(
    #         documents=documents,
    #         site_url=site_url,   # предполагается, что переменная определена выше по модулю/конфигу
    #         company_code=None,
    #         direction=None,
    #     )
    #     logger.info("[EXP] APSKAITA5 produced file=%s content_type=%s size=%d",
    #                 filename, content_type, len(content))
    #     response = HttpResponse(content, content_type=content_type)
    #     response['Content-Disposition'] = f'attachment; filename="{filename}"'
    #     response['X-Content-Type-Options'] = 'nosniff'
    #     export_success = True

    # ========================= AGNUM =========================
    elif export_type == 'agnum':
        logger.info("[EXP] AGNUM export started")
        assign_random_prekes_kodai(documents)

        files_to_export = []

        # 1) Pirkimai (Type="2")
        if pirkimai_docs:
            logger.info("[EXP] AGNUM exporting pirkimai: %d docs", len(pirkimai_docs))
            pirkimai_xml = export_pirkimai_group_to_agnum(pirkimai_docs, request.user, own_company_code=cp_key)
            files_to_export.append((f'{today_str}_pirkimai_agnum.xml', pirkimai_xml))

        # 2) Pardavimai (Type="4")
        if pardavimai_docs:
            logger.info("[EXP] AGNUM exporting pardavimai: %d docs", len(pardavimai_docs))
            pardavimai_xml = export_pardavimai_group_to_agnum(pardavimai_docs, request.user, own_company_code=cp_key)
            files_to_export.append((f'{today_str}_pardavimai_agnum.xml', pardavimai_xml))

        logger.info("[EXP] AGNUM files_to_export=%s", [n for n, _ in files_to_export])

        if len(files_to_export) > 1:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for filename, xml_content in files_to_export:
                    zf.writestr(filename, xml_content)
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename={today_str}_agnum.zip'
            export_success = True
        elif len(files_to_export) == 1:
            filename, xml_content = files_to_export[0]
            response = HttpResponse(
                xml_content,
                content_type='application/xml; charset=utf-8'
            )
            response['Content-Disposition'] = f'attachment; filename={filename}'
            export_success = True
        else:
            logger.warning("[EXP] AGNUM nothing to export")
            response = Response({"error": "No documents to export"}, status=400)


    # # ========================= DEBETAS =========================
    elif export_type == 'debetas':
        logger.info("[EXP] DEBETAS export started")
        assign_random_prekes_kodai(documents)

        # Берём только уже классифицированные документы (pirkimai + pardavimai)
        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])

        if not all_docs:
            logger.warning("[EXP] DEBETAS no documents to export (no pirkimai/pardavimai)")
            return Response({"error": "No documents to export"}, status=400)

        try:
            debetas_result = export_to_debetas(
                documents=all_docs,
                user=request.user,
                own_company_code=cp_key,
            )
        except FileNotFoundError as e:
            logger.exception("[EXP] DEBETAS template not found: %s", e)
            return Response(
                {
                    "error": "Debetas template not found",
                    "detail": str(e),
                },
                status=500,
            )
        except Exception as e:
            logger.exception("[EXP] DEBETAS export failed: %s", e)
            return Response(
                {
                    "error": "Debetas export failed",
                    "detail": str(e),
                },
                status=500,
            )

        logger.info("[EXP] DEBETAS export result keys: %s", list(debetas_result.keys()))

        # Если есть zip (и pirkimai, и pardavimai) — отдаём его
        if debetas_result.get("zip"):
            content = debetas_result["zip"]
            filename = debetas_result.get("zip_filename", f"Debetas_Import_{today_str}.zip")
            response = HttpResponse(content, content_type="application/zip")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            export_success = True

        # Если только pirkimai
        elif debetas_result.get("pirkimai"):
            content = debetas_result["pirkimai"]
            filename = debetas_result.get("pirkimai_filename", f"Debetas_Pirkimai_{today_str}.csv")
            response = HttpResponse(
                content,
                content_type='text/csv; charset=windows-1257'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            export_success = True

        # Если только pardavimai
        elif debetas_result.get("pardavimai"):
            content = debetas_result["pardavimai"]
            filename = debetas_result.get("pardavimai_filename", f"Debetas_Pardavimai_{today_str}.csv")
            response = HttpResponse(
                content,
                content_type='text/csv; charset=windows-1257'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            export_success = True

        else:
            logger.warning("[EXP] DEBETAS nothing to export (empty result dict)")
            response = Response({"error": "No documents to export"}, status=400)


    # ========================= APSA (i.SAF, Paulita XML) =========================
    elif export_type in ('apsa', 'isaf', 'paulita'):
        logger.info("[EXP] APSA export started")
        
        # Собираем документы с direction
        all_docs = []
        all_packs = []  # сохраняем pack для доступа к line_items
        
        for pack in prepared.get("pirkimai", []):
            doc = pack["doc"]
            doc.direction = "pirkimas"
            all_docs.append(doc)
            all_packs.append(pack)
        
        for pack in prepared.get("pardavimai", []):
            doc = pack["doc"]
            doc.direction = "pardavimas"
            all_docs.append(doc)
            all_packs.append(pack)
        
        if not all_docs:
            logger.warning("[EXP] APSA nothing to export")
            return Response({"error": "No documents to export"}, status=400)
        
        # RegistrationNumber = код выбранного контрагента из cp_key
        registration_number = ""
        cp_name = ""
        if cp_key:
            cp = cp_key.strip()
            if cp.lower().startswith("id:"):
                cp = cp.split(":", 1)[1].strip()

            # Если похоже на код/ПВМ (например LT123456789 или просто цифры)
            if re.match(r"^[A-Za-z]{0,2}\d{4,}$", cp):
                registration_number = cp
            else:
                cp_name = cp

        # Если cp_key был именем - ищем совпадение по имени и берём id_programoje
        if not registration_number and cp_name and all_docs:
            target = cp_name.strip().lower()
            for doc in all_docs:
                buyer_name = (getattr(doc, "buyer_name", "") or "").strip().lower()
                if buyer_name and buyer_name == target:
                    reg_from_name = (getattr(doc, "buyer_id_programoje", "") or "").strip()
                    if reg_from_name:
                        registration_number = reg_from_name
                        break

                seller_name = (getattr(doc, "seller_name", "") or "").strip().lower()
                if seller_name and seller_name == target:
                    reg_from_name = (getattr(doc, "seller_id_programoje", "") or "").strip()
                    if reg_from_name:
                        registration_number = reg_from_name
                        break

        
        # Fallback: берём из первого документа
        if not registration_number and all_docs:
            first_doc = all_docs[0]
            if first_doc.direction == "pirkimas":
                registration_number = (
                    getattr(first_doc, "buyer_id", "") or 
                    getattr(first_doc, "buyer_vat_code", "") or ""
                )
            else:
                registration_number = (
                    getattr(first_doc, "seller_id", "") or 
                    getattr(first_doc, "seller_vat_code", "") or ""
                )
        
        if not registration_number:
            logger.error("[EXP] APSA no registration number")
            return Response({
                "error": "Company registration number is required. Select counterparty or ensure documents have company data."
            }, status=400)
        
        # pvm_resolver из pack["line_items"] (CP данные с vat_percent и pvm_kodas)
        # Структура: {doc_id: {item_id: {"vat_percent": ..., "pvm_kodas": ...}}}
        pvm_resolver = {}
        for pack in all_packs:
            doc = pack["doc"]
            line_items_data = pack.get("line_items", [])
            
            if line_items_data:
                # DETALIAI - есть line_items
                item_map = {}
                for li in line_items_data:
                    item_id = li.get("id")
                    if item_id is not None:
                        item_map[item_id] = {
                            "vat_percent": li.get("vat_percent"),
                            "pvm_kodas": li.get("pvm_kodas"),
                        }
                if item_map:
                    pvm_resolver[doc.id] = item_map
            else:
                # SUMISKAI - нет line_items, берём из pack напрямую
                pvm_resolver[doc.id] = {
                    "pvm_kodas": pack.get("pvm_kodas"),
                    "vat_percent": pack.get("vat_percent"),
                }
        
        logger.info("[EXP] APSA docs=%d reg_num=%s pvm_resolver_docs=%d", 
                    len(all_docs), registration_number, len(pvm_resolver))
        
        try:
            result = export_to_apsa(
                documents=all_docs,
                registration_number=registration_number,
                pvm_resolver=pvm_resolver,
            )

            if len(result) == 1:
                # Один месяц — один файл (ключ = "isaf_2026-04")
                key, xml_bytes = next(iter(result.items()))
                filename = f"{key}.xml"
                response = HttpResponse(xml_bytes, content_type='application/xml')
                response['Content-Disposition'] = f'attachment; filename={filename}'
                export_success = True
                logger.info("[EXP] APSA export completed (single month), file=%s size=%d", filename, len(xml_bytes))
            else:
                # Несколько месяцев — ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    for key, xml_bytes in sorted(result.items()):
                        zf.writestr(f"{key}.xml", xml_bytes)
                        logger.info("[EXP] APSA added to ZIP: %s.xml size=%d", key, len(xml_bytes))
                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename=isaf_{today_str}.zip'
                export_success = True
                logger.info("[EXP] APSA export completed (%d months)", len(result))

        except ValueError as e:
            logger.error("[EXP] APSA export error: %s", str(e))
            return Response({"error": str(e)}, status=400)
        except Exception as e:
            logger.exception("[EXP] APSA export failed")
            return Response({"error": f"Export failed: {str(e)}"}, status=500)



    # ========================= RIVILĖ ERP (XLSX) =========================
    elif export_type == 'rivile_erp':
        logger.info("[EXP] RIVILE_ERP export started")
        assign_random_prekes_kodai(documents)
        rivile_defaults = getattr(request.user, "rivile_erp_extra_fields", None) or {}
        rivile_defaults = getattr(request.user, "rivile_erp_extra_fields", None) or {}
        user_extra_settings = getattr(request.user, "extra_settings", None)
        if not isinstance(user_extra_settings, dict):
            user_extra_settings = {}

        klientai = []
        seen = set()

        for pack in (prepared.get("pirkimai", []) + prepared.get("pardavimai", [])):
            doc = pack["doc"]
            dir_ = pack.get("direction")

            if dir_ == 'pirkimas':
                is_person = doc.seller_is_person
                klient_type = 'pirkimas'
                # Код клиента: id → vat_code → id_programoje (как в get_party_code)
                client_code = doc.seller_id or doc.seller_vat_code or doc.seller_id_programoje or ""
                client = {
                    'id': client_code,
                    'vat': doc.seller_vat_code or "",
                    'name': doc.seller_name or "",
                    'address': doc.seller_address or "",
                    'country_iso': doc.seller_country_iso or "",
                    'currency': doc.currency or "EUR",
                    'kodas_ds': 'PT001',
                    'type': klient_type,
                    'is_person': is_person,
                    'iban': doc.seller_iban or "",
                }
            elif dir_ == 'pardavimas':
                is_person = doc.buyer_is_person
                klient_type = 'pardavimas'
                # Код клиента: id → vat_code → id_programoje (как в get_party_code)
                client_code = doc.buyer_id or doc.buyer_vat_code or doc.buyer_id_programoje or ""
                client = {
                    'id': client_code,
                    'vat': doc.buyer_vat_code or "",
                    'name': doc.buyer_name or "",
                    'address': doc.buyer_address or "",
                    'country_iso': doc.buyer_country_iso or "",
                    'currency': doc.currency or "EUR",
                    'kodas_ds': 'PT001',
                    'type': klient_type,
                    'is_person': is_person,
                    'iban': doc.buyer_iban or "",
                }
            else:
                continue

            client_key = (client['id'], client['vat'], client['name'], client['type'])
            if client['id'] and client_key not in seen:
                klientai.append(client)
                seen.add(client_key)

        logger.info("[EXP] RIVILE_ERP klientai=%d docs_pirk=%d docs_pard=%d",
                    len(klientai), len(pirkimai_docs), len(pardavimai_docs))

        # --- Company name map для prekės/paslaugos файла ---
        company_name_map = {}
        if str(user_extra_settings.get("rivile_erp_add_company", "0")).strip() == "1":
            for pack in (prepared.get("pirkimai", []) + prepared.get("pardavimai", [])):
                doc = pack["doc"]
                dir_ = pack.get("direction")
                doc_pk = doc.pk
                if dir_ == "pirkimas":
                    company_name_map[doc_pk] = doc.buyer_name or ""
                elif dir_ == "pardavimas":
                    company_name_map[doc_pk] = doc.seller_name or ""

        files_to_export = []

        if klientai:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                export_clients_to_rivile_erp_xlsx(klientai, tmp.name)
                tmp.seek(0)
                klientai_xlsx_bytes = tmp.read()
            files_to_export.append((f'klientai_{today_str}.xlsx', klientai_xlsx_bytes))

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            export_prekes_and_paslaugos_to_rivile_erp_xlsx(
                documents,
                tmp.name,
                user=request.user,
                company_name_map=company_name_map,
            )
            tmp.seek(0)
            prekes_xlsx_bytes = tmp.read()
        files_to_export.append((f'prekes_paslaugos_{today_str}.xlsx', prekes_xlsx_bytes))

        # --- iSAF классификация: разделяем на formuoti / neformuoti ---
        merge_vat = str(user_extra_settings.get("merge_vat", "0")).strip() == "1"

        def _split_by_isaf(docs, doc_type):
            form, neform = [], []
            for d in docs:
                if classify_isaf_for_erp(d, doc_type, merge_vat) == "formuoti":
                    form.append(d)
                else:
                    neform.append(d)
            return form, neform

        pirk_form, pirk_neform = _split_by_isaf(pirkimai_docs, "pirkimai")
        pard_form, pard_neform = _split_by_isaf(pardavimai_docs, "pardavimai")

        logger.info(
            "[EXP] RIVILE_ERP isaf split: pirk_form=%d pirk_neform=%d pard_form=%d pard_neform=%d",
            len(pirk_form), len(pirk_neform), len(pard_form), len(pard_neform),
        )

        def _export_erp_xlsx(docs, doc_type, suffix):
            if not docs:
                return None
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                export_documents_to_rivile_erp_xlsx(
                    docs,
                    tmp.name,
                    doc_type=doc_type,
                    rivile_erp_extra_fields={
                        **rivile_defaults,
                        "user": {"extra_settings": user_extra_settings},
                    },
                    own_company_code=cp_key,
                )
                tmp.seek(0)
                return (f'{doc_type}_{suffix}_{today_str}.xlsx', tmp.read())

        for result in [
            _export_erp_xlsx(pirk_form, "pirkimai", "form_isaf"),
            _export_erp_xlsx(pirk_neform, "pirkimai", "neform_isaf"),
            _export_erp_xlsx(pard_form, "pardavimai", "form_isaf"),
            _export_erp_xlsx(pard_neform, "pardavimai", "neform_isaf"),
        ]:
            if result:
                files_to_export.append(result)

        logger.info("[EXP] RIVILE_ERP files_to_export=%s", [n for n, _ in files_to_export])

        if len(files_to_export) > 1:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for filename, file_bytes in files_to_export:
                    zf.writestr(filename, file_bytes)
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename={today_str}_rivile_erp.zip'
            export_success = True
        elif len(files_to_export) == 1:
            filename, file_bytes = files_to_export[0]
            response = HttpResponse(
                file_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={filename}'
            export_success = True
        else:
            logger.warning("[EXP] RIVILE_ERP nothing to export")
            response = Response({"error": "No clients or products to export"}, status=400)


    # ========================= STEKAS PLIUS (ZIP/JSON) =========================
    elif export_type == 'stekas':
        logger.info("[EXP] STEKAS_PLIUS export started")

        all_docs = []
        for pack in (prepared.get("pirkimai", []) + prepared.get("pardavimai", [])):
            doc = pack["doc"]
            doc.pirkimas_pardavimas = pack.get("direction", "")
            all_docs.append(doc)

        logger.info("[EXP] STEKAS_PLIUS docs=%d pirk=%d pard=%d",
                    len(all_docs),
                    len(prepared.get("pirkimai", [])),
                    len(prepared.get("pardavimai", [])))

        if not all_docs:
            logger.warning("[EXP] STEKAS_PLIUS nothing to export")
            response = Response({"error": "Nėra dokumentų eksportui"}, status=400)
        else:

            content, filename, content_type = export_documents_group_to_stekas_files(
                documents=all_docs,
                site_url=request.build_absolute_uri('/') if request else "",
                company_code=getattr(request.user, 'company_code', '') or '',
                direction=None,  # направление берётся из doc.pirkimas_pardavimas
            )

            response = HttpResponse(content, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename={filename}'
            export_success = True


    # ========================= OPTIMUM (API) =========================
    elif export_type == 'optimum':
        logger.info("[EXP] OPTIMUM API export started")
        assign_random_prekes_kodai(documents)

        from docscanner_app.models import ExportSession
        from .utils.api_key_resolver import resolve_api_key

        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])
        if not all_docs:
            logger.warning("[EXP] OPTIMUM no documents to export")
            return Response({"error": "No documents to export"}, status=400)

        # --- Определяем company_code для поиска API ключа ---
        if source == "invoice":
            opt_company = "__israsymas__"
        elif cp_key:
            cp = cp_key.strip()
            opt_company = cp.split(":", 1)[1].strip() if cp.lower().startswith("id:") else cp
        else:
            opt_company = str(getattr(user, "company_code", "") or "").strip()

        if source == "invoice":
            api_key_obj = resolve_api_key(user, "optimum", opt_company, strict=True)
        else:
            api_key_obj = resolve_api_key(user, "optimum", opt_company)
        if not api_key_obj:
            logger.warning("[EXP] OPTIMUM key missing for company=%s", opt_company)
            return Response(
                {"error": f"Optimum API raktas nerastas įmonei {opt_company}. "
                          "Pridėkite raktą Nustatymuose."},
                status=400,
            )

        doc_ids = [d.pk for d in all_docs]

        session = ExportSession.objects.create(
            user=request.user,
            program='optimum',
            stage=ExportSession.Stage.QUEUED,
            total_documents=len(doc_ids),
        )
        session.documents.set(doc_ids)

        task = export_to_optimum_task.delay(session.id, api_key_obj.pk)
        session.task_id = task.id
        session.save(update_fields=["task_id"])

        logger.info(
            "[EXP] OPTIMUM session=%s task=%s docs=%d company=%s key=%s",
            session.pk, task.id, len(doc_ids), opt_company, api_key_obj.pk,
        )

        return Response({
            "status": "ok",
            "session_id": session.pk,
            "total_documents": len(doc_ids),
            "message": "Export started",
        }, status=202)


    # ========================= DINETA (API) =========================
    elif export_type == 'dineta':
        logger.info("[EXP] DINETA API export started")
        assign_random_prekes_kodai(documents)

        from docscanner_app.models import ExportSession
        from .utils.api_key_resolver import resolve_api_key

        all_docs = (pirkimai_docs or []) + (pardavimai_docs or [])
        if not all_docs:
            logger.warning("[EXP] DINETA no documents to export")
            return Response({"error": "No documents to export"}, status=400)

        # --- Определяем company_code для поиска API ключа ---
        if source == "invoice":
            din_company = "__israsymas__"
        elif cp_key:
            cp = cp_key.strip()
            din_company = cp.split(":", 1)[1].strip() if cp.lower().startswith("id:") else cp
        else:
            din_company = str(getattr(user, "company_code", "") or "").strip()

        if source == "invoice":
            api_key_obj = resolve_api_key(user, "dineta", din_company, strict=True)
        else:
            api_key_obj = resolve_api_key(user, "dineta", din_company)
        if not api_key_obj:
            logger.warning("[EXP] DINETA key missing for company=%s", din_company)
            return Response(
                {"error": f"Dineta API raktas nerastas įmonei {din_company}. "
                          "Pridėkite raktą Nustatymuose."},
                status=400,
            )

        doc_ids = [d.pk for d in all_docs]

        session = ExportSession.objects.create(
            user=request.user,
            program='dineta',
            stage=ExportSession.Stage.QUEUED,
            total_documents=len(doc_ids),
        )
        session.documents.set(doc_ids)

        task = export_to_dineta_task.delay(session.id, api_key_obj.pk)
        session.task_id = task.id
        session.save(update_fields=["task_id"])

        logger.info(
            "[EXP] DINETA session=%s task=%s docs=%d company=%s key=%s",
            session.pk, task.id, len(doc_ids), din_company, api_key_obj.pk,
        )

        return Response({
            "status": "ok",
            "session_id": session.pk,
            "total_documents": len(doc_ids),
            "message": "Export started",
        }, status=202)


    else:
        logger.error("[EXP] unknown export type: %s", export_type)
        return Response({"error": "Unknown export type"}, status=400)

    # --- universal finalize ---
    if response is not None:
        try:
            if export_success and exported_ids:
                if source == "invoice":
                    from django.utils import timezone as tz
                    Invoice.objects.filter(pk__in=exported_ids).update(
                        exported=True,
                        exported_at=tz.now(),
                    )
                    logger.info("[EXP] Marked %d invoices as exported", len(exported_ids))

                    # --- Record inv export usage + add headers ---
                    try:
                        if inv_sub and inv_sub.status == "free" and inv_usage:
                            for inv_id in exported_ids:
                                inv_usage.record_export(inv_id)
                            inv_usage.refresh_from_db()
                            response["X-Inv-Exports-Used"] = str(inv_usage.exports_used)
                            response["X-Inv-Exports-Max"] = "10"
                            response["X-Inv-Status"] = "free"
                            response["Access-Control-Expose-Headers"] = "X-Inv-Exports-Used, X-Inv-Exports-Max, X-Inv-Status"
                    except Exception as e:
                        logger.warning("[EXP] Failed to record inv export usage: %s", e)
                else:
                    ScannedDocument.objects.filter(pk__in=exported_ids).update(status="exported")
                    logger.info("[EXP] Marked %d documents as exported", len(exported_ids))
        except Exception as e:
            logger.warning("[EXP] Failed to mark as exported: %s", e)
        return response

    # # --- universal finalize ---
    # if response is not None:
    #     try:
    #         if export_success and exported_ids:
    #             ScannedDocument.objects.filter(pk__in=exported_ids).update(status="exported")
    #             logger.info("[EXP] Marked %d documents as exported (universal)", len(exported_ids))
    #     except Exception as e:
    #         logger.warning("[EXP] Failed to mark documents as exported: %s", e)
    #     return response

    logger.warning("[EXP] fell through unexpectedly")
    return Response({"error": "No documents to export"}, status=400)



# Soxranenije user infy s Dineta
class DinetaSettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Вернуть текущие настройки Dineta этого пользователя.
        Пароль НЕ возвращается.
        Вместо server/client отдаём склеенный url для отображения на фронте.
        """
        user = request.user
        settings_dict = user.dineta_settings or {}

        serializer = DinetaSettingsSerializer(instance=settings_dict)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request):

        user = request.user

        serializer = DinetaSettingsSerializer(
            data=request.data,
            instance=user.dineta_settings,
        )
        serializer.is_valid(raise_exception=True)

        settings_to_store = serializer.build_settings_dict()

        user.dineta_settings = settings_to_store
        user.save(update_fields=["dineta_settings"])

        response_serializer = DinetaSettingsSerializer(instance=settings_to_store)
        response_data = response_serializer.data

        try:
            dineta_hello(
                server=settings_to_store.get("server", ""),
                client=settings_to_store.get("client", ""),
                username=settings_to_store.get("username", ""),
                password=decrypt_password(settings_to_store.get("password", "")),
            )
            response_data["connection_status"] = "ok"
            response_data["connection_message"] = "Prisijungimas sėkmingas."
        except DinetaError as e:
            response_data["connection_status"] = "warning"
            response_data["connection_message"] = str(e)
        except Exception:
            response_data["connection_status"] = "warning"
            response_data["connection_message"] = "Prisijungimo patikrinimą nepavyko atlikti."

        return Response(response_data, status=status.HTTP_200_OK)




# Soxranenije user infy s Optimum i do soxranenija delajet probnyj Hello test
class OptimumSettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Grąžina meta informaciją (be rakto) + užmaskuotą rakto galą."""
        settings = request.user.optimum_settings or {}
        raw_key = settings.get("key", "")
        return Response({
            "has_key": bool(raw_key),
            "key_suffix": settings.get("key_suffix", ""),
            "verified_at": settings.get("verified_at"),
            "last_ok": settings.get("last_ok"),
            "last_error_at": settings.get("last_error_at"),
            "last_error": settings.get("last_error", ""),
        })

    def put(self, request):
        """Išsaugoti naują raktą: Hello testas → jei OK saugom, jei klaida — nesaugom rakto."""
        user = request.user
        raw_key = (request.data.get("key") or "").strip()

        if not raw_key:
            return Response({"detail": "API Key yra privalomas."}, status=status.HTTP_400_BAD_REQUEST)

        now_iso = timezone.now().isoformat()

        try:
            optimum_hello(raw_key)
        except OptimumError as exc:
            # Rakto nesaugom, saugom klaidos metaduomenis
            current = user.optimum_settings or {}
            current["last_ok"] = False
            current["last_error_at"] = now_iso
            current["last_error"] = str(exc) or "Optimum: klaida"
            user.optimum_settings = current
            user.save(update_fields=["optimum_settings"])

            return Response({
                "detail": str(exc) or "Nepavyko patikrinti Optimum API Key.",
                "last_ok": False,
                "last_error": str(exc),
                "last_error_at": now_iso,
            }, status=status.HTTP_400_BAD_REQUEST)

        # Sėkmė: saugom raktą + metaduomenis
        user.optimum_settings = {
            "key": raw_key,
            "key_suffix": raw_key[-4:] if len(raw_key) >= 4 else raw_key,
            "verified_at": now_iso,
            "last_ok": True,
            "last_error": "",
            "last_error_at": None,
        }
        user.save(update_fields=["optimum_settings"])

        return Response({
            "has_key": True,
            "key_suffix": raw_key[-4:] if len(raw_key) >= 4 else "",
            "verified_at": now_iso,
            "last_ok": True,
            "last_error": "",
            "last_error_at": None,
        })

    def post(self, request):
        """Patikrinti jau išsaugotą raktą (Patikrinti API mygtukas)."""
        user = request.user
        settings = user.optimum_settings or {}
        raw_key = settings.get("key", "")

        if not raw_key:
            return Response({"detail": "API raktas nerastas. Pirmiausia išsaugokite raktą."}, status=status.HTTP_400_BAD_REQUEST)

        now_iso = timezone.now().isoformat()

        try:
            optimum_hello(raw_key)
        except OptimumError as exc:
            settings["last_ok"] = False
            settings["last_error_at"] = now_iso
            settings["last_error"] = str(exc) or "Optimum: klaida"
            user.optimum_settings = settings
            user.save(update_fields=["optimum_settings"])

            return Response({
                "detail": str(exc) or "Nepavyko patikrinti Optimum API Key.",
                "has_key": True,
                "key_suffix": raw_key[-4:] if len(raw_key) >= 4 else "",
                "verified_at": settings.get("verified_at"),
                "last_ok": False,
                "last_error": str(exc),
                "last_error_at": now_iso,
            }, status=status.HTTP_400_BAD_REQUEST)

        settings["verified_at"] = now_iso
        settings["last_ok"] = True
        settings["last_error"] = ""
        settings["last_error_at"] = None
        user.optimum_settings = settings
        user.save(update_fields=["optimum_settings"])

        return Response({
            "has_key": True,
            "key_suffix": raw_key[-4:] if len(raw_key) >= 4 else "",
            "verified_at": now_iso,
            "last_ok": True,
            "last_error": "",
            "last_error_at": None,
        })

    def delete(self, request):
        """Ištrinti raktą ir visus metaduomenis."""
        user = request.user
        user.optimum_settings = {}
        user.save(update_fields=["optimum_settings"])
        return Response({"detail": "Optimum API raktas ištrintas."})

    # def put(self, request):
    #     """
    #     Временно: сохраняем любой key без проверки.
    #     """
    #     user = request.user

    #     serializer = OptimumSettingsSerializer(data=request.data)
    #     serializer.is_valid(raise_exception=True)

    #     raw_key = (serializer.validated_data.get("key") or "").strip()
    #     now_iso = timezone.now().isoformat()

    #     # сохраняем key без проверки
    #     settings_to_store = {
    #         "key": encrypt_password(raw_key),
    #         "verified_at": now_iso,
    #         "last_ok": True,
    #         "last_error_at": None,
    #         "last_error": "",
    #     }

    #     user.optimum_settings = settings_to_store
    #     user.save(update_fields=["optimum_settings"])

    #     response_serializer = OptimumSettingsSerializer(instance=settings_to_store)
    #     return Response(response_serializer.data, status=status.HTTP_200_OK)





@api_view(['POST'])
@permission_classes([IsAuthenticated])
def process_image(request):
    raw_files = request.FILES.getlist("files")
    scan_type = request.data.get("scan_type", "sumiskai")

    if not raw_files:
        return Response({'error': 'Файлы не предоставлены'}, status=400)

    user = request.user

    # Выбираем цену за документ
    if scan_type == "detaliai":
        credits_per_doc = Decimal("1.3")
    else:
        credits_per_doc = Decimal("1")

    files_count = len(raw_files)

    # --- ПРОВЕРКА кредитов ДО обработки ---
    if user.credits < credits_per_doc * files_count:
        return Response({
            'error': f'Nepakanka kreditų. Liko – {user.credits}, reikia – {credits_per_doc * files_count}.'
        }, status=402)

    results = []
    for raw_file in raw_files:
        original_filename = raw_file.name

        # 1. Сохраняем запись в БД сразу!
        doc = ScannedDocument.objects.create(
            user=user,
            original_filename=original_filename,
            status='processing',
            scan_type=scan_type
        )
        doc.file.save(original_filename, raw_file)
        doc.save()

        # 2. Запускаем celery-task c ID
        process_uploaded_file_task.delay(
            user.id,
            doc.id,
            scan_type
        )

        results.append({
            "id": doc.id,
            "original_filename": original_filename,
            "status": "processing",
            "uploaded_at": doc.uploaded_at
        })

    return Response({
        'status': 'processing',
        'results': results,
        'msg': 'Dokumentai užregistruoti ir apdorojami. Po kelių sekundžių statusas atsinaujins.'
    })






# Obnovit company details v Nustatymai

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_own_company_details(request):
    user = request.user
    serializer = CustomUserSerializer(user, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=400)



# Udaliajet zapisi s dashboard i BD

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def bulk_delete_documents(request):
    ids = request.data.get('ids', [])
    if not ids:
        return Response({'error': 'No IDs provided'}, status=status.HTTP_400_BAD_REQUEST)

    queryset = ScannedDocument.objects.filter(id__in=ids, user=request.user)

    # --- audit log: помечаем удаление в CreditUsageLog ---
    doc_ids = list(queryset.values_list('id', flat=True))
    if doc_ids:
        CreditUsageLog.objects.filter(
            scanned_document_id__in=doc_ids,
            document_deleted_by_user=False,
        ).update(
            document_deleted_by_user=True,
            document_deleted_at=timezone.now(),
        )

    deleted, _ = queryset.delete()
    return Response({'deleted': deleted}, status=status.HTTP_200_OK)




# #poluciajem vsio infu iz BD dlia otobrazhenija v dashboard pri zagruzke

# /documents/
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_user_documents(request):
#     user = request.user
#     status = request.GET.get('status')
#     date_from = request.GET.get('from')
#     date_to = request.GET.get('to')

#     docs = ScannedDocument.objects.filter(user=user)

#     if status:
#         docs = docs.filter(status=status)
#     if date_from:
#         docs = docs.filter(created_at__date__gte=parse_date(date_from))
#     if date_to:
#         docs = docs.filter(created_at__date__lte=parse_date(date_to))

#     docs = docs.order_by('-uploaded_at').only(
#         "id","original_filename","status","uploaded_at","preview_url",
#         "document_number",
#         "seller_name","seller_id","seller_vat_code","seller_vat_val",
#         "buyer_name","buyer_id","buyer_vat_code","buyer_vat_val",
#         "pirkimas_pardavimas","scan_type","ready_for_export","math_validation_passed",
#     )

#     serializer = ScannedDocumentListSerializer(docs, many=True)
#     return Response(serializer.data)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_user_documents(request):
#     user = request.user
#     status = request.GET.get('status')
#     date_from = request.GET.get('from')
#     date_to = request.GET.get('to')

#     docs = ScannedDocument.objects.filter(user=user)
#     if status:
#         docs = docs.filter(status=status)
#     if date_from:
#         docs = docs.filter(created_at__date__gte=parse_date(date_from))
#     if date_to:
#         docs = docs.filter(created_at__date__lte=parse_date(date_to))

#     serializer = ScannedDocumentListSerializer(docs.order_by('-uploaded_at'), many=True)
#     return Response(serializer.data)



@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def get_user_documents(request):
    user = request.user

    if request.method == "GET":
        q = request.query_params
        status_param = q.get("status")
        date_from = q.get("from")
        date_to = q.get("to")
        search = q.get("search")
        cp = q.get("cp")
        
        # NEW: параметры для archive_warnings
        include_archive_warnings = q.get("include_archive_warnings", "").lower() == "true"
        session_id = q.get("session_id")

        qs = (
            ScannedDocument.objects
            .select_related("perkelta_i_company_profile")
            .filter(
                user=user,
                is_archive_container=False,
                is_multi_doc_container=False,
            )
        )

        if status_param:
            qs = qs.filter(status=status_param)

        tz = timezone.get_current_timezone()

        if date_from:
            d = parse_date(date_from)
            if d:
                dt_from = timezone.make_aware(datetime.combine(d, dt_time.min), tz)
                qs = qs.filter(uploaded_at__gte=dt_from)

        if date_to:
            d = parse_date(date_to)
            if d:
                dt_to = timezone.make_aware(datetime.combine(d, dt_time.min), tz) + timedelta(days=1)
                qs = qs.filter(uploaded_at__lt=dt_to)

        if search:
            search = search.strip()
            if search:
                qs = qs.filter(document_number__icontains=search)

        if cp:
            cp = cp.strip().lower()
            if cp.startswith("id:"):
                cp_id = cp.split("id:", 1)[1].strip()
                if cp_id:
                    qs = qs.filter(Q(seller_id=cp_id) | Q(buyer_id=cp_id))
            else:
                qs = qs.filter(
                    Q(seller_vat_code__iexact=cp) |
                    Q(buyer_vat_code__iexact=cp) |
                    Q(seller_name__icontains=cp) |
                    Q(buyer_name__icontains=cp)
                )

        # === Exportable count ===
        exportable_qs = qs.filter(
            status__in=["completed", "exported"],
            ready_for_export=True,
            math_validation_passed=True,
        )

        view_mode = getattr(user, "view_mode", "single")
        if view_mode != "multi":
            exportable_qs = exportable_qs.filter(pirkimas_pardavimas__in=["pirkimas", "pardavimas"])

        exportable_total = 0 if (view_mode == "multi" and not cp) else exportable_qs.count()

        qs = qs.order_by("-uploaded_at", "-id").only(
            "id",
            "original_filename",
            "status",
            "uploaded_at",
            "preview_url",

            "document_number",

            "seller_name",
            "seller_id",
            "seller_vat_code",
            "seller_vat_val",

            "buyer_name",
            "buyer_id",
            "buyer_vat_code",
            "buyer_vat_val",

            "separate_vat",

            "pirkimas_pardavimas",
            "scan_type",
            "ready_for_export",
            "math_validation_passed",

            "optimum_api_status",
            "optimum_last_try_date",

            "dineta_api_status",
            "dineta_last_try_date",

            "rivile_api_status",
            "rivile_api_last_try",

            "is_credit_invoice",
            "is_debit_invoice",

            "buyer_replaced_by_rule",
            "seller_replaced_by_rule",

            "is_long_term_asset_candidate",

            "perkelta_i_apskaita",
            "perkelta_i_apskaita_at",
            "perkelta_i_company_profile",
            "perkelta_i_company_profile__name",
        )

        paginator = DocumentsCursorPagination()
        page = paginator.paginate_queryset(qs, request)
        serializer = ScannedDocumentListSerializer(page, many=True)
        resp = paginator.get_paginated_response(serializer.data)
        resp.data["exportable_total"] = exportable_total

        # === NEW: Архивы с ошибками (только если запрошено) ===
        if include_archive_warnings and session_id:
            archive_warnings_qs = ScannedDocument.objects.filter(
                user=user,
                is_archive_container=True,
                upload_session_id=session_id,
                error_message__startswith="Praleista"
            )
            
            resp.data["archive_warnings"] = list(archive_warnings_qs.order_by("-uploaded_at").values(
                "id", "original_filename", "error_message", "uploaded_at"
            )[:50])

        return resp

    # POST upload
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"detail": "file is required"}, status=status.HTTP_400_BAD_REQUEST)

    doc = ScannedDocument.objects.create(
        user=user,
        file=uploaded_file,
        original_filename=uploaded_file.name,
        status="processing",
    )
    return Response(ScannedDocumentListSerializer(doc).data, status=status.HTTP_201_CREATED)

# @api_view(["GET", "POST"])
# @permission_classes([IsAuthenticated])
# @parser_classes([MultiPartParser, FormParser])
# def get_user_documents(request):
#     user = request.user

#     if request.method == "GET":
#         q = request.query_params
#         status_param = q.get("status")
#         date_from = q.get("from")
#         date_to = q.get("to")
#         search = q.get("search")
#         cp = q.get("cp")

#         qs = ScannedDocument.objects.filter(user=user, is_archive_container=False)

#         if status_param:
#             qs = qs.filter(status=status_param)

#         tz = timezone.get_current_timezone()

#         if date_from:
#             d = parse_date(date_from)
#             if d:
#                 dt_from = timezone.make_aware(datetime.combine(d, dt_time.min), tz)
#                 qs = qs.filter(uploaded_at__gte=dt_from)

#         if date_to:
#             d = parse_date(date_to)
#             if d:
#                 dt_to = timezone.make_aware(datetime.combine(d, dt_time.min), tz) + timedelta(days=1)
#                 qs = qs.filter(uploaded_at__lt=dt_to)

#         if search:
#             search = search.strip()
#             if search:
#                 qs = qs.filter(document_number__icontains=search)

#         if cp:
#             cp = cp.strip().lower()
#             if cp.startswith("id:"):
#                 cp_id = cp.split("id:", 1)[1].strip()
#                 if cp_id.isdigit():
#                     qs = qs.filter(Q(seller_id=int(cp_id)) | Q(buyer_id=int(cp_id)))
#             else:
#                 # пробуем как VAT (точное совпадение) или как имя (icontains)
#                 qs = qs.filter(
#                     Q(seller_vat_code__iexact=cp) |
#                     Q(buyer_vat_code__iexact=cp) |
#                     Q(seller_name__icontains=cp) |
#                     Q(buyer_name__icontains=cp)
#                 )

#         # === NEW: сколько документов реально экспортируемо по текущим фильтрам ===
#         exportable_qs = qs.filter(
#             status__in=["completed", "exported"],
#             ready_for_export=True,
#             math_validation_passed=True,
#         )

#         view_mode = getattr(user, "view_mode", "single")
#         if view_mode != "multi":
#             exportable_qs = exportable_qs.filter(pirkimas_pardavimas__in=["pirkimas", "pardavimas"])

#         # в multi без выбранного контрагента — считаем 0 (экспорт всё равно запрещён)
#         exportable_total = 0 if (view_mode == "multi" and not cp) else exportable_qs.count()

        

#         qs = qs.order_by("-uploaded_at", "-id").only(
#             "id","original_filename","status","uploaded_at","preview_url",
#             "document_number",
#             "seller_name","seller_id","seller_vat_code","seller_vat_val",
#             "buyer_name","buyer_id","buyer_vat_code","buyer_vat_val",
#             "pirkimas_pardavimas","scan_type","ready_for_export","math_validation_passed",
#         )

#         paginator = DocumentsCursorPagination()
#         page = paginator.paginate_queryset(qs, request)
#         serializer = ScannedDocumentListSerializer(page, many=True)
#         resp = paginator.get_paginated_response(serializer.data)
#         resp.data["exportable_total"] = exportable_total  # NEW
#         return resp

#     # POST upload
#     uploaded_file = request.FILES.get("file")
#     if not uploaded_file:
#         return Response({"detail": "file is required"}, status=status.HTTP_400_BAD_REQUEST)

#     doc = ScannedDocument.objects.create(
#         user=user,
#         file=uploaded_file,
#         original_filename=uploaded_file.name,
#         status="processing",
#     )
#     # фронт сможет сделать prepend
#     return Response(ScannedDocumentListSerializer(doc).data, status=status.HTTP_201_CREATED)


BIG_FIELDS = ("raw_text", "gpt_raw_json", "structured_json", "glued_raw_text")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_document_detail(request, pk):
    user = request.user

    if user.is_superuser:
        doc = get_object_or_404(
            ScannedDocument.objects.annotate(line_items_count=Count("line_items")),
            pk=pk
        )
        ser = ScannedDocumentAdminDetailSerializer(doc, context={"request": request})
        data = ser.data
        data["line_items_count"] = doc.line_items_count  # уже посчитано

        if getattr(user, "view_mode", None) == "multi":
            cp_key = request.query_params.get("cp_key")
            preview = build_preview(
                doc,
                user,
                cp_key=cp_key,
                view_mode="multi",
                base_vat_percent=data.get("vat_percent"),
                base_preke_paslauga=data.get("preke_paslauga"),
            )
            data["preview"] = preview

        return Response(data)

    qs = (
        ScannedDocument.objects
        .defer(*BIG_FIELDS)
        .annotate(line_items_count=Count("line_items"))
    )
    doc = get_object_or_404(qs, pk=pk, user=user)

    ser = ScannedDocumentDetailSerializer(doc, context={"request": request})
    data = ser.data
    data["line_items_count"] = doc.line_items_count  # уже посчитано

    if getattr(user, "view_mode", None) != "multi":
        return Response(data)

    cp_key = request.query_params.get("cp_key")
    preview = build_preview(
        doc,
        user,
        cp_key=cp_key,
        view_mode="multi",
        base_vat_percent=data.get("vat_percent"),
        base_preke_paslauga=data.get("preke_paslauga"),
    )
    data["preview"] = preview
    return Response(data)



# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def get_document_detail(request, pk):
#     user = request.user

#     line_items_prefetch = Prefetch(
#         "line_items",
#         queryset=LineItem.objects.order_by("id")
#     )

#     # --- Суперюзер: оставляем как есть (все поля) ---
#     if user.is_superuser:
#         doc = get_object_or_404(
#             ScannedDocument.objects.prefetch_related(line_items_prefetch),
#             pk=pk
#         )
#         ser = ScannedDocumentAdminDetailSerializer(doc, context={"request": request})
#         data = ser.data

#         if getattr(user, "view_mode", None) == "multi":
#             cp_key = request.query_params.get("cp_key")
#             preview = build_preview(
#                 doc,
#                 user,
#                 cp_key=cp_key,
#                 view_mode="multi",
#                 base_vat_percent=data.get("vat_percent"),
#                 base_preke_paslauga=data.get("preke_paslauga"),
#             )
#             data["preview"] = preview

#         return Response(data)

#     # --- Обычный пользователь: НЕ читаем большие поля ---
#     qs = (
#         ScannedDocument.objects
#         .prefetch_related(line_items_prefetch)
#         .defer(*BIG_FIELDS)
#     )

#     doc = get_object_or_404(qs, pk=pk, user=user)

#     ser = ScannedDocumentDetailSerializer(doc, context={"request": request})
#     data = ser.data

#     if getattr(user, "view_mode", None) != "multi":
#         return Response(data)

#     cp_key = request.query_params.get("cp_key")
#     preview = build_preview(
#         doc,
#         user,
#         cp_key=cp_key,
#         view_mode="multi",
#         base_vat_percent=data.get("vat_percent"),
#         base_preke_paslauga=data.get("preke_paslauga"),
#     )
#     data["preview"] = preview
#     return Response(data)



# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def get_document_lineitems(request, pk):
#     """Пагинированная загрузка line items для документа."""
#     user = request.user
    
#     if user.is_superuser:
#         doc = get_object_or_404(ScannedDocument, pk=pk)
#     else:
#         doc = get_object_or_404(ScannedDocument, pk=pk, user=user)
    
#     line_items_qs = doc.line_items.order_by("id")
    
#     paginator = LineItemPagination()
#     page = paginator.paginate_queryset(line_items_qs, request)
    
#     serializer = LineItemSerializer(page, many=True)
    
#     return paginator.get_paginated_response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_document_lineitems(request, pk):
    """Пагинированная загрузка line items для документа (+ PVM klasė по cp_key)."""
    user = request.user

    if user.is_superuser:
        doc = get_object_or_404(ScannedDocument, pk=pk)
    else:
        doc = get_object_or_404(ScannedDocument, pk=pk, user=user)

    # cp_key приходит из фронта так же, как на /documents/{id}/
    cp_key = request.query_params.get("cp_key") or None
    cp_selected = bool(cp_key)

    # строим ctx ровно как в preview (multi)
    ctx = ResolveContext(
        user=user,
        view_mode="multi",
        purpose="preview",
        overrides={},
        cp_key=cp_key,
    )
    direction = resolve_direction(doc, ctx)

    # базовые значения на случай, если в lineitem нет vat_percent / preke_paslauga
    # (если у тебя эти поля называются иначе — поправь getattr)
    base_vat_percent = getattr(doc, "vat_percent", None)
    base_preke_paslauga = getattr(doc, "preke_paslauga", None)

    line_items_qs = doc.line_items.order_by("id")

    paginator = LineItemPagination()
    page = paginator.paginate_queryset(line_items_qs, request)

    serializer = LineItemSerializer(page, many=True)
    data = list(serializer.data)  # <-- будем модифицировать

    # Если cp не выбран — как раньше: “Pasirinkite kontrahentą”
    if not cp_selected:
        for row in data:
            row["pvm_kodas"] = None
            row["pvm_kodas_label"] = "Pasirinkite kontrahentą"
        return paginator.get_paginated_response(data)

    # cp выбран — считаем PVM kodą на каждую строку страницы
    buyer_iso = _nz(doc.buyer_country_iso)
    seller_iso = _nz(doc.seller_country_iso)
    buyer_has_v = bool(_nz(doc.buyer_vat_code))
    seller_has_v = bool(_nz(doc.seller_vat_code))

    ps_doc = _normalize_ps(base_preke_paslauga)
    separate_vat = bool(doc.separate_vat)
    doc_96_str = bool(getattr(doc, "doc_96_str", False))

    # page и serializer.data должны быть в одном порядке — zip безопасен
    for li_obj, row in zip(page, data):
        li_vat = _normalize_vat_percent(
            li_obj.vat_percent if getattr(li_obj, "vat_percent", None) is not None else base_vat_percent
        )

        # если в модели есть preke_paslauga — используем её, иначе fallback на doc-level
        li_ps_val = getattr(li_obj, "preke_paslauga", None)
        li_ps = _normalize_ps(li_ps_val if li_ps_val is not None else ps_doc)
        li_ps_bin = _ps_to_bin(li_ps)

        if _need_geo(li_vat) and (direction is None or not (buyer_iso and seller_iso)):
            li_code = None
        else:
            li_code = auto_select_pvm_code(
                pirkimas_pardavimas=direction,
                buyer_country_iso=buyer_iso,
                seller_country_iso=seller_iso,
                preke_paslauga=li_ps_bin,
                vat_percent=li_vat,
                separate_vat=False,
                buyer_has_vat_code=buyer_has_v,
                seller_has_vat_code=seller_has_v,
                doc_96_str=doc_96_str,
            )

        row["pvm_kodas"] = li_code
        row["pvm_kodas_label"] = _pvm_label(li_code, cp_selected=True)

    return paginator.get_paginated_response(data)




#Obnovit extra field vzavisimosti ot vybranoj buh programy
@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_scanned_document_extra_fields(request, pk):
    import logging
    from django.db import transaction
    from .models import ScannedDocument, LineItem
    from .serializers import ScannedDocumentSerializer
    from .utils.pirkimas_pardavimas import determine_pirkimas_pardavimas
    from .validators.vat_klas import auto_select_pvm_code
    from .utils.save_document import _apply_sumiskai_defaults_from_user
    from .validators.required_fields_checker import check_required_fields_for_export

    log = logging.getLogger("docscanner_app.api.update_extra_fields")

    doc = ScannedDocument.objects.filter(pk=pk, user=request.user).first()
    if not doc:
        log.warning("PATCH extra_fields pk=%s: document not found for user=%s", pk, request.user.id)
        return Response({'error': 'Dokumentas nerastas'}, status=404)

    ALLOWED_FIELDS = [
        'buyer_id', 'buyer_name', 'buyer_vat_code', 'buyer_iban', 'buyer_address', 'buyer_country_iso', 'buyer_is_person',
        'seller_id', 'seller_name', 'seller_vat_code', 'seller_iban', 'seller_address', 'seller_country_iso', 'seller_is_person',
        'prekes_kodas', 'prekes_barkodas', 'prekes_pavadinimas', 'preke_paslauga',
        'vat_percent', 'scan_type', 'doc_96_str',
    ]

    def _is_cleared(prefix: str) -> bool:
        keys = [
            f"{prefix}_name", f"{prefix}_id", f"{prefix}_vat_code",
            f"{prefix}_iban", f"{prefix}_address", f"{prefix}_country_iso",
        ]
        provided = [k for k in keys if k in request.data]
        if not provided:
            return False
        return all(not str(request.data.get(k) or "").strip() for k in provided)

    def _to_bool_allow(x):
        if x is None: return None
        if isinstance(x, bool): return x
        s = str(x).strip().lower()
        if s in {"0","false","no","ne","off"}: return False
        if s in {"1","true","taip","yes","on"}: return True
        return None

    def _normalize_ps(v):
        if v is None: return None
        if isinstance(v, int): return v if v in (1,2,3,4) else None
        s = str(v).strip()
        return int(s) if s.isdigit() and int(s) in (1,2,3,4) else None

    def _normalize_vat_percent(v):
        if v is None: return None
        try:
            from decimal import Decimal
            if isinstance(v, Decimal): return float(v)
            s = str(v).strip().replace(",", ".")
            if not s: return None
            if s.endswith("%"): s = s[:-1]
            return float(Decimal(s))
        except Exception:
            return None

    def _nz(s):
        if s is None: return None
        s2 = str(s).strip()
        return s2 if s2 else None

    # применяем входные изменения (сыро), логируем
    fields_to_update = []
    for field in ALLOWED_FIELDS:
        if field in request.data:
            old_val = getattr(doc, field, None)
            new_val = request.data[field]
            setattr(doc, field, new_val)
            fields_to_update.append(field)
            if str(old_val) != str(new_val):
                log.info("pk=%s: field %s changed: %r -> %r", pk, field, old_val, new_val)

    buyer_cleared = _is_cleared("buyer")
    seller_cleared = _is_cleared("seller")
    if buyer_cleared or seller_cleared:
        log.info("pk=%s: clear detected: buyer_cleared=%s seller_cleared=%s", pk, buyer_cleared, seller_cleared)

    apply_defaults_req = _to_bool_allow(request.data.get("apply_defaults", None))
    log.info("pk=%s: apply_defaults_req=%r", pk, apply_defaults_req)

    with transaction.atomic():
        # 0) Сохранить присланные поля
        if fields_to_update:
            doc.save(update_fields=fields_to_update)

        # 1) Пересчитать pirkimas/pardavimas
        doc_struct = {
            "seller_id": doc.seller_id,
            "seller_vat_code": doc.seller_vat_code,
            "seller_name": doc.seller_name,
            "buyer_id": doc.buyer_id,
            "buyer_vat_code": doc.buyer_vat_code,
            "buyer_name": doc.buyer_name,
        }
        doc.pirkimas_pardavimas = determine_pirkimas_pardavimas(doc_struct, request.user)
        log.info("pk=%s: pirkimas_pardavimas=%r", pk, doc.pirkimas_pardavimas)

        # 1.1) Флаги наличия VAT кода
        buyer_has_vat_code = bool((doc.buyer_vat_code or "").strip())
        seller_has_vat_code = bool((doc.seller_vat_code or "").strip())
        if hasattr(doc, "buyer_has_vat_code"):
            doc.buyer_has_vat_code = buyer_has_vat_code
        if hasattr(doc, "seller_has_vat_code"):
            doc.seller_has_vat_code = seller_has_vat_code
        log.info("pk=%s: buyer_has_vat_code=%s seller_has_vat_code=%s", pk, buyer_has_vat_code, seller_has_vat_code)

        # 1.2) ЕСЛИ очищаем buyer/seller — чистим товарные поля и PVM И ВЫХОДИМ РАНО
        if buyer_cleared or seller_cleared:
            doc.prekes_pavadinimas = ""
            doc.prekes_kodas = ""
            doc.prekes_barkodas = ""
            doc.preke_paslauga = ""
            doc.pvm_kodas = None
            update_fields_now = ["prekes_pavadinimas","prekes_kodas","prekes_barkodas","preke_paslauga","pvm_kodas","pirkimas_pardavimas"]

            if hasattr(doc, "buyer_has_vat_code"): update_fields_now.append("buyer_has_vat_code")
            if hasattr(doc, "seller_has_vat_code"): update_fields_now.append("seller_has_vat_code")

            if (doc.scan_type or "").strip().lower() == "detaliai":
                cleared = LineItem.objects.filter(document=doc).update(pvm_kodas=None)
                log.info("pk=%s: cleared LineItem.pvm_kodas for %d items", pk, cleared)

            doc.save(update_fields=update_fields_now)
            
            # Валидация перед ранним возвратом
            try:
                is_valid = check_required_fields_for_export(doc)
                doc.ready_for_export = is_valid
                doc.save(update_fields=['ready_for_export'])
                log.info("pk=%s: validated after clear, ready_for_export=%s", pk, is_valid)
            except Exception as e:
                log.error("pk=%s: validation error after clear: %s", pk, str(e))
            
            log.info("pk=%s: PVM cleared due to party clear, early return", pk)
            return Response(ScannedDocumentSerializer(doc).data)

        # 2) Применить дефолты (sumiskai, если разрешено)
        scan_type = (doc.scan_type or "").strip().lower()
        allow_defaults = (scan_type == "sumiskai" and (apply_defaults_req is None or apply_defaults_req is True))
        if allow_defaults:
            changed = _apply_sumiskai_defaults_from_user(doc, request.user)
            log.info("pk=%s: defaults applied=%s", pk, changed)
            if changed:
                doc.save(update_fields=["prekes_pavadinimas","prekes_kodas","prekes_barkodas","preke_paslauga"])

        # 3) Нормализованные данные для расчёта
        buyer_iso = _nz(doc.buyer_country_iso)
        seller_iso = _nz(doc.seller_country_iso)
        doc_vat_norm = _normalize_vat_percent(doc.vat_percent)
        doc_ps = _normalize_ps(doc.preke_paslauga)

        log.info("pk=%s: buyer_iso=%r seller_iso=%r vat_percent_norm=%r preke_paslauga_norm=%r",
                 pk, buyer_iso, seller_iso, doc_vat_norm, doc_ps)

        # требуем страны/направление только если 0%
        need_countries_doc = (doc_vat_norm == 0.0)
        missing_crit = need_countries_doc and (
            doc.pirkimas_pardavimas not in ("pirkimas", "pardavimas") or not (buyer_iso and seller_iso)
        )
        log.info("pk=%s: need_countries_doc=%s missing_crit=%s", pk, need_countries_doc, missing_crit)

        # ============ СОХРАНЯЕМ ОРИГИНАЛЬНЫЙ vat_percent ============
        original_vat_percent = doc.vat_percent
        # ============================================================

        # 4) Пересчёт PVM
        if scan_type == "detaliai":
            items = list(LineItem.objects.filter(document=doc))
            pvm_codes = set()
            vat_percents = set()
            log.info("pk=%s: recalc items count=%d", pk, len(items))

            for item in items:
                item_vat = item.vat_percent if item.vat_percent is not None else doc.vat_percent
                item_vat_norm = _normalize_vat_percent(item_vat)
                item_ps = _normalize_ps(item.preke_paslauga)
                if item_ps is None:
                    item_ps = doc_ps

                item_pvm = auto_select_pvm_code(
                    pirkimas_pardavimas=doc.pirkimas_pardavimas,
                    buyer_country_iso=buyer_iso,
                    seller_country_iso=seller_iso,
                    preke_paslauga=item_ps,
                    vat_percent=item_vat_norm,
                    separate_vat=bool(doc.separate_vat),
                    buyer_has_vat_code=buyer_has_vat_code,
                    seller_has_vat_code=seller_has_vat_code,
                    doc_96_str=bool(getattr(doc, "doc_96_str", False)),
                )

                old = item.pvm_kodas
                item.pvm_kodas = item_pvm
                item.save(update_fields=["pvm_kodas"])
                log.info("pk=%s: item[%s] vat=%r ps=%r -> pvm %r (was %r)",
                         pk, item.id, item_vat_norm, item_ps, item_pvm, old)

                if item_pvm is not None: pvm_codes.add(item_pvm)
                if item_vat_norm is not None: vat_percents.add(item_vat_norm)

            if bool(doc.separate_vat):
                doc.pvm_kodas = "Keli skirtingi PVM"
                doc.vat_percent = None
                log.info("pk=%s: separate_vat=True -> doc.pvm_kodas='Keli skirtingi PVM'", pk)
            else:
                if len(pvm_codes) == 1 and len(vat_percents) == 1:
                    doc.pvm_kodas = next(iter(pvm_codes))
                    doc.vat_percent = next(iter(vat_percents))
                    log.info("pk=%s: unified items -> doc.pvm_kodas=%r vat_percent=%r",
                             pk, doc.pvm_kodas, doc.vat_percent)
                elif len(pvm_codes) == 0:
                    # ============ FIX: Не удалось рассчитать PVM - сохраняем оригинальный vat_percent ============
                    doc.pvm_kodas = ""
                    # НЕ трогаем vat_percent - оставляем как было
                    log.info("pk=%s: could not calculate PVM (no pvm_codes), keeping vat_percent=%r", 
                             pk, doc.vat_percent)
                    # ============================================================================================
                else:
                    doc.pvm_kodas = ""
                    doc.vat_percent = None
                    log.info("pk=%s: heterogeneous items -> doc.pvm_kodas cleared", pk)

        else:
            # sumiskai / detaliai без строк — документный расчёт
            doc_pvm = auto_select_pvm_code(
                pirkimas_pardavimas=doc.pirkimas_pardavimas,
                buyer_country_iso=buyer_iso,
                seller_country_iso=seller_iso,
                preke_paslauga=doc_ps,
                vat_percent=doc_vat_norm,
                separate_vat=bool(doc.separate_vat),
                buyer_has_vat_code=buyer_has_vat_code,
                seller_has_vat_code=seller_has_vat_code,
                doc_96_str=bool(getattr(doc, "doc_96_str", False)),
            )
            old_doc_pvm = doc.pvm_kodas
            doc.pvm_kodas = doc_pvm
            log.info("pk=%s: doc-level recalc -> pvm %r (was %r)", pk, doc_pvm, old_doc_pvm)

        # 5) Сохранить
        update_set = {"pirkimas_pardavimas","pvm_kodas","vat_percent"}
        if hasattr(doc, "buyer_has_vat_code"): update_set.add("buyer_has_vat_code")
        if hasattr(doc, "seller_has_vat_code"): update_set.add("seller_has_vat_code")

        doc.save(update_fields=list(update_set))
        log.info("pk=%s: saved fields=%s", pk, sorted(update_set))

    # Валидация в конце - ВСЕГДА проверяем после изменений
    try:
        is_valid = check_required_fields_for_export(doc)
        doc.ready_for_export = is_valid
        doc.save(update_fields=['ready_for_export'])
        log.info("pk=%s: validated after update, ready_for_export=%s", pk, is_valid)
    except Exception as e:
        log.error("pk=%s: validation error: %s", pk, str(e))

    return Response(ScannedDocumentSerializer(doc).data)




# Udaliajet produkt s doka
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def clear_document_product(request, pk):
    from .models import ScannedDocument
    from .serializers import ScannedDocumentSerializer

    doc = ScannedDocument.objects.filter(pk=pk, user=request.user).first()
    if not doc:
        return Response({'error': 'Not found'}, status=404)

    # Очищаем только поля товара
    doc.prekes_pavadinimas = ""
    doc.prekes_kodas = ""
    doc.prekes_barkodas = ""
    doc.preke_paslauga = ""
    doc.save(update_fields=["prekes_pavadinimas", "prekes_kodas", "prekes_barkodas", "preke_paslauga"])

    return Response(ScannedDocumentSerializer(doc).data)



# Udaliajet produkt s lineitem
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def clear_lineitem_product(request, pk, lineitem_id):
    from .models import ScannedDocument, LineItem
    from .serializers import LineItemSerializer

    doc = ScannedDocument.objects.filter(pk=pk, user=request.user).first()
    if not doc:
        return Response({'error': 'Document not found'}, status=404)

    item = LineItem.objects.filter(document=doc, pk=lineitem_id).first()
    if not item:
        return Response({'error': 'Line item not found'}, status=404)

    item.prekes_pavadinimas = ""
    item.prekes_kodas = ""
    item.prekes_barkodas = ""
    item.preke_paslauga = ""
    item.save(update_fields=["prekes_pavadinimas", "prekes_kodas", "prekes_barkodas", "preke_paslauga"])

    return Response(LineItemSerializer(item).data)



@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_view_mode(request):
    """
    PATCH /users/me/view-mode/
    Body: { "view_mode": "single" | "multi" }
    """
    serializer = ViewModeSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    request.user.view_mode = serializer.validated_data['view_mode']
    request.user.save(update_fields=['view_mode'])

    return Response({'view_mode': request.user.view_mode}, status=status.HTTP_200_OK)




@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_lineitem_fields(request, doc_id, lineitem_id):
    from .serializers import LineItemSerializer
    doc = get_object_or_404(ScannedDocument, pk=doc_id, user=request.user)
    lineitem = get_object_or_404(LineItem, pk=lineitem_id, document=doc)

    allowed = [
        'prekes_kodas', 'prekes_pavadinimas', 'prekes_barkodas',
        'pirkimo_saskaita', 'pardavimo_saskaita',
        'unit', 'preke_paslauga',
        'matched_prekes_pavadinimas', 'matched_prekes_kodas',
        'matched_prekes_barkodas', 'matched_unit', 'matched_preke_paslauga',
        'catalog_match_user_override',
    ]
    changed = []
    for field in allowed:
        if field in request.data:
            setattr(lineitem, field, request.data[field])
            changed.append(field)
    if changed:
        lineitem.save(update_fields=changed)
    return Response(LineItemSerializer(lineitem).data, status=200)





#Autocomplete funkcii

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def autocomplete_products(request):
    query = request.GET.get('q', '').strip()
    qs = ProductAutocomplete.objects.filter(user=request.user)
    if query:
        qs = qs.filter(
            Q(prekes_pavadinimas__icontains=query) |
            Q(prekes_kodas__icontains=query) |
            Q(prekes_barkodas__icontains=query)
        )
    qs = qs.order_by('prekes_pavadinimas')[:30]  # Ограничь 30, чтобы не грузить всё
    data = [
        {
            "id": prod.id,
            "prekes_pavadinimas": prod.prekes_pavadinimas,
            "prekes_kodas": prod.prekes_kodas,
            "prekes_barkodas": prod.prekes_barkodas,
            "preke_paslauga": prod.preke_paslauga,
            "unit": prod.unit,
        }
        for prod in qs
    ]
    return Response(data)




def _filled_fields_count(record: dict) -> int:
    fields = ["pavadinimas", "imones_kodas", "pvm_kodas", "address", "country_iso", "ibans"]
    return sum(1 for f in fields if record.get(f) and str(record[f]).strip())


def _dedup_results(results: list, show_richer_duplicates: bool = False) -> list:
    seen_codes = {}
    seen_vats = {}
    output = []
    duplicates = []

    source_priority = {"imported": 0, "document": 1, "company_db": 2}

    for r in results:
        code = (r.get("imones_kodas") or "").strip()
        vat = (r.get("pvm_kodas") or "").strip()
        r_priority = source_priority.get(r.get("source"), 9)

        existing = None
        if code and code in seen_codes:
            existing = seen_codes[code]
        elif vat and vat in seen_vats:
            existing = seen_vats[vat]

        if existing is None:
            output.append(r)
            if code:
                seen_codes[code] = r
            if vat:
                seen_vats[vat] = r
        else:
            existing_priority = source_priority.get(existing.get("source"), 9)
            existing_filled = _filled_fields_count(existing)
            r_filled = _filled_fields_count(r)

            if r_priority < existing_priority:
                output = [x for x in output if x is not existing]
                output.append(r)
                if code:
                    seen_codes[code] = r
                if vat:
                    seen_vats[vat] = r
                if show_richer_duplicates and existing_filled > r_filled:
                    duplicates.append(existing)
            else:
                if show_richer_duplicates and r_filled > existing_filled:
                    duplicates.append(r)

    if show_richer_duplicates:
        output.extend(duplicates)

    output.sort(key=lambda r: (
        source_priority.get(r.get("source"), 9),
        -(r.get("doc_count") or 0),
        (r.get("pavadinimas") or ""),
    ))

    return output


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def autocomplete_clients(request):
    from .models import ClientAutocomplete, Company

    query = request.GET.get('q', '').strip()
    show_richer = len(query) >= 3

    MAX_IMPORTED = 4
    MAX_DOCUMENT = 4
    MAX_COMPANY = 2

    # ── 1) Imported ──
    qs_imported = ClientAutocomplete.objects.filter(user=request.user, source="imported")
    if query:
        qs_imported = qs_imported.filter(
            Q(pavadinimas__icontains=query) |
            Q(imones_kodas__icontains=query) |
            Q(pvm_kodas__icontains=query)
        )
    qs_imported = qs_imported.order_by("-doc_count", "pavadinimas")[:MAX_IMPORTED]

    results = [
        {
            "id": c.id,
            "pavadinimas": c.pavadinimas,
            "imones_kodas": c.imones_kodas,
            "pvm_kodas": c.pvm_kodas,
            "address": c.address,
            "country_iso": c.country_iso,
            "ibans": c.ibans,
            "is_person": c.is_person,
            "source": "imported",
            "source_label": "Importuotas",
            "doc_count": c.doc_count or 0,
        }
        for c in qs_imported
    ]

    # ── 2) Document ──
    qs_doc = ClientAutocomplete.objects.filter(user=request.user, source="document")
    if query:
        qs_doc = qs_doc.filter(
            Q(pavadinimas__icontains=query) |
            Q(imones_kodas__icontains=query) |
            Q(pvm_kodas__icontains=query)
        )
    qs_doc = qs_doc.order_by("-doc_count", "pavadinimas")[:MAX_DOCUMENT]

    results.extend([
        {
            "id": c.id,
            "pavadinimas": c.pavadinimas,
            "imones_kodas": c.imones_kodas,
            "pvm_kodas": c.pvm_kodas,
            "address": c.address,
            "country_iso": c.country_iso,
            "ibans": c.ibans,
            "is_person": c.is_person,
            "source": "document",
            "source_label": "Iš dokumentų",
            "doc_count": c.doc_count or 0,
        }
        for c in qs_doc
    ])

    # ── 3) Company DB ──
    if query:
        companies = Company.objects.filter(
            Q(pavadinimas__icontains=query) |
            Q(im_kodas__icontains=query) |
            Q(pvm_kodas__icontains=query)
        )[:MAX_COMPANY]

        results.extend([
            {
                "id": f"company_{comp.id}",
                "pavadinimas": comp.pavadinimas,
                "imones_kodas": comp.im_kodas,
                "pvm_kodas": comp.pvm_kodas,
                "address": comp.adresas,
                "country_iso": "LT",
                "ibans": "",
                "is_person": False,
                "source": "company_db",
                "source_label": "Iš registro",
                "doc_count": 0,
            }
            for comp in companies
        ])

    # ── 4) Дедупликация ──
    results = _dedup_results(results, show_richer_duplicates=show_richer)

    return Response(results[:10])




























# --- Импорт товаров (products) ---
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_products_view(request):
    file = request.FILES.get('file')
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
    try:
        report = import_products_from_xlsx(request.user, file)
        return Response(report, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# --- Импорт клиентов (clients) ---
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_clients_view(request):
    file = request.FILES.get('file')
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
    try:
        report = import_clients_from_xlsx(request.user, file)
        return Response(report, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# --- Экспорт товаров ---
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_products_view(request):
    from openpyxl.styles import Font

    qs = ProductAutocomplete.objects.filter(user=request.user).order_by("prekes_pavadinimas")

    COLUMNS = [
        ("prekes_pavadinimas*", "prekes_pavadinimas"),
        ("prekes_kodas*", "prekes_kodas"),
        ("prekes_barkodas", "prekes_barkodas"),
        ("mato_vnt", "unit"),
        (
            "preke_paslauga_kodas* "
            "(galimos reikšmės: 1, 2, 3, 4)",
            "preke_paslauga",
        ),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Prekės"

    bold = Font(bold=True)
    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 18)

    for p in qs:
        row = []
        for _, field in COLUMNS:
            val = getattr(p, field, None)
            if val is None or str(val).strip() in ("", "None", "0"):
                row.append("")
            else:
                row.append(str(val).strip())
        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="prekes_eksportas.xlsx"'
    wb.save(response)
    return response


# --- Удалить все товары ---
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_all_products_view(request):
    count, _ = ProductAutocomplete.objects.filter(user=request.user).delete()
    return Response({"deleted": count})


# --- Счётчик товаров ---
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def products_count_view(request):
    count = ProductAutocomplete.objects.filter(user=request.user).count()
    return Response({"count": count})


# --- Экспорт клиентов ---
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_clients_view(request):
    from openpyxl.styles import Font

    qs = ClientAutocomplete.objects.filter(
        user=request.user, source="imported"
    ).order_by("pavadinimas")

    COLUMNS = [
        ("Pavadinimas*", "pavadinimas"),
        ("Kodas*", "imones_kodas"),
        ("Fizinis_asmuo*", "is_person"),
        ("PVM_kodas", "pvm_kodas"),
        ("IBAN", "ibans"),
        ("Adresas", "address"),
        ("Salies_kodas*", "country_iso"),
        ("Kodas_programoje", "kodas_programoje"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Klientai"

    bold = Font(bold=True)
    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 18)

    for c in qs:
        row = []
        for _, field in COLUMNS:
            val = getattr(c, field, None)
            if field == "is_person":
                row.append("Taip" if val else "")
            elif val is None or str(val).strip() in ("", "None"):
                row.append("")
            else:
                row.append(str(val).strip())
        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="klientu_eksportas.xlsx"'
    wb.save(response)
    return response


# --- Удалить всех клиентов ---
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_all_clients_view(request):
    count, _ = ClientAutocomplete.objects.filter(
        user=request.user, source="imported"
    ).delete()
    return Response({"deleted": count})


# --- Счётчик клиентов ---
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def clients_count_view(request):
    count = ClientAutocomplete.objects.filter(
        user=request.user, source="imported"
    ).count()
    return Response({"count": count})









#Proverka ili obnovlenija default accounting program

@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def user_profile(request):
    user = request.user
    if request.method == 'PATCH':
        old_company_code = user.company_code
        serializer = CustomUserSerializer(user, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()

            # Veidrodis: pakeitus programą Nustatymuose – įrašom ją ir į aktyvų profilį.
            # .update() rašo tiesiai į DB – aplenkia signalus ir instance cache.
            if 'default_accounting_program' in request.data and user.active_company_profile_id:
                CompanyProfile.objects.filter(
                    pk=user.active_company_profile_id
                ).update(accounting_program=user.default_accounting_program)

            new_company_code = serializer.validated_data.get('company_code', old_company_code)

            # Пытаемся найти старую строку по старому company_code
            ca = ClientAutocomplete.objects.filter(
                user=user,
                imones_kodas=old_company_code
            ).first()

            if not ca:
                # Если не нашли — пробуем по новому
                ca = ClientAutocomplete.objects.filter(
                    user=user,
                    imones_kodas=new_company_code
                ).first()

            if not ca:
                # Если всё равно не нашли — создаём новую
                ca = ClientAutocomplete(user=user, imones_kodas=new_company_code)

            # Теперь обновляем все поля:
            ca.pavadinimas = user.company_name
            ca.imones_kodas = new_company_code
            ca.pvm_kodas = user.vat_code
            ca.ibans = user.company_iban
            ca.address = user.company_address
            ca.country_iso = user.company_country_iso
            ca.save()
            
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    serializer = CustomUserSerializer(user)
    return Response(serializer.data)





@api_view(['POST'])
@permission_classes([IsAdminUser])
def update_currency_rates_view(request):
    d = request.data.get('date')
    if d:
        try:
            from datetime import datetime
            d = datetime.strptime(d, '%Y-%m-%d').date()
        except Exception:
            return Response({'error': 'Invalid date'}, status=400)
    else:
        d = date.today()
    count = update_currency_rates(d)
    return Response({'message': f'Updated {count} currency rates for {d}.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_me_view(request):
    serializer = CustomUserSerializer(request.user)
    data = serializer.data

    # ── Company profiles ──
    data["company_profiles"] = CompanyProfileSerializer(
        request.user.company_profiles.all(), many=True
    ).data
    data["active_company_profile_id"] = request.user.active_company_profile_id
    data["onboarding_completed"] = request.user.onboarding_completed

    # ── Subscription (вычисленный статус с lazy-expire — перезаписываем сырое поле) ──
    data["subscription_status"] = request.user.get_subscription_status()

    return Response(data)


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def user_me_view(request):
#     serializer = CustomUserSerializer(request.user)
#     data = serializer.data

#     # ── Company profiles ──
#     data["company_profiles"] = CompanyProfileSerializer(
#         request.user.company_profiles.all(), many=True
#     ).data
#     data["active_company_profile_id"] = request.user.active_company_profile_id
#     data["onboarding_completed"] = request.user.onboarding_completed

#     return Response(data)


class TrackAdClickView(generics.CreateAPIView):
    queryset = AdClick.objects.all()
    serializer_class = AdClickSerializer
    permission_classes = [permissions.AllowAny]  # даже гости могут

    def create(self, request, *args, **kwargs):
        user = request.user if request.user.is_authenticated else None
        ip = request.META.get("REMOTE_ADDR")
        ua = request.META.get("HTTP_USER_AGENT", "")

        ad_click = AdClick.objects.create(
            ad_name=request.data.get("ad_name", "Unknown"),
            user=user,
            ip_address=ip,
            user_agent=ua
        )
        return Response({"status": "ok"})


#Skacivanje Apskaita5 plugina

FILE_PATH = "/var/files/DokSkenas_apskaita5_adapteris.zip"

def _sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

@api_view(['GET'])
@permission_classes([IsAuthenticated])  # доступ только для авторизованных
def download_apskaita5_adapter(request):
    if not os.path.exists(FILE_PATH):
        raise Http404("Adapter not found")

    resp = FileResponse(open(FILE_PATH, "rb"))
    resp["Content-Type"] = "application/zip"
    resp["Content-Disposition"] = f'attachment; filename="{os.path.basename(FILE_PATH)}"'
    resp["X-Checksum-SHA256"] = _sha256(FILE_PATH)
    return resp




#JWT functions

class CustomTokenObtainPairView(TokenObtainPairView):
    def post(self, request, *args, **kwargs):
        
        try:
            response = super().post(request, *args, **kwargs)
            tokens = response.data

            access_token = tokens['access']
            refresh_token = tokens['refresh']

            token = AccessToken(access_token)
            user_id = token['user_id']
            from .models import CustomUser  
            user = CustomUser.objects.get(id=user_id)
            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])

            res = Response()

            res.data = {'success':True}

            res.set_cookie(
                key="access_token",
                value=access_token,
                httponly=True,
                secure=True,
                samesite='Lax',
                path='/'
            )

            res.set_cookie(
                key="refresh_token",
                value=refresh_token,
                httponly=True,
                secure=True,
                samesite='Lax',
                path='/'
            )

            return res


        except:
            return Response({'success':False})


class CustomRefreshTokenView(TokenRefreshView):
    def post(self, request, *args, **kwargs):
        try:
            refresh_token = request.COOKIES.get('refresh_token')
            request.data['refresh'] = refresh_token

            response = super().post(request, *args, **kwargs)
            tokens = response.data
            access_token = tokens['access']

            res = Response()
            res.data = {'refreshed': True}

            res.set_cookie(
                key='access_token',
                value=access_token,
                httponly=True,
                secure=True,
                samesite='Lax',
                path='/'
            )

            if 'refresh' in tokens:
                res.set_cookie(
                    key='refresh_token',
                    value=tokens['refresh'],
                    httponly=True,
                    secure=True,
                    samesite='Lax',
                    path='/'
                )

            return res

        except:
            return Response({'refreshed': False})


@api_view(['POST'])
def logout(request):
    try:
        res = Response()
        res.data = {'success':True}
        res.delete_cookie('access_token', path='/', samesite='Lax')
        res.delete_cookie('refresh_token', path='/', samesite='Lax')
        return res
    except:
        return Response({'success':False})
    

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def is_authenticated(request):
    return Response({'authenticated':True})


# Funkcija dlia sozdanija trial

@permission_classes([AllowAny])
def create_trial_subscription(user):
    """
    Создаёт пробную подписку для нового пользователя.
    """
    logger.info(f"Начинаем создание триал-подписки для пользователя: {user.email}")
    try:
        # Устанавливаем триал-подписку
        trial_start_date = timezone.now()
        trial_end_date = trial_start_date + timedelta(days=3000)

        user.subscription_status = 'trial'
        user.subscription_plan = 'trial'
        user.subscription_start_date = trial_start_date
        user.subscription_end_date = trial_end_date
        user.save()

        logger.info(f"Триал-подписка успешно создана для пользователя: {user.email}")
    except Exception as e:
        logger.error(f"Ошибка при создании триал-подписки для пользователя {user.email}: {str(e)}")
        raise e

#Naxodit IP usera pri registracii
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')

@api_view(['POST'])
@authentication_classes([])  # Отключение проверки аутентификации
@permission_classes([AllowAny])  # Разрешить всем пользователям доступ к этому эндпоинту
def register(request):
    """
    Регистрация нового пользователя и создание триал-подписки.
    """
    logger.info("Получен запрос на регистрацию нового пользователя.")

    # Удаление cookies с токенами
    if 'access_token' in request.COOKIES:
        logger.info("Удаляем access_token из cookies.")
        del request.COOKIES['access_token']

    if 'refresh_token' in request.COOKIES:
        logger.info("Удаляем refresh_token из cookies.")
        del request.COOKIES['refresh_token']

    try:
        serializer = CustomUserSerializer(data=request.data)
        if serializer.is_valid():
            logger.info("Данные пользователя валидны.")

            # Создаём пользователя
            user = serializer.save()
            logger.info(f"Пользователь {user.email} успешно зарегистрирован.")

            # Устанавливаем default extra_settings
            user.registration_ip = get_client_ip(request)
            user.extra_settings = {"fix_delta": 1}

            # --- Onboarding source ---
            reg_source = request.data.get("registration_source", "")
            if reg_source in ("skaitmenizavimas", "israsymas"):
                user.registration_source = reg_source

            user.save(update_fields=["extra_settings", "registration_ip", "registration_source"])

            user.ensure_inbox_token(save=True)

            InvSubscription.objects.create(user=user)

            # Создаём триал-подписку для нового пользователя
            create_trial_subscription(user)

            # 3️⃣ Ставим welcome email в очередь Celery (после коммита)

            try:
                t0 = perf_counter()
                siusti_sveikinimo_laiska(user)
                t1 = perf_counter()
                logger.info(f"Welcome email išsiųstas vartotojui {user.email} per {t1 - t0:.4f}s (be Celery).")
            except Exception as mail_err:
                logger.exception(f"Nepavyko išsiųsti welcome email be Celery: {mail_err}")

            # try:
            #     t_reg = perf_counter()

            #     def _enqueue():
            #         t0 = perf_counter()
            #         logger.info(f"[ENQUEUE] on_commit fired; start apply_async for {user.email}")
            #         try:
            #             task_siusti_sveikinimo_laiska.apply_async(args=[user.id], ignore_result=True)
            #         finally:
            #             t1 = perf_counter()
            #             logger.info(f"[ENQUEUE] apply_async_time={t1 - t0:.4f}s for {user.email}")

            #     transaction.on_commit(_enqueue)

            #     t_reg_done = perf_counter()
            #     logger.info(f"Queued welcome email for {user.email}. on_commit_registration_time={t_reg_done - t_reg:.4f}s")

            # except Exception as mail_err:
            #     logger.exception(f"Не удалось поставить welcome email в очередь: {mail_err}")

            return Response({
                "message": "Registracija sėkminga!",
                "user": {
                    "id": user.id,
                    "email": user.email,
                    "subscription_status": user.subscription_status,
                    "subscription_plan": user.subscription_plan,
                    "subscription_start_date": user.subscription_start_date,
                    "subscription_end_date": user.subscription_end_date
                }
            }, status=201)

        logger.warning(f"Ошибка в данных регистрации: {serializer.errors}")
        return Response(serializer.errors, status=400)

    except Exception as e:
        logger.error(f"Ошибка при регистрации пользователя: {str(e)}")
        return Response({"error": "An error occurred during registration."}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def onboarding_company_search(request):
    query = str(request.query_params.get("q") or "").strip()[:100]

    if len(query) < 2:
        return Response({"results": []})

    normalized_query = " ".join(query.upper().split())

    companies = (
        Company.objects
        .filter(
            Q(pavadinimas__icontains=query)
            | Q(normalized_pavadinimas__icontains=normalized_query)
            | Q(im_kodas__icontains=query)
            | Q(pvm_kodas__icontains=query)
        )
        .order_by("pavadinimas", "im_kodas")[:20]
    )

    return Response({
        "results": [
            {
                "id": c.id,
                "pavadinimas": c.pavadinimas or "",
                "im_kodas": c.im_kodas or "",
                "pvm_kodas": c.pvm_kodas or "",
                "adresas": c.adresas or "",
            }
            for c in companies
        ]
    })


# Proveriajem status subscriptiona usera
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def subscription_status(request):
    try:
        # Получаем текущего пользователя
        user = request.user

        # Получаем расширенную модель пользователя (CustomUser)
        user_profile = get_object_or_404(CustomUser, pk=user.pk)

        # Вызываем метод get_subscription_status из модели CustomUser
        status = user_profile.get_subscription_status()

        # Возвращаем статус подписки
        return Response({'status': status}, status=200)

    except Exception as e:
        # Обработка ошибок и возврат сообщения
        return Response({'error': str(e)}, status=500)
    




#DLIA SUPERUSEROV!!!:
#1) dlia admin-suvestine


def _ensure_dict(x):
    if x is None:
        return {}
    if isinstance(x, dict):
        return x
    if isinstance(x, str):
        try:
            return json.loads(x)
        except Exception:
            return {}
    return {}


def _ensure_dict(x):
    if x is None:
        return {}
    if isinstance(x, dict):
        return x
    if isinstance(x, str):
        try:
            return json.loads(x)
        except Exception:
            return {}
    return {}

def summarize_doc_issues(doc_struct):
    """
    Возвращает 'error' ТОЛЬКО если overall_status == "FAIL" из финальной математической валидации.
    Это единственный критерий для определения проблемных документов.
    """
    doc = _ensure_dict(doc_struct)

    # ✅ ЕДИНСТВЕННАЯ ПРОВЕРКА: overall_status из финальной валидации
    math_failed = False
    math_badge = None
    validation_type = None
    
    # Проверяем для detaliai (с line_items)
    final_validation = doc.get("_final_math_validation")
    if final_validation:
        overall = final_validation.get("summary", {}).get("overall_status")
        if overall == "FAIL":
            math_failed = True
            math_badge = "MATH✗"
            validation_type = "detaliai"
    
    # Проверяем для sumiskai (без line_items)
    sumiskai_validation = doc.get("_final_math_validation_sumiskai")
    if sumiskai_validation:
        overall = sumiskai_validation.get("overall_status")
        if overall == "FAIL":
            math_failed = True
            math_badge = "MATH✗"
            validation_type = "sumiskai"

    has_error = math_failed

    # --- оформление результата ---
    badges = []
    if math_badge:
        badges.append(math_badge)

    summary = " ".join(badges) if badges else ""
    if validation_type:
        summary = f"{summary} ({validation_type})".strip()

    issue_count = 1 if has_error else 0

    return {
        "has_issues": has_error,
        "severity": "error" if has_error else "ok",
        "issue_badges": " ".join(badges),
        "issue_summary": summary,
        "issue_count": issue_count,
    }

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_documents_with_errors(request):
    """
    Для superuser — документы всех пользователей с ошибками.
    Ошибка = math_validation_passed=False ИЛИ ready_for_export=False
             ИЛИ structured_json._global_validation_log содержит "OVERALL STATUS: FAIL"
    Курсорная пагинация с infinite scroll.
    """
    user = request.user
    if not user.is_superuser:
        return Response({"detail": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)

    # Только документы с ошибками
    qs = ScannedDocument.objects.select_related('user').filter(
        Q(math_validation_passed=False) | 
        Q(ready_for_export=False) |
        Q(structured_json___global_validation_log__icontains='OVERALL STATUS: FAIL')
    )

    # --- фильтры ---
    status_filter = request.GET.get('status')
    if status_filter:
        qs = qs.filter(status=status_filter)

    # Сортировка
    qs = qs.order_by('-uploaded_at', '-id')

    # --- курсорная пагинация ---
    paginator = DocumentsCursorPagination()
    page = paginator.paginate_queryset(qs, request)

    ser = ScannedDocumentListSerializer(page, many=True)

    # --- обогащение данных ---
    data = []
    for obj, row in zip(page, ser.data):
        r = dict(row)
        r["user_id"] = getattr(obj.user, "id", None)
        r["owner_email"] = getattr(obj.user, "email", None)
        
        # Показываем какая именно ошибка
        badges = []
        if not obj.math_validation_passed:
            badges.append("MATH✗")
        if not obj.ready_for_export:
            badges.append("NOT_READY")
        
        # Проверяем _global_validation_log в structured_json
        structured = obj.structured_json or {}
        validation_log = structured.get('_global_validation_log', '')
        if validation_log and 'OVERALL STATUS: FAIL' in validation_log:
            badges.append("VALIDATION✗")
        
        r["issue_badges"] = " ".join(badges)
        r["issue_has"] = True
        data.append(r)

    return paginator.get_paginated_response(data)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def admin_documents_with_errors(request):
#     """
#     Для superuser — документы всех пользователей с ошибками.
#     Ошибка = math_validation_passed=False ИЛИ ready_for_export=False
#     Курсорная пагинация с infinite scroll.
#     """
#     user = request.user
#     if not user.is_superuser:
#         return Response({"detail": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)

#     # Только документы с ошибками
#     qs = ScannedDocument.objects.select_related('user').filter(
#         Q(math_validation_passed=False) | Q(ready_for_export=False)
#     )

#     # --- фильтры ---
#     status_filter = request.GET.get('status')
#     if status_filter:
#         qs = qs.filter(status=status_filter)

#     # Сортировка
#     qs = qs.order_by('-uploaded_at', '-id')

#     # --- курсорная пагинация ---
#     paginator = DocumentsCursorPagination()
#     page = paginator.paginate_queryset(qs, request)

#     ser = ScannedDocumentListSerializer(page, many=True)

#     # --- обогащение данных ---
#     data = []
#     for obj, row in zip(page, ser.data):
#         r = dict(row)
#         r["user_id"] = getattr(obj.user, "id", None)
#         r["owner_email"] = getattr(obj.user, "email", None)
        
#         # Показываем какая именно ошибка
#         badges = []
#         if not obj.math_validation_passed:
#             badges.append("MATH✗")
#         if not obj.ready_for_export:
#             badges.append("NOT_READY")
        
#         r["issue_badges"] = " ".join(badges)
#         r["issue_has"] = True
#         data.append(r)

#     return paginator.get_paginated_response(data)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_all_documents(request):
    """
    Для superuser — сводный список ВСЕХ документов всех пользователей.
    Курсорная пагинация с infinite scroll.
    """
    user = request.user
    if not user.is_superuser:
        return Response({"detail": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)

    qs = ScannedDocument.objects.select_related('user').all()

    # --- фильтры ---
    status_filter = request.GET.get('status')
    if status_filter:
        qs = qs.filter(status=status_filter)

    owner = request.GET.get('owner')
    if owner:
        qs = qs.filter(user__email__icontains=owner)

    search = request.GET.get('search')
    if search:
        qs = qs.filter(document_number__icontains=search)

    from django.utils.dateparse import parse_date
    from datetime import timedelta

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if date_from:
        d = parse_date(date_from)
        if d:
            qs = qs.filter(uploaded_at__gte=d)

    if date_to:
        d = parse_date(date_to)
        if d:
            qs = qs.filter(uploaded_at__lt=d + timedelta(days=1))

    # --- сортировка (курсорная пагинация требует фиксированный order) ---
    qs = qs.order_by('-uploaded_at', '-id')

    # --- курсорная пагинация ---
    paginator = DocumentsCursorPagination()
    page = paginator.paginate_queryset(qs, request)

    from .serializers import ScannedDocumentListSerializer
    ser = ScannedDocumentListSerializer(page, many=True)

    # --- обогащение данных ---
    data = []
    for obj, row in zip(page, ser.data):
        doc_struct_raw = getattr(obj, 'structured_json', None) or getattr(obj, 'gpt_raw_json', None)
        issues = summarize_doc_issues(doc_struct_raw)

        enriched_row = {
            "user_id": getattr(obj.user, "id", None),
            "owner_email": getattr(obj.user, "email", None),
        }
        enriched_row.update(row)
        enriched_row["issue_has"] = issues["has_issues"]
        enriched_row["issue_badges"] = issues["issue_badges"]
        enriched_row["issue_summary"] = issues["issue_summary"]
        enriched_row["issue_count"] = issues["issue_count"]
        data.append(enriched_row)

    return paginator.get_paginated_response(data)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_all_waybills(request):
    """Superuser — vse važtaraščiai vsex userov. Cursornaja paginacija."""
    if not request.user.is_superuser:
        return Response({"detail": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)

    qs = ScannedWaybill.objects.select_related('user').filter(
        is_archive_container=False,
        is_multi_doc_container=False,
    )

    # Filtrai
    status_filter = request.GET.get('status')
    if status_filter:
        qs = qs.filter(status=status_filter)

    owner = request.GET.get('owner')
    if owner:
        qs = qs.filter(user__email__icontains=owner)

    search = request.GET.get('search')
    if search:
        qs = qs.filter(document_number__icontains=search)

    from django.utils.dateparse import parse_date
    from datetime import timedelta

    date_from = request.GET.get('date_from')
    if date_from:
        d = parse_date(date_from)
        if d:
            qs = qs.filter(uploaded_at__gte=d)

    date_to = request.GET.get('date_to')
    if date_to:
        d = parse_date(date_to)
        if d:
            qs = qs.filter(uploaded_at__lt=d + timedelta(days=1))

    qs = qs.order_by('-uploaded_at', '-id')

    paginator = DocumentsCursorPagination()
    page = paginator.paginate_queryset(qs, request)

    data = []
    for obj in page:
        data.append({
            "id": obj.id,
            "user_id": obj.user_id,
            "owner_email": getattr(obj.user, "email", ""),
            "original_filename": obj.original_filename,
            "status": obj.status,
            "error_message": obj.error_message or "",
            "document_number": obj.document_number or "",
            "document_date": str(obj.document_date) if obj.document_date else "",
            "airport": obj.airport or "",
            "buyer_name": obj.buyer_name or "",
            "buyer_iata_code": obj.buyer_iata_code or "",
            "payment_type": obj.payment_type or "",
            "flight_nature": obj.flight_nature or "",
            "from_city": obj.from_city or "",
            "from_airport_code": obj.from_airport_code or "",
            "to_city": obj.to_city or "",
            "to_airport_code": obj.to_airport_code or "",
            "quantity_liters_observed": str(obj.quantity_liters_observed) if obj.quantity_liters_observed else "",
            "quantity_kg_observed": str(obj.quantity_kg_observed) if obj.quantity_kg_observed else "",
            "preview_url": obj.preview_url or "",
            "uploaded_at": obj.uploaded_at.isoformat() if obj.uploaded_at else "",
        })

    return paginator.get_paginated_response(data)


#3) Dlia users
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_users_simple(request):
    """
    Для superuser — список всех пользователей (CustomUser).
    Курсорная пагинация с infinite scroll.
    """
    if not request.user.is_superuser:
        return Response({"detail": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)

    qs = CustomUser.objects.all().order_by("-date_joined", "-id")
    
    # --- фильтры (опционально) ---
    email = request.GET.get('email')
    if email:
        qs = qs.filter(email__icontains=email)
    
    # --- курсорная пагинация ---
    paginator = UsersCursorPagination()
    page = paginator.paginate_queryset(qs, request)
    
    ser = CustomUserAdminListSerializer(page, many=True, context={'request': request})
    return paginator.get_paginated_response(ser.data)



#Wagtail blog
class GuideCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    """
    /guides-api/v2/guide-categories/                 -> список категорий
    /guides-api/v2/guide-categories/<slug>/          -> категория + articles[] (детально)
    /guides-api/v2/guide-categories/<slug>/articles/ -> только список статей категории
    """
    permission_classes = [AllowAny]
    lookup_field = "slug"
    queryset = GuideCategoryPage.objects.live().public().order_by("order", "title")

    def get_serializer_class(self):
        # list -> короткий сериализатор
        # retrieve -> детальный (с вложенным массивом статей)
        return (
            GuideCategoryDetailSerializer
            if self.action == "retrieve"
            else GuideCategoryListSerializer
        )

    @action(detail=True, methods=["get"], url_path="articles")
    def articles(self, request, slug=None):
        """
        Вернёт список статей одной категории (удобно для пагинации фронта).
        GET-параметры: ?limit=12&offset=0
        """
        category = self.get_object()

        try:
            limit = int(request.query_params.get("limit", 100))
            offset = int(request.query_params.get("offset", 0))
        except ValueError:
            limit, offset = 100, 0

        qs = (
            GuidePage.objects.child_of(category)
            .live()
            .public()
            .specific()
            .order_by("-first_published_at")
        )
        total = qs.count()
        items = qs[offset : offset + limit]

        data = GuideArticleListSerializer(items, many=True, context={"request": request}).data
        return Response(
            {
                "count": total,
                "limit": limit,
                "offset": offset,
                "results": data,
            }
        )


class GuideArticleViewSet(viewsets.ReadOnlyModelViewSet):
    """
    /guides-api/v2/guides/           -> (опц.) список всех статей (короткие карточки)
    /guides-api/v2/guides/<slug>/    -> детальная статья
    """
    permission_classes = [AllowAny]
    lookup_field = "slug"

    def get_queryset(self):
        return GuidePage.objects.live().public().specific()

    def get_serializer_class(self):
        return (
            GuideArticleDetailSerializer
            if self.action == "retrieve"
            else GuideArticleListSerializer
        )


# Update doc and item field in Preview

# --- разрешённые поля ---
ALLOWED_DOC_FIELDS = {
    "invoice_date","due_date","operation_date","document_series","document_number","order_number",
    "amount_wo_vat","vat_amount","vat_percent","amount_with_vat","currency","paid_by_cash",
    "buyer_name","buyer_id","buyer_vat_code","seller_name","seller_id","seller_vat_code",
    "prekes_kodas","prekes_pavadinimas","prekes_barkodas", "invoice_discount_wo_vat", "invoice_discount_with_vat",
    "pirkimo_saskaita","pardavimo_saskaita",
}

ALLOWED_LINE_FIELDS = {
    "prekes_kodas","prekes_pavadinimas","prekes_barkodas",
    "unit","quantity","price","subtotal","vat","vat_percent","total","pirkimo_saskaita",
    "pardavimo_saskaita",
    # catalog matching
    "matched_prekes_pavadinimas","matched_prekes_kodas","matched_prekes_barkodas",
    "matched_unit","matched_preke_paslauga","catalog_match_user_override",
}

REQUIRED_FIELDS = {
    'invoice_date', 'document_number', 'amount_wo_vat', 'vat_amount', 
    'amount_with_vat', 'currency', 'seller_name', 'seller_vat_code', 
    'buyer_name', 'buyer_vat_code', 'seller_id', 'buyer_id'
}

MATH_FIELDS = {
    'amount_wo_vat', 'vat_amount', 'amount_with_vat', 'vat_percent',
    'invoice_discount_wo_vat', 'invoice_discount_with_vat'
}

LINE_MATH_FIELDS = {
    'quantity', 'price', 'subtotal', 'vat', 'vat_percent', 'total',
    'discount_wo_vat', 'discount_with_vat'
}


class InlineDocUpdateView(APIView):
    permission_classes = [IsOwner]

    def patch(self, request, doc_id):
        doc = get_object_or_404(ScannedDocument, pk=doc_id, user=request.user)
        field = request.data.get("field")
        value = request.data.get("value")

        if field not in ALLOWED_DOC_FIELDS:
            return Response({"detail": "Field not allowed"}, status=400)

        if value in ("", None):
            value = None

        setattr(doc, field, value)
        doc.save(update_fields=[field])

        # Валидация
        response_data = {
            "ok": True,
            "id": doc.id,
            field: getattr(doc, field),
        }
        
        try:
            if field in REQUIRED_FIELDS:
                is_valid = check_required_fields_for_export(doc)
                doc.ready_for_export = is_valid
                response_data['ready_for_export'] = is_valid
            
            if field in MATH_FIELDS:
                is_valid, _ = validate_document_math_for_export(doc)
                doc.math_validation_passed = is_valid
                response_data['math_validation_passed'] = is_valid
            
            if field in REQUIRED_FIELDS or field in MATH_FIELDS:
                update_fields = []
                if field in REQUIRED_FIELDS:
                    update_fields.append('ready_for_export')
                if field in MATH_FIELDS:
                    update_fields.append('math_validation_passed')
                if update_fields:
                    doc.save(update_fields=update_fields)
        except Exception as e:
            logger.error(f"Validation error: {str(e)}")
        
        return Response(response_data)


class InlineLineUpdateView(APIView):
    permission_classes = [IsOwner]

    def patch(self, request, doc_id, line_id):
        doc = get_object_or_404(ScannedDocument, pk=doc_id, user=request.user)
        line = get_object_or_404(LineItem, pk=line_id, document=doc)

        field = request.data.get("field")
        value = request.data.get("value")

        if field not in ALLOWED_LINE_FIELDS:
            return Response({"detail": "Field not allowed"}, status=400)

        if value in ("", None):
            value = None

        setattr(line, field, value)
        line.save(update_fields=[field])

        # Валидация
        response_data = {
            "ok": True,
            "id": line.id,
            field: getattr(line, field),
        }
        
        try:
            if field in LINE_MATH_FIELDS:
                is_valid, _ = validate_document_math_for_export(doc)
                doc.math_validation_passed = is_valid
                doc.save(update_fields=['math_validation_passed'])
                response_data['math_validation_passed'] = is_valid
        except Exception as e:
            logger.error(f"Validation error: {str(e)}")
        
        return Response(response_data)


# Add / delete lineitem in Preview

class ScannedDocumentViewSet(viewsets.ModelViewSet):
    queryset = ScannedDocument.objects.all()
    serializer_class = ScannedDocumentDetailSerializer
    permission_classes = [IsAuthenticated]

    # --- ДОБАВИТЬ ПУСТОЙ LINE ITEM ---
    @action(detail=True, methods=["post"], url_path="add-lineitem")
    def add_lineitem(self, request, pk=None):
        doc = self.get_object()
        line = LineItem.objects.create(document=doc)
        return Response(LineItemSerializer(line).data, status=status.HTTP_201_CREATED)

    # --- УДАЛИТЬ LINE ITEM ---
    @action(detail=True, methods=["delete"], url_path="delete-lineitem/(?P<line_id>[^/.]+)")
    def delete_lineitem(self, request, pk=None, line_id=None):
        doc = self.get_object()
        try:
            line = doc.line_items.get(id=line_id)
            line.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except LineItem.DoesNotExist:
            return Response({"detail": "Line item not found"}, status=status.HTTP_404_NOT_FOUND)
        


# contact email sender
@api_view(["POST"])
@permission_classes([AllowAny])
def contact_form(request):
    vardas  = (request.data.get("name") or "").strip()
    email   = (request.data.get("email") or "").strip()
    zinute  = (request.data.get("message") or "").strip()
    # subject nėra formoje – paliekame None (bus generinė)

    if not vardas or not email or len(zinute) < 10:
        return Response({"detail": "Klaida formoje"}, status=status.HTTP_400_BAD_REQUEST)

    ok = siusti_kontakto_laiska(vardas=vardas, email=email, zinute=zinute, tema=None)
    if ok:
        return Response({"detail": "Žinutė sėkmingai išsiųsta. Ačiū!"})
    return Response({"detail": "Nepavyko išsiųsti žinutės. Pabandykite vėliau."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR)





def send_newsletter():
    text_tpl = (
        "Sveiki,\n\n"
        "nuo šiol sąskaitas į DokSkeną galite siųsti ir el. paštu.\n"
        "Šiame video parodome, kaip tai vyksta: https://youtu.be/6zcjzTEiK1A\n\n"
        "Jei nematote \"Kiti būdai\" suvestinėje, savo klaviatūroje paspauskite CTRL, SHIFT, R klavišus kartu.\n\n"
        "Iš ateities planų:\n"
        "- integracija su Google Drive ir Dropbox (prijungiate savo paskyras, pasidalinate prieiga su klientais, kiekvienam klientui sukuriamas aplankas į kurį jie kelia failus, o DokSkenas automatiškai juos pasiėma)\n\n"
        "- sąskaitų išrašymas ir duomenų eksportas į 16 apskaitos programų (pardavimo sąskaitas galėsite formuoti tiesiai DokSkene, greitai ir patogiai, o sąskaitas nereikės skaitmenizuoti, nes duomenis tiesiog eksportuosite į savo apskaitos programą)\n\n"
        "Jei turite klausimų ar pastebėjimų, klauskite.\n\n"
        "Su pagarba,\n"
        "DokSkeno komanda\n"
        "Denis"
    )

    siusti_masini_laiska_visiems(
        subject="Svarbus atnaujinimas",
        text_template=text_tpl,
        html_template_name=None,      # ← НЕ используем HTML вообще
        extra_context=None,           # можно опустить
        exclude_user_ids=[46, 2, 16, 24, 31, 69, 105, 133, 202, 233, 283, 284, 289, 322, 351, 360 ],   # кого исключить (опционально)
        tik_aktyviems=True,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_mobile_key(request):
    """
    POST /api/mobile/generate-key/

    Новая версия: создаёт ОТДЕЛЬНЫЙ MobileAccessKey для этого пользователя.

    Ожидает (опционально):
    - email: для какого отправителя (по умолчанию user.email)
    - label: человекопонятное имя (pvz. "Jonas (ofisas)", "Kasa #2")

    НИЧЕГО не шлёт по email – просто генерирует и возвращает.
    """
    user = request.user

    raw_email = (request.data.get("email") or "").strip().lower()
    label = (request.data.get("label") or "").strip()

    if not raw_email:
        # Если email не пришёл – пробуем взять email самого пользователя
        raw_email = (user.email or "").strip().lower()

    if not raw_email:
        return Response({"error": "EMAIL_REQUIRED"}, status=400)

    # создаём MobileAccessKey и получаем raw_key (строка, которую покажем/отправим)
    access_key, raw_key = MobileAccessKey.create_for_user(
        user=user,
        sender_email=raw_email,
        label=label or None,
    )

    play_store_link = build_mobile_play_store_link(raw_key)

    return Response({
        "id": access_key.id,
        "mobile_key": raw_key,          # ПОЛНЫЙ ключ – покажем один раз
        "key_last4": access_key.key_last4,
        "sender_email": access_key.sender_email,
        "label": access_key.label,
        "play_store_link": play_store_link,
    })



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_mobile_invitation(request):
    """
    POST /api/mobile/send-invitation/

    Ожидает:
    - email: получатель приглашения
    - (опционально) label: pvz. "Jonas (ofisas)", "Kasa #2"

    Действия:
    - создаём новый MobileAccessKey для этого email
    - строим Play Store ссылку с этим ключом
    - шлём письмо (siusti_mobilios_apps_kvietima)
    """
    user = request.user

    raw_email = (request.data.get("email") or "").strip().lower()
    label = (request.data.get("label") or "").strip()

    if not raw_email:
        return Response({"error": "EMAIL_REQUIRED"}, status=400)

    # создаём отдельный ключ под этот email
    access_key, raw_key = MobileAccessKey.create_for_user(
        user=user,
        sender_email=raw_email,
        label=label or None,
    )

    play_store_link = build_mobile_play_store_link(raw_key)

    ok = siusti_mobilios_apps_kvietima(
        kvietejas=user,
        gavejo_email=raw_email,
        play_store_link=play_store_link,
        mobile_key=raw_key,  # важный момент: сюда кладём СЫРОЙ ключ
    )

    if not ok:
        return Response(
            {"error": "EMAIL_SEND_FAILED"},
            status=500,
        )

    return Response({
        "status": "OK",
        "email": raw_email,
        "label": access_key.label,
        "key_last4": access_key.key_last4,
        "play_store_link": play_store_link,
        "id": access_key.id,
    })


User = get_user_model()


@api_view(['POST'])
@authentication_classes([])   # авторизация только по мобильному ключу
@permission_classes([AllowAny])
def mobile_upload_documents(request: HttpRequest):
    """
    POST /api/mobile/upload/

    Headers:
      X-Mobile-Key: <pilnas mobilus raktas>

    Body (multipart/form-data):
      files: <pdf1>, files: <pdf2>, ...
      (neprivaloma) sender_email: jei nori perrašyti (dažniausiai nereikės)
    """

    raw_key = (
        request.META.get("HTTP_X_MOBILE_KEY")
        or request.data.get("mobile_key")
    )

    if not raw_key:
        return Response(
            {"error": "MOBILE_KEY_REQUIRED"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    key_hash = MobileAccessKey.make_hash(raw_key)

    try:
        access_key = MobileAccessKey.objects.select_related("user").get(
            key_hash=key_hash,
            is_active=True,
        )
    except MobileAccessKey.DoesNotExist:
        return Response(
            {"error": "INVALID_OR_REVOKED_MOBILE_KEY"},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    user = access_key.user

    files = request.FILES.getlist("files")
    if not files:
        return Response(
            {"error": "NO_FILES"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    sender_email = request.data.get("sender_email") or access_key.sender_email or None

    access_key.last_used_at = timezone.now()
    access_key.save(update_fields=["last_used_at"])

    created_docs = []
    for f in files:
        doc = MobileInboxDocument.objects.create(
            user=user,
            access_key=access_key,
            uploaded_file=f,
            original_filename=f.name,
            size_bytes=getattr(f, "size", 0) or 0,
            page_count=None,          # page_count v budushchem mozhno budet peredavat iz mobile
            sender_email=sender_email,
            is_processed=False,       # v inbox po umolchaniyu neperenesennyj
        )

        # posle soxranenija u polja uploaded_file uzhe est .url
        doc.preview_url = f"{settings.SITE_URL_BACKEND}{doc.uploaded_file.url}"
        doc.save(update_fields=["preview_url"])

        created_docs.append({
            "id": doc.id,
            "original_filename": doc.original_filename,
            "size_bytes": doc.size_bytes,
            "page_count": doc.page_count,
            "sender_email": doc.sender_email,
            "created_at": doc.created_at.isoformat(),
            # bez URL-ov, kak dogovorilis'
        })

    return Response(
        {
            "status": "OK",
            "count": len(created_docs),
            "documents": created_docs,
        },
        status=status.HTTP_201_CREATED,
    )



@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def mobile_access_keys_list_create(request):
    """
    GET  /api/mobile/keys/   -> sąrašas visų MobileAccessKey šitam user'iui
    POST /api/mobile/keys/   -> sukuria naują MobileAccessKey ir išsiunčia kvietimą el. paštu

    Body (POST):
      - email (required)
      - label (optional)
    """
    user = request.user

    if request.method == "GET":
        qs = MobileAccessKey.objects.filter(user=user).order_by("-created_at")
        serializer = MobileAccessKeySerializer(qs, many=True)
        return Response(serializer.data)

    # POST
    raw_email = (request.data.get("email") or "").strip().lower()
    label = (request.data.get("label") or "").strip()

    if not raw_email:
        return Response({"error": "EMAIL_REQUIRED"}, status=status.HTTP_400_BAD_REQUEST)

    # создаём отдельный ключ под этот email
    access_key, raw_key = MobileAccessKey.create_for_user(
        user=user,
        sender_email=raw_email,
        label=label or None,
    )

    play_store_link = build_mobile_play_store_link(raw_key)

    ok = siusti_mobilios_apps_kvietima(
        kvietejas=user,
        gavejo_email=raw_email,
        play_store_link=play_store_link,
        mobile_key=raw_key,
    )

    if not ok:
        return Response(
            {"error": "EMAIL_SEND_FAILED"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    serializer = MobileAccessKeySerializer(access_key)
    data = serializer.data
    # play_store_link мы можем вернуть только здесь (когда ещё есть raw_key)
    data["play_store_link"] = play_store_link

    return Response(data, status=status.HTTP_201_CREATED)



@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def mobile_access_key_detail(request, pk: int):
    """
    PATCH  /api/mobile/keys/<id>/   -> keičiam is_active (toggle)
    DELETE /api/mobile/keys/<id>/   -> ištrinam raktą

    PATCH body:
      { "is_active": true/false }
    """
    try:
        access_key = MobileAccessKey.objects.get(pk=pk, user=request.user)
    except MobileAccessKey.DoesNotExist:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        access_key.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # PATCH
    new_is_active = request.data.get("is_active", None)

    # поддержим строки "true"/"false" на всякий случай
    if isinstance(new_is_active, str):
        new_is_active = new_is_active.lower() in ("1", "true", "yes", "on")

    if new_is_active is None:
        return Response(
            {"detail": "is_active is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not new_is_active:
        # выключаем аккуратно через метод модели (ставит revoked_at)
        access_key.revoke()
    else:
        # включаем обратно, чистим revoked_at
        if not access_key.is_active:
            access_key.is_active = True
            access_key.revoked_at = None
            access_key.save(update_fields=["is_active", "revoked_at"])

    serializer = MobileAccessKeySerializer(access_key)
    return Response(serializer.data)




@api_view(["GET"])
@permission_classes([IsAuthenticated])
def web_mobile_inbox(request):
    """
    GET /api/web/mobile-inbox/

    Список мобильных документов для текущего пользователя (WEB).
    Курсорная пагинация с infinite scroll.
    """
    user = request.user

    qs = (
        MobileInboxDocument.objects
        .filter(user=user, is_processed=False)
        .select_related("processed_document", "access_key")
        .order_by("-created_at", "-id")
    )

    paginator = MobileInboxCursorPagination()
    page = paginator.paginate_queryset(qs, request)

    serializer = MobileInboxDocumentSerializer(page, many=True)
    return paginator.get_paginated_response(serializer.data)


#Udaliajem faily s IsKlientu spiska
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def web_mobile_inbox_bulk_delete(request):
    user = request.user
    select_all = request.data.get("select_all", False)
    exclude_ids = request.data.get("exclude_ids", [])

    if select_all:
        qs = MobileInboxDocument.objects.filter(user=user, is_processed=False)
        source = request.data.get("source")
        client_id = request.data.get("client_id")
        if source:
            qs = qs.filter(source=source)
        if client_id:
            qs = qs.filter(cloud_client_id=client_id)
        if exclude_ids:
            qs = qs.exclude(id__in=exclude_ids)
        docs = qs
    else:
        ids = request.data.get("ids") or []
        if not isinstance(ids, list) or not ids:
            return Response(
                {"error": "NO_IDS", "detail": "Pateikite bent vieną ID."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        docs = MobileInboxDocument.objects.filter(
            user=user, is_processed=False, id__in=ids,
        )

    file_paths = []
    for d in docs:
        if d.uploaded_file and d.uploaded_file.name:
            try:
                file_paths.append(d.uploaded_file.path)
            except Exception:
                pass

    deleted_count = docs.count()
    deleted_ids = list(docs.values_list("id", flat=True))
    docs.delete()

    for path in file_paths:
        try:
            if path and os.path.exists(path):
                os.remove(path)
        except Exception:
            continue

    return Response(
        {"status": "OK", "count": deleted_count, "deleted_ids": deleted_ids},
        status=status.HTTP_200_OK,
    )


ARCHIVE_EXTS_PROMOTE = {".zip", ".rar", ".7z", ".tar", ".tgz", ".tar.gz", ".tar.bz2", ".tar.xz", ".tbz2"}

def _ext_lower(name):
    n = (name or "").lower()
    if n.endswith(".tar.gz"): return ".tar.gz"
    if n.endswith(".tar.bz2"): return ".tar.bz2"
    if n.endswith(".tar.xz"): return ".tar.xz"
    return os.path.splitext(n)[1]


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def web_mobile_inbox_promote(request):
    user = request.user
    scan_type = request.data.get("scan_type", "sumiskai")

    # 1. Blocked session check
    if UploadSession.objects.filter(user=user, stage="blocked").exists():
        return Response({
            "error": "BLOCKED_SESSION_EXISTS",
            "detail": "Turite neapmokėtą užduotį. Papildykite kreditus arba panaikinkite užduotį.",
        }, status=409)

    # 2. Get documents (support select_all)
    select_all = request.data.get("select_all", False)
    exclude_ids = request.data.get("exclude_ids", [])

    if select_all:
        qs = MobileInboxDocument.objects.filter(user=user, is_processed=False)
        source = request.data.get("source")
        client_id = request.data.get("client_id")
        if source:
            qs = qs.filter(source=source)
        if client_id:
            qs = qs.filter(cloud_client_id=client_id)
        if exclude_ids:
            qs = qs.exclude(id__in=exclude_ids)
        mobile_docs = list(qs)
    else:
        ids = request.data.get("ids") or []
        if not isinstance(ids, list) or not ids:
            return Response(
                {"error": "NO_IDS", "detail": "Pateikite bent vieną ID."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        mobile_docs = list(
            MobileInboxDocument.objects
            .filter(user=user, is_processed=False, id__in=ids)
        )

    if not mobile_docs:
        return Response({"status": "OK", "count": 0, "processed_ids": []})

    # 3. Create UploadSession
    session = UploadSession.objects.create(
        user=user,
        scan_type=scan_type,
        stage="uploading",
        client_total_files=len(mobile_docs),
    )

    # 4. Copy files to ScannedDocuments
    processed_ids = []
    for mobile_doc in mobile_docs:
        if not mobile_doc.uploaded_file:
            continue
        try:
            with transaction.atomic():
                original_file = mobile_doc.uploaded_file
                original_file.open("rb")
                content = original_file.read()
                original_file.close()

                ext = _ext_lower(mobile_doc.original_filename)
                is_archive = ext in ARCHIVE_EXTS_PROMOTE

                scanned = ScannedDocument(
                    user=user,
                    original_filename=mobile_doc.original_filename,
                    status="pending",
                    scan_type=scan_type,
                    upload_session=session,
                    is_archive_container=is_archive,
                    uploaded_size_bytes=len(content),
                )
                scanned.file.save(
                    original_file.name.split("/")[-1],
                    ContentFile(content),
                    save=True,
                )
                scanned.save()

                mobile_doc.processed_document = scanned
                mobile_doc.processed_at = timezone.now()
                mobile_doc.is_processed = True
                mobile_doc.save(
                    update_fields=["processed_document", "processed_at", "is_processed"]
                )

                processed_ids.append(mobile_doc.id)

        except Exception as e:
            logger.error("Promote failed for doc %s: %s", mobile_doc.id, e)
            continue

    if not processed_ids:
        session.delete()
        return Response({"status": "OK", "count": 0, "processed_ids": []})

    # 5. Update session counters
    session.uploaded_files = len(processed_ids)
    session.save(update_fields=["uploaded_files"])

    # 6. Reserve credits + check
    session = reserve_and_queue(str(session.id), user.id)

    # 7. Start if not blocked
    if session.stage == "processing":
        start_session_processing.delay(str(session.id))

    return Response({
        "status": "OK",
        "count": len(processed_ids),
        "processed_ids": processed_ids,
        "session_id": str(session.id),
        "session_stage": session.stage,
        "error_message": session.error_message or None,
    })



@api_view(["POST"])
@permission_classes([IsAuthenticated])
def retry_blocked_session(request, session_id):
    """POST /api/web/sessions/<id>/retry/"""
    try:
        session = UploadSession.objects.get(id=session_id, user=request.user)
    except UploadSession.DoesNotExist:
        return Response(status=404)

    if session.stage != "blocked":
        return Response({"error": "Session is not blocked"}, status=400)

    # Reset rejected docs back to pending
    ScannedDocument.objects.filter(
        upload_session=session,
        status="rejected",
        error_message="Nepakanka kreditų",
    ).update(status="pending", error_message=None)

    # Reset counters, go straight to credit_check (NOT uploading)
    session.stage = "credit_check"
    session.error_message = ""
    session.processed_items = 0
    session.done_items = 0
    session.failed_items = 0
    session.save(update_fields=[
        "stage", "error_message",
        "processed_items", "done_items", "failed_items", "updated_at",
    ])

    try:
        session = reserve_and_queue(str(session.id), request.user.id)
    except Exception as e:
        logger.error("[RETRY] reserve_and_queue failed for session %s: %s", session.id, e)
        session.stage = "blocked"
        session.error_message = f"Klaida bandant pakartoti: {str(e)[:200]}"
        session.save(update_fields=["stage", "error_message", "updated_at"])
        return Response({
            "id": str(session.id),
            "stage": session.stage,
            "error_message": session.error_message,
        }, status=500)

    if session.stage == "processing":
        start_session_processing.delay(str(session.id))

    return Response({
        "id": str(session.id),
        "stage": session.stage,
        "error_message": session.error_message or None,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def cancel_blocked_session(request, session_id):
    """POST /api/web/sessions/<id>/cancel/ — отмена, возврат файлов в inbox"""
    try:
        session = UploadSession.objects.get(id=session_id, user=request.user)
    except UploadSession.DoesNotExist:
        return Response(status=404)

    if session.stage != "blocked":
        return Response({"error": "Session is not blocked"}, status=400)

    # Находим ScannedDocuments этой сессии
    scanned_docs = ScannedDocument.objects.filter(upload_session=session)
    scanned_ids = list(scanned_docs.values_list("id", flat=True))

    # Возвращаем файлы в inbox
    MobileInboxDocument.objects.filter(
        user=request.user,
        processed_document_id__in=scanned_ids,
    ).update(
        is_processed=False,
        processed_document=None,
        processed_at=None,
    )

    # Удаляем физические файлы + ScannedDocuments
    for doc in scanned_docs:
        if doc.file:
            try:
                doc.file.delete(save=False)
            except Exception:
                pass
    scanned_docs.delete()

    # Удаляем сессию
    session.delete()

    return Response({"status": "cancelled"})




@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def payments_list(request):
    """
    /api/payments/
    Возвращает историю платежей текущего пользователя.
    GET и POST делают одно и то же, чтобы не ломать твой привычный паттерн.
    """
    qs = (
        Payments.objects
        .filter(user=request.user)
        .order_by('-paid_at')
    )

    serializer = PaymentSerializer(
        qs,
        many=True,
        context={'request': request},  # важно для invoice_url в сериализаторе
    )
    return Response(serializer.data)



@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def payment_invoice(request, pk):
    """
    /api/payments/<pk>/invoice/
    Dati skirti PDF sąskaitai.
    """
    payment = get_object_or_404(Payments, pk=pk, user=request.user)
    user = request.user  # CustomUser

    # Pardavėjas 
    seller = {
        "pavadinimas": "Denis Orlov - DokSkenas",
        "iv_numeris": "1292165",
        "imonesKodas": "",  
        "pvmKodas": "",
        "adresas": "Kreivasis skg. 18-19, Vilnius",
        "telefonas": "",
        "bankoPavadinimas": "",
        "iban": "",
        "swift": "",
    }

    # Pirkėjas – klientas из CustomUser
    buyer = {
        "pavadinimas": user.company_name or user.email,
        "imonesKodas": user.company_code or "",
        "pvmKodas": user.vat_code or "",
        "adresas": user.company_address or "",
        "telefonas": "",
        "bankoPavadinimas": "",
        "iban": user.company_iban or "",
        "swift": "",
        "salis": user.company_country_iso or "",
    }

    data = {
        "id": payment.id,
        "dok_number": payment.dok_number,
        "paid_at": payment.paid_at,
        "credits_purchased": payment.credits_purchased,
        "net_amount": payment.net_amount,
        "currency": (payment.currency or "EUR").upper(),
        "buyer_email": payment.buyer_email,
        "buyer_address": payment.buyer_address_json,
        "seller": seller,
        "buyer": buyer,
    }

    return Response(data)


#Pagination dlia DocumentsTable

# Optimizacija skorosti zagruzki

def company_key(name, vat, cp_id):
    cp_id = (cp_id or "").strip()
    if cp_id:
        return f"id:{cp_id}"
    vat = (vat or "").strip().lower()
    if vat:
        return vat
    name = (name or "").strip().lower()
    return name


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_user_counterparties(request):
    user = request.user
    q = request.query_params

    status_param = q.get("status")
    date_from = q.get("from")
    date_to = q.get("to")
    search = (q.get("q") or "").strip().lower()
    limit = int(q.get("limit") or 200)

    qs = ScannedDocument.objects.filter(user=user, is_archive_container=False)

    if status_param:
        qs = qs.filter(status=status_param)

    tz = timezone.get_current_timezone()

    if date_from:
        d = parse_date(date_from)
        if d:
            dt_from = timezone.make_aware(datetime.combine(d, dt_time.min), tz)
            qs = qs.filter(uploaded_at__gte=dt_from)

    if date_to:
        d = parse_date(date_to)
        if d:
            dt_to = timezone.make_aware(datetime.combine(d, dt_time.min), tz) + timedelta(days=1)
            qs = qs.filter(uploaded_at__lt=dt_to)

    # агрегируем sellers
    sellers = (
        qs.exclude(seller_name__isnull=True, seller_name__exact="")
          .values("seller_id", "seller_name", "seller_vat_code")
          .annotate(docs_count=Count("id"))
    )

    # агрегируем buyers
    buyers = (
        qs.exclude(buyer_name__isnull=True, buyer_name__exact="")
          .values("buyer_id", "buyer_name", "buyer_vat_code")
          .annotate(docs_count=Count("id"))
    )

    merged = {}

    def upsert(cp_id, name, vat, cnt):
        key = company_key(name, vat, cp_id)
        if not key:
            return
        item = merged.get(key)
        if not item:
            merged[key] = {
                "key": key,
                "id": (cp_id or "").strip() or None,
                "name": name or "",
                "vat": vat or "",
                "docs_count": int(cnt or 0),
            }
        else:
            item["docs_count"] += int(cnt or 0)
            # “улучшаем” данные, если раньше было пусто
            if not item["id"] and cp_id:
                item["id"] = (cp_id or "").strip() or None
            if not item["vat"] and vat:
                item["vat"] = vat or ""
            if not item["name"] and name:
                item["name"] = name or ""

    for r in sellers:
        upsert(r.get("seller_id"), r.get("seller_name"), r.get("seller_vat_code"), r.get("docs_count"))

    for r in buyers:
        upsert(r.get("buyer_id"), r.get("buyer_name"), r.get("buyer_vat_code"), r.get("docs_count"))

    items = list(merged.values())

    # поиск по контрагентам (по имени/ват/id)
    if search:
        def match(it):
            return (
                search in (it.get("name") or "").lower()
                or search in (it.get("vat") or "").lower()
                or search in (it.get("id") or "").lower()
            )
        items = [x for x in items if match(x)]

    items.sort(key=lambda x: (-(x.get("docs_count") or 0), (x.get("name") or "").lower()))
    items = items[:limit]

    ser = CounterpartySerializer(items, many=True)
    return Response({"results": ser.data})




#Создать сессию

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_session(request):
    has_blocked = UploadSession.objects.filter(
        user=request.user,
        stage="blocked",
    ).exists()
    if has_blocked:
        return Response({
            "error": "BLOCKED_SESSION_EXISTS",
            "detail": "Turite neužbaigtą užduotį. Papildykite kreditus ir spauskite PAKARTOTI arba panaikinkite užduotį.",
        }, status=409)

    scan_type = (request.data.get("scan_type") or "sumiskai").strip()
    client_total_files = int(request.data.get("client_total_files") or 0)

    s = UploadSession.objects.create(
        user=request.user,
        scan_type=scan_type,
        stage="uploading",
        client_total_files=max(0, client_total_files),
    )
    return Response({"id": str(s.id), "stage": s.stage})

# @api_view(["POST"])
# @permission_classes([IsAuthenticated])
# def create_session(request):
#     stuck = UploadSession.objects.filter(
#         user=request.user,
#         stage__in=["blocked", "uploading"],
#     ).exists()
#     if stuck:
#         return Response({
#             "error": "BLOCKED_SESSION_EXISTS",
#             "detail": "Turite neužbaigtą užduotį. Papildykite kreditus ir spauskite PAKARTOTI arba panaikinkite užduotį.",
#         }, status=409)

#     scan_type = (request.data.get("scan_type") or "sumiskai").strip()
#     client_total_files = int(request.data.get("client_total_files") or 0)

#     multi_doc = bool(request.data.get("multi_doc", False))

#     s = UploadSession.objects.create(
#         user=request.user,
#         scan_type=scan_type,
#         stage="uploading",
#         client_total_files=max(0, client_total_files),
#         multi_doc=multi_doc,
#     )
#     return Response({"id": str(s.id), "stage": s.stage})


#Статус сессии (для progress bar)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def session_status(request, session_id):
    s = UploadSession.objects.get(id=session_id, user=request.user)

    return Response({
        "id": str(s.id),
        "stage": s.stage,
        "client_total_files": s.client_total_files,
        "uploaded_files": s.uploaded_files,
        "uploaded_bytes": s.uploaded_bytes,
        "expected_items": s.expected_items,
        "actual_items": s.actual_items,
        "processed_items": s.processed_items,
        "done_items": s.done_items,
        "failed_items": s.failed_items,
        "pending_archives": s.pending_archives,
        "reserved_credits": str(s.reserved_credits),
        "error_message": s.error_message,
        "created_at": s.created_at,
        "updated_at": s.updated_at,
        "multi_doc": s.multi_doc,
    })



#Upload обычных файлов батчами

MAX_BATCH_BYTES = 300 * 1024 * 1024
MAX_BATCH_FILES = 60

ARCHIVE_EXTS = {".zip",".rar",".7z",".tar",".tgz",".tar.gz",".tar.bz2",".tar.xz",".tbz2"}

def _ext(name: str) -> str:
    n = (name or "").lower()
    if n.endswith(".tar.gz"): return ".tar.gz"
    if n.endswith(".tar.bz2"): return ".tar.bz2"
    if n.endswith(".tar.xz"): return ".tar.xz"
    import os
    return os.path.splitext(n)[1]

def _get_archive_format(filename: str) -> str:
    """Определяет формат архива по расширению"""
    lower = filename.lower()
    if lower.endswith('.tar.gz') or lower.endswith('.tgz'):
        return 'tar.gz'
    if lower.endswith('.tar.bz2') or lower.endswith('.tbz2'):
        return 'tar.bz2'
    if lower.endswith('.tar.xz'):
        return 'tar.xz'
    if lower.endswith('.tar'):
        return 'tar'
    if lower.endswith('.zip'):
        return 'zip'
    if lower.endswith('.rar'):
        return 'rar'
    if lower.endswith('.7z'):
        return '7z'
    return ''

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_batch(request, session_id):
    s = UploadSession.objects.get(id=session_id, user=request.user)
    if s.stage not in ("uploading",):
        return Response({"error": "Session is not in uploading stage"}, status=400)

    files = list(request.FILES.getlist("files") or [])
    if not files:
        return Response({"error": "No files"}, status=400)

    if len(files) > MAX_BATCH_FILES:
        return Response({"error": f"Too many files in batch (max {MAX_BATCH_FILES})"}, status=400)

    total_bytes = sum(getattr(f, "size", 0) or 0 for f in files)
    if total_bytes > MAX_BATCH_BYTES:
        return Response({"error": f"Batch too large (max {MAX_BATCH_BYTES} bytes)"}, status=400)

    # запретим архивы тут — архивы только через chunk upload
    for f in files:
        if _ext(f.name) in ARCHIVE_EXTS:
            return Response({"error": "Archives must be uploaded via chunk upload"}, status=400)

    created_ids = []
    with transaction.atomic():
        for idx, f in enumerate(files, start=1):
            doc = ScannedDocument.objects.create(
                user=request.user,
                upload_session=s,
                status="pending",
                original_filename=f.name,
                scan_type=s.scan_type,
                uploaded_size_bytes=int(getattr(f, "size", 0) or 0),
            )
            doc.file.save(f.name, f, save=True)
            created_ids.append(doc.id)

        s.uploaded_files = s.uploaded_files + len(files)
        s.uploaded_bytes = s.uploaded_bytes + int(total_bytes)
        s.save(update_fields=["uploaded_files","uploaded_bytes","updated_at"])

    return Response({"ok": True, "created": len(created_ids)})



#Chunk upload для архивов (ZIP/RAR/7Z/TAR), max 2GB

MAX_ARCHIVE_SIZE = 2 * 1024 * 1024 * 1024  # 2GB

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def chunk_init(request, session_id):
    s = UploadSession.objects.get(id=session_id, user=request.user)
    if s.stage != "uploading":
        return Response({"error":"Session not uploading"}, status=400)

    filename = (request.data.get("filename") or "").strip()
    total_size = int(request.data.get("total_size") or 0)
    chunk_size = int(request.data.get("chunk_size") or 0)
    total_chunks = int(request.data.get("total_chunks") or 0)

    if not filename or total_size <= 0 or chunk_size <= 0 or total_chunks <= 0:
        return Response({"error":"Bad init params"}, status=400)

    if total_size > MAX_ARCHIVE_SIZE:
        return Response({"error":"Archive too large (max 2GB)"}, status=400)

    if _ext(filename) not in ARCHIVE_EXTS:
        return Response({"error":"Not an archive filename"}, status=400)

    # создаём tmp file path
    import os, tempfile
    tmp_dir = os.path.join(tempfile.gettempdir(), "doksken_chunks")
    os.makedirs(tmp_dir, exist_ok=True)

    cu = ChunkedUpload.objects.create(
        user=request.user,
        session=s,
        filename=filename,
        total_size=total_size,
        chunk_size=chunk_size,
        total_chunks=total_chunks,
        received=[],
        status="uploading",
        tmp_path=os.path.join(tmp_dir, f"{uuid.uuid4().hex}.part"),
    )

    return Response({"upload_id": str(cu.id)})


#upload chunk

@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def upload_chunk(request, session_id, upload_id, index):
    s = UploadSession.objects.get(id=session_id, user=request.user)
    cu = ChunkedUpload.objects.get(id=upload_id, user=request.user, session=s)

    if cu.status != "uploading":
        return Response({"error":"Not uploading"}, status=400)

    index = int(index)
    if index < 0 or index >= cu.total_chunks:
        return Response({"error":"Bad index"}, status=400)

    data = request.body or b""
    if not data:
        return Response({"error":"Empty chunk"}, status=400)

    # проверки размера чанка
    is_last = (index == cu.total_chunks - 1)
    if not is_last and len(data) != cu.chunk_size:
        return Response({"error":"Bad chunk size"}, status=400)
    if is_last and len(data) > cu.chunk_size:
        return Response({"error":"Bad last chunk size"}, status=400)

    offset = index * cu.chunk_size
    if offset + len(data) > cu.total_size:
        return Response({"error":"Out of bounds"}, status=400)

    # пишем по смещению
    import os
    os.makedirs(os.path.dirname(cu.tmp_path), exist_ok=True)
    with open(cu.tmp_path, "ab") as f:
        pass  # ensure file exists
    with open(cu.tmp_path, "r+b") as f:
        f.seek(offset)
        f.write(data)

    # отметить чанк полученным (атомарно)
    with transaction.atomic():
        cu = ChunkedUpload.objects.select_for_update().get(id=cu.id)
        got = set(cu.received or [])
        if index not in got:
            got.add(index)
            cu.received = sorted(got)
            cu.save(update_fields=["received","updated_at"])

        received_count = len(cu.received)

    return Response({"ok": True, "received": received_count, "total": cu.total_chunks})



#status (resume)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def chunk_status(request, session_id, upload_id):
    s = UploadSession.objects.get(id=session_id, user=request.user)
    cu = ChunkedUpload.objects.get(id=upload_id, user=request.user, session=s)
    return Response({
        "upload_id": str(cu.id),
        "status": cu.status,
        "received": cu.received,
        "total_chunks": cu.total_chunks,
    })


#complete → создать архив-контейнер ScannedDocument

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def chunk_complete(request, session_id, upload_id):
    s = UploadSession.objects.get(id=session_id, user=request.user)
    cu = ChunkedUpload.objects.get(id=upload_id, user=request.user, session=s)

    if cu.status != "uploading":
        return Response({"error":"Bad state"}, status=400)

    got = set(cu.received or [])
    if len(got) != cu.total_chunks:
        return Response({"error":"Not all chunks uploaded"}, status=400)

    if not os.path.exists(cu.tmp_path):
        return Response({"error":"Missing tmp file"}, status=400)

    if os.path.getsize(cu.tmp_path) != cu.total_size:
        return Response({"error":"Size mismatch"}, status=400)

    # НОВОЕ: определяем формат архива
    archive_fmt = _get_archive_format(cu.filename)

    # атомарно создаём архив-документ
    with transaction.atomic():
        # пометим upload complete
        cu.status = "complete"
        cu.save(update_fields=["status","updated_at"])

        # создаём ScannedDocument container
        doc = ScannedDocument.objects.create(
            user=request.user,
            upload_session=s,
            status="pending",
            original_filename=cu.filename,
            scan_type=s.scan_type,
            is_archive_container=True,
            uploaded_size_bytes=cu.total_size,
        )

        # переносим tmp файл в FileField (через open)
        with open(cu.tmp_path, "rb") as fp:
            doc.file.save(cu.filename, File(fp), save=True)

        # счётчики upload
        s.uploaded_files = s.uploaded_files + 1
        s.uploaded_bytes = s.uploaded_bytes + int(cu.total_size)
        
        # НОВОЕ: сохраняем формат архива в сессии
        if archive_fmt:
            current_formats = s.archive_formats or []
            if archive_fmt not in current_formats:
                current_formats.append(archive_fmt)
            s.archive_formats = current_formats

        s.save(update_fields=["uploaded_files", "uploaded_bytes", "archive_formats", "updated_at"])

    # можно удалить tmp_path после успешного save (если storage локальный)
    try:
        os.remove(cu.tmp_path)
    except Exception:
        pass

    return Response({"ok": True, "doc_id": doc.id})


###Finalize: reserve credits + поставить в queued/processing

#Подсчёт expected_items

COST = {"sumiskai": Decimal("1.00"), "detaliai": Decimal("1.30")}

def compute_expected_items(session: UploadSession) -> int:
    # обычные файлы
    base = ScannedDocument.objects.filter(upload_session=session, is_archive_container=False).count()

    # архивы (минимальный preflight)
    archives = ScannedDocument.objects.filter(upload_session=session, is_archive_container=True)
    total_inside = 0

    for a in archives:
        path = a.file.path
        ext = _ext(a.original_filename)
        count = 0

        if ext == ".zip":
            with zipfile.ZipFile(path) as zf:
                for zi in zf.infolist():
                    if zi.is_dir():
                        continue
                    if not zi.filename:
                        continue
                    count += 1

        elif ext in {".tar", ".tgz", ".tar.gz", ".tar.bz2", ".tar.xz", ".tbz2"}:
            with tarfile.open(path, mode="r:*") as tf:
                for m in tf.getmembers():
                    if not m.isfile():
                        continue
                    count += 1

        elif ext == ".7z":
            try:
                import py7zr
                with py7zr.SevenZipFile(path, mode='r') as sz:
                    for name in sz.getnames():
                        if not name.endswith('/'):
                            count += 1
            except Exception as e:
                logger.warning(f"Failed to read 7z archive {a.original_filename}: {e}")
                count = 1

        elif ext == ".rar":
            try:
                import rarfile
                with rarfile.RarFile(path) as rf:
                    for ri in rf.infolist():
                        if ri.isdir():
                            continue
                        count += 1
            except Exception as e:
                logger.warning(f"Failed to read rar archive {a.original_filename}: {e}")
                count = 1

        else:
            count = 1

        a.archive_file_count = count
        a.save(update_fields=["archive_file_count"])

        total_inside += count
    return base + total_inside




#reserve + stage

@transaction.atomic
def reserve_and_queue(session_id, user_id):
    s = UploadSession.objects.select_for_update().get(id=session_id, user_id=user_id)
    u = CustomUser.objects.select_for_update().get(id=user_id)

    if s.stage not in ("uploading", "credit_check"):
        return s

    s.stage = "credit_check"
    s.save(update_fields=["stage","updated_at"])

    expected = compute_expected_items(s)
    s.expected_items = expected
    s.reserved_items = expected

    cost = COST.get(s.scan_type, Decimal("1.00"))
    needed = cost * Decimal(expected)

    available = (u.credits or Decimal("0")) - (u.credits_reserved or Decimal("0"))
    if available < needed:
        s.stage = "blocked"
        s.error_message = f"Nepakanka kreditų. Turite: {available:.0f} | Reikia: {needed:.0f}"
        s.reserved_credits = Decimal("0.00")
        s.save(update_fields=["stage","error_message","expected_items","reserved_items","reserved_credits","updated_at"])

        # Reject all pending documents in this session
        blocked_count = ScannedDocument.objects.filter(
            upload_session=s,
            status="pending",
        ).update(
            status="rejected",
            error_message="Nepakanka kreditų",
        )
        logger.info(
            "[SESSION] Blocked session %s: rejected %d pending docs (available=%.0f, needed=%.0f)",
            s.id, blocked_count, available, needed,
        )

        return s

    # reserve
    u.credits_reserved = (u.credits_reserved or Decimal("0")) + needed
    u.save(update_fields=["credits_reserved"])

    s.reserved_credits = needed

    has_processing = UploadSession.objects.filter(user_id=u.id, stage="processing").exists()
    s.stage = "queued" if has_processing else "processing"
    if s.stage == "processing" and not s.started_at:
        s.started_at = timezone.now()

    s.save(update_fields=["stage","expected_items","reserved_items","reserved_credits","started_at","updated_at"])
    return s



#finalize endpoint

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def finalize_session(request, session_id):
    try:
        s = UploadSession.objects.get(id=session_id, user=request.user)
    except UploadSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)
    
    if s.stage not in ("uploading", "credit_check"):
        return Response({
            "id": str(s.id),
            "stage": s.stage,
            "error": "Session already finalized"
        }, status=400)
    
    if UploadSession.objects.filter(user=request.user, stage="blocked").exists():
        return Response({
            "error": "BLOCKED_SESSION_EXISTS",
            "detail": "Turite neapmokėtą užduotį. Papildykite kreditus arba panaikinkite užduotį.",
        }, status=409)
    
    # Проверить что есть файлы
    docs_count = ScannedDocument.objects.filter(upload_session=s).count()
    if docs_count == 0:
        return Response({"error": "No files uploaded"}, status=400)
    
    s = reserve_and_queue(session_id, request.user.id)
    
    if s.stage == "processing":
        start_session_processing.delay(str(s.id))
    
    return Response({
        "id": str(s.id),
        "stage": s.stage,
        "expected_items": s.expected_items,
        "reserved_credits": str(s.reserved_credits),
        "error_message": s.error_message or None,
    })



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def active_sessions(request):
    """Получить все активные сессии пользователя + недавно завершённые"""
    now = timezone.now()
    
    # Активные сессии
    active_qs = UploadSession.objects.filter(
        user=request.user,
        stage__in=["processing", "queued", "credit_check", "blocked"]
    )
    
    # Недавно завершённые (за последние 10 секунд) — чтобы показать финальный статус
    recently_done_qs = UploadSession.objects.filter(
        user=request.user,
        stage="done",
        finished_at__gte=now - timedelta(seconds=10)
    )
    
    sessions = list(active_qs) + list(recently_done_qs)
    sessions.sort(key=lambda s: s.created_at)
    
    result = []
    for s in sessions:
        result.append({
            "id": str(s.id),
            "stage": s.stage,
            "scan_type": s.scan_type,
            "uploaded_files": s.uploaded_files,
            "expected_items": s.expected_items,
            "actual_items": s.actual_items,
            "processed_items": s.processed_items,
            "done_items": s.done_items,
            "failed_items": s.failed_items,
            "created_at": s.created_at.isoformat(),
            "finished_at": s.finished_at.isoformat() if s.finished_at else None,
        })
    
    return Response({"sessions": result})



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def swap_buyer_seller(request, pk):
    """
    Меняет местами данные buyer и seller.
    - Обычный юзер может менять только свои документы
    - Superuser может менять документы любого юзера
    """
    try:
        doc = ScannedDocument.objects.get(pk=pk)
    except ScannedDocument.DoesNotExist:
        return Response({'error': 'Dokumentas nerastas.'}, status=status.HTTP_404_NOT_FOUND)
    
    # Проверка прав: либо владелец, либо superuser
    if doc.user != request.user and not request.user.is_superuser:
        return Response(
            {'error': 'Neturite teisės keisti šio dokumento.'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    # Swap всех полей
    swap_pairs = [
        ('seller_id_programoje', 'buyer_id_programoje'),
        ('seller_id', 'buyer_id'),
        ('seller_name', 'buyer_name'),
        ('seller_vat_code', 'buyer_vat_code'),
        ('seller_address', 'buyer_address'),
        ('seller_country', 'buyer_country'),
        ('seller_country_iso', 'buyer_country_iso'),
        ('seller_iban', 'buyer_iban'),
        ('seller_is_person', 'buyer_is_person'),
        ('seller_name_normalized', 'buyer_name_normalized'),
        ('seller_vat_val', 'buyer_vat_val'),
    ]
    
    for seller_field, buyer_field in swap_pairs:
        seller_val = getattr(doc, seller_field)
        buyer_val = getattr(doc, buyer_field)
        setattr(doc, seller_field, buyer_val)
        setattr(doc, buyer_field, seller_val)
    
    doc.save()
    
    return Response({
        'success': True,
        'seller_name': doc.seller_name,
        'buyer_name': doc.buyer_name,
    })


#proveriajet status zadachi exporta cerez API

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_sessions_active(request):
    """
    Возвращает активные и недавно завершённые ExportSession для текущего юзера.
    Фронт поллит каждые 2 секунды для progress bar.
    """
    from docscanner_app.models import ExportSession
    from django.utils import timezone
    from datetime import timedelta

    # Активные сессии (queued + processing)
    active = ExportSession.objects.filter(
        user=request.user,
        stage__in=[ExportSession.Stage.QUEUED, ExportSession.Stage.PROCESSING],
    )

    # Недавно завершённые (за последние 10 секунд) — чтобы фронт увидел финальное состояние
    recent_done = ExportSession.objects.filter(
        user=request.user,
        stage=ExportSession.Stage.DONE,
        finished_at__gte=timezone.now() - timedelta(seconds=10),
    )

    sessions = list(active) + list(recent_done)

    data = {
        "sessions": [
            {
                "id": s.pk,
                "program": s.program,
                "stage": s.stage,
                "total_documents": s.total_documents,
                "processed_documents": s.processed_documents,
                "success_count": s.success_count,
                "partial_count": s.partial_count,
                "error_count": s.error_count,
                "created_at": s.created_at.isoformat() if s.created_at else None,
                "started_at": s.started_at.isoformat() if s.started_at else None,
                "finished_at": s.finished_at.isoformat() if s.finished_at else None,
                "total_time_seconds": s.total_time_seconds,
                "has_invoices": s.invoice_documents.exists(),
            }
            for s in sessions
        ]
    }

    return Response(data, status=200)


# Proverka errors exportirovanyx dokumentov cerez API (dlia documentstable)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_log_detail(request, document_id):
    """
    Возвращает последний APIExportLog для документа + вложенные article_logs.
    Для popup при клике на статус в DocumentsTable.
    Query params:
      ?program=optimum (по умолчанию optimum)
    """
    from docscanner_app.models import APIExportLog

    program = request.query_params.get("program", "optimum")

    export_log = (
        APIExportLog.objects
        .filter(
            document_id=document_id,
            user=request.user,
            program=program,
        )
        .prefetch_related("article_logs")
        .order_by("-created_at")
        .first()
    )

    if not export_log:
        return Response({"error": "No export log found"}, status=404)

    data = {
        "id": export_log.pk,
        "status": export_log.status,
        "created_at": export_log.created_at.isoformat(),
        "partner_status": getattr(export_log, "partner_status", None),
        "partner_error": getattr(export_log, "partner_error", None),
        "invoice_type": export_log.invoice_type,
        "invoice_status": export_log.invoice_status,
        "invoice_result": export_log.invoice_result,
        "invoice_error": export_log.invoice_error,
        "articles": [
            {
                "article_name": a.article_name,
                "article_code": a.article_code,
                "status": a.status,
                "result": a.result,
                "error": a.error,
            }
            for a in export_log.article_logs.all()
        ],
    }

    return Response(data, status=200)




#Dokumenty iz emailov

MAX_MAILGUN_FILE_SIZE = 25 * 1024 * 1024  # 25 MB


@csrf_exempt
@require_POST
def mailgun_inbound(request):
    """
    POST /api/mailgun/inbound/
    Webhook от Mailgun — принимает входящие email и сохраняет вложения
    в MobileInboxDocument с source='email'.
    """

    # 1. Проверяем подпись Mailgun
    signing_key = getattr(settings, 'MAILGUN_WEBHOOK_SIGNING_KEY', '')
    if not signing_key:
        logger.error("MAILGUN_WEBHOOK_SIGNING_KEY not configured")
        return HttpResponseForbidden('Webhook not configured')

    token = request.POST.get('token', '')
    timestamp = request.POST.get('timestamp', '')
    signature = request.POST.get('signature', '')

    try:
        import time as time_module
        if abs(time_module.time() - int(timestamp)) > 300:
            logger.warning("Mailgun webhook: stale timestamp")
            return HttpResponseForbidden('Stale request')
    except (ValueError, TypeError):
        return HttpResponseForbidden('Invalid timestamp')

    expected = hmac.new(
        key=signing_key.encode('utf-8'),
        msg=f'{timestamp}{token}'.encode('utf-8'),
        digestmod=hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(signature, expected):
        logger.warning("Mailgun webhook: invalid signature")
        return HttpResponseForbidden('Invalid signature')  # ВРЕМЕННО ОТКЛЮЧЕНО

    # 2. Находим пользователя по inbox-токену
    recipient = request.POST.get('recipient', '')
    inbox_token = recipient.split('@')[0].lower().strip()

    if not inbox_token:
        return HttpResponse('ok', status=200)

    try:
        user = CustomUser.objects.get(email_inbox_token=inbox_token)
    except CustomUser.DoesNotExist:
        logger.info(f"Mailgun inbound: unknown token '{inbox_token}'")
        return HttpResponse('ok', status=200)

    sender_email = request.POST.get('sender', '')
    subject = request.POST.get('subject', '')

    # 3. Сохраняем вложения
    saved_count = 0
    skipped_count = 0

    for key, uploaded_file in request.FILES.items():
        # Проверка расширения
        ext = os.path.splitext(uploaded_file.name)[1].lower()
        if ext not in SUPPORTED_EXTS:
            logger.debug(f"Mailgun inbound: skip unsupported ext '{ext}' ({uploaded_file.name})")
            skipped_count += 1
            continue

        # Проверка размера
        file_size = getattr(uploaded_file, 'size', 0) or 0
        if file_size > MAX_MAILGUN_FILE_SIZE:
            logger.warning(f"Mailgun inbound: skip too large file ({file_size} bytes): {uploaded_file.name}")
            skipped_count += 1
            continue

        doc = MobileInboxDocument.objects.create(
            user=user,
            uploaded_file=uploaded_file,
            original_filename=uploaded_file.name,
            size_bytes=file_size,
            sender_email=sender_email,
            source='email',
            sender_subject=subject,
            is_processed=False,
        )

        doc.preview_url = f"{settings.SITE_URL_BACKEND}{doc.uploaded_file.url}"
        doc.save(update_fields=["preview_url"])

        saved_count += 1

    logger.info(
        f"Mailgun inbound: user={user.email}, token={inbox_token}, "
        f"sender={sender_email}, saved={saved_count}, skipped={skipped_count}"
    )

    return HttpResponse('ok', status=200)




#NEW - saskaitu israsymas
"""
DokSkenas — Sąskaitų išrašymas
Views: CRUD для Invoice, Counterparty, InvoiceSettings + бизнес-действия.
"""

from datetime import timedelta
from decimal import Decimal

from django.db import transaction
from django.db.models import Q, Sum, Count, Case, When, BooleanField, Value
from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response

from .models import Counterparty, InvoiceSettings, Invoice, InvoiceLineItem
from .serializers import (
    InvoiceCounterpartySerializer,
    InvoiceCounterpartyListSerializer,
    InvoiceSettingsSerializer,
    InvoiceListSerializer,
    InvoiceDetailSerializer,
    InvoiceWriteSerializer,
    InvoicePublicSerializer,
    InvoiceLineItemSerializer,
)


# ════════════════════════════════════════════════════════════
# Counterparty
# ════════════════════════════════════════════════════════════

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def counterparty_list_create(request):
    """
    GET  — список контрагентов (с поиском ?q=..., фильтром ?role=...)
           Пагинация: ?limit=25&offset=0
    POST — создать нового контрагента
    """
    user = request.user
    active_id = getattr(user, "active_company_profile_id", None)

    if request.method == "GET":
        qs = Counterparty.objects.filter(user=user)
        if active_id:
            qs = qs.filter(company_profile_id=active_id)

        q = request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(company_code__icontains=q)
                | Q(vat_code__icontains=q)
            )

        role = request.GET.get("role")
        if role in ("buyer", "seller"):
            qs = qs.filter(Q(default_role=role) | Q(default_role="both"))

        # Новые сверху
        qs = qs.order_by("-id")
        total = qs.count()

        try:
            limit = int(request.GET.get("limit", 25))
            offset = int(request.GET.get("offset", 0))
        except (ValueError, TypeError):
            limit, offset = 25, 0
        limit = max(1, min(limit, 250))
        offset = max(offset, 0)

        page = qs[offset : offset + limit]
        serializer = InvoiceCounterpartyListSerializer(page, many=True)
        return Response({"results": serializer.data, "count": total})

    # POST
    serializer = InvoiceCounterpartySerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save(user=user, company_profile_id=active_id)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["GET", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def counterparty_detail(request, pk):
    """GET / PUT / DELETE одного контрагента."""
    cp = get_object_or_404(Counterparty, pk=pk, user=request.user)

    if request.method == "GET":
        return Response(InvoiceCounterpartySerializer(cp).data)

    if request.method == "PUT":
        serializer = InvoiceCounterpartySerializer(cp, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    # DELETE
    cp.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


# ════════════════════════════════════════════════════════════
# InvoiceSettings
# ════════════════════════════════════════════════════════════

def _build_seller_from_profile(profile):
    """Реквизиты продавца из CompanyProfile. Ветвление ИВ / юр.лицо."""
    if profile is None:
        return {
            "id": None, "entity_type": None,
            "name": "", "company_code": "", "vat_code": "", "address": "",
            "email": "", "phone": "", "bank_name": "", "bank_swift": "",
            "bank_account": "", "country_iso": "",
        }

    is_iv = getattr(profile, "entity_type", None) == "iv"

    if is_iv:
        name = (profile.owner_name or profile.name or "")
        company_code = (profile.iv_certificate_nr or "")
    else:
        name = (profile.name or "")
        company_code = (profile.company_code or "")

    return {
        "id": profile.id,
        "entity_type": getattr(profile, "entity_type", None),
        "name": name,
        "company_code": company_code,
        "vat_code": profile.vat_code or "",
        "address": profile.address or "",
        "email": profile.email or "",
        "phone": profile.phone or "",
        "bank_name": profile.bank_name or "",
        "bank_swift": profile.swift or "",
        "bank_account": profile.iban or "",
        "country_iso": profile.country_iso or "",
    }


@api_view(["GET", "PUT"])
@permission_classes([IsAuthenticated])
def invoice_settings(request):
    active_profile = None
    active_id = getattr(request.user, "active_company_profile_id", None)
    if active_id:
        from .models import CompanyProfile
        active_profile = CompanyProfile.objects.filter(id=active_id).first()

    obj, _ = InvoiceSettings.objects.get_or_create(
        user=request.user,
        company_profile=active_profile,
    )

    if request.method == "GET":
        data = InvoiceSettingsSerializer(obj, context={"request": request}).data
        data["payment_providers"] = request.user.payment_providers or {}
        data["seller"] = _build_seller_from_profile(active_profile)
        return Response(data)

    serializer = InvoiceSettingsSerializer(
        obj, data=request.data, partial=True, context={"request": request}
    )
    serializer.is_valid(raise_exception=True)
    serializer.save()

    resp_data = serializer.data
    resp_data["payment_providers"] = request.user.payment_providers or {}
    resp_data["seller"] = _build_seller_from_profile(active_profile)
    return Response(resp_data)


# ════════════════════════════════════════════════════════════
# Invoice List — с поддержкой категорий
# ════════════════════════════════════════════════════════════

from django.db.models import Count, Exists, OuterRef, Q

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_list(request):
    user = request.user
    today = date.today()

    qs = (
        Invoice.objects
        .filter(user=user)
        .select_related("scanned_document")
        .annotate(
            line_items_count=Count("line_items"),
            _has_proposed=Exists(
                PaymentAllocation.objects.filter(
                    invoice=OuterRef("pk"),
                    status="proposed",
                )
            ),
        )
    )


    # ── Company profile filter ──
    company_id = request.GET.get("company_profile")
    if company_id:
        qs = qs.filter(company_profile_id=company_id)
    else:
        active_id = getattr(user, "active_company_profile_id", None)
        if active_id:
            qs = qs.filter(company_profile_id=active_id)

    category = (request.GET.get("category") or "").strip().lower()
    status_param = (request.GET.get("status") or "").strip().lower()

    # ── Category filter ──
    if category == "israsytos":
        # Base: issued + sent, NOT overdue
        qs = qs.filter(status__in=["issued", "sent"]).filter(
            Q(due_date__gte=today) | Q(due_date__isnull=True)
        )
        # Sub-filter within israsytos
        if status_param in ("issued", "sent"):
            qs = qs.filter(status=status_param)

    elif category == "veluojancios":
        qs = qs.filter(status__in=["issued", "sent"], due_date__lt=today)
        if status_param in ("issued", "sent"):
            qs = qs.filter(status=status_param)

    elif category == "apmoketos":
        qs = qs.filter(status__in=["paid", "partially_paid"])

    elif category == "juodrasciai":
        qs = qs.filter(status="draft")

    elif category == "cancelled":
        qs = qs.filter(status="cancelled")

    else:
        # No category — old-style status filter
        if status_param:
            statuses = [s.strip() for s in status_param.split(",") if s.strip()]
            qs = qs.filter(status__in=statuses)

    # ── Exported filter ──
    exported_param = (request.GET.get("exported") or "").strip().lower()
    if exported_param == "true":
        qs = qs.filter(exported=True)
    elif exported_param == "false":
        qs = qs.filter(exported=False)

    # ── Invoice type ──
    invoice_type = request.GET.get("invoice_type")
    if invoice_type:
        qs = qs.filter(invoice_type=invoice_type)

    # ── Search ──
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(buyer_name__icontains=q)
            | Q(buyer_id__icontains=q)
            | Q(document_number__icontains=q)
            | Q(document_series__icontains=q)
        )

    # ── Date range ──
    date_from = request.GET.get("date_from")
    if date_from:
        qs = qs.filter(invoice_date__gte=date_from)
    date_to = request.GET.get("date_to")
    if date_to:
        qs = qs.filter(invoice_date__lte=date_to)

    # ── Sort & paginate ──
    qs = qs.order_by("-created_at")

    limit = min(int(request.GET.get("limit", 50)), 200)
    offset = int(request.GET.get("offset", 0))
    total = qs.count()

    page = qs[offset : offset + limit]
    serializer = InvoiceListSerializer(
        page,
        many=True,
        context={"request": request},
    )

    return Response({
        "count": total,
        "limit": limit,
        "offset": offset,
        "results": serializer.data,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_create(request):
    """Создать новый счёт (draft)."""
    serializer = InvoiceWriteSerializer(
        data=request.data, context={"request": request}
    )
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_detail(request, pk):
    """Получить полный счёт с line items."""
    qs = Invoice.objects.select_related("scanned_document")

    if request.user.is_superuser:
        invoice = get_object_or_404(qs, pk=pk)
    else:
        invoice = get_object_or_404(qs, pk=pk, user=request.user)

    serializer = InvoiceDetailSerializer(invoice, context={"request": request})
    return Response(serializer.data)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def invoice_update(request, pk):
    """Обновить счёт."""
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)
    serializer = InvoiceWriteSerializer(
        invoice, data=request.data, partial=True, context={"request": request}
    )
    serializer.is_valid(raise_exception=True)
    instance = serializer.save()

    # ── Пересоздать JE если invoice уже issued ────────
    if instance.status not in ("draft", "cancelled") and instance.invoice_type != "isankstine":
        from .services.accounting_transfer import recreate_je_for_invoice
        try:
            recreate_je_for_invoice(instance)
        except Exception as e:
            logger.warning("[InvoiceUpdate] Recreate JE failed for %s: %s", instance.id, e)

    return Response(serializer.data)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def invoice_delete(request, pk):
    """Ištrinti sąskaitą. Galima tik juodraščiams ir iš skaitmenizavimo perkeltoms."""
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

    if not invoice.can_delete:
        return Response(
            {"detail": "Galima ištrinti tik juodraščius arba iš skaitmenizavimo perkeltas sąskaitas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    with transaction.atomic():
        # ── Удалить JE ──
        from .services.accounting_transfer import delete_je_for_invoice
        delete_je_for_invoice(invoice)

        # ── Сбросить флаг переноса на ScannedDocument ──
        if invoice.scanned_document:
            doc = invoice.scanned_document
            doc.perkelta_i_apskaita = False
            doc.perkelta_i_apskaita_at = None
            doc.perkelta_i_company_profile = None
            doc.save(update_fields=[
                "perkelta_i_apskaita",
                "perkelta_i_apskaita_at",
                "perkelta_i_company_profile",
            ])

        # ── Удалить PaymentAllocations и вернуть транзакции в unmatched ──
        for alloc in invoice.payment_allocations.all():
            from .services.accounting_transfer import delete_je_for_allocation
            delete_je_for_allocation(alloc)

            txn = alloc.transaction
            alloc.delete()

            if txn:
                from django.db.models import Sum
                new_total = txn.allocations.aggregate(t=Sum("amount"))["t"] or 0
                txn.allocated_amount = new_total
                if not txn.allocations.exists():
                    txn.match_status = "unmatched"
                    txn.match_confidence = 0
                    txn.transaction_category = ""
                txn.save(update_fields=[
                    "allocated_amount", "match_status",
                    "match_confidence", "transaction_category", "updated_at",
                ])
                if txn.bank_statement:
                    txn.bank_statement.refresh_stats()

        invoice.delete()

    return Response(status=status.HTTP_204_NO_CONTENT)


# ════════════════════════════════════════════════════════════
# Invoice — Line Items (отдельный endpoint для lazy loading)
# ════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_line_items(request, pk):
    """Line items для конкретного счёта (lazy load)."""
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)
    qs = invoice.line_items.order_by("sort_order", "id")
    serializer = InvoiceLineItemSerializer(qs, many=True)
    return Response(serializer.data)


# ════════════════════════════════════════════════════════════
# Invoice — Бизнес-действия
# ════════════════════════════════════════════════════════════

def _recalculate_credit_draft_totals(invoice):
    """
    Perskaičiuoja kreditinės sumas prieš išrašymą.

    Eilučių subtotal lieka prieš bendrą dokumento nuolaidą.
    Bendra nuolaida proporcingai paskirstoma pagal PVM tarifus.
    """
    from decimal import Decimal, ROUND_HALF_UP

    MONEY = Decimal("0.01")
    ZERO = Decimal("0")

    def dec(value):
        if value in (None, ""):
            return ZERO
        return Decimal(str(value))

    def round_money(value):
        return Decimal(value).quantize(
            MONEY,
            rounding=ROUND_HALF_UP,
        )

    invoice_lines = list(
        invoice.line_items.order_by("sort_order", "id")
    )

    sum_net = sum(
        (
            abs(dec(line.subtotal))
            for line in invoice_lines
        ),
        ZERO,
    )

    invoice_discount = min(
        abs(dec(invoice.invoice_discount_wo_vat)),
        sum_net,
    )

    amount_wo_vat = round_money(
        sum_net - invoice_discount
    )

    vat_groups = {}

    for line in invoice_lines:
        line_net = abs(dec(line.subtotal))

        vat_percent = dec(
            line.vat_percent
            if line.vat_percent is not None
            else invoice.vat_percent or 0
        )

        vat_percent = abs(vat_percent)

        if vat_percent not in vat_groups:
            vat_groups[vat_percent] = ZERO

        vat_groups[vat_percent] += line_net

    vat_amount = ZERO

    if (
        invoice.pvm_tipas == "taikoma"
        and sum_net > ZERO
    ):
        for vat_percent, group_net in vat_groups.items():
            ratio = group_net / sum_net

            discounted_group_net = round_money(
                group_net -
                invoice_discount * ratio
            )

            group_vat = round_money(
                discounted_group_net *
                vat_percent /
                Decimal("100")
            )

            vat_amount += group_vat

    vat_amount = round_money(vat_amount)
    amount_with_vat = round_money(
        amount_wo_vat + vat_amount
    )

    invoice.amount_wo_vat = amount_wo_vat
    invoice.vat_amount = vat_amount
    invoice.amount_with_vat = amount_with_vat

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_issue(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

    if invoice.status != "draft":
        return Response(
            {"detail": "Galima išrašyti tik juodraštį."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    errors = {}
    if not invoice.seller_name:
        errors["seller_name"] = "Pardavėjo pavadinimas privalomas."
    if not invoice.buyer_name:
        errors["buyer_name"] = "Pirkėjo pavadinimas privalomas."
    if not invoice.invoice_date:
        errors["invoice_date"] = "Sąskaitos data privaloma."
    if not invoice.line_items.exists():
        errors["line_items"] = "Būtina bent viena eilutė."
    if errors:
        return Response(errors, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        series_obj = None
        series_profile_id = invoice.company_profile_id or getattr(
            request.user, "active_company_profile_id", None
        )

        def _series_qs():
            qs = InvoiceSeries.objects.select_for_update().filter(
                user=request.user,
                invoice_type=invoice.invoice_type,
                is_active=True,
            )
            if series_profile_id is not None:
                qs = qs.filter(company_profile_id=series_profile_id)
            return qs

        if invoice.document_series:
            series_obj = _series_qs().filter(prefix=invoice.document_series).first()

        if not series_obj:
            series_obj = _series_qs().filter(is_default=True).first()

        if not series_obj:
            series_obj = _series_qs().first()

        # Если юзер уже задал номер вручную — проверяем уникальность
        if invoice.document_number:
            exists = Invoice.objects.filter(
                user=request.user,
                company_profile_id=series_profile_id,
                document_series=series_obj.prefix,
                document_number=invoice.document_number,
                invoice_type=invoice.invoice_type,
            ).exclude(pk=invoice.pk).exists()
            if exists:
                return Response(
                    {"detail": f"Numeris {series_obj.prefix}-{invoice.document_number} jau užimtas."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            prefix = series_obj.prefix
            number_str = invoice.document_number
        else:
            # Авто-присвоение
            prefix, number_str, number_int = series_obj.allocate_number()

        invoice.document_series = prefix
        invoice.document_number = number_str

        update_fields = [
            "document_series",
            "document_number",
            "status",
            "pvm_kodas",
        ]

        # Kreditinė: prieš keičiant ženklus perskaičiuojame bendrą
        # nuolaidą ir PVM pagal dabartines kreditinės eilutes.
        if invoice.invoice_type == "kreditine":
            _recalculate_credit_draft_totals(invoice)

            from decimal import Decimal, InvalidOperation

            def _negative(value):
                if value in (None, ""):
                    return value

                try:
                    decimal_value = Decimal(str(value))
                except (InvalidOperation, TypeError, ValueError):
                    return value

                return -abs(decimal_value) if decimal_value != 0 else decimal_value

            invoice_negative_fields = (
                "amount_wo_vat",
                "vat_amount",
                "amount_with_vat",
                "invoice_discount_wo_vat",
                "invoice_discount_with_vat",
                "delivery_fee",
            )

            for field in invoice_negative_fields:
                setattr(invoice, field, _negative(getattr(invoice, field)))

            update_fields.extend(invoice_negative_fields)

            line_negative_fields = (
                "quantity",
                "subtotal",
                "vat",
                "total",
                "discount_wo_vat",
                "discount_with_vat",
            )

            for line in invoice.line_items.select_for_update():
                for field in line_negative_fields:
                    setattr(line, field, _negative(getattr(line, field)))

                line.save(update_fields=list(line_negative_fields))

        invoice.status = "issued"
        invoice.assign_pvm_codes()

        if not invoice.company_profile_id:
            active_id = getattr(request.user, "active_company_profile_id", None)
            if active_id:
                invoice.company_profile_id = active_id
                update_fields.append("company_profile")

        invoice.save(update_fields=update_fields)

        # ── Auto JE: pardavimo SF → DK įrašas ────────────
        from .services.accounting_transfer import create_je_for_invoice
        try:
            create_je_for_invoice(invoice)
        except Exception as e:
            logger.warning("[InvoiceIssue] Auto JE failed for %s: %s", invoice.id, e)

    serializer = InvoiceDetailSerializer(invoice, context={"request": request})
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_send(request, pk):
    """
    Отправить счёт по email покупателю.
    Body: {"email": "buyer@example.com"} (опционально, иначе buyer_email)
    """
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

    if not invoice.can_be_sent:
        return Response(
            {"detail": "Sąskaita turi būti išrašyta prieš siunčiant."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    email = request.data.get("email") or invoice.buyer_email
    if not email:
        return Response(
            {"detail": "Nurodykite pirkėjo el. paštą."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # TODO: отправить email с PDF вложением
    # send_invoice_email(invoice, email)

    invoice.status = "sent"
    invoice.sent_at = timezone.now()
    invoice.sent_to_email = email
    invoice.save(update_fields=["status", "sent_at", "sent_to_email"])

    serializer = InvoiceDetailSerializer(invoice, context={"request": request})
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_mark_paid(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

    if invoice.status not in ("issued", "sent", "partially_paid"):
        return Response(
            {"detail": "Galima pažymėti tik išrašytą/išsiųstą/dalinai apmokėtą sąskaitą."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Validate input
    amount = request.data.get("amount")
    payment_date = request.data.get("payment_date")

    if not amount or not payment_date:
        return Response(
            {"detail": "Privalomi laukai: amount, payment_date."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    from decimal import Decimal, InvalidOperation
    try:
        amount = Decimal(str(amount))
    except (InvalidOperation, ValueError):
        return Response(
            {"detail": "Neteisinga suma."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    note = request.data.get("note", "")

    # Create PaymentAllocation
    from .services.payment_service import PaymentService
    svc = PaymentService(request.user)
    svc.mark_paid_manual(
        invoice=invoice,
        amount=amount,
        payment_date=payment_date,
        note=note,
    )

    # Auto SF creation
    invoice.refresh_from_db()
    from .services.auto_sf import maybe_auto_create_sf
    created_sf = maybe_auto_create_sf(invoice)

    data = InvoiceDetailSerializer(invoice, context={"request": request}).data
    if created_sf:
        data["auto_created_sf"] = {
            "id": created_sf.id,
            "full_number": created_sf.full_number,
            "status": created_sf.status,
        }

    return Response(data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_cancel(request, pk):
    """Atšaukti sąskaitą + каскад на связанные."""
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

    if invoice.status in ("cancelled", "draft"):
        return Response(
            {"detail": "Ši sąskaita negali būti anuliuota."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    now = timezone.now()

    with transaction.atomic():
        invoice.status = "cancelled"
        invoice.cancelled_at = now
        invoice.save(update_fields=["status", "cancelled_at", "updated_at"])

        # ── Удалить JE ────────────────────────────────────
        from .services.accounting_transfer import delete_je_for_invoice
        delete_je_for_invoice(invoice)

        # Каскад: išankstinė → derived SF/PVM SF
        if invoice.invoice_type == "isankstine":
            derived = list(
                invoice.derived_invoices.filter(
                    invoice_type__in=["pvm_saskaita", "saskaita"],
                ).exclude(status="cancelled")
            )
            for derived_inv in derived:
                derived_inv.status = "cancelled"
                derived_inv.cancelled_at = now
                derived_inv.save(update_fields=["status", "cancelled_at", "updated_at"])
                delete_je_for_invoice(derived_inv)

        # Каскад: SF/PVM SF → source išankstinė
        if (
            invoice.invoice_type in ("pvm_saskaita", "saskaita")
            and invoice.source_invoice_id
        ):
            Invoice.objects.filter(
                pk=invoice.source_invoice_id,
                invoice_type="isankstine",
                user=request.user,
            ).exclude(
                status="cancelled",
            ).update(status="cancelled", cancelled_at=now, updated_at=now)

    serializer = InvoiceDetailSerializer(invoice, context={"request": request})
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_duplicate(request, pk):
    """
    Скопировать счёт как новый draft.
    Копируются все данные и line items, но без номера/серии/статуса.
    """
    source = get_object_or_404(Invoice, pk=pk, user=request.user)

    # Поля которые НЕ копируем
    skip_fields = {
        "id", "uuid", "status", "document_series", "document_number",
        "pdf_file", "sent_at", "sent_to_email", "paid_at", "cancelled_at",
        "optimum_api_status", "optimum_last_try_date",
        "dineta_api_status", "dineta_last_try_date",
        "created_at", "updated_at",
    }

    new_data = {}
    for field in Invoice._meta.get_fields():
        if not hasattr(field, "attname"):
            continue
        name = field.attname
        if name in skip_fields:
            continue
        new_data[name] = getattr(source, name)

    new_data["status"] = "draft"
    new_data["invoice_date"] = timezone.now().date()

    with transaction.atomic():
        new_invoice = Invoice.objects.create(**new_data)

        # Копируем line items
        for li in source.line_items.order_by("sort_order", "id"):
            li_data = {}
            for field in InvoiceLineItem._meta.get_fields():
                if not hasattr(field, "attname"):
                    continue
                if field.attname in ("id", "invoice_id"):
                    continue
                li_data[field.attname] = getattr(li, field.attname)
            InvoiceLineItem.objects.create(invoice=new_invoice, **li_data)

    serializer = InvoiceDetailSerializer(new_invoice, context={"request": request})
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_create_pvm_sf(request, pk):
    source = get_object_or_404(Invoice, pk=pk, user=request.user)

    if not source.can_create_pvm_sf:
        if source.invoice_type != "isankstine":
            msg = "SF galima sukurti tik iš išankstinės sąskaitos."
        elif source.status not in ("issued", "sent", "paid"):
            msg = "Išankstinė sąskaita turi būti išrašyta."
        else:
            msg = "SF jau sukurta iš šios išankstinės sąskaitos."
        return Response({"detail": msg}, status=status.HTTP_400_BAD_REQUEST)

    from .services.auto_sf import create_sf_from_isankstine

    try:
        new_invoice = create_sf_from_isankstine(
            source=source,
            user=request.user,
            series_prefix=None,
        )
    except ValueError as e:
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    if not new_invoice:
        return Response({"detail": "SF jau sukurta."}, status=status.HTTP_400_BAD_REQUEST)

    # DK įrašas jau sukurtas viduje create_sf_from_isankstine().

    serializer = InvoiceDetailSerializer(new_invoice, context={"request": request})
    return Response(serializer.data, status=status.HTTP_201_CREATED)


class CreateCreditInvoiceView(APIView):
    """
    POST /api/invoices/<id>/create-credit/

    Создаёт kreditinę sąskaitą faktūrą из существующей SF/PVM SF.
    Копирует все данные и строки. Возвращает draft kreditinę
    для редактирования (юзер может убрать строки, изменить количество).

    Также можно создать kreditinę вручную через обычный create endpoint
    с invoice_type="kreditine" без source_invoice.
    """
    permission_classes = [permissions.IsAuthenticated]

    # Поля для копирования из оригинала
    COPY_FIELDS = [
        # Seller
        "seller_counterparty_id", "seller_id_programoje",
        "seller_name", "seller_name_normalized", "seller_id", "seller_vat_code",
        "seller_address", "seller_country", "seller_country_iso",
        "seller_phone", "seller_email", "seller_bank_name",
        "seller_iban", "seller_swift", "seller_is_person",
        "seller_vat_val", "seller_extra_info",
        # Buyer
        "buyer_counterparty_id", "buyer_id_programoje",
        "buyer_name", "buyer_name_normalized", "buyer_id", "buyer_vat_code",
        "buyer_address", "buyer_country", "buyer_country_iso",
        "buyer_phone", "buyer_email", "buyer_bank_name",
        "buyer_iban", "buyer_swift", "buyer_is_person",
        "buyer_vat_val", "buyer_extra_info", "buyer_delivery_address",
        # Суммы
        "currency", "pvm_tipas", "vat_percent",
        "amount_wo_vat", "vat_amount", "amount_with_vat",
        "invoice_discount_with_vat", "invoice_discount_wo_vat",
        "delivery_fee", "separate_vat", "doc_96_str",
        # Мета
        "document_type", "document_type_code",
        "pirkimas_pardavimas", "report_to_isaf",
        "issued_by", "received_by",
        # Экспорт
        "prekes_kodas", "prekes_barkodas", "prekes_pavadinimas",
        "prekes_tipas", "preke_paslauga", "pvm_kodas",
        # Company profile
        "company_profile_id",
    ]

    LINE_COPY_FIELDS = [
        "line_id", "prekes_kodas", "prekes_barkodas",
        "prekes_pavadinimas", "prekes_tipas", "preke_paslauga",
        "unit", "quantity", "price", "subtotal", "vat", "vat_percent",
        "total", "discount_with_vat", "discount_wo_vat",
        "pvm_kodas", "sort_order",
    ]

    def post(self, request, pk):
        from .models import Invoice, InvoiceLineItem, InvoiceSeries
        from .serializers import InvoiceDetailSerializer

        original = get_object_or_404(Invoice, pk=pk, user=request.user)

        # ── Validations ────────────────────────────────────
        if original.invoice_type not in ("pvm_saskaita", "saskaita"):
            return Response(
                {"detail": "Kreditinę galima išrašyti tik iš PVM SF arba SF."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if original.status not in ("issued", "sent", "partially_paid", "paid"):
            return Response(
                {"detail": "Sąskaita turi būti išrašyta, kad galima būtų kurti kreditinę."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Кредитка принадлежит той же фирме, что и оригинал — профиль и серию
        # берём из оригинала, а не из активной фирмы.
        credit_profile_id = original.company_profile_id or getattr(
            request.user, "active_company_profile_id", None
        )
        # гарантируем, что у этой фирмы есть серии (в т.ч. кредитная)
        if credit_profile_id:
            InvoiceSeries.create_defaults_for_user(request.user, credit_profile_id)
        series = InvoiceSeries.get_default_for_type(
            request.user, "kreditine", credit_profile_id
        )
        series_prefix = series.prefix if series else "KS"

        # ── Копировать invoice ─────────────────────────────
        credit = Invoice(
            user=request.user,
            invoice_type="kreditine",
            is_credit_invoice=True,
            source_invoice=original,
            status="draft",
            document_series=series_prefix,
            # document_number — пусто, присвоится при issue
        )

        for field in self.COPY_FIELDS:
            setattr(credit, field, getattr(original, field))

        # Даты: invoice_date = сегодня, operation_date = из оригинала
        from django.utils import timezone
        credit.invoice_date = timezone.localdate()
        credit.operation_date = original.operation_date or original.invoice_date
        credit.due_date = None  # у кредитной нет срока оплаты

        # Примечание со ссылкой на оригинал
        credit.note = (
            f"Kreditinė sąskaita pagal {original.full_number}"
            f"{(' nuo ' + str(original.invoice_date)) if original.invoice_date else ''}"
        )

        # Korespondencija: pardavimo pajamų sąskaita (5000/5001) ir PVM sąskaita
        # gyvena EILUTĖSE (InvoiceLineItem.kredito_saskaita / pvm_saskaita),
        # o ne dokumento antraštėje. Debetas pardavimui visada 2410.
        credit.debeto_saskaita = original.debeto_saskaita or "2410"
        credit.kredito_saskaita = original.kredito_saskaita
        credit.pvm_saskaita = original.pvm_saskaita
        credit.pirkimo_saskaita = original.pirkimo_saskaita

        credit.save()

        # ── Копировать строки вместе с корреспонденцией ──
        # Juodraštyje kiekiai ir sumos lieka teigiami.
        for li in original.line_items.order_by("sort_order", "id"):
            new_li = InvoiceLineItem(invoice=credit)

            for field in self.LINE_COPY_FIELDS:
                setattr(new_li, field, getattr(li, field))

            new_li.kredito_saskaita = li.kredito_saskaita
            new_li.pvm_saskaita = li.pvm_saskaita
            new_li.pirkimo_saskaita = li.pirkimo_saskaita
            new_li.save()

        return Response(
            InvoiceDetailSerializer(credit, context={"request": request}).data,
            status=status.HTTP_201_CREATED,
        )


# ════════════════════════════════════════════════════════════
# Invoice Summary — counts для табов (overdue на лету)
# ════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_summary(request):
    user = request.user
    today = date.today()

    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")

    base_qs = Invoice.objects.filter(user=user).exclude(source_invoice__isnull=False)

    company_id = request.query_params.get("company_profile")
    if company_id:
        base_qs = base_qs.filter(company_profile_id=company_id)
    else:
        active_id = getattr(user, "active_company_profile_id", None)
        if active_id:
            base_qs = base_qs.filter(company_profile_id=active_id)

    if date_from:
        base_qs = base_qs.filter(invoice_date__gte=date_from)
    if date_to:
        base_qs = base_qs.filter(invoice_date__lte=date_to)

    israsytos_qs = base_qs.filter(
        status__in=["issued", "sent"]
    ).filter(Q(due_date__gte=today) | Q(due_date__isnull=True))

    veluojancios_qs = base_qs.filter(
        status__in=["issued", "sent"], due_date__lt=today,
    )

    apmoketos_qs = base_qs.filter(status="paid")
    juodrasciai_qs = base_qs.filter(status="draft")
    cancelled_qs = base_qs.filter(status="cancelled")
    exported_qs = base_qs.filter(exported=True)

    def _agg(qs):
        agg = qs.aggregate(total=Sum("amount_with_vat"), count=Count("id"))
        return {
            "count": agg["count"] or 0,
            "total": str(agg["total"] or Decimal("0.00")),
        }

    return Response({
        "israsytos": _agg(israsytos_qs),
        "veluojancios": _agg(veluojancios_qs),
        "apmoketos": _agg(apmoketos_qs),
        "juodrasciai": _agg(juodrasciai_qs),
        "cancelled": _agg(cancelled_qs),
        "exported": _agg(exported_qs),
    })


# ════════════════════════════════════════════════════════════
# Public — Просмотр счёта покупателем (без авторизации)
# ════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([AllowAny])
def invoice_public(request, uuid):
    invoice = get_object_or_404(Invoice, uuid=uuid)

    if not invoice.public_link_enabled:
        return Response(
            {"detail": "Ši sąskaita nėra vieša."},
            status=status.HTTP_404_NOT_FOUND,
        )

    if invoice.status == "draft":
        return Response(
            {"detail": "Sąskaita dar neišrašyta."},
            status=status.HTTP_404_NOT_FOUND,
        )

    serializer = InvoicePublicSerializer(invoice, context={'request': request})
    data = serializer.data

    # --- Watermark for free plan ---
    show_watermark = False
    try:
        sub = getattr(invoice.user, "inv_subscription", None)
        if sub:
            sub.check_and_expire()
            show_watermark = sub.status == "free"
    except Exception:
        pass
    data["show_watermark"] = show_watermark

    return Response(data)


@api_view(["GET"])
@permission_classes([AllowAny])
def invoice_public_pdf(request, uuid):
    invoice = get_object_or_404(Invoice, uuid=uuid)

    if not invoice.public_link_enabled or invoice.status == "draft":
        return Response(status=status.HTTP_404_NOT_FOUND)

    # Логотип
    logo_path = None
    try:
        settings = invoice.user.invoice_settings
        if settings.logo and settings.logo.storage.exists(settings.logo.name):
            logo_path = settings.logo.path
    except Exception:
        pass

    # --- Watermark for free plan ---
    watermark = False
    try:
        from docscanner_app.models import InvSubscription
        sub = InvSubscription.objects.filter(user=invoice.user).first()
        if sub:
            sub.check_and_expire()
            watermark = sub.status == "free"
    except Exception:
        pass

    from .utils.invoice_pdf import generate_invoice_pdf
    pdf_bytes = generate_invoice_pdf(invoice, logo_path=logo_path, watermark=watermark)

    filename = f"saskaita-{invoice.document_series}-{invoice.document_number}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response




"""
Hybrid search: Counterparty (сохранённые клиенты) + Company (все фирмы ЛТ).
Приоритет — сохранённые клиенты. Дедупликация по im_kodas.

Добавить в urls.py:
    path('invoicing/search-companies/', invoice_search_companies, name='invoice-search-companies'),

Добавить в views.py импорт:
    from .views import invoice_search_companies
    (или в тот же файл где остальные invoicing views)
"""

from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_search_companies(request):
    """
    GET /api/invoicing/search-companies/?q=UAB+Senukai&limit=20

    Возвращает:
    [
      {
        "source": "saved",          // или "company"
        "id": 123,                  // PK
        "name": "UAB Senukai",
        "company_code": "234376520",
        "vat_code": "LT234376520",
        "address": "...",
        "phone": "...",
        "email": "...",
        "bank_name": "...",
        "iban": "...",
        "swift": "...",
        "is_person": false,
      }
    ]
    """
    from .models import Counterparty, Company

    q = (request.GET.get("q") or "").strip()
    limit = min(int(request.GET.get("limit") or 20), 50)

    if len(q) < 2:
        return Response([])

    results = []
    seen_codes = set()

    # ── 1. Сохранённые клиенты (Counterparty) — приоритет ──
    cp_qs = Counterparty.objects.filter(user=request.user)
    _active_id = getattr(request.user, "active_company_profile_id", None)
    if _active_id:
        cp_qs = cp_qs.filter(company_profile_id=_active_id)
    cp_qs = cp_qs.filter(
        Q(name__icontains=q)
        | Q(company_code__icontains=q)
        | Q(vat_code__icontains=q)
    )[:limit]

    for cp in cp_qs:
        code = (cp.company_code or "").strip()
        results.append({
            "source": "saved",
            "id": cp.id,
            "name": cp.name or "",
            "company_code": code,
            "vat_code": cp.vat_code or "",
            "address": cp.address or "",
            "phone": cp.phone or "",
            "email": cp.email or "",
            "bank_name": cp.bank_name or "",
            "iban": cp.iban or "",
            "swift": cp.swift or "",
            "is_person": getattr(cp, "is_person", False),
        })
        if code:
            seen_codes.add(code)

    # ── 2. Company (все фирмы ЛТ) — дополняем до limit ──
    remaining = limit - len(results)
    if remaining > 0:
        co_qs = Company.objects.filter(
            Q(pavadinimas__icontains=q)
            | Q(im_kodas__icontains=q)
            | Q(pvm_kodas__icontains=q)
        ).only(
            "id", "pavadinimas", "im_kodas", "pvm_kodas"
        )[:remaining + len(seen_codes)]

        count = 0
        for co in co_qs:
            if count >= remaining:
                break
            code = (co.im_kodas or "").strip()
            if code in seen_codes:
                continue
            seen_codes.add(code)

            results.append({
                "source": "company",
                "id": co.id,
                "name": co.pavadinimas or "",
                "company_code": code,
                "vat_code": co.pvm_kodas or "",
                "address": getattr(co, "adresas", "") or "",
                "phone": "",
                "email": "",
                "bank_name": "",
                "iban": "",
                "swift": "",
                "is_person": False,
            })
            count += 1

    return Response(results)








# ═══════════════════════════════════════════════════════════
# MeasurementUnit views
# ═══════════════════════════════════════════════════════════

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def measurement_unit_list(request):
    """
    GET  — список единиц (auto-создаёт дефолтные при первом вызове)
    POST — создать новую единицу
    """
    user = request.user
    active_id = getattr(user, "active_company_profile_id", None)

    # Авто-создание дефолтных при первом обращении (в разрезе активной фирмы)
    MeasurementUnit.create_defaults_for_user(user, active_id)

    if request.method == "GET":
        qs = MeasurementUnit.objects.filter(user=user, is_active=True)
        if active_id:
            qs = qs.filter(company_profile_id=active_id)
        return Response(MeasurementUnitSerializer(qs, many=True).data)

    # POST
    ser = MeasurementUnitSerializer(data=request.data)
    ser.is_valid(raise_exception=True)

    code = ser.validated_data["code"]
    # Проверка на дубликат (re-activate если был удалён) — в разрезе фирмы
    existing = MeasurementUnit.objects.filter(
        user=user, company_profile_id=active_id, code=code
    ).first()
    if existing:
        if not existing.is_active:
            existing.is_active = True
            existing.name = ser.validated_data.get("name", existing.name)
            existing.save(update_fields=["is_active", "name"])
            return Response(MeasurementUnitSerializer(existing).data, status=status.HTTP_200_OK)
        return Response(
            {"detail": f"Matavimo vienetas '{code}' jau egzistuoja."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    max_order = MeasurementUnit.objects.filter(
        user=user, company_profile_id=active_id
    ).count()
    ser.save(user=user, company_profile_id=active_id, sort_order=max_order)
    return Response(ser.data, status=status.HTTP_201_CREATED)


@api_view(["PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def measurement_unit_detail(request, pk):
    try:
        unit = MeasurementUnit.objects.get(pk=pk, user=request.user)
    except MeasurementUnit.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        unit.is_active = False
        unit.is_default = False
        unit.save(update_fields=["is_active", "is_default"])
        return Response(status=status.HTTP_204_NO_CONTENT)

    # PUT
    ser = MeasurementUnitSerializer(unit, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)

    new_code = ser.validated_data.get("code", unit.code)
    if new_code != unit.code:
        if MeasurementUnit.objects.filter(
            user=request.user, company_profile_id=unit.company_profile_id, code=new_code
        ).exclude(pk=pk).exists():
            return Response(
                {"detail": f"Matavimo vienetas '{new_code}' jau egzistuoja."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    obj = ser.save()

    if obj.is_default:
        MeasurementUnit.ensure_only_one_default(request.user, obj.id)

    return Response(MeasurementUnitSerializer(obj).data)


# ═══════════════════════════════════════════════════════════
# InvoiceSeries views
# ═══════════════════════════════════════════════════════════

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def invoice_series_list(request):
    """
    GET  — список серий. ?invoice_type=pvm_saskaita для фильтрации
    POST — создать серию
    """
    user = request.user
    active_id = getattr(user, "active_company_profile_id", None)

    if request.method == "GET":
        # Ленивый досев: если у активной фирмы нет серий — создать дефолтные
        if active_id:
            InvoiceSeries.create_defaults_for_user(user, active_id)

        qs = InvoiceSeries.objects.filter(user=user, is_active=True)
        if active_id:
            qs = qs.filter(company_profile_id=active_id)
        invoice_type = request.query_params.get("invoice_type")
        if invoice_type:
            qs = qs.filter(invoice_type=invoice_type)
        return Response(InvoiceSeriesSerializer(qs, many=True).data)

    # POST
    ser = InvoiceSeriesSerializer(data=request.data)
    ser.is_valid(raise_exception=True)

    prefix = ser.validated_data["prefix"].strip().upper()
    invoice_type = ser.validated_data["invoice_type"]

    # Уникальность prefix в пределах user + фирмы
    if InvoiceSeries.objects.filter(
        user=user, company_profile_id=active_id, prefix=prefix
    ).exists():
        return Response(
            {"detail": f"Serija '{prefix}' jau egzistuoja."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    is_default = ser.validated_data.get("is_default", False)
    # Если это первая серия для типа в этой фирме — сделать default
    if not InvoiceSeries.objects.filter(
        user=user, company_profile_id=active_id,
        invoice_type=invoice_type, is_active=True,
    ).exists():
        is_default = True

    obj = ser.save(user=user, company_profile_id=active_id, prefix=prefix, is_default=is_default)

    # Ensure only one default
    if is_default:
        InvoiceSeries.ensure_only_one_default(user, invoice_type, obj.id, active_id)

    return Response(InvoiceSeriesSerializer(obj).data, status=status.HTTP_201_CREATED)


@api_view(["PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def invoice_series_detail(request, pk):
    """
    PUT    — обновить серию (prefix, next_number, padding, is_default, is_active)
    DELETE — soft-delete (is_active=False)
    """
    try:
        series = InvoiceSeries.objects.get(pk=pk, user=request.user)
    except InvoiceSeries.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        series.is_active = False
        series.save(update_fields=["is_active"])
        return Response(status=status.HTTP_204_NO_CONTENT)

    # PUT
    data = request.data.copy()
    # Prefix всегда uppercase
    if "prefix" in data:
        data["prefix"] = data["prefix"].strip().upper()

    ser = InvoiceSeriesSerializer(series, data=data, partial=True)
    ser.is_valid(raise_exception=True)

    # Проверка уникальности prefix (в разрезе фирмы записи)
    new_prefix = ser.validated_data.get("prefix", series.prefix)
    if new_prefix != series.prefix:
        if InvoiceSeries.objects.filter(
            user=request.user,
            company_profile_id=series.company_profile_id,
            prefix=new_prefix,
        ).exclude(pk=pk).exists():
            return Response(
                {"detail": f"Serija '{new_prefix}' jau egzistuoja."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    obj = ser.save()

    # Ensure only one default
    if ser.validated_data.get("is_default"):
        InvoiceSeries.ensure_only_one_default(
            request.user, obj.invoice_type, obj.id, obj.company_profile_id
        )

    return Response(InvoiceSeriesSerializer(obj).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_series_check_number(request):
    """
    GET /api/invoicing/series/check-number/?prefix=AA&number=001

    Проверяет, существует ли документ с такой серией и номером.
    Возвращает: { "exists": true/false, "invoice_id": 123 }
    """
    prefix = (request.GET.get("prefix") or "").strip()
    number = (request.GET.get("number") or "").strip()

    if not prefix or not number:
        return Response({"exists": False, "invoice_id": None})

    qs = Invoice.objects.filter(
        user=request.user,
        document_series=prefix,
        document_number=number,
    )
    _active_id = getattr(request.user, "active_company_profile_id", None)
    if _active_id:
        qs = qs.filter(company_profile_id=_active_id)
    invoice = qs.first()

    return Response({
        "exists": invoice is not None,
        "invoice_id": invoice.id if invoice else None,
    })




from django.db.models import Q, Value, CharField
from django.db.models.functions import Concat


# ────────────────────────────────────────────────────────────
# 1. GET /invoicing/next-number/?series=SF&invoice_type=pvm_saskaita
#    → { "next_number": "001", "preview": "SF-001", "prefix": "SF", "padding": 3 }
# ────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_next_number(request):
    """
    Возвращает следующий свободный номер для указанной серии.
    Если серия не указана — берёт default для типа.
    """
    series_prefix = request.query_params.get("series", "").strip()
    invoice_type = request.query_params.get("invoice_type", "pvm_saskaita").strip()

    active_id = getattr(request.user, "active_company_profile_id", None)

    # Найти серию
    if series_prefix:
        qs = InvoiceSeries.objects.filter(
            user=request.user,
            prefix=series_prefix,
            is_active=True,
        )
        if active_id:
            qs = qs.filter(company_profile_id=active_id)
        series = qs.first()
    else:
        series = InvoiceSeries.get_default_for_type(request.user, invoice_type, active_id)

    if not series:
        return Response({"next_number": "", "preview": "", "prefix": "", "padding": 3})

    next_num = series.format_number()

    return Response({
        "next_number": next_num,
        "preview": f"{series.prefix}-{next_num}",
        "prefix": series.prefix,
        "padding": series.padding,
        "raw_next": series.next_number,
    })


# ────────────────────────────────────────────────────────────
# 2. GET /invoicing/check-number/?number=001&series=SF&invoice_type=pvm_saskaita
#    → { "exists": true/false }
# ────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_check_number(request):
    """
    Проверяет, существует ли уже счёт с таким номером+серией у пользователя.
    Учитывает только не-отменённые счета.
    """
    number = request.query_params.get("number", "").strip()
    series = request.query_params.get("series", "").strip()
    invoice_type = request.query_params.get("invoice_type", "").strip()

    if not number or not series:
        return Response({"exists": False})

    qs = Invoice.objects.filter(
        user=request.user,
        document_series=series,
        document_number=number,
    ).exclude(
        status="cancelled",
    )
    _active_id = getattr(request.user, "active_company_profile_id", None)
    if _active_id:
        qs = qs.filter(company_profile_id=_active_id)

    # Опционально фильтр по типу
    if invoice_type:
        qs = qs.filter(invoice_type=invoice_type)

    return Response({"exists": qs.exists()})


# ────────────────────────────────────────────────────────────
# 3. GET /invoicing/search-products/?q=konsultacij&limit=15
#    → [{ prekes_pavadinimas, prekes_kodas, prekes_barkodas, price, unit, vat_percent }, ...]
#
#    Ищет по ранее использованным товарам/услугам в InvoiceLineItem.
#    Поиск по pavadinimas, kodas, barkodas.
# ────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_search_products(request):
    q = request.query_params.get("q", "").strip()
    limit = min(int(request.query_params.get("limit", 15)), 50)

    if len(q) < 2:
        return Response([])

    qs = Product.objects.filter(user=request.user)
    _active_id = getattr(request.user, "active_company_profile_id", None)
    if _active_id:
        qs = qs.filter(company_profile_id=_active_id)
    qs = qs.filter(
        Q(pavadinimas__icontains=q) |
        Q(kodas__icontains=q) |
        Q(barkodas__icontains=q)
    ).select_related("measurement_unit").order_by("pavadinimas")[:limit]

    results = [
        {
            "id": p.id,
            "prekes_pavadinimas": p.pavadinimas,
            "prekes_kodas": p.kodas,
            "prekes_barkodas": p.barkodas,
            "preke_paslauga": p.preke_paslauga,
            "price": float(p.pardavimo_kaina) if p.pardavimo_kaina is not None else None,
            "unit": str(p.measurement_unit) if p.measurement_unit else "vnt",
            "vat_percent": p.pvm_procentas,
        }
        for p in qs
    ]

    return Response(results)

# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def invoice_search_products(request):
#     """
#     Ищет по ранее использованным товарам/услугам пользователя.
#     Источник — InvoiceLineItem из счетов этого пользователя.
#     Дедупликация по (pavadinimas, kodas, barkodas).
#     """
#     q = request.query_params.get("q", "").strip()
#     limit = min(int(request.query_params.get("limit", 15)), 50)

#     if len(q) < 2:
#         return Response([])

#     # Все строки из счетов этого пользователя
#     qs = InvoiceLineItem.objects.filter(
#         invoice__user=request.user,
#     ).exclude(
#         prekes_pavadinimas="",
#     )

#     # Поиск по трём полям
#     q_upper = q.upper()
#     q_filter = (
#         Q(prekes_pavadinimas__icontains=q) |
#         Q(prekes_kodas__icontains=q) |
#         Q(prekes_barkodas__icontains=q)
#     )
#     qs = qs.filter(q_filter)

#     # Берём уникальные комбинации, приоритет — последние использованные
#     qs = qs.order_by("-invoice__created_at")

#     # Дедупликация в Python (чтобы сохранить последнюю цену)
#     seen = set()
#     results = []
#     for item in qs.iterator():
#         key = (
#             item.prekes_pavadinimas.strip().upper(),
#             item.prekes_kodas.strip().upper(),
#         )
#         if key in seen:
#             continue
#         seen.add(key)
#         results.append({
#             "id": item.id,
#             "prekes_pavadinimas": item.prekes_pavadinimas,
#             "prekes_kodas": item.prekes_kodas,
#             "prekes_barkodas": item.prekes_barkodas,
#             "price": float(item.price) if item.price is not None else None,
#             "unit": item.unit or "vnt",
#             "vat_percent": float(item.vat_percent) if item.vat_percent is not None else None,
#         })
#         if len(results) >= limit:
#             break

#     return Response(results)





@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def product_list_create(request):
    """
    GET  — список товаров/услуг (с поиском ?q=... и фильтром ?type=preke|paslauga)
           Пагинация: ?limit=25&offset=0
    POST — создать
    """
    user = request.user
    active_id = getattr(user, "active_company_profile_id", None)

    if request.method == "GET":
        qs = Product.objects.filter(user=user).select_related("measurement_unit")
        if active_id:
            qs = qs.filter(company_profile_id=active_id)

        q = request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                Q(pavadinimas__icontains=q)
                | Q(kodas__icontains=q)
                | Q(barkodas__icontains=q)
            )

        typ = request.GET.get("type")
        if typ in ("preke", "paslauga"):
            qs = qs.filter(preke_paslauga=typ)

        # Новые сверху
        qs = qs.order_by("-id")
        total = qs.count()

        try:
            limit = int(request.GET.get("limit", 25))
            offset = int(request.GET.get("offset", 0))
        except (ValueError, TypeError):
            limit, offset = 25, 0
        limit = max(1, min(limit, 250))
        offset = max(offset, 0)

        page = qs[offset : offset + limit]
        serializer = ProductListSerializer(page, many=True)
        return Response({"results": serializer.data, "count": total})

    # POST
    serializer = ProductSerializer(data=request.data, context={"request": request})
    serializer.is_valid(raise_exception=True)
    serializer.save(user=user, company_profile_id=active_id)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["GET", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def product_detail(request, pk):
    """GET / PUT / DELETE одного товара."""
    obj = get_object_or_404(
        Product.objects.select_related("measurement_unit"),
        pk=pk, user=request.user,
    )

    if request.method == "GET":
        return Response(ProductSerializer(obj, context={"request": request}).data)

    if request.method == "PUT":
        serializer = ProductSerializer(
            obj, data=request.data, partial=True, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    obj.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)




#Dlia generacii PDF dlia skacivanja iz israsymas

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_pdf(request, pk):
    """Сгенерировать PDF на лету и отдать."""
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

    if invoice.status == "draft":
        return Response(
            {"detail": "PDF galimas tik išrašytai sąskaitai."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Логотип
    logo_path = None
    try:
        settings = invoice.user.invoice_settings
        if settings.logo and settings.logo.storage.exists(settings.logo.name):
            logo_path = settings.logo.path
    except Exception:
        pass

    # --- Watermark for free plan ---
    watermark = False
    try:
        from docscanner_app.models import InvSubscription
        sub = InvSubscription.objects.filter(user=request.user).first()
        if sub:
            sub.check_and_expire()
            watermark = sub.status == "free"
    except Exception:
        pass

    from .utils.invoice_pdf import generate_invoice_pdf
    pdf_bytes = generate_invoice_pdf(invoice, logo_path=logo_path, watermark=watermark)

    filename = f"{invoice.full_number or invoice.pk}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



# ────────────────────────────────────────────────────────────
# Recurring invoices
# ────────────────────────────────────────────────────────────

class RecurringInvoiceViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = RecurringInvoice.objects.prefetch_related("line_items")
        if not self.request.user.is_superuser:
            qs = qs.filter(user=self.request.user)

        company_id = self.request.query_params.get("company_profile")
        if company_id:
            qs = qs.filter(company_profile_id=company_id)
        else:
            active_id = getattr(self.request.user, "active_company_profile_id", None)
            if active_id:
                qs = qs.filter(company_profile_id=active_id)

        return qs

    def get_serializer_class(self):
        if self.action == "list":
            return RecurringInvoiceListSerializer
        if self.action in ("create", "update", "partial_update"):
            return RecurringInvoiceWriteSerializer
        return RecurringInvoiceDetailSerializer

    def _finalize_if_no_future_runs(self, obj):
        """
        Если после пересчета запусков больше нет, переводим в finished.
        Пропущенные во время pause даты НЕ считаются использованными.
        max_count считаем только по generation_count.
        """
        if obj.max_count and obj.generation_count >= obj.max_count:
            obj.status = "finished"
            obj.next_run_at = None
            return

        if not obj.next_run_at:
            obj.status = "finished"
            obj.next_run_at = None
            return

        if obj.end_date and obj.next_run_at.date() > obj.end_date:
            obj.status = "finished"
            obj.next_run_at = None
            return


    def _build_future_dates(self, obj, count):
        """
        Будущие даты только для active + next_run_at.
        Никакого backfill.
        """
        if obj.status != "active" or not obj.next_run_at:
            return []

        if obj.max_count and obj.generation_count >= obj.max_count:
            return []

        future = []
        current_dt = obj.next_run_at
        remaining = None

        if obj.max_count:
            remaining = max(obj.max_count - obj.generation_count, 0)

        for _ in range(count):
            if not current_dt:
                break

            if obj.end_date and current_dt.date() > obj.end_date:
                break

            if remaining is not None and len(future) >= remaining:
                break

            future.append(current_dt.date().isoformat())
            current_dt = obj.compute_next_run_after(from_dt=current_dt)

        return future

    @action(detail=True, methods=["post"])
    def pause(self, request, pk=None):
        obj = self.get_object()
        if obj.status != "active":
            return Response(
                {"detail": "Galima pristabdyti tik aktyvią periodinę sąskaitą."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        obj.status = "paused"
        obj.save(update_fields=["status", "updated_at"])
        return Response(RecurringInvoiceDetailSerializer(obj).data)

    @action(detail=True, methods=["post"])
    def resume(self, request, pk=None):
        obj = self.get_object()
        if obj.status != "paused":
            return Response(
                {"detail": "Galima tęsti tik pristabdytą periodinę sąskaitą."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        obj.status = "active"

        # Пересчитываем next_run только вперед, без backfill
        obj.refresh_next_run_at()

        # Если future run больше нет, сразу finished
        self._finalize_if_no_future_runs(obj)

        obj.save(update_fields=["status", "next_run_at", "updated_at"])
        return Response(RecurringInvoiceDetailSerializer(obj).data)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        obj = self.get_object()
        if obj.status in ("finished", "cancelled"):
            return Response(
                {"detail": "Periodinė sąskaita jau baigta arba atšaukta."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        obj.status = "cancelled"
        obj.next_run_at = None
        obj.save(update_fields=["status", "next_run_at", "updated_at"])
        return Response(RecurringInvoiceDetailSerializer(obj).data)

    @action(detail=True, methods=["get"])
    def preview_next(self, request, pk=None):
        """Показать когда будут следующие N запусков."""
        obj = self.get_object()
        count = min(int(request.query_params.get("count", 5)), 12)
        dates = self._build_future_dates(obj, count)
        return Response({"dates": dates})

    @action(detail=True, methods=["get"])
    def plan_history(self, request, pk=None):
        """Прошлые runs + будущие даты."""
        obj = self.get_object()
        count = min(int(request.query_params.get("count", 12)), 24)

        # Прошлые (из RecurringInvoiceRun)
        past_runs = obj.runs.order_by("-scheduled_for")[:12].values(
            "scheduled_for", "status", "invoice_id", "error_text", "created_at"
        )
        past = []
        for run in past_runs:
            past.append({
                "date": run["scheduled_for"].date().isoformat() if run["scheduled_for"] else None,
                "status": run["status"],
                "invoice_id": run["invoice_id"],
                "error": run["error_text"][:200] if run["error_text"] else "",
            })

        future = self._build_future_dates(obj, count)

        return Response({
            "past": list(reversed(past)),
            "future": future,
        })




import os
import openpyxl
from django.conf import settings as django_settings
from django.http import FileResponse


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def counterparty_import_template(request):
    """Скачать XLSX шаблон для импорта контрагентов."""
    file_path = os.path.join(
        django_settings.BASE_DIR, "templates", "israsymas", "kontrahentu_sablonas.xlsx"
    )
    return FileResponse(
        open(file_path, "rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        filename="kontrahentu_sablonas.xlsx",
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def counterparty_import_xlsx(request):
    """Импорт контрагентов из XLSX файла."""
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "Failas nerastas."}, status=status.HTTP_400_BAD_REQUEST)
    if not file.name.endswith((".xlsx", ".xls")):
        return Response({"detail": "Tik .xlsx failai palaikomi."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return Response({"detail": "Nepavyko atidaryti failo."}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return Response({"detail": "Failas tuščias."}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    active_id = getattr(user, "active_company_profile_id", None)
    created = 0
    updated = 0
    errors = []

    # Case-insensitive, с/без диакритиков
    ROLE_MAP = {
        "pirkėjas": "buyer", "pirkejas": "buyer", "buyer": "buyer",
        "pardavėjas": "seller", "pardavejas": "seller", "seller": "seller",
        "abu": "both", "both": "both",
    }

    PERSON_TRUE = {"taip", "yes", "true", "1", "fizinis", "t"}
    PERSON_FALSE = {"ne", "no", "false", "0", "juridinis", "f", ""}


    for row_idx, row in enumerate(rows, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell_val(idx):
            if idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        # Колонки по шаблону
        name = cell_val(0)
        company_code = cell_val(1)
        vat_code = cell_val(2)
        address = cell_val(3)
        country_iso = cell_val(4).upper()
        phone = cell_val(5)
        email = cell_val(6)
        bank_name = cell_val(7)
        iban = cell_val(8)
        swift = cell_val(9)
        is_person_raw = cell_val(10).lower().strip()
        role_raw = cell_val(11).lower().strip()
        delivery_address = cell_val(12)
        extra_info = cell_val(13)

        # Валидация required
        row_errors = []
        if not name:
            row_errors.append("Pavadinimas privalomas")
        if not company_code:
            row_errors.append("Įmonės kodas privalomas")
        if row_errors:
            errors.append({"row": row_idx, "name": name or "—", "errors": row_errors})
            continue

        # Парсинг — всё case-insensitive
        is_person = is_person_raw in PERSON_TRUE
        default_role = ROLE_MAP.get(role_raw, "buyer")

        # Country — если указан ISO, берём его; иначе LT по умолчанию
        if not country_iso:
            country_iso = "LT"
        country_name = COUNTRY_NAME_LT.get(country_iso, country_iso)

        try:
            cp, is_new = Counterparty.objects.update_or_create(
                user=user,
                company_profile_id=active_id,
                company_code=company_code,
                defaults={
                    "name": name,
                    "name_normalized": name.strip().upper(),
                    "vat_code": vat_code,
                    "address": address,
                    "country": country_name,
                    "country_iso": country_iso,
                    "phone": phone,
                    "email": email,
                    "bank_name": bank_name,
                    "iban": iban,
                    "swift": swift,
                    "is_person": is_person,
                    "default_role": default_role,
                    "delivery_address": delivery_address,
                    "extra_info": extra_info,
                },
            )
            if is_new:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append({"row": row_idx, "name": name, "errors": [str(e)[:200]]})

    return Response({
        "created": created,
        "updated": updated,
        "total_rows": len(rows),
        "errors": errors,
    })





@api_view(["GET"])
@permission_classes([IsAuthenticated])
def product_import_template(request):
    """Скачать XLSX шаблон для импорта товаров/услуг."""
    file_path = os.path.join(
        django_settings.BASE_DIR, "templates", "israsymas", "prekiu_sablonas.xlsx"
    )
    return FileResponse(
        open(file_path, "rb"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        filename="prekiu_sablonas.xlsx",
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def product_import_xlsx(request):
    """Импорт товаров/услуг из XLSX файла."""
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "Failas nerastas."}, status=status.HTTP_400_BAD_REQUEST)
    if not file.name.endswith((".xlsx", ".xls")):
        return Response({"detail": "Tik .xlsx failai palaikomi."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return Response({"detail": "Nepavyko atidaryti failo."}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return Response({"detail": "Failas tuščias."}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    active_id = getattr(user, "active_company_profile_id", None)
    created = 0
    updated = 0
    errors = []

    TYPE_MAP = {
        "prekė": "preke", "preke": "preke", "prekė": "preke",
        "paslauga": "paslauga", "paslaugą": "paslauga",
        "product": "preke", "service": "paslauga",
    }

    # Загрузить measurement units для маппинга по коду
    from .models import MeasurementUnit
    unit_map = {}
    _unit_qs = MeasurementUnit.objects.filter(user=user)
    if active_id:
        _unit_qs = _unit_qs.filter(company_profile_id=active_id)
    for u in _unit_qs:
        unit_map[u.code.lower()] = u
        if u.name:
            unit_map[u.name.lower()] = u

    for row_idx, row in enumerate(rows, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell_val(idx):
            if idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        pavadinimas = cell_val(0)
        kodas = cell_val(1)
        barkodas = cell_val(2)
        type_raw = cell_val(3).lower()
        unit_raw = cell_val(4)
        price_raw = cell_val(5)
        pvm_raw = cell_val(6)

        # Валидация required
        row_errors = []
        if not pavadinimas:
            row_errors.append("Pavadinimas privalomas")
        if not kodas:
            row_errors.append("Kodas privalomas")
        if row_errors:
            errors.append({"row": row_idx, "name": pavadinimas or "—", "errors": row_errors})
            continue

        # Тип
        preke_paslauga = TYPE_MAP.get(type_raw, "preke")

        # Mato vienetas
        measurement_unit = None
        if unit_raw:
            measurement_unit = unit_map.get(unit_raw.lower())
            if not measurement_unit:
                existing_codes = ", ".join(sorted(set(
                    u.code for u in unit_map.values()
                )))
                row_errors.append(
                    f'Mato vienetas "{unit_raw}" nerastas. '
                    f'Sukurkite tokį matavimo vienetą arba naudokite vieną iš jau sukurtų: {existing_codes}'
                )

        # Цена
        pardavimo_kaina = 0
        if price_raw:
            try:
                pardavimo_kaina = float(price_raw.replace(",", ".").replace(" ", ""))
            except ValueError:
                row_errors.append(f"Neteisinga kaina: {price_raw}")

        # PVM
        pvm_procentas = None
        if pvm_raw:
            try:
                pvm_val = int(float(pvm_raw.replace(",", ".").replace("%", "").strip()))
                if 0 <= pvm_val <= 100:
                    pvm_procentas = pvm_val
                else:
                    row_errors.append(f"PVM turi būti 0-100: {pvm_raw}")
            except (ValueError, TypeError):
                row_errors.append(f"Neteisingas PVM: {pvm_raw}")

        if row_errors:
            errors.append({"row": row_idx, "name": pavadinimas, "errors": row_errors})
            continue

        try:
            obj, is_new = Product.objects.update_or_create(
                user=user,
                company_profile_id=active_id,
                kodas=kodas,
                defaults={
                    "pavadinimas": pavadinimas,
                    "barkodas": barkodas,
                    "preke_paslauga": preke_paslauga,
                    "measurement_unit": measurement_unit,
                    "pardavimo_kaina": pardavimo_kaina,
                    "pvm_procentas": pvm_procentas,
                },
            )
            if is_new:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append({"row": row_idx, "name": pavadinimas, "errors": [str(e)[:200]]})

    return Response({
        "created": created,
        "updated": updated,
        "total_rows": len(rows),
        "errors": errors,
    })












# ────────────────────────────────────────────────────────────
# Banko israso importas
# ────────────────────────────────────────────────────────────
"""
API Views для банковского импорта и платежей.

Bank Import:
  POST   /api/bank-import/upload/
  GET    /api/bank-import/statements/
  GET    /api/bank-import/statements/<id>/
  DELETE /api/bank-import/statements/<id>/
  POST   /api/bank-import/statements/<id>/re-match/

Payment Management:
  GET    /api/bank-import/invoice/<id>/payments/     ← PaymentProofDialog data
  POST   /api/bank-import/invoice/<id>/mark-paid/    ← MarkPaidDialog
  POST   /api/bank-import/invoice/<id>/remove-payment/<alloc_id>/

Matching Actions:
  POST   /api/bank-import/confirm/
  POST   /api/bank-import/bulk-confirm/
  POST   /api/bank-import/reject/

Dashboard:
  GET    /api/bank-import/stats/
"""

from django.shortcuts import get_object_or_404
from rest_framework import generics, permissions, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import BankStatement, PaymentAllocation, IncomingTransaction, OutgoingTransaction, BankTransactionRule
from .serializers import (
    BankStatementListSerializer,
    BankStatementUploadSerializer,
    BulkConfirmSerializer,
    ConfirmAllocationSerializer,
    InvoicePaymentDetailsSerializer,
    MarkPaidSerializer,
    BankTransactionRuleSerializer,
)
from .services.payment_service import BankImportService, PaymentService, BankImportError


class Pagination50(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 200


# ────────────────────────────────────────────────────────────
# Bank Statement Upload & List
# ────────────────────────────────────────────────────────────


class StatementUploadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        # --- Inv subscription: bank_import check ---
        allowed, err = check_inv_feature(request.user, "bank_import")
        if not allowed:
            return Response(err, status=status.HTTP_403_FORBIDDEN)

        ser = BankStatementUploadSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        svc = BankImportService(request.user)
        try:
            stmt = svc.import_statement(
                file=ser.validated_data["file"],
                bank_name=ser.validated_data.get("bank_name", ""),
                file_format=ser.validated_data.get("file_format", ""),
                original_filename=ser.validated_data["file"].name,
            )
        except BankImportError as e:

            from .celery_signals import _send_telegram
            _send_telegram(
                f"🏦 <b>Bank import failed</b>\n"
                f"User: {request.user.email}\n"
                f"File: {ser.validated_data['file'].name}\n"
                f"Error: {str(e)[:300]}"
            )

            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            BankStatementListSerializer(stmt).data,
            status=status.HTTP_201_CREATED,
        )


class StatementListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = BankStatementListSerializer
    pagination_class = Pagination50

    def get_queryset(self):
        return BankStatement.objects.filter(user=self.request.user)


class StatementDetailView(generics.RetrieveDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = BankStatementListSerializer
 
    def get_queryset(self):
        return BankStatement.objects.filter(user=self.request.user)
 
    def perform_destroy(self, instance):
        from .models import PaymentAllocation, JournalEntry
        from django.db import transaction

        stmt_id = instance.id  # сохраняем до удаления

        with transaction.atomic():
            # ── 1. Собрать затронутые документы ──
            affected_invoice_ids = set()
            affected_purchase_ids = set()
            je_ids_to_delete = set()

            for alloc in PaymentAllocation.objects.filter(
                incoming_transaction__bank_statement=instance,
            ).select_related("invoice", "purchase"):
                if alloc.invoice_id:
                    affected_invoice_ids.add(alloc.invoice_id)
                if alloc.purchase_id:
                    affected_purchase_ids.add(alloc.purchase_id)
                if alloc.journal_entry_id:
                    je_ids_to_delete.add(alloc.journal_entry_id)

            for alloc in PaymentAllocation.objects.filter(
                outgoing_transaction__bank_statement=instance,
            ).select_related("invoice", "purchase"):
                if alloc.invoice_id:
                    affected_invoice_ids.add(alloc.invoice_id)
                if alloc.purchase_id:
                    affected_purchase_ids.add(alloc.purchase_id)
                if alloc.journal_entry_id:
                    je_ids_to_delete.add(alloc.journal_entry_id)

            for txn in instance.incoming_transactions.filter(journal_entry__isnull=False):
                je_ids_to_delete.add(txn.journal_entry_id)
            for txn in instance.outgoing_transactions.filter(journal_entry__isnull=False):
                je_ids_to_delete.add(txn.journal_entry_id)

            # ── 2. Удалить JE ──
            if je_ids_to_delete:
                JournalEntry.objects.filter(id__in=je_ids_to_delete).delete()

            # ── 3. Удалить statement (CASCADE удалит txn → allocations) ──
            instance.delete()

            # ── 4. Пересчитать документы ──
            if affected_invoice_ids:
                from .models import Invoice
                for inv in Invoice.objects.filter(id__in=affected_invoice_ids):
                    inv.recalc_payment_status()

            if affected_purchase_ids:
                from .models import Purchase
                for p in Purchase.objects.filter(id__in=affected_purchase_ids):
                    p.recalc_from_allocations()

        logger.info(
            "[BankStatement] Deleted stmt %s: %d JEs, %d invoices, %d purchases recalculated",
            stmt_id, len(je_ids_to_delete),
            len(affected_invoice_ids), len(affected_purchase_ids),
        )


class StatementReMatchView(APIView):
    permission_classes = [permissions.IsAuthenticated]
 
    def post(self, request, pk):
        stmt = get_object_or_404(BankStatement, pk=pk, user=request.user)

        if stmt.status != "processed":
            return Response(
                {"detail": "Pakartotinis susiejimas galimas tik apdorotiems išrašams."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        svc = BankImportService(request.user)
        svc.re_match_statement(stmt)

        stmt.refresh_stats()

        # Auto SF for any newly auto_matched incoming invoices
        for txn in stmt.incoming_transactions.filter(match_status="auto_matched"):
            for alloc in txn.allocations.filter(status="auto"):
                try:
                    from .services.auto_sf import maybe_auto_create_sf
                    created_sf = maybe_auto_create_sf(alloc.invoice)
                    if created_sf:
                        logger.info(
                            "[ReMatch] Auto SF created: %s for invoice %s",
                            created_sf.full_number,
                            alloc.invoice.full_number,
                        )
                except Exception as e:
                    logger.warning("[ReMatch] Auto SF failed: %s", e)

        return Response(BankStatementListSerializer(stmt).data)



class BankAccountMappingView(APIView):
    """
    GET  /api/invoicing/bank-accounts/ — список банковских счетов
    POST /api/invoicing/bank-accounts/ — обновить kor. sąskaitą
         Body: { key, account, label? }
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import CompanyProfile
        cp = CompanyProfile.objects.filter(
            user=request.user, is_active=True,
        ).first()
        if not cp:
            return Response([])
        return Response(cp.get_all_bank_accounts())

    def post(self, request):
        from .models import CompanyProfile
        cp = CompanyProfile.objects.filter(
            user=request.user, is_active=True,
        ).first()
        if not cp:
            return Response(
                {"detail": "Įmonės profilis nerastas."},
                status=status.HTTP_404_NOT_FOUND,
            )

        key = request.data.get("key", "").strip().upper()
        account = request.data.get("account", "").strip()
        label = request.data.get("label", "").strip()
        new_iban = request.data.get("iban", "").strip().upper()
        bank = request.data.get("bank", "").strip().lower()
        currency = request.data.get("currency", "EUR").strip().upper()

        if not key or not account:
            return Response(
                {"detail": "key ir account privalomi."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        mapping = cp.bank_accounts_mapping or {}

        if key in mapping:
            entry = dict(mapping[key])
            entry["account"] = account

            if label:
                entry["label"] = label

            if bank:
                entry["bank"] = bank

            if currency:
                entry["currency"] = currency

            # Jei pridėtas IBAN — perkelti iš seno rakto į naują
            if new_iban and len(new_iban) >= 10 and new_iban != key:
                del mapping[key]
                mapping[new_iban] = entry
            else:
                mapping[key] = entry
        else:
            mapping[key] = {
                "account": account,
                "bank": bank,
                "label": label,
                "currency": currency,
            }

        cp.bank_accounts_mapping = mapping
        cp.save(update_fields=["bank_accounts_mapping"])

        return Response({"status": "updated", "accounts": cp.get_all_bank_accounts()})



# ────────────────────────────────────────────────────────────
# Bank Transactions (unified list + actions)
# ────────────────────────────────────────────────────────────


# ── Transaction type mapping ──────────────────────────────

def get_tx_type_display(bank_operation_code: str) -> str:
    """Человекочитаемый тип транзакции из bank_operation_code."""
    code = (bank_operation_code or "").upper()
    if not code:
        return ""
    # SEB / ISO 20022 patterns
    if "PMNTCCRD" in code:
        return "Mokėjimas kortele"
    if "PMNTICDTBOOK" in code:
        return "Banko pavedimas"
    if "PMNTICDTESCT" in code or "PMNTRCDTESCT" in code:
        return "Momentinis mokėjimas"
    if "ACMTMDOP" in code:
        return "Banko mokestis"
    if "PMNTMCOP" in code:
        return "Korektūra"
    if "PMNTRCDTBOOK" in code:
        return "Vidinis pervedimas"
    # Swedbank patterns
    if code == "K":
        return "Korespondentinis"
    if code == "MK":
        return "Memorialinis"
    if code == "TT":
        return "Tarptautinis pervedimas"
    if code == "M":
        return "Mokestis"
    return ""


def _build_txn_allocations(allocs):
    """Сериализовать allocations для таблицы."""
    return [
        {
            "id": a.id,
            "amount": str(a.amount),
            "confidence": str(a.confidence),
            "status": a.status,
            "invoice_id": a.invoice_id,
            "invoice_number": a.invoice.full_number if a.invoice else None,
            "purchase_id": a.purchase_id,
            "purchase_number": (
                f"{a.purchase.document_series or ''}{a.purchase.document_number or ''}"
                if a.purchase else None
            ),
            "document_preview_url": _get_alloc_preview_url(a),
        }
        for a in allocs
    ]


def _get_alloc_preview_url(alloc):
    """Preview URL документа из allocation."""
    try:
        if alloc.purchase_id and alloc.purchase:
            scan = alloc.purchase.scanned_document
            if scan:
                if scan.preview_url:
                    return scan.preview_url
                if scan.file:
                    return scan.file.url
        if alloc.invoice_id and alloc.invoice:
            if alloc.invoice.pdf_file:
                return alloc.invoice.pdf_file.url
    except Exception:
        return None
    return None


def _build_txn_light(txn, direction_str, allocs):
    """Light dict для таблицы."""
    return {
        "id": txn.id,
        "direction": direction_str,
        "transaction_date": txn.transaction_date,
        "counterparty_name": txn.counterparty_name or "",
        "counterparty_code": txn.counterparty_code or "",
        "amount": txn.amount,
        "currency": txn.currency,
        "tx_type": get_tx_type_display(txn.bank_operation_code),
        "match_status": txn.match_status,
        "transaction_category": txn.transaction_category or "",
        "category_display": (
            txn.get_transaction_category_display()
            if txn.transaction_category else ""
        ),
        "statement_id": txn.bank_statement_id,
        "bank_name": (
            txn.bank_statement.get_bank_name_display()
            if txn.bank_statement else ""
        ),
        "allocations": _build_txn_allocations(allocs),
    }


def _build_txn_full(txn, direction_str, allocs):
    """Full dict для drawer."""
    data = _build_txn_light(txn, direction_str, allocs)
    data.update({
        "uuid": txn.uuid,
        "value_date": txn.value_date,
        "counterparty_account": txn.counterparty_account or "",
        "payment_purpose": txn.payment_purpose or "",
        "bank_operation_code": txn.bank_operation_code or "",
        "doc_number": txn.doc_number or "",
        "reference_number": txn.reference_number or "",
        "match_confidence": txn.match_confidence,
        "match_details": txn.match_details or {},
        "allocated_amount": txn.allocated_amount,
        "category_account_debit": txn.category_account_debit or "",
        "category_account_credit": txn.category_account_credit or "",
    })
    return data


# ═══════════════════════════════════════════════════════
# Замены в views.py — банковские транзакции
# ═══════════════════════════════════════════════════════


# ── 1. TransactionListView.get() ──────────────────────
# Заменить apply_filters и stats целиком


class TransactionListView(APIView):
    """
    GET /api/invoicing/bank-transactions/
    Light data для таблицы (без payment_purpose).
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        limit = min(int(request.query_params.get("limit", 50)), 200)
        offset = int(request.query_params.get("offset", 0))

        stmt_id = request.query_params.get("statement_id")
        direction = request.query_params.get("direction", "")
        match_status = request.query_params.get("match_status", "")
        category = request.query_params.get("category", "")
        q = request.query_params.get("q", "").strip()

        def apply_filters(qs, include_match_status=True):
            if stmt_id:
                qs = qs.filter(bank_statement_id=stmt_id)
            if include_match_status and match_status:
                if match_status == "needs_action":
                    from django.db.models import Q
                    qs = qs.filter(
                        Q(match_status="likely_matched")
                        | Q(match_status="unmatched", transaction_category="")
                    )
                elif match_status == "processed":
                    qs = qs.filter(
                        match_status__in=[
                            "auto_matched", "confirmed",
                            "manually_matched", "classified",
                        ]
                    )
                else:
                    qs = qs.filter(match_status=match_status)
            if category:
                if category == "uncategorized":
                    qs = qs.filter(transaction_category="")
                else:
                    qs = qs.filter(transaction_category=category)
            if q:
                from django.db.models import Q
                qs = qs.filter(
                    Q(counterparty_name__icontains=q)
                    | Q(counterparty_code__icontains=q)
                    | Q(doc_number__icontains=q)
                )
            return qs

        # Для результатов таблицы — с match_status фильтром
        inc_qs = apply_filters(
            IncomingTransaction.objects.filter(user=user).select_related("bank_statement")
        )
        out_qs = apply_filters(
            OutgoingTransaction.objects.filter(user=user).select_related("bank_statement")
        )

        # Для stats — без match_status фильтра
        inc_stats = apply_filters(
            IncomingTransaction.objects.filter(user=user),
            include_match_status=False,
        )
        out_stats = apply_filters(
            OutgoingTransaction.objects.filter(user=user),
            include_match_status=False,
        )

        if direction == "incoming":
            out_qs = OutgoingTransaction.objects.none()
        elif direction == "outgoing":
            inc_qs = IncomingTransaction.objects.none()

        total = inc_qs.count() + out_qs.count()

        fetch_size = offset + limit + 10
        inc_list = list(inc_qs.order_by("-transaction_date", "-id")[:fetch_size])
        out_list = list(out_qs.order_by("-transaction_date", "-id")[:fetch_size])

        merged = [("incoming", t) for t in inc_list] + [("outgoing", t) for t in out_list]
        merged.sort(key=lambda x: (x[1].transaction_date, x[1].id), reverse=True)

        # likely_matched вверху при фильтре "Reikia veiksmų"
        if match_status == "needs_action":
            merged.sort(key=lambda x: (0 if x[1].match_status == "likely_matched" else 1))

        page = merged[offset:offset + limit]

        results = []
        for dir_str, txn in page:
            results.append({
                "id": txn.id,
                "direction": dir_str,
                "transaction_date": txn.transaction_date,
                "counterparty_name": txn.counterparty_name or "",
                "counterparty_code": txn.counterparty_code or "",
                "amount": txn.amount,
                "currency": txn.currency,
                "tx_type": get_tx_type_display(txn.bank_operation_code),
                "match_status": txn.match_status,
                "transaction_category": txn.transaction_category or "",
                "category_display": (
                    txn.get_transaction_category_display()
                    if txn.transaction_category else ""
                ),
                "matched_document_number": txn.matched_document_number or "",
                "match_confidence": txn.match_confidence,
                "statement_id": txn.bank_statement_id,
                "bank_name": (
                    txn.bank_statement.get_bank_name_display()
                    if txn.bank_statement else ""
                ),
            })

        # ── Stats: 3 карточки ──
        def count_status(qs, status):
            return qs.filter(match_status=status).count()

        processed_statuses = [
            "auto_matched", "confirmed", "manually_matched", "classified",
        ]

        stats = {
            "total": inc_stats.count() + out_stats.count(),
            "processed": sum(
                count_status(inc_stats, s) + count_status(out_stats, s)
                for s in processed_statuses
            ),
            "needs_action": (
                count_status(inc_stats, "likely_matched")
                + count_status(out_stats, "likely_matched")
                + inc_stats.filter(
                    match_status="unmatched", transaction_category="",
                ).count()
                + out_stats.filter(
                    match_status="unmatched", transaction_category="",
                ).count()
            ),
        }

        return Response({
            "count": total,
            "stats": stats,
            "results": results,
        })


# ── 2. TransactionDetailView.get() ───────────────────
# Добавить select_related("purchase__scanned_document")


class TransactionDetailView(APIView):
    """
    GET /api/invoicing/bank-transactions/<id>/?direction=outgoing
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        direction = request.query_params.get("direction", "")
        txn = None
        direction_str = direction

        if direction == "incoming":
            txn = IncomingTransaction.objects.filter(
                id=pk, user=request.user,
            ).select_related("bank_statement").first()
        elif direction == "outgoing":
            txn = OutgoingTransaction.objects.filter(
                id=pk, user=request.user,
            ).select_related("bank_statement").first()
        else:
            txn = OutgoingTransaction.objects.filter(
                id=pk, user=request.user,
            ).select_related("bank_statement").first()
            if txn:
                direction_str = "outgoing"
            else:
                txn = IncomingTransaction.objects.filter(
                    id=pk, user=request.user,
                ).select_related("bank_statement").first()
                direction_str = "incoming"

        if not txn:
            return Response({"detail": "Nerasta."}, status=status.HTTP_404_NOT_FOUND)

        fk = "incoming_transaction" if direction_str == "incoming" else "outgoing_transaction"
        allocs = list(
            PaymentAllocation.objects.filter(
                **{fk: txn}
            ).select_related("invoice", "purchase", "purchase__scanned_document")
        )

        data = _build_txn_full(txn, direction_str, allocs)
        return Response(data)


# ── 3. _build_txn_full — добавить journal_entry_id ────


def _build_txn_full(txn, direction_str, allocs):
    """Full dict для dialog."""
    data = _build_txn_light(txn, direction_str, allocs)
    data.update({
        "uuid": txn.uuid,
        "value_date": txn.value_date,
        "counterparty_account": txn.counterparty_account or "",
        "payment_purpose": txn.payment_purpose or "",
        "bank_operation_code": txn.bank_operation_code or "",
        "doc_number": txn.doc_number or "",
        "reference_number": txn.reference_number or "",
        "match_confidence": txn.match_confidence,
        "match_details": txn.match_details or {},
        "allocated_amount": txn.allocated_amount,
        "category_account_debit": txn.category_account_debit or "",
        "category_account_credit": txn.category_account_credit or "",
        "journal_entry_id": txn.journal_entry_id,
    })
    return data


# ── 4. _build_txn_allocations + preview URL ──────────


def _build_txn_allocations(allocs):
    """Сериализовать allocations для dialog."""
    return [
        {
            "id": a.id,
            "amount": str(a.amount),
            "confidence": str(a.confidence),
            "status": a.status,
            "invoice_id": a.invoice_id,
            "invoice_number": a.invoice.full_number if a.invoice else None,
            "purchase_id": a.purchase_id,
            "purchase_number": (
                f"{a.purchase.document_series or ''}{a.purchase.document_number or ''}"
                if a.purchase else None
            ),
            "document_preview_url": _get_alloc_preview_url(a),
        }
        for a in allocs
    ]


def _get_alloc_preview_url(alloc):
    """Preview URL документа из allocation."""
    try:
        if alloc.purchase_id and alloc.purchase:
            scan = alloc.purchase.scanned_document
            if scan:
                if scan.preview_url:
                    return scan.preview_url
                if scan.file:
                    return scan.file.url
        if alloc.invoice_id and alloc.invoice:
            if alloc.invoice.pdf_file:
                return alloc.invoice.pdf_file.url
    except Exception:
        return None
    return None


class BankMatchingDebugView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        statement_id = request.query_params.get("statement_id")
        only_unmatched = str(request.query_params.get("only_unmatched", "true")).lower() == "true"

        if not statement_id:
            return Response(
                {"detail": "statement_id is required"},
                status=400,
            )

        stmt = get_object_or_404(
            BankStatement,
            id=statement_id,
            user=request.user,
        )

        qs = (
            stmt.outgoing_transactions
            .all()
            .order_by("transaction_date", "id")
        )

        # Debug page: показываем всё, кроме уже вручную подтверждённых,
        # чтобы видеть и unmatched, и likely, и auto.
        if only_unmatched:
            qs = qs.exclude(match_status__in=[
                "confirmed",
                "manually_matched",
            ])

        from .utils.purchase_matching_signals import SignalPurchaseMatchingEngine

        signal_engine = SignalPurchaseMatchingEngine(request.user)

        # Загружаем candidates один раз, чтобы не делать query на каждую txn
        signal_candidates = [
            signal_engine._purchase_to_candidate(p)
            for p in signal_engine._load_purchases()
        ]

        items = []

        summary = {
            "transactions": 0,

            "actual_auto_matched": 0,
            "actual_likely_matched": 0,
            "actual_unmatched": 0,

            "signal_auto_matched": 0,
            "signal_likely_matched": 0,
            "signal_unmatched": 0,
            "signal_skipped": 0,
        }

        for txn in qs:
            summary["transactions"] += 1

            actual_match = self._build_actual_match(txn)

            try:
                signal_result = signal_engine._match_one(txn, signal_candidates)
                signal_match = self._build_signal_match(signal_result)
            except Exception as e:
                logger.exception("[BankMatchingDebug] signal dry-run failed txn=%s", txn.id)
                signal_match = {
                    "status": "error",
                    "confidence": "0",
                    "confidence_pct": 0,
                    "error": str(e),
                    "purchase": None,
                    "amount": "0",
                    "matched_document_number": "",
                    "reasons": {},
                    "signals": {},
                }

            actual_status = actual_match.get("status") or "unmatched"
            signal_status = signal_match.get("status") or "unmatched"

            if actual_status == "auto_matched":
                summary["actual_auto_matched"] += 1
            elif actual_status == "likely_matched":
                summary["actual_likely_matched"] += 1
            else:
                summary["actual_unmatched"] += 1

            if signal_match.get("signals", {}).get("skip_matching"):
                summary["signal_skipped"] += 1

            if signal_status == "auto_matched":
                summary["signal_auto_matched"] += 1
            elif signal_status == "likely_matched":
                summary["signal_likely_matched"] += 1
            else:
                summary["signal_unmatched"] += 1

            # Backward-compatible fields, чтобы старый frontend не падал
            display_match = actual_match if actual_status != "unmatched" else signal_match
            display_purchase = display_match.get("purchase")

            best_candidate = None
            if display_purchase:
                best_candidate = {
                    **display_purchase,
                    "score": display_match.get("confidence_pct", 0),
                    "decision": display_match.get("status", "unmatched"),
                    "reasons": self._short_reasons(display_match),
                    "warnings": [],
                }

            items.append({
                "transaction": self._serialize_txn(txn),
                "actual_match": actual_match,
                "signal_match": signal_match,

                # deprecated, only for old UI compatibility
                "best_candidate": best_candidate,
                "candidates": [],
            })

        return Response({
            "statement": {
                "id": stmt.id,
                "bank_name": stmt.bank_name,
                "filename": stmt.original_filename,
                "period_from": stmt.period_from.isoformat() if stmt.period_from else None,
                "period_to": stmt.period_to.isoformat() if stmt.period_to else None,
                "currency": stmt.currency,
            },
            "summary": summary,
            "items": items,
        })

    def _build_actual_match(self, txn):
        alloc = (
            PaymentAllocation.objects
            .filter(outgoing_transaction=txn)
            .select_related("purchase")
            .order_by("-confidence", "-created_at")
            .first()
        )

        purchase = None
        if alloc and alloc.purchase:
            purchase = self._serialize_purchase(alloc.purchase)

        details = txn.match_details or {}
        signals = details.get("signals") or {}

        return {
            "status": txn.match_status or "unmatched",
            "confidence": str(txn.match_confidence or "0"),
            "confidence_pct": self._confidence_pct(txn.match_confidence),
            "matched_document_number": txn.matched_document_number or "",
            "allocated_amount": str(txn.allocated_amount or "0"),
            "purchase": purchase,
            "allocation": self._serialize_allocation(alloc) if alloc else None,
            "match_details": details,
            "signals": signals,
        }

    def _build_signal_match(self, result):
        purchase = None

        if result.purchase_id:
            try:
                purchase = self._serialize_purchase(
                    Purchase.objects.get(id=result.purchase_id, user=self.request.user)
                )
            except Purchase.DoesNotExist:
                purchase = None

        return {
            "status": result.status or "unmatched",
            "confidence": str(result.confidence or "0"),
            "confidence_pct": self._confidence_pct(result.confidence),
            "matched_document_number": result.matched_document_number or "",
            "amount": str(result.amount or "0"),
            "purchase": purchase,
            "reasons": result.reasons or {},
            "signals": result.signals or {},
        }

    def _serialize_txn(self, txn):
        return {
            "id": txn.id,
            "transaction_date": txn.transaction_date.isoformat() if txn.transaction_date else None,
            "amount": str(txn.amount),
            "currency": txn.currency,
            "counterparty_name": txn.counterparty_name or "",
            "counterparty_code": txn.counterparty_code or "",
            "counterparty_account": txn.counterparty_account or "",
            "payment_purpose": txn.payment_purpose or "",
            "reference_number": txn.reference_number or "",
            "doc_number": txn.doc_number or "",
            "bank_operation_code": txn.bank_operation_code or "",
            "match_status": txn.match_status or "unmatched",
            "match_confidence": str(txn.match_confidence or "0"),
            "matched_document_number": txn.matched_document_number or "",
            "transaction_category": txn.transaction_category or "",
            "allocated_amount": str(txn.allocated_amount or "0"),
        }

    def _serialize_purchase(self, p):
        series = getattr(p, "document_series", "") or ""
        number = getattr(p, "document_number", "") or ""
        full_number = getattr(p, "full_number", "") or f"{series}{number}".strip()

        return {
            "id": p.id,
            "full_number": full_number,
            "document_series": series,
            "document_number": number,
            "invoice_date": p.invoice_date.isoformat() if p.invoice_date else None,
            "seller_name": getattr(p, "seller_name", "") or "",
            "seller_id": (
                getattr(p, "seller_id", "") or
                getattr(p, "seller_code", "") or
                ""
            ),
            "seller_iban": getattr(p, "seller_iban", "") or "",
            "amount_with_vat": str(getattr(p, "amount_with_vat", "") or "0"),
            "paid_amount": str(getattr(p, "paid_amount", "") or "0"),
            "currency": getattr(p, "currency", "") or "EUR",
            "payment_status": getattr(p, "payment_status", "") or "",
        }

    def _serialize_allocation(self, alloc):
        if not alloc:
            return None

        return {
            "id": alloc.id,
            "status": alloc.status,
            "source": alloc.source,
            "amount": str(alloc.amount),
            "confidence": str(alloc.confidence or "0"),
            "payment_date": alloc.payment_date.isoformat() if alloc.payment_date else None,
            "purchase_id": alloc.purchase_id,
            "match_reasons": alloc.match_reasons or {},
        }

    def _confidence_pct(self, value):
        try:
            d = Decimal(str(value or "0"))
            if d <= Decimal("1"):
                return int((d * Decimal("100")).quantize(Decimal("1")))
            return int(d.quantize(Decimal("1")))
        except Exception:
            return 0

    def _short_reasons(self, match):
        reasons = []

        signals = match.get("signals") or {}
        if signals.get("merchant_name_clean"):
            reasons.append(f"Merchant: {signals.get('merchant_name_clean')}")

        refs = signals.get("references") or []
        if refs:
            refs_text = ", ".join(
                r.get("value", "")
                for r in refs[:3]
                if r.get("value")
            )
            if refs_text:
                reasons.append(f"Refs: {refs_text}")

        if signals.get("original_amount"):
            reasons.append(
                f"Original amount: {signals.get('original_amount')} {signals.get('original_currency')}"
            )

        if signals.get("conversion_fee"):
            reasons.append(
                f"FX fee: {signals.get('conversion_fee')} {signals.get('settled_currency') or ''}"
            )

        if not reasons and match.get("matched_document_number"):
            reasons.append(f"Matched: {match.get('matched_document_number')}")

        return reasons


class TransactionClassifyView(APIView):
    """
    POST /api/bank-import/transactions/<id>/classify/
    Body: { category, debit_account, credit_account, create_rule?, apply_to_similar? }
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        from .serializers import TransactionClassifySerializer
        from .models import (
            IncomingTransaction, OutgoingTransaction,
            BankTransactionRule, CompanyProfile,
        )
        from .services.accounting_transfer import (
            create_je_for_classified_transaction,
        )

        ser = TransactionClassifySerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        # Find transaction
        txn = None
        direction_str = ""
        try:
            txn = OutgoingTransaction.objects.select_related(
                "bank_statement",
            ).get(id=pk, user=request.user)
            direction_str = "outgoing"
        except OutgoingTransaction.DoesNotExist:
            try:
                txn = IncomingTransaction.objects.select_related(
                    "bank_statement",
                ).get(id=pk, user=request.user)
                direction_str = "incoming"
            except IncomingTransaction.DoesNotExist:
                return Response(
                    {"detail": "Operacija nerasta."},
                    status=status.HTTP_404_NOT_FOUND,
                )

        # Apply classification
        CATEGORY_DEFAULTS = {
            "bank_fee": "6880",
            "tax_vmi": "4481",
            "tax_sodra": "4482",
            "salary": "4491",
            "owner_withdrawal": "3120",
            "owner_deposit": "3120",
            "loan_payment": "3011",
            "loan_received": "4010",
            "provider_payout": "2719",
        }

        category = data["category"]
        debit = data.get("debit_account", "").strip()
        credit = data.get("credit_account", "").strip()

        # Auto-fill default если юзер не указал
        if not debit and category in CATEGORY_DEFAULTS:
            debit = CATEGORY_DEFAULTS[category]

        txn.transaction_category = category
        txn.category_account_debit = debit
        txn.category_account_credit = credit
        txn.save(update_fields=[
            "transaction_category",
            "category_account_debit",
            "category_account_credit",
            "updated_at",
        ])

        # Auto JE
        cp = CompanyProfile.objects.filter(
            user=request.user, is_active=True,
        ).first()
        if cp and data.get("debit_account"):
            try:
                create_je_for_classified_transaction(txn, cp)
            except Exception as e:
                logger.warning("[Classify] JE failed: %s", e)

        # Create rule
        if data.get("create_rule") and cp:
            rule_name = data.get("rule_name") or (
                f"{txn.get_transaction_category_display()} – "
                f"{txn.counterparty_name or 'be pavadinimo'}"
            )
            BankTransactionRule.objects.create(
                user=request.user,
                company_profile=cp,
                name=rule_name[:255],
                match_field="counterparty_name",
                match_operator="contains",
                match_value=(txn.counterparty_name or "")[:500],
                direction="debit" if direction_str == "outgoing" else "credit",
                category=data["category"],
                debit_account=data.get("debit_account", ""),
                credit_account=data.get("credit_account", ""),
                auto_create_je=True,
            )

        # Apply to similar
        applied_count = 0
        if data.get("apply_to_similar") and txn.counterparty_name:
            from .utils.transaction_classifier import find_similar_transactions

            if direction_str == "outgoing":
                all_txns = list(OutgoingTransaction.objects.filter(
                    user=request.user,
                    match_status="unmatched",
                    transaction_category="",
                ))
            else:
                all_txns = list(IncomingTransaction.objects.filter(
                    user=request.user,
                    match_status="unmatched",
                    transaction_category="",
                ))

            similar = find_similar_transactions(txn, all_txns)
            for sim in similar:
                sim.transaction_category = data["category"]
                sim.category_account_debit = data.get("debit_account", "")
                sim.category_account_credit = data.get("credit_account", "")
                sim.save(update_fields=[
                    "transaction_category",
                    "category_account_debit",
                    "category_account_credit",
                    "updated_at",
                ])
                if cp and data.get("debit_account"):
                    try:
                        create_je_for_classified_transaction(sim, cp)
                    except Exception as e:
                        logger.warning("[Classify] JE for similar %s failed: %s", sim.id, e)
                applied_count += 1

        return Response({
            "status": "classified",
            "category": data["category"],
            "applied_to_similar": applied_count,
        })


class TransactionManualMatchView(APIView):
    """
    POST /api/bank-import/transactions/<id>/match/
    Body: { invoice_id?, purchase_id?, amount? }
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        from .serializers import TransactionManualMatchSerializer
        from .models import (
            IncomingTransaction, OutgoingTransaction,
            PaymentAllocation, Invoice, Purchase,
        )
        from .services.accounting_transfer import create_je_for_allocation

        ser = TransactionManualMatchSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        invoice_id = data.get("invoice_id")
        purchase_id = data.get("purchase_id")

        if not invoice_id and not purchase_id:
            return Response(
                {"detail": "Nurodykite invoice_id arba purchase_id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Find transaction
        txn = None
        is_incoming = False
        try:
            txn = OutgoingTransaction.objects.get(id=pk, user=request.user)
        except OutgoingTransaction.DoesNotExist:
            try:
                txn = IncomingTransaction.objects.get(id=pk, user=request.user)
                is_incoming = True
            except IncomingTransaction.DoesNotExist:
                return Response(
                    {"detail": "Operacija nerasta."},
                    status=status.HTTP_404_NOT_FOUND,
                )

        # Validate direction
        if is_incoming and purchase_id:
            return Response(
                {"detail": "Įplauka negali būti susieta su pirkimo dokumentu."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not is_incoming and invoice_id:
            return Response(
                {"detail": "Išlaida negali būti susieta su pardavimo sąskaita."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get document
        invoice = None
        purchase = None
        if invoice_id:
            invoice = get_object_or_404(Invoice, pk=invoice_id, user=request.user)
        if purchase_id:
            purchase = get_object_or_404(Purchase, pk=purchase_id, user=request.user)

        amount = data.get("amount") or txn.amount

        # Create allocation
        alloc_kwargs = {
            "amount": amount,
            "source": "bank_import",
            "status": "manual",
            "confidence": Decimal("1.00"),
            "match_reasons": {"manual_match": True},
            "payment_date": txn.transaction_date,
            "confirmed_at": timezone.now(),
            "confirmed_by": request.user,
        }

        if is_incoming:
            alloc, _ = PaymentAllocation.objects.update_or_create(
                incoming_transaction=txn,
                invoice=invoice,
                defaults=alloc_kwargs,
            )
        else:
            alloc, _ = PaymentAllocation.objects.update_or_create(
                outgoing_transaction=txn,
                purchase=purchase,
                defaults=alloc_kwargs,
            )

        # Update transaction
        from django.db.models import Sum
        total_alloc = txn.allocations.aggregate(t=Sum("amount"))["t"] or Decimal("0")
        txn.allocated_amount = total_alloc
        if invoice:
            txn.matched_document_number = invoice.full_number
        elif purchase:
            txn.matched_document_number = f"{purchase.document_series or ''}{purchase.document_number or ''}".strip()
        txn.transaction_category = "customer_receipt" if is_incoming else "supplier_payment"
        txn.save(update_fields=[
            "allocated_amount", "match_status",
            "transaction_category", "updated_at",
        ])

        # Recalc document
        if invoice:
            invoice.recalc_payment_status()
        if purchase:
            purchase.recalc_from_allocations()

        # Auto JE
        try:
            create_je_for_allocation(alloc)
        except Exception as e:
            logger.warning("[ManualMatch] JE failed: %s", e)

        # Refresh statement
        if txn.bank_statement:
            txn.bank_statement.refresh_stats()

        return Response({
            "status": "matched",
            "allocation_id": alloc.id,
        })


class BankTransactionRuleListView(generics.ListCreateAPIView):
    """GET/POST /api/bank-import/rules/"""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = BankTransactionRuleSerializer

    def get_queryset(self):
        return BankTransactionRule.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        from .models import CompanyProfile
        cp = CompanyProfile.objects.filter(
            user=self.request.user, is_active=True,
        ).first()
        if not cp:
            from rest_framework.exceptions import ValidationError
            raise ValidationError({"detail": "Reikia sukurti įmonės profilį."})
        serializer.save(user=self.request.user, company_profile=cp)


class BankTransactionRuleDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PUT/DELETE /api/bank-import/rules/<id>/"""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = BankTransactionRuleSerializer

    def get_queryset(self):
        return BankTransactionRule.objects.filter(user=self.request.user)


class AllocationPreviewView(APIView):
    """
    GET /api/invoicing/allocations/<id>/preview/
    Возвращает документ + matching criteria для preview dialog.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        from django.db.models import Q

        alloc = PaymentAllocation.objects.select_related(
            "invoice", "purchase",
            "incoming_transaction", "outgoing_transaction",
        ).filter(
            Q(invoice__user=request.user) | Q(purchase__user=request.user),
            id=pk,
        ).first()

        if not alloc:
            return Response({"detail": "Nerasta."}, status=status.HTTP_404_NOT_FOUND)

        result = {
            "id": alloc.id,
            "amount": alloc.amount,
            "confidence": alloc.confidence,
            "status": alloc.status,
            "source": alloc.source,
            "payment_date": alloc.effective_payment_date,
            "match_reasons": alloc.match_reasons or {},
            "direction": alloc.direction,
        }

        # Document info
        if alloc.invoice:
            inv = alloc.invoice
            result["document_type"] = "invoice"
            result["document"] = {
                "id": inv.id,
                "full_number": inv.full_number,
                "invoice_type": inv.invoice_type,
                "invoice_date": inv.invoice_date,
                "due_date": inv.due_date,
                "buyer_name": inv.buyer_name or "",
                "buyer_id": inv.buyer_id or "",
                "seller_name": inv.seller_name or "",
                "seller_id": inv.seller_id or "",
                "amount_with_vat": inv.amount_with_vat,
                "amount_wo_vat": inv.amount_wo_vat,
                "vat_amount": inv.vat_amount,
                "currency": inv.currency or "EUR",
                "status": inv.status,
                "payment_status": inv.payment_status,
                "is_from_scan": inv.is_from_scan,
                "scanned_document_id": inv.scanned_document_id,
            }
        elif alloc.purchase:
            p = alloc.purchase
            result["document_type"] = "purchase"
            result["document"] = {
                "id": p.id,
                "full_number": f"{p.document_series or ''}{p.document_number or ''}".strip(),
                "document_series": p.document_series or "",
                "document_number": p.document_number or "",
                "invoice_date": p.invoice_date,
                "due_date": p.due_date,
                "seller_name": p.seller_name or "",
                "seller_id": p.seller_id or "",
                "seller_iban": p.seller_iban or "",
                "amount_with_vat": p.amount_with_vat,
                "amount_wo_vat": p.amount_wo_vat,
                "vat_amount": p.vat_amount,
                "currency": p.currency or "EUR",
                "payment_status": p.payment_status,
                "scanned_document_id": p.scanned_document_id,
            }

        # Transaction info
        txn = alloc.transaction
        if txn:
            result["transaction"] = {
                "id": txn.id,
                "transaction_date": txn.transaction_date,
                "counterparty_name": txn.counterparty_name or "",
                "amount": txn.amount,
                "payment_purpose": txn.payment_purpose or "",
            }

        return Response(result)


class TransactionDKTemplatesView(APIView):
    """
    GET /api/bank-import/transactions/<id>/dk-templates/
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        from .services.bank_dk_register import BankDKRegisterService

        txn, direction = self._find_txn(pk, request.user)
        if not txn:
            return Response({"detail": "Operacija nerasta."}, status=status.HTTP_404_NOT_FOUND)

        cp = CompanyProfile.objects.filter(user=request.user, is_active=True).first()
        svc = BankDKRegisterService(request.user, cp)
        data = svc.get_templates_for_transaction(txn, direction)

        return Response(data)

    @staticmethod
    def _find_txn(pk, user):
        try:
            return OutgoingTransaction.objects.select_related("bank_statement").get(id=pk, user=user), "outgoing"
        except OutgoingTransaction.DoesNotExist:
            pass
        try:
            return IncomingTransaction.objects.select_related("bank_statement").get(id=pk, user=user), "incoming"
        except IncomingTransaction.DoesNotExist:
            pass
        return None, ""


class TransactionRegisterDKView(APIView):
    """
    POST /api/bank-import/transactions/<id>/register-dk/
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        from .services.bank_dk_register import BankDKRegisterService

        txn, direction = self._find_txn(pk, request.user)
        if not txn:
            return Response({"detail": "Operacija nerasta."}, status=status.HTTP_404_NOT_FOUND)

        lines = request.data.get("lines", [])
        description = request.data.get("description", "")
        category = request.data.get("category", "")

        cp = CompanyProfile.objects.filter(user=request.user, is_active=True).first()
        if not cp:
            return Response({"detail": "Reikia sukurti įmonės profilį."}, status=status.HTTP_400_BAD_REQUEST)

        svc = BankDKRegisterService(request.user, cp)

        try:
            entry = svc.register_dk(txn, direction, lines, description)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Category
        if category:
            txn.refresh_from_db()
            txn.transaction_category = category
            txn.save(update_fields=["transaction_category", "updated_at"])

        # Rule
        if request.data.get("create_rule") and cp:
            rule_name = request.data.get("rule_name") or (txn.counterparty_name or "Operacija")
            debit_codes = [l["account_code"] for l in lines if l.get("side") == "debit"]
            credit_codes = [l["account_code"] for l in lines if l.get("side") == "credit"]

            BankTransactionRule.objects.create(
                user=request.user,
                company_profile=cp,
                name=rule_name[:255],
                match_field="counterparty_name",
                match_operator="contains",
                match_value=(txn.counterparty_name or "")[:500],
                direction="debit" if direction == "outgoing" else "credit",
                category=category or "other_expense",
                debit_account=debit_codes[0] if debit_codes else "",
                credit_account=credit_codes[0] if credit_codes else "",
                auto_create_je=True,
            )

        # Apply to similar
        applied_count = 0
        if request.data.get("apply_to_similar") and txn.counterparty_name and category:
            from .utils.transaction_classifier import find_similar_transactions

            Model = OutgoingTransaction if direction == "outgoing" else IncomingTransaction
            all_txns = list(Model.objects.filter(
                user=request.user, match_status="unmatched", transaction_category="",
            ))
            similar = find_similar_transactions(txn, all_txns)

            for sim in similar:
                try:
                    svc.register_dk(sim, direction, lines, description)
                    if category:
                        sim.refresh_from_db()
                        sim.transaction_category = category
                        sim.save(update_fields=["transaction_category", "updated_at"])
                    applied_count += 1
                except Exception as e:
                    logger.warning("[RegisterDK] Similar txn %s failed: %s", sim.id, e)

        return Response({
            "status": "registered",
            "journal_entry_id": entry.id,
            "applied_to_similar": applied_count,
        })

    @staticmethod
    def _find_txn(pk, user):
        try:
            return OutgoingTransaction.objects.select_related("bank_statement").get(id=pk, user=user), "outgoing"
        except OutgoingTransaction.DoesNotExist:
            pass
        try:
            return IncomingTransaction.objects.select_related("bank_statement").get(id=pk, user=user), "incoming"
        except IncomingTransaction.DoesNotExist:
            pass
        return None, ""

# ────────────────────────────────────────────────────────────
# Invoice Payment Details (для PaymentProofDialog)
# ────────────────────────────────────────────────────────────


class InvoicePaymentDetailsView(APIView):
    """
    GET /api/bank-import/invoice/<id>/payments/

    Возвращает полную информацию о платежах invoice.
    Используется PaymentProofDialog на фронте.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        from .models import Invoice
        invoice = get_object_or_404(Invoice, pk=pk, user=request.user)
        svc = PaymentService(request.user)
        data = svc.get_invoice_payment_details(invoice)
        serializer = InvoicePaymentDetailsSerializer(data)
        return Response(serializer.data)


# ────────────────────────────────────────────────────────────
# Mark Paid (refactored — для MarkPaidDialog)
# ────────────────────────────────────────────────────────────


class InvoiceMarkPaidView(APIView):
    """
    POST /api/bank-import/invoice/<id>/mark-paid/

    Ручная пометка invoice как оплаченный.
    Создаёт PaymentAllocation с source="manual".

    Body: { amount, payment_date, note? }
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        from .models import Invoice

        invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

        if invoice.status not in ("issued", "sent", "partially_paid"):
            return Response(
                {"detail": "Galima pažymėti tik išrašytą/išsiųstą/dalinai apmokėtą sąskaitą."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = MarkPaidSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        svc = PaymentService(request.user)
        alloc = svc.mark_paid_manual(
            invoice=invoice,
            amount=ser.validated_data["amount"],
            payment_date=ser.validated_data["payment_date"],
            note=ser.validated_data.get("note", ""),
        )

        # Auto SF creation (existing logic)
        from .services.auto_sf import maybe_auto_create_sf
        created_sf = maybe_auto_create_sf(invoice)

        # Reload invoice
        invoice.refresh_from_db()

        # Build response
        from .serializers import InvoiceDetailSerializer
        data = InvoiceDetailSerializer(invoice, context={"request": request}).data
        data["allocation_id"] = alloc.id
        if created_sf:
            data["auto_created_sf"] = {
                "id": created_sf.id,
                "full_number": created_sf.full_number,
                "status": created_sf.status,
            }

        return Response(data)


class RemoveManualPaymentView(APIView):
    """
    POST /api/bank-import/invoice/<invoice_id>/remove-payment/<alloc_id>/

    Удаление ручной пометки оплаты.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk, alloc_id):
        svc = PaymentService(request.user)
        try:
            svc.remove_manual_payment(int(alloc_id))
            return Response({"status": "removed"})
        except PaymentAllocation.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)


# ────────────────────────────────────────────────────────────
# Matching Actions
# ────────────────────────────────────────────────────────────


class ConfirmAllocationView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ser = ConfirmAllocationSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        svc = PaymentService(request.user)
        try:
            alloc = svc.confirm_allocation(ser.validated_data["allocation_id"])
            return Response({"status": "confirmed", "allocation_id": alloc.id})
        except PaymentAllocation.DoesNotExist:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)


class BulkConfirmView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ser = BulkConfirmSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        svc = PaymentService(request.user)
        confirmed, errors = [], []
        for aid in ser.validated_data["allocation_ids"]:
            try:
                alloc = svc.confirm_allocation(aid)
                confirmed.append(alloc.id)
            except PaymentAllocation.DoesNotExist:
                errors.append(aid)

        return Response({"confirmed": confirmed, "errors": errors})


class RejectAllocationView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        aid = request.data.get("allocation_id")
        if not aid:
            return Response({"error": "allocation_id required"}, status=400)

        svc = PaymentService(request.user)
        try:
            svc.reject_allocation(int(aid))
            return Response({"status": "rejected"})
        except PaymentAllocation.DoesNotExist:
            return Response({"error": "Not found"}, status=404)


# ────────────────────────────────────────────────────────────
# Stats
# ────────────────────────────────────────────────────────────


class ImportStatsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import IncomingTransaction, OutgoingTransaction
        from django.db.models import Sum

        user = request.user
        stmts = BankStatement.objects.filter(user=user)
        inc = IncomingTransaction.objects.filter(user=user)
        out = OutgoingTransaction.objects.filter(user=user)

        return Response({
            "total_statements": stmts.count(),
            "total_incoming": inc.count(),
            "total_outgoing": out.count(),
            "auto_matched": (
                inc.filter(match_status="auto_matched").count()
                + out.filter(match_status="auto_matched").count()
            ),
            "likely_matched": (
                inc.filter(match_status="likely_matched").count()
                + out.filter(match_status="likely_matched").count()
            ),
            "confirmed": (
                inc.filter(match_status="confirmed").count()
                + out.filter(match_status="confirmed").count()
            ),
            "unmatched": (
                inc.filter(match_status="unmatched").count()
                + out.filter(match_status="unmatched").count()
            ),
            "total_credit_amount": (
                inc.aggregate(t=Sum("amount"))["t"] or 0
            ),
            "total_debit_amount": (
                out.aggregate(t=Sum("amount"))["t"] or 0
            ),
            "total_allocated_amount": (
                PaymentAllocation.objects
                .filter(
                    models.Q(invoice__user=user) | models.Q(purchase__user=user),
                    status__in=["confirmed", "auto", "manual"],
                )
                .aggregate(t=Sum("amount"))["t"] or 0
            ),
        })





# ────────────────────────────────────────────────────────────
# Dlia direct payment linkov v invoicax
# ────────────────────────────────────────────────────────────
"""
Payment-provider endpoints.

Endpoints:
  POST /api/invoicing/invoices/{id}/generate-payment-link/   (auth)
  GET  /api/invoicing/payment-providers/                      (auth)
  POST /api/invoicing/payment-providers/connect/              (auth)
  POST /api/invoicing/payment-providers/disconnect/           (auth)
  *    /api/invoicing/payment-webhook/{provider}/{invoice_id}/ (public)
"""

import hashlib
import json
import logging
import time as pytime
from datetime import time as dt_time
import urllib.parse
from base64 import b64encode

import jwt
import requests as http_requests
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.utils import timezone as tz
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Invoice
from .services.payment_link_service import PaymentLinkService

logger = logging.getLogger("docscanner_app")

PAYMENT_ENV = getattr(settings, "PAYMENT_ENVIRONMENT", "sandbox")

MONTONIO_BASE_URLS = {
    "sandbox": "https://sandbox-stargate.montonio.com/api",
    "production": "https://stargate.montonio.com/api",
}


# ────────────────────────────────────────────────────────────
# 1. Generate payment link for an invoice
# ────────────────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def generate_payment_link(request, invoice_id):
    provider_name = request.data.get("provider", "montonio")

    try:
        invoice = Invoice.objects.get(id=invoice_id, user=request.user)
    except Invoice.DoesNotExist:
        return Response({"detail": "Sąskaita nerasta"}, status=404)

    if invoice.status == "cancelled":
        return Response(
            {"detail": "Negalima sukurti mokėjimo nuorodos atšauktai sąskaitai"},
            status=400,
        )

    service = PaymentLinkService(request.user)

    try:
        result = service.create_for_invoice(invoice, provider_name)
    except ValueError as e:
        return Response({"detail": str(e)}, status=400)
    except Exception as e:
        logger.exception("generate_payment_link error: invoice=%s", invoice_id)

        from .celery_signals import _send_telegram
        _send_telegram(
            f"💳 <b>Payment link failed</b>\n"
            f"Invoice: {invoice_id}\n"
            f"Provider: {provider_name}\n"
            f"User: {request.user.email}\n"
            f"Error: {str(e)[:300]}"
        )

        return Response(
            {"detail": f"Klaida kuriant mokėjimo nuorodą: {e}"},
            status=500,
        )

    return Response({
        "payment_url": result.url,
        "provider": provider_name,
        "provider_payment_id": result.provider_payment_id,
    })


# ────────────────────────────────────────────────────────────
# 2. Available providers
# ────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def available_payment_providers(request):
    service = PaymentLinkService(request.user)
    return Response(service.get_available_providers())


# ────────────────────────────────────────────────────────────
# 3. Connect: save + test + return result
# ────────────────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def connect_payment_provider(request):
    """
    POST /api/invoicing/payment-providers/connect/

    Saves credentials, runs a live test, saves result to DB.
    """
    provider_name = request.data.get("provider")
    if not provider_name:
        return Response({"detail": "Nenurodytas teikėjas"}, status=400)

    data = request.data
    error = _validate_provider_fields(provider_name, data)
    if error:
        return Response({"connected": False, "error": error}, status=400)

    save_data = {**data, "environment": PAYMENT_ENV}
    if provider_name == "montonio":
        save_data["access_key"] = save_data.get("access_key", "").strip()
        save_data["secret_key"] = save_data.get("secret_key", "").strip()
    elif provider_name == "paysera":
        save_data["project_id"] = save_data.get("project_id", "").strip()
        save_data["sign_password"] = save_data.get("sign_password", "").strip()

    # Save credentials first
    try:
        PaymentLinkService.save_provider_config(
            user=request.user,
            provider_name=provider_name,
            data=save_data,
        )
    except ValueError as e:
        return Response({"connected": False, "error": str(e)}, status=400)

    # Test connection
    logger.info("=" * 60)
    logger.info("Testing %s connection for user=%s", provider_name, request.user.id)

    connected, test_error, provider_response, raw_response = _test_provider_connection(
        provider_name, data
    )

    logger.info(
        "Test result: connected=%s, error=%s, methods=%s",
        connected, test_error,
        len(provider_response) if provider_response else 0,
    )
    logger.info("=" * 60)

    # Build test result
    last_test_result = {
        "tested_at": tz.now().isoformat(),
        "connected": connected,
        "error": test_error,
        "methods_count": len(provider_response) if provider_response else 0,
        "raw_response": raw_response,
    }

    # Always save test result
    final_data = {
        **save_data,
        "last_test_result": last_test_result,
    }
    if connected and provider_response:
        final_data["available_methods"] = provider_response

    logger.info(
        "Saving to DB for %s: keys=%s",
        provider_name, list(final_data.keys()),
    )

    try:
        PaymentLinkService.save_provider_config(
            user=request.user,
            provider_name=provider_name,
            data=final_data,
        )
        logger.info("Saved test result to DB for %s", provider_name)
    except Exception as e:
        logger.exception("FAILED to save test result for %s: %s", provider_name, e)

    return Response({
        "connected": connected,
        "error": test_error,
        "available_methods": provider_response,
    })


# ────────────────────────────────────────────────────────────
# 4. Disconnect (delete keys)
# ────────────────────────────────────────────────────────────

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def disconnect_payment_provider(request):
    provider_name = request.data.get("provider")
    if not provider_name:
        return Response({"detail": "Nenurodytas teikėjas"}, status=400)

    try:
        PaymentLinkService.save_provider_config(
            user=request.user,
            provider_name=provider_name,
            data={"provider": provider_name},  # empty = cleared
        )
    except ValueError as e:
        return Response({"detail": str(e)}, status=400)

    return Response({"status": "ok"})


# ────────────────────────────────────────────────────────────
# 5. Webhook (public)
# ────────────────────────────────────────────────────────────

@csrf_exempt
def payment_webhook(request, provider_name, invoice_id):
    """
    POST/GET /api/invoicing/payment-webhook/{provider}/{invoice_id}/
    Called by Montonio/Paysera after payment. No Django auth — verified via provider signature.
    """
    logger.info("=" * 60)
    logger.info(
        "[Webhook] Incoming: provider=%s, invoice=%s, method=%s",
        provider_name, invoice_id, request.method,
    )
    logger.info("[Webhook] GET params: %s", dict(request.GET.items()))
 
    if request.method == "POST":
        ct = request.content_type or ""
        logger.info("[Webhook] Content-Type: %s", ct)
        logger.info("[Webhook] POST body (first 1000): %s", request.body[:1000].decode("utf-8", errors="replace"))
 
    request_data = _extract_webhook_data(request)
    logger.info("[Webhook] Parsed data keys: %s", list(request_data.keys()))
 
    # Log key fields per provider
    if provider_name == "montonio":
        token = request_data.get("payment_token") or request_data.get("orderToken") or ""
        logger.info("[Webhook][Montonio] Token present: %s, length: %d", bool(token), len(token))
    elif provider_name == "paysera":
        logger.info(
            "[Webhook][Paysera] data=%s..., ss1=%s",
            (request_data.get("data", ""))[:60],
            request_data.get("ss1", ""),
        )
 
    try:
        allocation = PaymentLinkService.handle_webhook(
            provider_name, int(invoice_id), request_data
        )
 
        if allocation:
            logger.info(
                "[Webhook] SUCCESS: allocation_id=%s, amount=%s, invoice=%s → status=%s",
                allocation.id, allocation.amount, invoice_id,
                allocation.invoice.status if hasattr(allocation, 'invoice') else "?",
            )
        else:
            logger.info("[Webhook] No allocation created (duplicate or not finalized)")
 
        logger.info("=" * 60)
 
        if provider_name == "paysera":
            return HttpResponse("OK", content_type="text/plain")
 
        status_str = "ok" if allocation else "ignored"
        return JsonResponse({"status": status_str})
 
    except Exception as e:
        logger.exception(
            "[Webhook] ERROR: provider=%s invoice=%s error=%s",
            provider_name, invoice_id, e,
        )

        from .celery_signals import _send_telegram
        _send_telegram(
            f"💳 <b>Payment webhook error</b>\n"
            f"Provider: {provider_name}\n"
            f"Invoice: {invoice_id}\n"
            f"Error: {str(e)[:300]}"
        )

        logger.info("=" * 60)
 
        # Always 200 — prevent infinite retries from provider
        if provider_name == "paysera":
            return HttpResponse("OK", content_type="text/plain")
        return JsonResponse({"status": "error", "detail": str(e)}, status=200)


# ════════════════════════════════════════════════════════════
# Internal helpers
# ════════════════════════════════════════════════════════════

def _validate_provider_fields(provider_name: str, data: dict) -> str | None:
    if provider_name == "montonio":
        if not data.get("access_key"):
            return "Įveskite Montonio Access Key"
        if not data.get("secret_key"):
            return "Įveskite Montonio Secret Key"
    elif provider_name == "paysera":
        if not data.get("project_id"):
            return "Įveskite Paysera projekto ID"
        if not data.get("sign_password"):
            return "Įveskite Paysera parašo slaptažodį"
    else:
        return f"Nežinomas teikėjas: {provider_name}"
    return None


def _test_provider_connection(
    provider_name: str, data: dict
) -> tuple[bool, str | None, list | None, dict | None]:
    """Returns (connected, error, methods, raw_response)."""
    if provider_name == "montonio":
        return _test_montonio(data)
    elif provider_name == "paysera":
        return _test_paysera(data)
    return False, f"Nežinomas teikėjas: {provider_name}", None, None


# ── Montonio ─────────────────────────────────────────────────

def _test_montonio(data: dict) -> tuple[bool, str | None, list | None, dict | None]:
    access_key = data.get("access_key", "").strip()
    secret_key = data.get("secret_key", "").strip()
    base_url = MONTONIO_BASE_URLS.get(PAYMENT_ENV, MONTONIO_BASE_URLS["sandbox"])

    try:
        now = int(pytime.time())
        payload = {"accessKey": access_key, "iat": now, "exp": now + 600}
        token = jwt.encode(payload, secret_key, algorithm="HS256")

        url = f"{base_url}/stores/payment-methods"
        logger.info("[Montonio] GET %s (access_key=%s...)", url, access_key[:8])

        resp = http_requests.get(
            url,
            headers={"Authorization": f"Bearer {token}"},
            params={"currency": "EUR"},
            timeout=15,
        )

        logger.info(
            "[Montonio] Response: status=%s body=%s",
            resp.status_code, resp.text[:500],
        )

        raw = {
            "status_code": resp.status_code,
            "body": resp.text[:1000],
        }

        if resp.status_code == 200:
            methods = resp.json()
            logger.info("[Montonio] OK — %d payment methods", len(methods))
            return True, None, methods, raw

        # Parse error from JSON
        try:
            err_body = resp.json()
            montonio_msg = err_body.get("message", "")
            raw["provider_error"] = montonio_msg
        except Exception:
            montonio_msg = ""

        error_map = {
            "ACCESS_KEY_NOT_FOUND": "Montonio Access Key nerastas – patikrinkite raktą",
            "INVALID_SIGNATURE": "Neteisingas Secret Key – patikrinkite raktą",
            "UNAUTHORIZED": "Neteisingi Montonio raktai",
        }

        if montonio_msg in error_map:
            return False, error_map[montonio_msg], None, raw

        if resp.status_code in (400, 401, 403):
            return False, f"Montonio klaida: {montonio_msg or f'HTTP {resp.status_code}'}", None, raw

        return False, f"Montonio netikėtas atsakymas (HTTP {resp.status_code})", None, raw

    except http_requests.Timeout:
        return False, "Montonio neatsako – bandykite vėliau", None, {"timeout": True}
    except Exception as e:
        logger.exception("[Montonio] Test error")
        return False, f"Klaida jungiantis prie Montonio: {e}", None, {"exception": str(e)}

# ── Paysera ──────────────────────────────────────────────────

def _test_paysera(data: dict) -> tuple[bool, str | None, list | None, dict | None]:
    """
    Paysera test: build a signed payment request and POST to Paysera.
    Follow redirects to see the final page — check for errors in HTML.
    """
    project_id = data.get("project_id", "").strip()
    sign_password = data.get("sign_password", "").strip()

    if not project_id.isdigit():
        return False, "Paysera projekto ID turi būti skaičius", None, None
    if len(sign_password) < 5:
        return False, "Paysera parašo slaptažodis per trumpas", None, None

    try:
        # 1. Build params
        params = {
            "projectid": project_id,
            "orderid": "test_connection_check",
            "amount": "100",
            "currency": "EUR",
            "country": "LT",
            "test": "0" if PAYMENT_ENV == "production" else "1",
            "version": "1.6",
            "accepturl": "https://localhost/accept",
            "cancelurl": "https://localhost/cancel",
            "callbackurl": "https://localhost/callback",
        }

        # 2. URL-encode → url-safe base64
        query_string = urllib.parse.urlencode(params)
        b64_data = b64encode(query_string.encode()).decode()
        b64_data_safe = b64_data.replace("+", "-").replace("/", "_")

        # 3. sign = md5(data + password)
        sign = hashlib.md5((b64_data_safe + sign_password).encode()).hexdigest()

        logger.info("[Paysera] POST https://www.paysera.com/pay/ (project=%s)", project_id)
        logger.info("[Paysera] data=%s..., sign=%s", b64_data_safe[:60], sign)

        # 4. POST — follow redirects to get final page
        resp = http_requests.post(
            "https://www.paysera.com/pay/",
            data={"data": b64_data_safe, "sign": sign},
            timeout=20,
            allow_redirects=True,  # follow all redirects
        )

        final_url = resp.url
        status = resp.status_code
        body = resp.text[:2000] if resp.text else ""
        body_lower = body.lower()

        logger.info("[Paysera] Final URL: %s", final_url)
        logger.info("[Paysera] Final status: %s", status)
        logger.info("[Paysera] Body (first 1000): %s", body[:1000])

        raw = {
            "final_url": final_url,
            "status_code": status,
        }

        # 5. Analyze final page — check URL first, then body
        # Error patterns in URL (Paysera puts error info in URL path)
        url_lower = final_url.lower()
        url_has_error = any(kw in url_lower for kw in [
            "error", "bad_request", "pick_payment_error",
        ])

        if url_has_error:
            # Extract error_code from URL if present (e.g. error_code/0x6)
            error_code = ""
            if "error_code/" in final_url:
                parts = final_url.split("error_code/")
                if len(parts) > 1:
                    error_code = parts[1].split("/")[0]

            logger.warning(
                "[Paysera] Error in URL: %s (error_code=%s)", final_url, error_code,
            )

            raw["error_code"] = error_code

            # Map known error codes
            error_messages = {
                "0x1": "Neteisingi parametrai – patikrinkite projekto nustatymus",
                "0x2": "Neteisingi parametrai",
                "0x3": "Netinkama valiuta arba suma",
                "0x4": "Neteisingas parašas (sign) – patikrinkite parašo slaptažodį",
                "0x6": "Neteisingas projekto ID arba parašo slaptažodis",
            }

            err_msg = error_messages.get(
                error_code,
                "Paysera atmetė užklausą – patikrinkite projekto ID ir parašo slaptažodį",
            )
            return False, err_msg, None, raw

        # No error in URL — check body for payment page
        is_payment_page = any(kw in body_lower for kw in [
            "payment", "mokėjim", "choose", "bank", "pasirink",
        ])

        if is_payment_page:
            logger.info("[Paysera] Payment page reached — credentials OK")
            methods = _fetch_paysera_methods(project_id)
            return True, None, methods, raw

        # Not clearly error, not clearly payment — log everything
        logger.warning(
            "[Paysera] Unclear response. URL=%s status=%s body_snippet=%s",
            final_url, status, body[:300],
        )
        return (
            False,
            f"Neaiškus Paysera atsakymas (HTTP {status}). "
            "Patikrinkite projekto ID ir parašo slaptažodį.",
            None,
            raw,
        )

    except http_requests.Timeout:
        return False, "Paysera neatsako – bandykite vėliau", None, {"timeout": True}
    except Exception as e:
        logger.exception("[Paysera] Test error")
        return False, f"Klaida jungiantis prie Paysera: {e}", None, {"exception": str(e)}


def _fetch_paysera_methods(project_id: str) -> list | None:
    """Fetch available payment methods XML (public endpoint)."""
    import xml.etree.ElementTree as ET

    url = f"https://www.paysera.com/new/api/paymentMethods/{project_id}/currency:EUR"

    try:
        logger.info("[Paysera] Fetching methods: %s", url)
        resp = http_requests.get(url, timeout=10)
        logger.info("[Paysera] Methods response: status=%s, size=%d", resp.status_code, len(resp.text))

        if resp.status_code != 200:
            return None

        methods = []
        root = ET.fromstring(resp.text)
        for method_el in root.iter("method"):
            name = method_el.get("title", method_el.get("key", ""))
            logo = method_el.get("logo_url", "")
            if name:
                methods.append({"name": name, "logo_url": logo})

        logger.info("[Paysera] Parsed %d payment methods", len(methods))
        return methods or None

    except Exception:
        logger.warning("[Paysera] Could not fetch methods for project %s", project_id)
        return None


# ── Webhook data extraction ──────────────────────────────────


def _extract_webhook_data(request) -> dict:
    """Universal parsing of webhook data from GET and POST."""
    data = {}
 
    if request.method == "GET":
        data = {k: v for k, v in request.GET.items()}
    elif request.method == "POST":
        ct = request.content_type or ""
        if "json" in ct:
            try:
                data = json.loads(request.body)
            except (json.JSONDecodeError, ValueError):
                data = {}
        else:
            data = {k: v for k, v in request.POST.items()}
 
    # Montonio may send token in query string even on POST
    for key in ("payment_token", "orderToken"):
        if key in request.GET and key not in data:
            data[key] = request.GET[key]
 
    return data




from .models import InvoiceEmail, InvoiceSettings
from .serializers import InvoiceEmailSerializer, ReminderSettingsSerializer



# ════════════════════════════════════════════════════════════
#  Invoice Email — отправка, напоминания, tracking
# ════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_email_list(request, invoice_id):
    """GET /api/invoicing/invoices/<id>/emails/ — список отправленных email."""
    if request.user.is_superuser:
        invoice = get_object_or_404(Invoice, id=invoice_id)
    else:
        invoice = get_object_or_404(Invoice, id=invoice_id, user=request.user)
    emails = invoice.emails.all().order_by("-sent_at")
    serializer = InvoiceEmailSerializer(emails, many=True)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_send_email_view(request, invoice_id):
    """POST /api/invoicing/invoices/<id>/send-email/"""
    invoice = get_object_or_404(Invoice, id=invoice_id, user=request.user)

    if invoice.status not in ("issued", "sent", "partially_paid", "paid"):
        return Response(
            {"detail": "Sąskaitą galima siųsti tik išrašytą arba išsiųstą."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    recipient = request.data.get("email") or invoice.buyer_email or invoice.sent_to_email
    if not recipient:
        return Response(
            {"detail": "Nenurodytas gavėjo el. pašto adresas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    force = request.data.get("force", False)
    requested_type = request.data.get("email_type", "")

    # -- Determine email_type --
    if requested_type in ("invoice", "invoice_paid", "invoice_info"):
        email_type = requested_type
    elif invoice.status == "paid" and invoice.source_invoice_id:
        email_type = "invoice_paid"
    else:
        email_type = "invoice"

    # -- Inv subscription: email limit check --
    allowed, err = check_inv_email_limit(request.user, invoice.id)
    if not allowed:
        return Response(err, status=status.HTTP_403_FORBIDDEN)

    # -- Limits: max 3 invoice, max 1 invoice_paid/invoice_info per invoice --
    max_counts = {"invoice": 3, "invoice_paid": 1, "invoice_info": 2}
    max_count = max_counts.get(email_type, 3)

    total_sent = InvoiceEmail.objects.filter(
        invoice=invoice, email_type=email_type, status="sent",
    ).count()

    if total_sent >= max_count:
        return Response(
            {"detail": f"Pasiektas išsiųstų laiškų limitas ({max_count}) vienai sąskaitai."},
            status=status.HTTP_429_TOO_MANY_REQUESTS,
        )

    # -- Confirm check (skip for first send) --
    if total_sent > 0 and not force:
        last_email = InvoiceEmail.objects.filter(
            invoice=invoice, email_type=email_type, status="sent",
        ).order_by("-sent_at").first()

        # 5 min cooldown
        if last_email and last_email.sent_at:
            from datetime import timedelta as td
            diff = timezone.now() - last_email.sent_at
            if diff < td(minutes=5):
                return Response(
                    {"detail": "Prašome palaukti 5 min. prieš siunčiant pakartotinį laišką."},
                    status=status.HTTP_429_TOO_MANY_REQUESTS,
                )

        return Response({
            "needs_confirm": True,
            "last_sent_at": last_email.sent_at if last_email else None,
            "last_sent_to": last_email.to_email if last_email else "",
            "total_sent": total_sent,
            "max_count": max_count,
        }, status=status.HTTP_200_OK)

    # -- Update counters immediately --
    update_fields = {
        "email_sent_count": models.F("email_sent_count") + 1,
        "email_last_status": "sent",
    }
    if email_type == "invoice" and invoice.status == "issued":
        update_fields.update({
            "status": "sent",
            "sent_at": timezone.now(),
            "sent_to_email": recipient,
        })
    Invoice.objects.filter(id=invoice.id).update(**update_fields)

    # -- Record inv email usage --
    inv_email_info = None
    try:
        inv_email_info = record_inv_email(request.user, invoice.id)
    except Exception as e:
        logger.warning(f"Failed to record inv email usage: {e}")

    # -- Send in background --
    import threading
    from .services.invoice_email_service import send_invoice_email

    def _send():
        import django
        django.db.connections.close_all()
        send_invoice_email(
            invoice_id=invoice.id,
            email_type=email_type,
            recipient_email=recipient,
            skip_counter=True,
        )

    threading.Thread(target=_send, daemon=True).start()

# -- Build response with usage info --
    resp_data = {"detail": "El. laiškas siunčiamas.", "to": recipient}
    if inv_email_info:
        emails_used, emails_max, inv_status, was_new = inv_email_info
        resp_data["emails_used"] = emails_used
        resp_data["emails_max"] = emails_max
        resp_data["inv_status"] = inv_status
        resp_data["was_new_email"] = was_new

    return Response(resp_data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def invoice_send_reminder_view(request, invoice_id):
    """POST /api/invoicing/invoices/<id>/send-reminder/"""
    invoice = get_object_or_404(Invoice, id=invoice_id, user=request.user)

    if invoice.status not in ("issued", "sent", "partially_paid"):
        return Response(
            {"detail": "Priminimą galima siųsti tik neapmokėtai sąskaitai."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if invoice.send_payment_reminders:
        return Response(
            {"detail": "Automatiniai priminimai įjungti. Išjunkite juos, jei norite siųsti rankinį priminimą."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    recipient = request.data.get("email") or invoice.buyer_email or invoice.sent_to_email
    if not recipient:
        return Response(
            {"detail": "Nenurodytas gavėjo el. pašto adresas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # -- Inv subscription: email limit check --
    allowed, err = check_inv_email_limit(request.user, invoice.id)
    if not allowed:
        return Response(err, status=status.HTTP_403_FORBIDDEN)

    # Инкрементим счётчик сразу
    Invoice.objects.filter(id=invoice.id).update(
        email_sent_count=models.F("email_sent_count") + 1,
        email_last_status="sent",
    )

    # -- Record inv email usage --
    inv_email_info = None
    try:
        inv_email_info = record_inv_email(request.user, invoice.id)
    except Exception as e:
        logger.warning(f"Failed to record inv email usage: {e}")

    import threading
    from .services.invoice_email_service import send_invoice_email

    def _send():
        import django
        django.db.connections.close_all()
        send_invoice_email(
            invoice_id=invoice.id,
            email_type="manual_reminder",
            recipient_email=recipient,
            skip_counter=True,
        )

    threading.Thread(target=_send, daemon=True).start()

    # -- Build response with usage info --
    resp_data = {"detail": "Priminimas siunčiamas.", "to": recipient}
    if inv_email_info:
        emails_used, emails_max, inv_status, was_new = inv_email_info
        resp_data["emails_used"] = emails_used
        resp_data["emails_max"] = emails_max
        resp_data["inv_status"] = inv_status
        resp_data["was_new_email"] = was_new

    return Response(resp_data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def invoice_email_summary(request):
    """
    GET /api/invoicing/invoices/email-summary/?ids=1,2,3
    Краткая сводка для иконок в InvoiceList.
    """
    ids_param = request.query_params.get("ids", "")
    if not ids_param:
        return Response({})

    try:
        invoice_ids = [int(x) for x in ids_param.split(",") if x.strip()]
    except ValueError:
        return Response({})

    valid_ids = set(
        Invoice.objects.filter(id__in=invoice_ids, user=request.user)
        .values_list("id", flat=True)
    )

    from django.db.models import Count, Max, Q as DQ

    email_stats = (
        InvoiceEmail.objects.filter(invoice_id__in=valid_ids)
        .values("invoice_id")
        .annotate(
            total=Count("id"),
            sent_count=Count("id", filter=DQ(status="sent")),
            failed_count=Count("id", filter=DQ(status__in=["failed", "bounced"])),
            opened_count=Count("id", filter=DQ(opened_at__isnull=False)),
            last_sent=Max("sent_at"),
        )
    )

    result = {}
    for stat in email_stats:
        inv_id = stat["invoice_id"]
        if stat["opened_count"] > 0:
            icon_status = "opened"
        elif stat["failed_count"] > 0 and stat["sent_count"] == 0:
            icon_status = "failed"
        elif stat["sent_count"] > 0:
            icon_status = "sent"
        else:
            icon_status = "none"

        result[str(inv_id)] = {
            "total": stat["total"],
            "icon_status": icon_status,
            "last_sent": stat["last_sent"],
        }

    return Response(result)


@api_view(["GET", "PATCH"])
@permission_classes([IsAuthenticated])
def reminder_settings_view(request):
    """GET/PATCH /api/invoicing/reminder-settings/"""
    inv_settings, _ = InvoiceSettings.objects.get_or_create(user=request.user)

    if request.method == "GET":
        return Response({
            "reminder_enabled": inv_settings.reminder_enabled,
            "invoice_reminder_days": inv_settings.invoice_reminder_days or [-7, -1, 3],
        })

    serializer = ReminderSettingsSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    if "reminder_enabled" in serializer.validated_data:
        inv_settings.reminder_enabled = serializer.validated_data["reminder_enabled"]
    if "invoice_reminder_days" in serializer.validated_data:
        inv_settings.invoice_reminder_days = serializer.validated_data["invoice_reminder_days"]

    inv_settings.save(update_fields=["reminder_enabled", "invoice_reminder_days"])

    return Response({
        "reminder_enabled": inv_settings.reminder_enabled,
        "invoice_reminder_days": inv_settings.invoice_reminder_days,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def reminder_settings_reset_view(request):
    """POST /api/invoicing/reminder-settings/reset/"""
    inv_settings, _ = InvoiceSettings.objects.get_or_create(user=request.user)
    inv_settings.invoice_reminder_days = [-7, -1, 3]
    inv_settings.save(update_fields=["invoice_reminder_days"])

    return Response({
        "reminder_enabled": inv_settings.reminder_enabled,
        "invoice_reminder_days": inv_settings.invoice_reminder_days,
    })


# ════════════════════════════════════════════════════════════
#  Mailgun invoice tracking webhook
# ════════════════════════════════════════════════════════════

@csrf_exempt
@require_POST
def mailgun_invoice_tracking_webhook(request):
    """POST /api/webhooks/mailgun/invoice-tracking/"""
    import time as time_module

    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return HttpResponse("bad json", status=400)

    event_data = payload.get("event-data", {})
    signature_data = payload.get("signature", {})

    timestamp = str(signature_data.get("timestamp", ""))
    token = signature_data.get("token", "")
    signature = signature_data.get("signature", "")

    # Verify signature
    signing_key = getattr(django_settings, "MAILGUN_INVOICE_WEBHOOK_SIGNING_KEY", "")
    if not signing_key:
        logger.error("MAILGUN_INVOICE_WEBHOOK_SIGNING_KEY not configured")
        return HttpResponseForbidden("Not configured")

    try:
        if abs(time_module.time() - int(timestamp)) > 300:
            return HttpResponseForbidden("Stale request")
    except (ValueError, TypeError):
        return HttpResponseForbidden("Invalid timestamp")

    expected = hmac.new(
        key=signing_key.encode("utf-8"),
        msg=f"{timestamp}{token}".encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(signature, expected):
        logger.warning("Mailgun invoice tracking: invalid signature")
        return HttpResponseForbidden("Invalid signature")

    event_type = event_data.get("event", "")
    message_headers = event_data.get("message", {}).get("headers", {})
    message_id = message_headers.get("message-id", "").strip("<>")

    if not message_id:
        return HttpResponse("ok", status=200)

    try:
        email_log = InvoiceEmail.objects.get(mailgun_message_id=message_id)
    except InvoiceEmail.DoesNotExist:
        return HttpResponse("ok", status=200)
    except InvoiceEmail.MultipleObjectsReturned:
        email_log = InvoiceEmail.objects.filter(mailgun_message_id=message_id).first()

    if event_type == "opened":
        from django.db.models import F
        update_fields = ["open_count"]
        email_log.open_count = F("open_count") + 1
        if not email_log.opened_at:
            email_log.opened_at = timezone.now()
            update_fields.append("opened_at")
        email_log.save(update_fields=update_fields)
        logger.info(f"Invoice email {email_log.id} opened (invoice {email_log.invoice_id})")

    elif event_type == "delivered":
        if email_log.status == "pending":
            email_log.status = "sent"
            email_log.save(update_fields=["status"])

    elif event_type in ("failed", "rejected"):
        delivery_status = event_data.get("delivery-status", {})
        error_msg = delivery_status.get("message", "")[:500] or delivery_status.get("description", "")[:500]
        email_log.status = "failed"
        email_log.error_text = error_msg
        email_log.save(update_fields=["status", "error_text"])
        logger.warning(f"Invoice email {email_log.id} failed: {error_msg}")

        from .celery_signals import _send_telegram
        _send_telegram(
            f"📧 <b>Email failed</b>\n"
            f"Invoice: {email_log.invoice_id}\n"
            f"To: {email_log.to_email}\n"
            f"Type: {email_log.email_type}\n"
            f"Error: {error_msg[:300]}"
        )

    elif event_type == "complained":
        email_log.status = "bounced"
        email_log.error_text = "Spam complaint"
        email_log.save(update_fields=["status", "error_text"])

    return HttpResponse("ok", status=200)





# ════════════════════════════════════════════════════════════
#  Subscriptions - israsymas only
# ════════════════════════════════════════════════════════════

def get_inv_access(user):
    """
    Возвращает полный статус подписки Išrašymas + лимиты + фичи.
    Lazy-expire: если trial/active истёк — переводит в free.
    """
    sub = getattr(user, "inv_subscription", None)
    if sub is None:
        # На случай если запись не создана (старый user до миграции)
        from .models import InvSubscription
        sub, _ = InvSubscription.objects.get_or_create(user=user)

    # Lazy expire
    sub.check_and_expire()

    features = sub.get_features()
    days_left = sub.days_left

    # Баннер-логика
    banner = None
    if sub.status == "free" and not sub.trial_used:
        banner = "trial_available"       # «Начните 14-дневный trial»
    elif sub.status == "trial" and sub.show_trial_banner:
        banner = "trial_ending"          # «Осталось X дней»
    elif sub.status == "free" and sub.trial_used:
        banner = "trial_expired"         # «Trial закончился, купите план»

    # Лимиты (только для free)
    limits = None
    if sub.status == "free":
        from .models import InvMonthlyUsage
        usage = InvMonthlyUsage.get_current(user)
        limits = {
            "exports_max": 30,
            "exports_used": usage.exports_used,
            "emails_max": 10,
            "emails_used": usage.emails_used,
        }

    return {
        "status": sub.status,
        "trial_used": sub.trial_used,
        "days_left": days_left,
        "banner": banner,
        "features": features,
        "limits": limits,
    }


def check_inv_feature(user, feature_name):
    """
    Быстрая проверка одной фичи. Для использования в view перед действием.
    Возвращает (allowed: bool, error_response_data: dict|None)
    """
    sub = getattr(user, "inv_subscription", None)
    if sub is None:
        from .models import InvSubscription
        sub, _ = InvSubscription.objects.get_or_create(user=user)

    sub.check_and_expire()
    features = sub.get_features()

    if feature_name in features and not features[feature_name]:
        return False, {
            "error": "feature_locked",
            "feature": feature_name,
            "message": "Ši funkcija leidžiama tik turintiems mokamą planą.",
        }
    return True, None


def check_inv_export_limit(user, invoice_id):
    """
    Проверка лимита экспорта. Вызывать перед экспортом фактуры.
    Возвращает (allowed, error_data).
    """
    sub = getattr(user, "inv_subscription", None)
    if sub is None:
        from .models import InvSubscription
        sub, _ = InvSubscription.objects.get_or_create(user=user)

    sub.check_and_expire()
    if sub.status in ("trial", "active"):
        return True, None

    from .models import InvMonthlyUsage
    usage = InvMonthlyUsage.get_current(user)
    if not usage.can_export(invoice_id):
        return False, {
            "error": "limit_reached",
            "feature": "export",
            "message": "Pasiektas mėnesio eksporto limitas (30 sąskaitų).",
            "exports_used": usage.exports_used,
            "exports_max": 30,
        }
    return True, None


def check_inv_email_limit(user, invoice_id):
    from .models import InvSubscription, InvMonthlyUsage
    sub = InvSubscription.objects.filter(user=user).first()
    if sub is None:
        return True, None

    sub.check_and_expire()
    if sub.status in ("trial", "active"):
        return True, None

    usage = InvMonthlyUsage.get_current(user)
    if not usage.can_email(invoice_id):
        return False, {
            "error": "limit_reached",
            "feature": "email",
            "message": (
                f"Mėnesio el. pašto limitas: {usage.emails_used}/10 panaudota. "
                f"Įsigykite mokamą planą neribotam naudojimui."
            ),
            "emails_used": usage.emails_used,
            "emails_max": 10,
        }
    return True, None


def record_inv_email(user, invoice_id):
    from .models import InvSubscription, InvMonthlyUsage
    sub = InvSubscription.objects.filter(user=user).first()
    if sub is None or sub.status != "free":
        return None

    usage = InvMonthlyUsage.get_current(user)
    was_new = invoice_id not in usage.emailed_invoice_ids
    usage.record_email(invoice_id)
    usage.refresh_from_db()
    return usage.emails_used, 10, "free", was_new




# --- Inv Subscription endpoints ---

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def inv_subscription_status(request):
    """
    GET /api/inv/subscription/
    Возвращает полный статус подписки Išrašymas.
    """
    data = get_inv_access(request.user)

    # Check Stripe cancel_at_period_end
    sub = InvSubscription.objects.filter(user=request.user).first()
    cancel_at_period_end = False
    plan_end_display = ""

    if sub and sub.stripe_subscription_id and sub.status == "active":
        try:
            import stripe
            stripe.api_key = django_settings.STRIPE_SECRET_KEY
            stripe_sub = stripe.Subscription.retrieve(sub.stripe_subscription_id)
            cancel_at_period_end = stripe_sub.get("cancel_at_period_end", False)
        except Exception as e:
            logger.warning("[InvSub] Stripe check failed: %s", e)

    if sub and sub.plan_end:
        plan_end_display = sub.plan_end.strftime("%Y-%m-%d")

    data["cancel_at_period_end"] = cancel_at_period_end
    data["plan_end_display"] = plan_end_display

    return Response(data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def inv_start_trial(request):
    """
    POST /api/inv/start-trial/
    Активирует 14-дневный trial Išrašymas.
    """
    sub = getattr(request.user, "inv_subscription", None)
    if sub is None:
        sub, _ = InvSubscription.objects.get_or_create(user=request.user)

    if sub.trial_used:
        return Response(
            {"error": "trial_already_used",
             "message": "Bandomasis laikotarpis jau buvo panaudotas."},
            status=400,
        )

    if sub.status == "active":
        return Response(
            {"error": "already_active",
             "message": "Jūs jau turite aktyvų planą."},
            status=400,
        )

    sub.start_trial()
    logger.info(f"Inv trial started for user {request.user.email}, ends {sub.trial_end}")

    return Response(get_inv_access(request.user), status=200)



@api_view(["POST"])
@permission_classes([IsAuthenticated])
def inv_cancel_subscription(request):
    """POST /api/inv/cancel-subscription/"""
    import stripe
    stripe.api_key = django_settings.STRIPE_SECRET_KEY

    sub = InvSubscription.objects.filter(user=request.user).first()
    if not sub or sub.status != "active" or not sub.stripe_subscription_id:
        return Response(
            {"error": "Neturite aktyvaus PRO plano."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        stripe.Subscription.modify(
            sub.stripe_subscription_id,
            cancel_at_period_end=True,
        )
    except Exception as e:
        logger.error("[InvCancel] Stripe error: %s", e)
        return Response(
            {"error": "Nepavyko atšaukti plano."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    plan_end_display = sub.plan_end.strftime("%Y-%m-%d") if sub.plan_end else ""

    logger.info("[InvCancel] User %s cancelled PRO, active until %s", request.user.email, plan_end_display)

    return Response({
        "status": "active",
        "cancel_at_period_end": True,
        "plan_end": str(sub.plan_end) if sub.plan_end else None,
        "plan_end_display": plan_end_display,
    })

# ════════════════════════════════════════════════════════════
#  END --- Subscriptions - israsymas only
# ════════════════════════════════════════════════════════════





# ════════════════════════════════════════════════════════════
# ─── Rivile GAMA API Key Views ───
# ════════════════════════════════════════════════════════════
class RivileGamaAPIKeyListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = RivileGamaAPIKey.objects.filter(user=request.user)
        return Response(RivileGamaAPIKeySerializer(qs, many=True).data)

    def post(self, request):
        ser = RivileGamaAPIKeyCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        user = request.user
        company_code = ser.validated_data["company_code"]
        raw_key = ser.validated_data["api_key"]
        label = ser.validated_data.get("label", "")

        if RivileGamaAPIKey.objects.filter(user=user, company_code=company_code).exists():
            return Response(
                {"detail": f"API raktas įmonei {company_code} jau egzistuoja. Naudokite redagavimą."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        verify_result = verify_api_key(raw_key)

        obj = RivileGamaAPIKey(
            user=user,
            company_code=company_code,
            label=label,
            is_active=ser.validated_data.get("is_active", True),
        )
        obj.set_api_key(raw_key)
        obj.save()

        obj.mark_verified(
            success=verify_result.success,
            error="" if verify_result.success else (verify_result.error_message or "Verification failed"),
        )

        logger.info(
            "[RIVILE_API_KEY] Created key for user=%s company=%s verified=%s",
            user.id, company_code, verify_result.success,
        )

        return Response(RivileGamaAPIKeySerializer(obj).data, status=status.HTTP_201_CREATED)


class RivileGamaAPIKeyDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_obj(self, request, pk):
        try:
            return RivileGamaAPIKey.objects.get(pk=pk, user=request.user)
        except RivileGamaAPIKey.DoesNotExist:
            return None

    def get(self, request, pk):
        obj = self._get_obj(request, pk)
        if not obj:
            return Response({"detail": "Raktas nerastas."}, status=status.HTTP_404_NOT_FOUND)
        return Response(RivileGamaAPIKeySerializer(obj).data)

    def patch(self, request, pk):
        obj = self._get_obj(request, pk)
        if not obj:
            return Response({"detail": "Raktas nerastas."}, status=status.HTTP_404_NOT_FOUND)

        ser = RivileGamaAPIKeyUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        if "label" in ser.validated_data:
            obj.label = ser.validated_data["label"]

        if "is_active" in ser.validated_data:
            obj.is_active = ser.validated_data["is_active"]

        if "company_code" in ser.validated_data:
            new_code = ser.validated_data["company_code"].strip()
            if new_code and new_code != obj.company_code:
                if RivileGamaAPIKey.objects.filter(
                    user=request.user, company_code=new_code
                ).exclude(pk=obj.pk).exists():
                    return Response(
                        {"detail": f"API raktas įmonei {new_code} jau egzistuoja."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                obj.company_code = new_code

        raw_key = ser.validated_data.get("api_key", "").strip()
        if raw_key:
            obj.set_api_key(raw_key)
            obj.save()
            verify_result = verify_api_key(raw_key)
            obj.mark_verified(
                success=verify_result.success,
                error="" if verify_result.success else (verify_result.error_message or ""),
            )
        else:
            obj.save()

        return Response(RivileGamaAPIKeySerializer(obj).data)

    def delete(self, request, pk):
        obj = self._get_obj(request, pk)
        if not obj:
            return Response({"detail": "Raktas nerastas."}, status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class RivileGamaAPIKeyVerifyView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            obj = RivileGamaAPIKey.objects.get(pk=pk, user=request.user)
        except RivileGamaAPIKey.DoesNotExist:
            return Response({"detail": "Raktas nerastas."}, status=status.HTTP_404_NOT_FOUND)

        raw_key = obj.get_api_key()
        result = verify_api_key(raw_key)

        obj.mark_verified(
            success=result.success,
            error="" if result.success else (result.error_message or "Patikrinimas nepavyko"),
        )

        logger.info("[RIVILE_API_KEY] Verify key=%s company=%s ok=%s", pk, obj.company_code, result.success)

        return Response(RivileGamaAPIKeySerializer(obj).data)

# ════════════════════════════════════════════════════════════
# END ─── Rivile GAMA API Key Views ───
# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
# ─── Extra fields v nustatymai ───
# ════════════════════════════════════════════════════════════

"""
Views для per-company extra fields.

Добавить в docscanner_app/views.py (или в отдельный файл и импортировать).

Endpoint'ы:
  GET    /api/extra-fields/<program_key>/                    — список профилей (пагинация)
  GET    /api/extra-fields/<program_key>/<company_code>/     — полные данные профиля
  PATCH  /api/extra-fields/<program_key>/<company_code>/     — создать/обновить профиль
  DELETE /api/extra-fields/<program_key>/<company_code>/     — удалить профиль
  POST   /api/extra-fields/<program_key>/check-duplicate/    — проверка дубликата
"""

import logging
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .utils.extra_fields import (
    VALID_PROGRAM_KEYS,
    get_field_name,
    get_profiles_summary,
    count_non_empty_fields,
    get_non_empty_field_keys,
)

logger = logging.getLogger("docscanner_app")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def extra_fields_list(request, program_key):
    """
    GET /api/extra-fields/<program_key>/?offset=0&limit=5

    Возвращает лёгкий список профилей (без полных полей).
    """
    if program_key not in VALID_PROGRAM_KEYS:
        return Response(
            {"detail": f"Nežinoma programa: {program_key}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    field_name = get_field_name(program_key)
    data = getattr(request.user, field_name, None) or {}

    all_profiles = get_profiles_summary(data)
    total = len(all_profiles)

    # Пагинация
    try:
        offset = max(0, int(request.query_params.get("offset", 0)))
        limit = max(1, min(100, int(request.query_params.get("limit", 5))))
    except (ValueError, TypeError):
        offset, limit = 0, 5

    page = all_profiles[offset : offset + limit]

    return Response({
        "total": total,
        "offset": offset,
        "limit": limit,
        "items": page,
    })


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def extra_fields_detail(request, program_key, company_code):
    """
    GET    — полные данные одного профиля
    PATCH  — создать или обновить профиль (мердж на бэкенде)
    DELETE — удалить профиль
    """
    if program_key not in VALID_PROGRAM_KEYS:
        return Response(
            {"detail": f"Nežinoma programa: {program_key}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user = request.user
    field_name = get_field_name(program_key)
    data = getattr(user, field_name, None) or {}

    # Нормализуем data в nested формат если ещё плоский (legacy)
    data = _ensure_nested(data)

    # ──── GET ────
    if request.method == "GET":
        profile = data.get(company_code)
        if profile is None:
            return Response(
                {"detail": "Profilis nerastas."},
                status=status.HTTP_404_NOT_FOUND,
            )
        # Возвращаем без служебных ключей
        clean = {k: v for k, v in profile.items() if not k.startswith("__")}
        return Response({
            "company_code": company_code,
            "company_name": profile.get("__name__", ""),
            "fields": clean,
            "fields_count": count_non_empty_fields(profile),
        })

    # ──── DELETE ────
    if request.method == "DELETE":
        if company_code not in data:
            return Response(
                {"detail": "Profilis nerastas."},
                status=status.HTTP_404_NOT_FOUND,
            )
        del data[company_code]
        setattr(user, field_name, data)
        user.save(update_fields=[field_name])
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ──── PATCH (create or update) ────
    body = request.data
    if not isinstance(body, dict):
        return Response(
            {"detail": "Turinys turi būti JSON objektas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    fields = body.get("fields", {})
    company_name = body.get("company_name", "")

    if not isinstance(fields, dict):
        return Response(
            {"detail": "fields turi būti JSON objektas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Валидация: для не-__all__ нужен company_code
    if company_code != "__all__" and not company_code.strip():
        return Response(
            {"detail": "Įmonės kodas privalomas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Сохраняем __name__ для отображения в списке
    if company_code != "__all__" and company_name:
        fields["__name__"] = str(company_name).strip()
    elif company_code == "__all__":
        fields.pop("__name__", None)

    # Мердж: обновляем только переданные поля, не трогая остальные
    existing = data.get(company_code, {})
    if isinstance(existing, dict):
        existing.update(fields)
    else:
        existing = fields

    data[company_code] = existing
    setattr(user, field_name, data)
    user.save(update_fields=[field_name])

    clean = {k: v for k, v in existing.items() if not k.startswith("__")}
    return Response({
        "company_code": company_code,
        "company_name": existing.get("__name__", ""),
        "fields": clean,
        "fields_count": count_non_empty_fields(existing),
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def extra_fields_check_duplicate(request, program_key):
    """
    POST /api/extra-fields/<program_key>/check-duplicate/
    Body: { "company_code": "123456" }

    Проверяет, есть ли уже профиль с таким company_code.
    Если есть — возвращает список непустых полей (для предупреждения).
    """
    if program_key not in VALID_PROGRAM_KEYS:
        return Response(
            {"detail": f"Nežinoma programa: {program_key}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    company_code = (request.data.get("company_code", "") or "").strip()
    if not company_code:
        return Response(
            {"detail": "company_code privalomas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    field_name = get_field_name(program_key)
    data = getattr(request.user, field_name, None) or {}
    data = _ensure_nested(data)

    existing = data.get(company_code)
    if existing and isinstance(existing, dict):
        non_empty = get_non_empty_field_keys(existing)
        return Response({
            "exists": True,
            "company_name": existing.get("__name__", ""),
            "fields_count": count_non_empty_fields(existing),
            "non_empty_fields": non_empty,
        })

    return Response({"exists": False})


def _ensure_nested(data):
    """
    Если data — плоский dict (legacy), конвертируем в nested.
    Если уже nested — возвращаем как есть.
    """
    if not data or not isinstance(data, dict):
        return {}

    # Уже nested
    if "__all__" in data or "__israsymas__" in data:
        return data

    # Проверяем: если есть ключи с "_" — это плоский формат
    has_flat_keys = any(
        "_" in k and not k.startswith("__")
        for k in data
    )

    if has_flat_keys:
        # Конвертируем на лету (в памяти, не сохраняем — сохранит PATCH)
        return {"__all__": dict(data)}

    # Может быть пустой nested или уже с company codes
    return data

# ════════════════════════════════════════════════════════════
# END ─── Extra fields v nustatymai ───
# ════════════════════════════════════════════════════════════


# ════════════════════════════════════════════════════════════
# ─── Dlia ADMIN israsymas ───
# ════════════════════════════════════════════════════════════

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from django.utils.dateparse import parse_date
from datetime import timedelta
 
 
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_all_invoices(request):
    """
    Dlia superuser — spisok VSEX sčotov vsex polzovatelej.
    Offset/limit paginacija dlia infinite scroll.
    """
    if not request.user.is_superuser:
        return Response({"detail": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)
 
    qs = Invoice.objects.select_related("user").all()
 
    # --- filtry ---
 
    st = request.GET.get("status")
    if st:
        qs = qs.filter(status=st)
 
    invoice_type = request.GET.get("invoice_type")
    if invoice_type:
        qs = qs.filter(invoice_type=invoice_type)
 
    q = request.GET.get("q")
    if q:
        qs = qs.filter(
            Q(document_number__icontains=q)
            | Q(document_series__icontains=q)
            | Q(buyer_name__icontains=q)
            | Q(buyer_email__icontains=q)
            | Q(seller_name__icontains=q)
            | Q(user__email__icontains=q)
        )
 
    date_from = request.GET.get("date_from")
    if date_from:
        d = parse_date(date_from)
        if d:
            qs = qs.filter(invoice_date__gte=d)
 
    date_to = request.GET.get("date_to")
    if date_to:
        d = parse_date(date_to)
        if d:
            qs = qs.filter(invoice_date__lte=d)
 
    # --- sortirovka ---
    qs = qs.order_by("-created_at", "-id")
 
    # --- offset/limit paginacija ---
    try:
        offset = int(request.GET.get("offset", 0))
    except (ValueError, TypeError):
        offset = 0
    try:
        limit = int(request.GET.get("limit", 50))
    except (ValueError, TypeError):
        limit = 50
    limit = min(limit, 100)  # max 100
 
    total = qs.count()
    results = qs[offset : offset + limit]
 
    serializer = InvoiceAdminListSerializer(results, many=True)
 
    return Response({
        "count": total,
        "results": serializer.data,
    })




@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_all_recurring_invoices(request):
    """
    Dlia superuser — spisok VSEX periodičeskich sčotov vsex polzovatelej.
    Offset/limit paginacija.
    """
    if not request.user.is_superuser:
        return Response({"detail": "Not authorized"}, status=status.HTTP_403_FORBIDDEN)
 
    qs = RecurringInvoice.objects.select_related("user").prefetch_related("line_items").all()
 
    # --- filtry ---
    q = request.GET.get("q")
    if q:
        qs = qs.filter(
            Q(document_series__icontains=q)
            | Q(buyer_name__icontains=q)
            | Q(buyer_email__icontains=q)
            | Q(seller_name__icontains=q)
            | Q(user__email__icontains=q)
        )
 
    st = request.GET.get("status")
    if st:
        qs = qs.filter(status=st)
 
    qs = qs.order_by("-created_at", "-id")
 
    # --- offset/limit ---
    try:
        offset = int(request.GET.get("offset", 0))
    except (ValueError, TypeError):
        offset = 0
    try:
        limit = int(request.GET.get("limit", 50))
    except (ValueError, TypeError):
        limit = 50
    limit = min(limit, 100)
 
    total = qs.count()
    results = qs[offset : offset + limit]
 
    serializer = RecurringInvoiceAdminListSerializer(results, many=True)
 
    return Response({
        "count": total,
        "results": serializer.data,
    })

# ════════════════════════════════════════════════════════════
# END ─── Dlia ADMIN israsymas ───
# ════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════
# Funkcii sviazany s exportom po API
# ═══════════════════════════════════════════════════

"""
Views for APIProviderKey — универсальный CRUD + verify.

URL pattern:
  /api/settings/api-keys/<provider>/          — GET list, POST create
  /api/settings/api-keys/<provider>/<pk>/     — GET detail, PATCH update, DELETE
  /api/settings/api-keys/<provider>/<pk>/verify/ — POST verify

Добавить в docscanner_app/views.py (или отдельный файл).
"""

import logging
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

logger = logging.getLogger("docscanner_app")

VALID_PROVIDERS = {"rivile_gama_api", "dineta", "optimum"}


def _verify_key(provider, credentials):
    """
    Проверяет credentials для конкретного провайдера.
    Возвращает (success: bool, error: str).
    """
    try:
        if provider == "rivile_gama_api":
            from docscanner_app.exports.rivile_gama_api import verify_api_key
            result = verify_api_key(credentials.get("api_key", ""))
            if result.success:
                return True, ""
            return False, result.error_message or "API raktas neteisingas."

        elif provider == "dineta":
            from docscanner_app.exports.dineta import (
                dineta_hello, parse_dineta_url, build_api_base_url, build_auth_header,
            )
            from docscanner_app.utils.password_encryption import decrypt_password

            url = credentials.get("url", "")
            username = credentials.get("username", "")
            password = credentials.get("password", "")

            server, client = parse_dineta_url(url)
            dineta_hello(
                server=server,
                client=client,
                username=username,
                password=password,
            )
            return True, ""

        elif provider == "optimum":
            from docscanner_app.exports.optimum import optimum_hello
            optimum_hello(credentials.get("api_key", ""))
            return True, ""

    except Exception as e:
        return False, str(e) or f"{provider}: klaida"

    return False, "Nežinomas teikėjas."


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def api_provider_keys_list(request, provider):
    """
    GET  — список ключей для провайдера.
    POST — создать новый ключ.
    """
    from docscanner_app.models import APIProviderKey

    if provider not in VALID_PROVIDERS:
        return Response(
            {"detail": f"Nežinomas teikėjas: {provider}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user = request.user

    # ─── GET ───
    if request.method == "GET":
        keys = APIProviderKey.objects.filter(user=user, provider=provider)
        data = []
        for k in keys:
            data.append({
                "id": k.pk,
                "provider": k.provider,
                "label": k.label,
                "company_code": k.company_code,
                "key_suffix": k.key_suffix,
                "is_active": k.is_active,
                "use_for_all": k.use_for_all,
                "verified_at": k.verified_at,
                "last_ok": k.last_ok,
                "last_error": k.last_error,
                "created_at": k.created_at,
            })
        return Response(data)

    # ─── POST (create) ───
    body = request.data or {}
    use_for_all = bool(body.get("use_for_all", False))
    company_code = "__all__" if use_for_all else (body.get("company_code") or "").strip()
    label = (body.get("label") or "").strip()

    if not company_code:
        return Response(
            {"detail": "Įmonės kodas yra privalomas."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Проверка дубликата
    if APIProviderKey.objects.filter(user=user, provider=provider, company_code=company_code).exists():
        if company_code == "__all__":
            return Response(
                {"detail": "Bendras raktas \"Visoms įmonėms\" jau egzistuoja. Redaguokite esamą profilį arba ištrinkite ir sukurkite naują."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(
            {"detail": f"Raktas įmonei '{company_code}' jau egzistuoja. Redaguokite esamą profilį arba ištrinkite ir sukurkite naują."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Собираем credentials
    creds = _extract_credentials(body, provider)
    if not creds:
        return Response(
            {"detail": "Prisijungimo duomenys yra privalomi."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Валидация credentials
    missing = _validate_credentials(creds, provider)
    if missing:
        return Response(
            {"detail": f"Trūksta laukų: {', '.join(missing)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Создаём
    key_obj = APIProviderKey(
        user=user,
        provider=provider,
        label=label,
        company_code=company_code,
        use_for_all=use_for_all,
    )
    key_obj.set_credentials(creds)
    key_obj.save()

    # Верифицируем
    success, error = _verify_key(provider, creds)
    key_obj.mark_verified(success, error)

    logger.info(
        "[API_KEYS] Created %s key=%s company=%s verified=%s",
        provider, key_obj.pk, company_code, success,
    )

    return Response(_serialize_key(key_obj), status=status.HTTP_201_CREATED)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def api_provider_keys_detail(request, provider, pk):
    """
    GET    — детали одного ключа.
    PATCH  — обновить (label, company_code, credentials, is_active, use_for_all).
    DELETE — удалить.
    """
    from docscanner_app.models import APIProviderKey

    if provider not in VALID_PROVIDERS:
        return Response(
            {"detail": f"Nežinomas teikėjas: {provider}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        key_obj = APIProviderKey.objects.get(pk=pk, user=request.user, provider=provider)
    except APIProviderKey.DoesNotExist:
        return Response({"detail": "Nerastas."}, status=status.HTTP_404_NOT_FOUND)

    # ─── GET ───
    if request.method == "GET":
        return Response(_serialize_key(key_obj))

    # ─── DELETE ───
    if request.method == "DELETE":
        key_obj.delete()
        logger.info("[API_KEYS] Deleted %s key=%s", provider, pk)
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ─── PATCH ───
    body = request.data or {}

    if "label" in body:
        key_obj.label = (body["label"] or "").strip()

    if "company_code" in body:
        new_code = (body["company_code"] or "").strip()
        if new_code and new_code != key_obj.company_code:
            # Проверка дубликата
            if APIProviderKey.objects.filter(
                user=request.user, provider=provider, company_code=new_code
            ).exclude(pk=pk).exists():
                return Response(
                    {"detail": f"Raktas įmonei '{new_code}' jau egzistuoja."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            key_obj.company_code = new_code

    if "is_active" in body:
        key_obj.is_active = bool(body["is_active"])

    if "use_for_all" in body:
        key_obj.use_for_all = bool(body["use_for_all"])
        if key_obj.use_for_all:
            key_obj.company_code = "__all__"

    # Credentials update (только непустые поля)
    creds_update = _extract_credentials(body, provider)
    if creds_update and any(v.strip() for v in creds_update.values()):
        existing_creds = key_obj.get_credentials()
        for k, v in creds_update.items():
            if v.strip():
                existing_creds[k] = v.strip()
        key_obj.set_credentials(existing_creds)

    key_obj.save()

    # Ре-верификация если credentials были обновлены
    if creds_update and any(v.strip() for v in creds_update.values()):
        creds = key_obj.get_credentials()
        success, error = _verify_key(provider, creds)
        key_obj.mark_verified(success, error)
        logger.info(
            "[API_KEYS] Re-verified after update %s key=%s success=%s",
            provider, pk, success,
        )

    logger.info("[API_KEYS] Updated %s key=%s", provider, pk)
    return Response(_serialize_key(key_obj))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def api_provider_keys_verify(request, provider, pk):
    """
    POST — проверить существующий ключ.
    """
    from docscanner_app.models import APIProviderKey

    if provider not in VALID_PROVIDERS:
        return Response(
            {"detail": f"Nežinomas teikėjas: {provider}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        key_obj = APIProviderKey.objects.get(pk=pk, user=request.user, provider=provider)
    except APIProviderKey.DoesNotExist:
        return Response({"detail": "Nerastas."}, status=status.HTTP_404_NOT_FOUND)

    creds = key_obj.get_credentials()
    success, error = _verify_key(provider, creds)
    key_obj.mark_verified(success, error)

    logger.info("[API_KEYS] Verify %s key=%s success=%s", provider, pk, success)
    return Response(_serialize_key(key_obj))


# ═══════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════

PROVIDER_CRED_FIELDS = {
    "rivile_gama_api": ["api_key"],
    "dineta": ["url", "username", "password"],
    "optimum": ["api_key"],
}


def _extract_credentials(body, provider):
    """Извлекает credential поля из request body."""
    fields = PROVIDER_CRED_FIELDS.get(provider, [])
    creds = {}
    for f in fields:
        val = body.get(f)
        if val is not None:
            creds[f] = str(val)
    return creds


def _validate_credentials(creds, provider):
    """Возвращает список отсутствующих обязательных полей."""
    fields = PROVIDER_CRED_FIELDS.get(provider, [])
    return [f for f in fields if not (creds.get(f) or "").strip()]


def _serialize_key(key_obj):
    """Сериализует APIProviderKey в dict для Response."""
    return {
        "id": key_obj.pk,
        "provider": key_obj.provider,
        "label": key_obj.label,
        "company_code": key_obj.company_code,
        "key_suffix": key_obj.key_suffix,
        "is_active": key_obj.is_active,
        "use_for_all": key_obj.use_for_all,
        "verified_at": key_obj.verified_at,
        "last_ok": key_obj.last_ok,
        "last_error": key_obj.last_error,
        "created_at": key_obj.created_at,
    }

# ═══════════════════════════════════════════════════
# END - Funkcii sviazany s exportom po API
# ═══════════════════════════════════════════════════



# ──────────────────────────────────────────────────────────────
# Individualios veiklos žurnalas
# ──────────────────────────────────────────────────────────────
import io
import logging
from decimal import Decimal

from django.db.models import Count
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .models import ScannedDocument, Invoice


logger = logging.getLogger("docscanner_app")


def get_currency_rate(currency_code, date_obj):
    """Получить курс для валюты на заданную дату (к EUR)."""
    if not currency_code or currency_code.upper() == "EUR":
        return 1.0
    from .models import CurrencyRate
    obj = CurrencyRate.objects.filter(currency=currency_code.upper(), date=date_obj).first()
    if obj:
        return obj.rate
    obj = CurrencyRate.objects.filter(currency=currency_code.upper(), date__lt=date_obj).order_by("-date").first()
    return obj.rate if obj else None


def _company_key(name, vat, cp_id):
    """Тот же ключ что в get_user_counterparties."""
    cp_id = (cp_id or "").strip()
    if cp_id:
        return f"id:{cp_id}"
    vat = (vat or "").strip().lower()
    if vat:
        return vat
    name = (name or "").strip().lower()
    return name


def _dedup_key(series, number, amount):
    """Ключ дедупликации: (series_lower, number_lower, amount_rounded)."""
    s = (series or "").strip().lower()
    n = (number or "").strip().lower()
    try:
        a = str(Decimal(str(amount)).quantize(Decimal('0.01'))) if amount is not None else ''
    except Exception:
        a = ''
    return (s, n, a)


class VeiklosContractorSearchView(APIView):
    """Поиск контрагентов для выбора своей ИВ."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = (request.query_params.get('q') or '').strip().lower()
        if len(search) < 2:
            return Response([])

        merged = {}

        def upsert(cp_id, name, vat, cnt):
            key = _company_key(name, vat, cp_id)
            if not key:
                return
            if key in merged:
                merged[key]['count'] += int(cnt or 0)
                if not merged[key]['code'] and cp_id:
                    merged[key]['code'] = (cp_id or '').strip() or None
                if not merged[key]['vat'] and vat:
                    merged[key]['vat'] = vat or ''
                if not merged[key]['display_name'] and name:
                    merged[key]['display_name'] = name or ''
            else:
                merged[key] = {
                    'key': key,
                    'display_name': name or '',
                    'code': (cp_id or '').strip() or None,
                    'vat': vat or '',
                    'count': int(cnt or 0),
                }

        # ── ScannedDocument ──
        sd_qs = ScannedDocument.objects.filter(
            user=request.user,
            status__in=['completed', 'exported'],
            is_archive_container=False,
            ready_for_export=True,
            math_validation_passed=True,
        )
        for r in sd_qs.exclude(seller_name__isnull=True).exclude(seller_name__exact='') \
                .values('seller_id', 'seller_name', 'seller_vat_code').annotate(cnt=Count('id')):
            upsert(r['seller_id'], r['seller_name'], r['seller_vat_code'], r['cnt'])
        for r in sd_qs.exclude(buyer_name__isnull=True).exclude(buyer_name__exact='') \
                .values('buyer_id', 'buyer_name', 'buyer_vat_code').annotate(cnt=Count('id')):
            upsert(r['buyer_id'], r['buyer_name'], r['buyer_vat_code'], r['cnt'])

        # ── Invoice ──
        inv_qs = Invoice.objects.filter(
            user=request.user,
            status__in=['issued', 'sent', 'partially_paid', 'paid'],
            invoice_type__in=['saskaita', 'pvm_saskaita'],
        )
        for r in inv_qs.exclude(seller_name='') \
                .values('seller_id', 'seller_name', 'seller_vat_code').annotate(cnt=Count('id')):
            upsert(r['seller_id'], r['seller_name'], r['seller_vat_code'], r['cnt'])
        for r in inv_qs.exclude(buyer_name='') \
                .values('buyer_id', 'buyer_name', 'buyer_vat_code').annotate(cnt=Count('id')):
            upsert(r['buyer_id'], r['buyer_name'], r['buyer_vat_code'], r['cnt'])

        # ── Фильтр ──
        items = []
        for item in merged.values():
            if (
                search in (item['display_name'] or '').lower()
                or search in (item['vat'] or '').lower()
                or search in (item['code'] or '').lower()
                or search in (item['key'] or '').lower()
            ):
                items.append(item)

        items.sort(key=lambda x: -x['count'])
        return Response(items[:30])


class VeiklosZurnalasGenerateView(APIView):
    """Генерация журнала с пагинацией."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        contractor_keys = request.data.get('contractor_keys', [])
        pvm_moketojas = request.data.get('pvm_moketojas', False)
        date_from = request.data.get('date_from')
        date_to = request.data.get('date_to')
        sources = request.data.get('sources', ['skaitmenizavimas', 'israsymas'])
        offset = int(request.data.get('offset', 0))
        limit = int(request.data.get('limit', 25))

        if not contractor_keys:
            return Response({'error': 'Nepasirinktas nė vienas kontrahentas'}, status=400)

        entries, summary = self._build_journal(
            request.user, contractor_keys, pvm_moketojas, date_from, date_to, sources,
        )

        total_count = len(entries)
        page = entries[offset:offset + limit]

        return Response({
            'summary': summary,
            'total_count': total_count,
            'offset': offset,
            'entries': page,
        })

    @staticmethod
    def _build_journal(user, contractor_keys, pvm_moketojas, date_from, date_to, sources):
        contractor_keys_set = set(contractor_keys)
        raw_rows = []

        # ── ScannedDocument ──
        if 'skaitmenizavimas' in sources:
            sd_qs = ScannedDocument.objects.filter(
                user=user,
                status__in=['completed', 'exported'],
                is_archive_container=False,
                invoice_date__isnull=False,
                ready_for_export=True,
                math_validation_passed=True,
            )
            if date_from:
                sd_qs = sd_qs.filter(invoice_date__gte=date_from)
            if date_to:
                sd_qs = sd_qs.filter(invoice_date__lte=date_to)

            for doc in sd_qs.iterator():
                seller_key = _company_key(doc.seller_name, doc.seller_vat_code, doc.seller_id)
                buyer_key = _company_key(doc.buyer_name, doc.buyer_vat_code, doc.buyer_id)

                if seller_key in contractor_keys_set:
                    is_pardavimas = True
                elif buyer_key in contractor_keys_set:
                    is_pardavimas = False
                else:
                    continue

                raw_rows.append({
                    'source': 'skaitmenizavimas',
                    'doc_id': doc.id,
                    'invoice_date': doc.invoice_date,
                    'document_series': doc.document_series or '',
                    'document_number': doc.document_number or '',
                    'amount_with_vat': doc.amount_with_vat,
                    'amount_wo_vat': doc.amount_wo_vat,
                    'currency': doc.currency or 'EUR',
                    'is_pardavimas': is_pardavimas,
                    'buyer_name': doc.buyer_name or '',
                    'buyer_id': doc.buyer_id or '',
                    'seller_name': doc.seller_name or '',
                    'seller_id': doc.seller_id or '',
                    'prekes_pavadinimas': doc.prekes_pavadinimas or '',
                })

        # ── Invoice ──
        if 'israsymas' in sources:
            inv_qs = Invoice.objects.filter(
                user=user,
                status__in=['issued', 'sent', 'partially_paid', 'paid'],
                invoice_type__in=['saskaita', 'pvm_saskaita'],
                invoice_date__isnull=False,
            )
            if date_from:
                inv_qs = inv_qs.filter(invoice_date__gte=date_from)
            if date_to:
                inv_qs = inv_qs.filter(invoice_date__lte=date_to)

            for inv in inv_qs.iterator():
                seller_key = _company_key(inv.seller_name, inv.seller_vat_code, inv.seller_id)
                buyer_key = _company_key(inv.buyer_name, inv.buyer_vat_code, inv.buyer_id)

                if seller_key in contractor_keys_set:
                    is_pardavimas = True
                elif buyer_key in contractor_keys_set:
                    is_pardavimas = False
                else:
                    continue

                raw_rows.append({
                    'source': 'israsymas',
                    'doc_id': inv.id,
                    'invoice_date': inv.invoice_date,
                    'document_series': inv.document_series or '',
                    'document_number': inv.document_number or '',
                    'amount_with_vat': inv.amount_with_vat,
                    'amount_wo_vat': inv.amount_wo_vat,
                    'currency': inv.currency or 'EUR',
                    'is_pardavimas': is_pardavimas,
                    'buyer_name': inv.buyer_name or '',
                    'buyer_id': inv.buyer_id or '',
                    'seller_name': inv.seller_name or '',
                    'seller_id': inv.seller_id or '',
                    'prekes_pavadinimas': inv.prekes_pavadinimas or '',
                })

        # ── Дедупликация: israsymas приоритетнее ──
        raw_rows.sort(key=lambda r: (0 if r['source'] == 'israsymas' else 1))
        seen_keys = {}
        deduped = []

        for row in raw_rows:
            dk = _dedup_key(row['document_series'], row['document_number'], row['amount_with_vat'])
            if dk == ('', '', ''):
                deduped.append(row)
                continue
            if dk in seen_keys:
                continue
            seen_keys[dk] = True
            deduped.append(row)

        deduped.sort(key=lambda r: r['invoice_date'], reverse=True)

        # ── Формирование entries ──
        entries = []
        pajamu_suma = Decimal('0')
        islaidu_suma = Decimal('0')
        pardavimo_cnt = 0
        pirkimo_cnt = 0

        for row in deduped:
            is_pardavimas = row['is_pardavimas']
            raw_amount = row['amount_wo_vat'] if pvm_moketojas else row['amount_with_vat']

            if raw_amount is None:
                amount_eur = Decimal('0')
            else:
                currency = (row['currency'] or 'EUR').upper()
                if currency == 'EUR':
                    amount_eur = Decimal(str(raw_amount))
                else:
                    rate = get_currency_rate(currency, row['invoice_date'])
                    if rate:
                        amount_eur = (Decimal(str(raw_amount)) / Decimal(str(rate))).quantize(Decimal('0.01'))
                    else:
                        amount_eur = Decimal(str(raw_amount))

            series = row['document_series']
            number = row['document_number']
            serija_nr = f"{series}-{number}" if series else number

            if is_pardavimas:
                counterparty = row['buyer_name'] or row['buyer_id'] or ''
                pardavimo_cnt += 1
                pajamu_suma += amount_eur
            else:
                counterparty = row['seller_name'] or row['seller_id'] or ''
                pirkimo_cnt += 1
                islaidu_suma += amount_eur

            turinys = row['prekes_pavadinimas']
            if not turinys:
                turinys = 'Pajamos' if is_pardavimas else 'Išlaidos'

            entries.append({
                'doc_id': row['doc_id'],
                'source': row['source'],
                'invoice_date': row['invoice_date'].strftime('%Y-%m-%d') if row['invoice_date'] else '',
                'serija_nr': serija_nr,
                'counterparty': counterparty,
                'turinys': turinys,
                'pajamos': str(amount_eur) if is_pardavimas else None,
                'islaidos': str(amount_eur) if not is_pardavimas else None,
                'currency': row['currency'] or 'EUR',
                'converted': (row['currency'] or 'EUR').upper() != 'EUR',
            })

        summary = {
            'pardavimo_operacijos': pardavimo_cnt,
            'pirkimo_operacijos': pirkimo_cnt,
            'pajamu_suma': str(pajamu_suma),
            'islaidu_suma': str(islaidu_suma),
        }

        return entries, summary


class VeiklosZurnalasExportView(APIView):
    """XLSX экспорт журнала."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        contractor_keys = request.data.get('contractor_keys', [])
        pvm_moketojas = request.data.get('pvm_moketojas', False)
        date_from = request.data.get('date_from')
        date_to = request.data.get('date_to')
        sources = request.data.get('sources', ['skaitmenizavimas', 'israsymas'])

        if not contractor_keys:
            return Response({'error': 'Nepasirinktas nė vienas kontrahentas'}, status=400)

        entries, summary = VeiklosZurnalasGenerateView._build_journal(
            request.user, contractor_keys, pvm_moketojas, date_from, date_to, sources,
        )

        wb = Workbook()
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # ── Sheet 1: Operacijos ──
        ws1 = wb.active
        ws1.title = "Operacijos"

        headers = [
            'Eil. Nr.', 'Sąskaitos data', 'Serija ir numeris',
            'Pirkėjas/pardavėjas', 'Turinys', 'Pajamos, EUR', 'Išlaidos, EUR',
        ]

        for col_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, entry in enumerate(entries, 1):
            r = row_idx + 1
            ws1.cell(row=r, column=1, value=row_idx).border = thin_border
            ws1.cell(row=r, column=2, value=entry['invoice_date']).border = thin_border
            ws1.cell(row=r, column=3, value=entry['serija_nr']).border = thin_border
            ws1.cell(row=r, column=4, value=entry['counterparty']).border = thin_border
            ws1.cell(row=r, column=5, value=entry['turinys']).border = thin_border

            paj_cell = ws1.cell(row=r, column=6)
            paj_cell.border = thin_border
            if entry['pajamos']:
                paj_cell.value = float(entry['pajamos'])
                paj_cell.number_format = '#,##0.00'

            isl_cell = ws1.cell(row=r, column=7)
            isl_cell.border = thin_border
            if entry['islaidos']:
                isl_cell.value = float(entry['islaidos'])
                isl_cell.number_format = '#,##0.00'

        ws1.column_dimensions['A'].width = 10
        ws1.column_dimensions['B'].width = 16
        ws1.column_dimensions['C'].width = 22
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['E'].width = 30
        ws1.column_dimensions['F'].width = 16
        ws1.column_dimensions['G'].width = 16

        # ── Sheet 2: Apžvalga ──
        ws2 = wb.create_sheet("Apžvalga")

        apzvalga_headers = ['Rodiklis', 'Reikšmė']
        for col_idx, h in enumerate(apzvalga_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        summary_rows = [
            ('Pardavimo operacijos', summary['pardavimo_operacijos']),
            ('Pirkimo operacijos', summary['pirkimo_operacijos']),
            ('Pajamos, EUR', float(summary['pajamu_suma'])),
            ('Išlaidos, EUR', float(summary['islaidu_suma'])),
        ]

        for i, (label, val) in enumerate(summary_rows, 2):
            ws2.cell(row=i, column=1, value=label).border = thin_border
            c = ws2.cell(row=i, column=2, value=val)
            c.border = thin_border
            if isinstance(val, float):
                c.number_format = '#,##0.00'

        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 18

        # ── Response ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ind_veiklos_zurnalas.xlsx"'
        return response


# ──────────────────────────────────────────────────────────────
# END - Individualios veiklos žurnalas
# ──────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────
# OSS žurnalas
# ──────────────────────────────────────────────────────────────

import io
import logging
from decimal import Decimal
from collections import defaultdict

from django.db.models import Count, Q
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .models import ScannedDocument, Invoice


logger = logging.getLogger("docscanner_app")


EU_COUNTRIES = {
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR",
    "DE", "GR", "HU", "IE", "IT", "LV", "LT", "LU", "MT", "NL",
    "PL", "PT", "RO", "SK", "SI", "ES", "SE",
}
EU_COUNTRIES_NO_LT = EU_COUNTRIES - {"LT"}

EU_COUNTRY_NAMES_LT = {
    "AT": "Austrija", "BE": "Belgija", "BG": "Bulgarija", "HR": "Kroatija",
    "CY": "Kipras", "CZ": "Čekija", "DK": "Danija", "EE": "Estija",
    "FI": "Suomija", "FR": "Prancūzija", "DE": "Vokietija", "GR": "Graikija",
    "HU": "Vengrija", "IE": "Airija", "IT": "Italija", "LV": "Latvija",
    "LU": "Liuksemburgas", "MT": "Malta", "NL": "Nyderlandai",
    "PL": "Lenkija", "PT": "Portugalija", "RO": "Rumunija", "SK": "Slovakija",
    "SI": "Slovėnija", "ES": "Ispanija", "SE": "Švedija",
}


def _safe_d(v):
    if v is None:
        return Decimal("0")
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _to_eur(amount, currency, invoice_date):
    """Конвертирует сумму в EUR, если валюта не EUR."""
    if not amount:
        return Decimal("0")
    cur = (currency or "EUR").upper()
    if cur == "EUR":
        return _safe_d(amount)
    rate = get_currency_rate(cur, invoice_date)
    if rate:
        return (_safe_d(amount) / Decimal(str(rate))).quantize(Decimal("0.01"))
    return _safe_d(amount)


def _oss_dedup_key(series, number, amount_with_vat, invoice_date):
    """Ключ дедупликации: (series, number, amount, date)."""
    s = (series or "").strip().lower()
    n = (number or "").strip().lower()
    try:
        a = str(Decimal(str(amount_with_vat)).quantize(Decimal("0.01"))) if amount_with_vat is not None else ""
    except Exception:
        a = ""
    d = str(invoice_date) if invoice_date else ""
    return (s, n, a, d)


def _is_buyer_no_vat(vat_code):
    """Покупатель без PVM кода → физлицо / B2C."""
    return not (vat_code or "").strip()


class OSSContractorSearchView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = (request.query_params.get("q") or "").strip().lower()
        if len(search) < 2:
            return Response([])

        merged = {}

        def upsert(cp_id, name, vat, cnt, oss_cnt=0):
            key = _company_key(name, vat, cp_id)
            if not key:
                return
            if key in merged:
                merged[key]["count"] += int(cnt or 0)
                merged[key]["oss_count"] += int(oss_cnt or 0)
                if not merged[key]["code"] and cp_id:
                    merged[key]["code"] = (cp_id or "").strip() or None
                if not merged[key]["vat"] and vat:
                    merged[key]["vat"] = vat or ""
                if not merged[key]["display_name"] and name:
                    merged[key]["display_name"] = name or ""
            else:
                merged[key] = {
                    "key": key,
                    "display_name": name or "",
                    "code": (cp_id or "").strip() or None,
                    "vat": vat or "",
                    "count": int(cnt or 0),
                    "oss_count": int(oss_cnt or 0),
                }

        oss_buyer_q = (
            Q(buyer_country_iso__in=EU_COUNTRIES_NO_LT)
            & (Q(buyer_vat_code__isnull=True) | Q(buyer_vat_code=""))
            & Q(buyer_is_person=True)
        )

        # ── ScannedDocument ──
        sd_qs = ScannedDocument.objects.filter(
            user=request.user,
            status__in=["completed", "exported"],
            is_archive_container=False,
        )
        # Все sellers/buyers — для поиска
        for r in sd_qs.exclude(seller_name__isnull=True).exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            upsert(r["seller_id"], r["seller_name"], r["seller_vat_code"], r["cnt"])
        for r in sd_qs.exclude(buyer_name__isnull=True).exclude(buyer_name="") \
                .values("buyer_id", "buyer_name", "buyer_vat_code").annotate(cnt=Count("id")):
            upsert(r["buyer_id"], r["buyer_name"], r["buyer_vat_code"], r["cnt"])

        # OSS count — только где этот контрагент = seller + buyer подходит
        for r in sd_qs.filter(oss_buyer_q) \
                .exclude(seller_name__isnull=True).exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            key = _company_key(r["seller_name"], r["seller_vat_code"], r["seller_id"])
            if key and key in merged:
                merged[key]["oss_count"] += int(r["cnt"] or 0)

        # ── Invoice ──
        inv_qs = Invoice.objects.filter(
            user=request.user,
            status__in=["issued", "sent", "partially_paid", "paid"],
            invoice_type__in=["saskaita", "pvm_saskaita"],
        )
        for r in inv_qs.exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            upsert(r["seller_id"], r["seller_name"], r["seller_vat_code"], r["cnt"])
        for r in inv_qs.exclude(buyer_name="") \
                .values("buyer_id", "buyer_name", "buyer_vat_code").annotate(cnt=Count("id")):
            upsert(r["buyer_id"], r["buyer_name"], r["buyer_vat_code"], r["cnt"])

        for r in inv_qs.filter(oss_buyer_q) \
                .exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            key = _company_key(r["seller_name"], r["seller_vat_code"], r["seller_id"])
            if key and key in merged:
                merged[key]["oss_count"] += int(r["cnt"] or 0)

        items = [
            item for item in merged.values()
            if search in (item["display_name"] or "").lower()
            or search in (item["vat"] or "").lower()
            or search in (item["code"] or "").lower()
        ]
        items.sort(key=lambda x: (-x["oss_count"], -x["count"]))
        return Response(items[:30])

# ──────────────────────────────────────────────────────────────
# Generate
# ──────────────────────────────────────────────────────────────

class OSSReportGenerateView(APIView):
    """Генерация OSS отчёта с пагинацией."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        contractor_keys = request.data.get("contractor_keys", [])
        date_from = request.data.get("date_from")
        date_to = request.data.get("date_to")
        sources = request.data.get("sources", ["skaitmenizavimas", "israsymas"])
        offset = int(request.data.get("offset", 0))
        limit = int(request.data.get("limit", 25))

        if not contractor_keys:
            return Response({"error": "Nepasirinktas nė vienas kontrahentas"}, status=400)

        summary, entries, grand_totals = self._build_report(
            request.user, contractor_keys, date_from, date_to, sources,
        )

        total_count = len(entries)
        page = entries[offset : offset + limit]

        return Response({
            "summary": summary,
            "grand_totals": grand_totals,
            "total_count": total_count,
            "offset": offset,
            "entries": page,
        })

    @staticmethod
    def _build_report(user, contractor_keys, date_from, date_to, sources):
        contractor_keys_set = set(contractor_keys)
        raw_rows = []

        # ── ScannedDocument ──
        if "skaitmenizavimas" in sources:
            sd_qs = ScannedDocument.objects.filter(
                user=user,
                status__in=["completed", "exported"],
                is_archive_container=False,
                buyer_country_iso__in=EU_COUNTRIES_NO_LT,
                buyer_is_person=True,
                invoice_date__isnull=False,
                ready_for_export=True,
                math_validation_passed=True,
            ).filter(
                Q(buyer_vat_code__isnull=True) | Q(buyer_vat_code="")
            )
            if date_from:
                sd_qs = sd_qs.filter(invoice_date__gte=date_from)
            if date_to:
                sd_qs = sd_qs.filter(invoice_date__lte=date_to)

            for doc in sd_qs.iterator():
                seller_key = _company_key(doc.seller_name, doc.seller_vat_code, doc.seller_id)
                buyer_key = _company_key(doc.buyer_name, doc.buyer_vat_code, doc.buyer_id)

                # Направление определяем по совпадению с выбранным контрагентом
                if seller_key in contractor_keys_set:
                    pass  # pardavimas — наша фирма = seller, подходит для OSS
                elif buyer_key in contractor_keys_set:
                    continue  # pirkimas — наша фирма = buyer, не OSS
                else:
                    continue

                scan_type = getattr(doc, "scan_type", "sumiskai") or "sumiskai"
                has_separate_vat = bool(doc.separate_vat)

                base = {
                    "source": "skaitmenizavimas",
                    "doc_id": doc.id,
                    "invoice_date": doc.invoice_date,
                    "document_series": doc.document_series or "",
                    "document_number": doc.document_number or "",
                    "amount_with_vat": doc.amount_with_vat,
                    "buyer_name": doc.buyer_name or "",
                    "buyer_country_iso": doc.buyer_country_iso or "",
                    "currency": doc.currency or "EUR",
                }

                if has_separate_vat and scan_type == "sumiskai":
                    base.update({
                        "vat_percent": None,
                        "taxable_amount_eur": Decimal("0"),
                        "vat_amount_eur": Decimal("0"),
                        "warning": "Keli skirtingi PVM – suminis režimas, reikia peržiūrėti",
                    })
                    raw_rows.append(base)

                elif has_separate_vat and scan_type == "detaliai":
                    rate_groups = defaultdict(lambda: {"subtotal": Decimal("0"), "vat": Decimal("0")})
                    for li in doc.line_items.all():
                        vp = li.vat_percent
                        if vp is None or _safe_d(vp) <= 0:
                            continue
                        vp_key = _safe_d(vp)
                        rate_groups[vp_key]["subtotal"] += _safe_d(li.subtotal)
                        rate_groups[vp_key]["vat"] += _safe_d(li.vat)

                    if not rate_groups:
                        continue

                    for vp_key, amounts in rate_groups.items():
                        row = dict(base)
                        row.update({
                            "vat_percent": str(vp_key),
                            "taxable_amount_eur": _to_eur(amounts["subtotal"], doc.currency, doc.invoice_date),
                            "vat_amount_eur": _to_eur(amounts["vat"], doc.currency, doc.invoice_date),
                            "warning": None,
                        })
                        raw_rows.append(row)
                else:
                    vp = _safe_d(doc.vat_percent)
                    if vp <= 0:
                        continue
                    base.update({
                        "vat_percent": str(vp),
                        "taxable_amount_eur": _to_eur(doc.amount_wo_vat, doc.currency, doc.invoice_date),
                        "vat_amount_eur": _to_eur(doc.vat_amount, doc.currency, doc.invoice_date),
                        "warning": None,
                    })
                    raw_rows.append(base)

        # ── Invoice ──
        if "israsymas" in sources:
            inv_qs = Invoice.objects.filter(
                user=user,
                status__in=["issued", "sent", "partially_paid", "paid"],
                invoice_type__in=["saskaita", "pvm_saskaita"],
                buyer_country_iso__in=EU_COUNTRIES_NO_LT,
                buyer_is_person=True,
                invoice_date__isnull=False,
            ).filter(
                Q(buyer_vat_code__isnull=True) | Q(buyer_vat_code="")
            )
            if date_from:
                inv_qs = inv_qs.filter(invoice_date__gte=date_from)
            if date_to:
                inv_qs = inv_qs.filter(invoice_date__lte=date_to)

            for inv in inv_qs.iterator():
                seller_key = _company_key(inv.seller_name, inv.seller_vat_code, inv.seller_id)
                buyer_key = _company_key(inv.buyer_name, inv.buyer_vat_code, inv.buyer_id)

                if seller_key in contractor_keys_set:
                    pass  # pardavimas
                elif buyer_key in contractor_keys_set:
                    continue  # pirkimas — не OSS
                else:
                    continue

                has_separate_vat = bool(inv.separate_vat)

                base = {
                    "source": "israsymas",
                    "doc_id": inv.id,
                    "invoice_date": inv.invoice_date,
                    "document_series": inv.document_series or "",
                    "document_number": inv.document_number or "",
                    "amount_with_vat": inv.amount_with_vat,
                    "buyer_name": inv.buyer_name or "",
                    "buyer_country_iso": inv.buyer_country_iso or "",
                    "currency": inv.currency or "EUR",
                }

                if has_separate_vat:
                    rate_groups = defaultdict(lambda: {"subtotal": Decimal("0"), "vat": Decimal("0")})
                    for li in inv.line_items.all():
                        vp = li.vat_percent
                        if vp is None or _safe_d(vp) <= 0:
                            continue
                        vp_key = _safe_d(vp)
                        rate_groups[vp_key]["subtotal"] += _safe_d(li.subtotal)
                        rate_groups[vp_key]["vat"] += _safe_d(li.vat)

                    if not rate_groups:
                        continue

                    for vp_key, amounts in rate_groups.items():
                        row = dict(base)
                        row.update({
                            "vat_percent": str(vp_key),
                            "taxable_amount_eur": _to_eur(amounts["subtotal"], inv.currency, inv.invoice_date),
                            "vat_amount_eur": _to_eur(amounts["vat"], inv.currency, inv.invoice_date),
                            "warning": None,
                        })
                        raw_rows.append(row)
                else:
                    vp = _safe_d(inv.vat_percent)
                    if vp <= 0:
                        continue
                    base.update({
                        "vat_percent": str(vp),
                        "taxable_amount_eur": _to_eur(inv.amount_wo_vat, inv.currency, inv.invoice_date),
                        "vat_amount_eur": _to_eur(inv.vat_amount, inv.currency, inv.invoice_date),
                        "warning": None,
                    })
                    raw_rows.append(base)

        # ── Дедупликация: israsymas приоритетнее ──
        raw_rows.sort(key=lambda r: (0 if r["source"] == "israsymas" else 1))
        seen_keys = {}
        entries = []

        for row in raw_rows:
            dk = _oss_dedup_key(
                row["document_series"], row["document_number"],
                row["amount_with_vat"], row["invoice_date"],
            )
            is_duplicate = False
            if dk != ("", "", "", ""):
                if dk in seen_keys:
                    if row["source"] == "skaitmenizavimas" and seen_keys[dk] == "israsymas":
                        is_duplicate = True
                    elif row["source"] == "israsymas" and seen_keys[dk] == "skaitmenizavimas":
                        for prev in entries:
                            prev_dk = _oss_dedup_key(
                                prev["document_series"], prev["document_number"],
                                prev.get("_amount_with_vat_raw"), prev.get("_invoice_date_raw"),
                            )
                            if prev_dk == dk and prev["source"] == "skaitmenizavimas":
                                prev["is_duplicate"] = True
                        is_duplicate = False
                        seen_keys[dk] = "israsymas"
                    else:
                        is_duplicate = True
                else:
                    seen_keys[dk] = row["source"]

            series = row["document_series"]
            number = row["document_number"]
            serija_nr = f"{series}-{number}" if series else number
            country_iso = row["buyer_country_iso"]

            entries.append({
                "doc_id": row["doc_id"],
                "source": row["source"],
                "invoice_date": row["invoice_date"].strftime("%Y-%m-%d") if row["invoice_date"] else "",
                "serija_nr": serija_nr,
                "buyer_name": row["buyer_name"],
                "buyer_country_iso": country_iso,
                "buyer_country_name": EU_COUNTRY_NAMES_LT.get(country_iso, country_iso),
                "vat_percent": row["vat_percent"],
                "taxable_amount": str(row["taxable_amount_eur"].quantize(Decimal("0.01"))),
                "vat_amount": str(row["vat_amount_eur"].quantize(Decimal("0.01"))),
                "is_duplicate": is_duplicate,
                "warning": row.get("warning"),
                "document_series": row["document_series"],
                "document_number": row["document_number"],
                "_amount_with_vat_raw": row["amount_with_vat"],
                "_invoice_date_raw": row["invoice_date"],
            })

        entries.sort(key=lambda r: r["invoice_date"], reverse=True)

        # ── Сводка по (страна, ставка) — только не-дубликаты и без warnings ──
        summary_map = defaultdict(lambda: {"taxable": Decimal("0"), "vat": Decimal("0"), "doc_count": 0})
        for entry in entries:
            if entry["is_duplicate"] or entry["warning"]:
                continue
            key = (entry["buyer_country_iso"], entry["vat_percent"])
            summary_map[key]["taxable"] += _safe_d(entry["taxable_amount"])
            summary_map[key]["vat"] += _safe_d(entry["vat_amount"])
            summary_map[key]["doc_count"] += 1

        summary = []
        grand_taxable = Decimal("0")
        grand_vat = Decimal("0")
        grand_doc_count = 0

        for (country_iso, vat_pct), amounts in sorted(summary_map.items(), key=lambda x: (x[0][0], x[0][1])):
            t = amounts["taxable"].quantize(Decimal("0.01"))
            v = amounts["vat"].quantize(Decimal("0.01"))
            summary.append({
                "buyer_country_iso": country_iso,
                "buyer_country_name": EU_COUNTRY_NAMES_LT.get(country_iso, country_iso),
                "vat_percent": vat_pct,
                "taxable_amount": str(t),
                "vat_amount": str(v),
                "doc_count": amounts["doc_count"],
            })
            grand_taxable += t
            grand_vat += v
            grand_doc_count += amounts["doc_count"]

        grand_totals = {
            "taxable_amount": str(grand_taxable.quantize(Decimal("0.01"))),
            "vat_amount": str(grand_vat.quantize(Decimal("0.01"))),
            "documents_count": grand_doc_count,
            "warnings_count": sum(1 for e in entries if e["warning"]),
            "duplicates_count": sum(1 for e in entries if e["is_duplicate"]),
        }

        clean_entries = []
        for entry in entries:
            clean_entries.append({
                "doc_id": entry["doc_id"],
                "source": entry["source"],
                "invoice_date": entry["invoice_date"],
                "serija_nr": entry["serija_nr"],
                "buyer_name": entry["buyer_name"],
                "buyer_country_iso": entry["buyer_country_iso"],
                "buyer_country_name": entry["buyer_country_name"],
                "vat_percent": entry["vat_percent"],
                "taxable_amount": entry["taxable_amount"],
                "vat_amount": entry["vat_amount"],
                "is_duplicate": entry["is_duplicate"],
                "warning": entry["warning"],
            })

        return summary, clean_entries, grand_totals

# ──────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────

class OSSReportExportView(APIView):
    """XLSX экспорт OSS отчёта."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        contractor_keys = request.data.get("contractor_keys", [])
        date_from = request.data.get("date_from")
        date_to = request.data.get("date_to")
        sources = request.data.get("sources", ["skaitmenizavimas", "israsymas"])

        if not contractor_keys:
            return Response({"error": "Nepasirinktas nė vienas kontrahentas"}, status=400)

        summary, entries, grand_totals = OSSReportGenerateView._build_report(
            request.user, contractor_keys, date_from, date_to, sources,
        )

        wb = Workbook()
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # ── Sheet 1: OSS suvestinė ──
        ws1 = wb.active
        ws1.title = "OSS suvestinė"

        s_headers = [
            "Vartojimo valstybė narė", "PVM tarifas, %",
            "Apmokestinamoji vertė (EUR)", "PVM suma (EUR)",
        ]
        for col_idx, h in enumerate(s_headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, s_row in enumerate(summary, 2):
            ws1.cell(row=row_idx, column=1, value=s_row["buyer_country_name"]).border = thin_border
            c = ws1.cell(row=row_idx, column=2, value=float(s_row["vat_percent"]))
            c.border = thin_border
            c.number_format = "0.00"
            c = ws1.cell(row=row_idx, column=3, value=float(s_row["taxable_amount"]))
            c.border = thin_border
            c.number_format = "#,##0.00"
            c = ws1.cell(row=row_idx, column=4, value=float(s_row["vat_amount"]))
            c.border = thin_border
            c.number_format = "#,##0.00"

        # Итоговая строка
        total_row = len(summary) + 2
        c = ws1.cell(row=total_row, column=1, value="Viso:")
        c.font = header_font
        c.fill = total_fill
        c.border = thin_border
        ws1.cell(row=total_row, column=2, value="").border = thin_border
        c = ws1.cell(row=total_row, column=3, value=float(grand_totals["taxable_amount"]))
        c.font = header_font
        c.fill = total_fill
        c.border = thin_border
        c.number_format = "#,##0.00"
        c = ws1.cell(row=total_row, column=4, value=float(grand_totals["vat_amount"]))
        c.font = header_font
        c.fill = total_fill
        c.border = thin_border
        c.number_format = "#,##0.00"

        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 16
        ws1.column_dimensions["C"].width = 26
        ws1.column_dimensions["D"].width = 20

        # ── Sheet 2: Dokumentai (без дубликатов) ──
        ws2 = wb.create_sheet("Dokumentai")

        d_headers = [
            "Eil. Nr.", "Sąskaitos data", "Serija ir numeris",
            "Pirkėjas", "Šalis", "PVM %",
            "Apmokestinamoji vertė (EUR)", "PVM suma (EUR)", "Šaltinis",
        ]
        for col_idx, h in enumerate(d_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num = 0
        for entry in entries:
            if entry["is_duplicate"]:
                continue
            row_num += 1
            r = row_num + 1
            ws2.cell(row=r, column=1, value=row_num).border = thin_border
            ws2.cell(row=r, column=2, value=entry["invoice_date"]).border = thin_border
            ws2.cell(row=r, column=3, value=entry["serija_nr"]).border = thin_border
            ws2.cell(row=r, column=4, value=entry["buyer_name"]).border = thin_border
            ws2.cell(row=r, column=5, value=entry["buyer_country_name"]).border = thin_border

            c = ws2.cell(row=r, column=6)
            c.border = thin_border
            if entry["vat_percent"]:
                c.value = float(entry["vat_percent"])
                c.number_format = "0.00"

            c = ws2.cell(row=r, column=7, value=float(entry["taxable_amount"]))
            c.border = thin_border
            c.number_format = "#,##0.00"
            c = ws2.cell(row=r, column=8, value=float(entry["vat_amount"]))
            c.border = thin_border
            c.number_format = "#,##0.00"

            src_label = "Išrašymas" if entry["source"] == "israsymas" else "Skaitmenizavimas"
            if entry["warning"]:
                src_label += f" ⚠ {entry['warning']}"
            ws2.cell(row=r, column=9, value=src_label).border = thin_border

        ws2.column_dimensions["A"].width = 10
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 22
        ws2.column_dimensions["D"].width = 30
        ws2.column_dimensions["E"].width = 18
        ws2.column_dimensions["F"].width = 10
        ws2.column_dimensions["G"].width = 26
        ws2.column_dimensions["H"].width = 20
        ws2.column_dimensions["I"].width = 20

        # ── Response ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="oss_ataskaita.xlsx"'
        return response

# ──────────────────────────────────────────────────────────────
# END - OSS žurnalas
# ──────────────────────────────────────────────────────────────

# ────────────────────────────────────────────────────────────
# ─── SVS
# ────────────────────────────────────────────────────────────

import io
import logging
from decimal import Decimal
from collections import defaultdict

from django.db.models import Count, Q
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill



# ── SVS traded_type filter (vključaja NULL dlia obratnoj sovmestimosti) ──
_SVS_TRADED_Q = (
    Q(traded_type="services")
    | Q(traded_type__isnull=True)
    | Q(traded_type="")
)

CODE_LABELS = {
    "140": "Iš kitų ES valstybių narių įsigytų paslaugų apmokestinamoji vertė",
    "141": "Iš ne ES šalių įsigytų paslaugų apmokestinamoji vertė",
    "043": "Paslaugų, kurių teikimo vieta yra kita valstybė narė, apmokestinamoji vertė",
}


# ──────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────

def _svs_dedup_key(series, number, amount, invoice_date):
    s = (series or "").strip().lower()
    n = (number or "").strip().lower()
    try:
        a = str(Decimal(str(amount)).quantize(Decimal("0.01"))) if amount is not None else ""
    except Exception:
        a = ""
    d = str(invoice_date) if invoice_date else ""
    return (s, n, a, d)


def _strip_vat_prefix(vat_code, country_iso):
    """Nuima šalies prefiksą iš PVM kodo (FR0564 reikalauja be prefikso)."""
    if not vat_code:
        return ""
    vat = vat_code.strip()
    prefix = (country_iso or "").strip().upper()
    if prefix and vat.upper().startswith(prefix):
        return vat[len(prefix):]
    return vat


def _classify_svs_doc(
    source, doc_id,
    seller_key, buyer_key,
    seller_country_iso, buyer_country_iso,
    buyer_vat_code, buyer_name, seller_name,
    vat_percent, vat_amount,
    amount_wo_vat, amount_with_vat,
    traded_type,
    invoice_date, document_series, document_number,
    currency,
    contractor_keys_set,
):
    """
    Klasifikuoja dokumentą kaip SVS pirkimą (140/141) arba pardavimą (043).
    Grąžina dict arba None jei dokumentas netinka SVS.
    """
    seller_iso = (seller_country_iso or "").strip().upper()
    buyer_iso = (buyer_country_iso or "").strip().upper()
    b_vat = (buyer_vat_code or "").strip()
    t_type = (traded_type or "").strip().lower()

    # traded_type filtras: tik paslaugos/mišrus (arba tuščias — atgalinis suderinamumas)
    is_services = t_type in ("services")

    # ── Pirkimas: mano įmonė = buyer, seller užsienietis, PVM ≈ 0% ──
    if buyer_key in contractor_keys_set and seller_key not in contractor_keys_set:
        if not seller_iso or seller_iso == "LT":
            return None
        if not is_services:
            return None

        # Tikriname ar pardavėjas nepritaikė LT PVM (pvz. per OSS)
        vp = _safe_d(vat_percent)
        va = _safe_d(vat_amount)
        if vp > 0 and va > 0:
            return None

        svs_code = "140" if seller_iso in EU_COUNTRIES_NO_LT else "141"

        # Apmokestinamoji vertė: amount_wo_vat, fallback į amount_with_vat (nes PVM=0)
        taxable = _to_eur(amount_wo_vat, currency, invoice_date)
        if taxable == Decimal("0") and amount_with_vat:
            taxable = _to_eur(amount_with_vat, currency, invoice_date)

        return {
            "source": source,
            "doc_id": doc_id,
            "svs_code": svs_code,
            "invoice_date": invoice_date,
            "document_series": document_series or "",
            "document_number": document_number or "",
            "counterparty_name": seller_name or "",
            "counterparty_country_iso": seller_iso,
            "counterparty_vat_code": "",
            "amount_wo_vat_eur": taxable,
            "vat_amount_eur": Decimal("0"),
            "amount_with_vat": amount_with_vat,
        }

    # ── Pardavimas: mano įmonė = seller, buyer yra ES PVM mokėtojas ──
    if seller_key in contractor_keys_set and buyer_key not in contractor_keys_set:
        if buyer_iso not in EU_COUNTRIES_NO_LT:
            return None
        if not b_vat:
            return None
        if not is_services:
            return None

        taxable = _to_eur(amount_wo_vat, currency, invoice_date)
        if taxable == Decimal("0") and amount_with_vat:
            taxable = _to_eur(amount_with_vat, currency, invoice_date)

        return {
            "source": source,
            "doc_id": doc_id,
            "svs_code": "043",
            "invoice_date": invoice_date,
            "document_series": document_series or "",
            "document_number": document_number or "",
            "counterparty_name": buyer_name or "",
            "counterparty_country_iso": buyer_iso,
            "counterparty_vat_code": b_vat,
            "amount_wo_vat_eur": taxable,
            "vat_amount_eur": Decimal("0"),
            "amount_with_vat": amount_with_vat,
        }

    return None


# ──────────────────────────────────────────────────────────────
# Contractor search
# ──────────────────────────────────────────────────────────────

class SVSContractorSearchView(APIView):
    """Kontrahentų paieška SVS ataskaitai — vartotojas renkasi savo įmonės variacijas."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = (request.query_params.get("q") or "").strip().lower()
        if len(search) < 2:
            return Response([])

        merged = {}

        def upsert(cp_id, name, vat, cnt, svs_cnt=0):
            key = _company_key(name, vat, cp_id)
            if not key:
                return
            if key in merged:
                merged[key]["count"] += int(cnt or 0)
                merged[key]["svs_count"] += int(svs_cnt or 0)
                if not merged[key]["code"] and cp_id:
                    merged[key]["code"] = (cp_id or "").strip() or None
                if not merged[key]["vat"] and vat:
                    merged[key]["vat"] = vat or ""
                if not merged[key]["display_name"] and name:
                    merged[key]["display_name"] = name or ""
            else:
                merged[key] = {
                    "key": key,
                    "display_name": name or "",
                    "code": (cp_id or "").strip() or None,
                    "vat": vat or "",
                    "count": int(cnt or 0),
                    "svs_count": int(svs_cnt or 0),
                }

        # ── SVS pirkimų filtras: pardavėjas užsienietis, PVM ≈ 0, paslaugos ──
        svs_purchase_q = (
            ~Q(seller_country_iso="LT")
            & Q(seller_country_iso__isnull=False)
            & ~Q(seller_country_iso="")
            & (Q(vat_percent__isnull=True) | Q(vat_percent=0) | Q(vat_amount=0) | Q(vat_amount__isnull=True))
            & _SVS_TRADED_Q
        )

        # ── SVS pardavimų filtras: pirkėjas yra ES PVM mokėtojas ──
        svs_sale_q = (
            Q(buyer_country_iso__in=EU_COUNTRIES_NO_LT)
            & ~Q(buyer_vat_code__isnull=True)
            & ~Q(buyer_vat_code="")
            & _SVS_TRADED_Q
        )

        # ── ScannedDocument ──
        sd_qs = ScannedDocument.objects.filter(
            user=request.user,
            status__in=["completed", "exported"],
            is_archive_container=False,
        )

        for r in sd_qs.exclude(seller_name__isnull=True).exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            upsert(r["seller_id"], r["seller_name"], r["seller_vat_code"], r["cnt"])
        for r in sd_qs.exclude(buyer_name__isnull=True).exclude(buyer_name="") \
                .values("buyer_id", "buyer_name", "buyer_vat_code").annotate(cnt=Count("id")):
            upsert(r["buyer_id"], r["buyer_name"], r["buyer_vat_code"], r["cnt"])

        # SVS count — pirkimai (kontrahentas = buyer)
        for r in sd_qs.filter(svs_purchase_q) \
                .exclude(buyer_name__isnull=True).exclude(buyer_name="") \
                .values("buyer_id", "buyer_name", "buyer_vat_code").annotate(cnt=Count("id")):
            key = _company_key(r["buyer_name"], r["buyer_vat_code"], r["buyer_id"])
            if key and key in merged:
                merged[key]["svs_count"] += int(r["cnt"] or 0)

        # SVS count — pardavimai (kontrahentas = seller)
        for r in sd_qs.filter(svs_sale_q) \
                .exclude(seller_name__isnull=True).exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            key = _company_key(r["seller_name"], r["seller_vat_code"], r["seller_id"])
            if key and key in merged:
                merged[key]["svs_count"] += int(r["cnt"] or 0)

        # ── Invoice ──
        inv_qs = Invoice.objects.filter(
            user=request.user,
            status__in=["issued", "sent", "partially_paid", "paid"],
            invoice_type__in=["saskaita", "pvm_saskaita"],
        )
        for r in inv_qs.exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            upsert(r["seller_id"], r["seller_name"], r["seller_vat_code"], r["cnt"])
        for r in inv_qs.exclude(buyer_name="") \
                .values("buyer_id", "buyer_name", "buyer_vat_code").annotate(cnt=Count("id")):
            upsert(r["buyer_id"], r["buyer_name"], r["buyer_vat_code"], r["cnt"])

        # SVS count iš Invoice — pardavimai (kontrahentas = seller, buyer ES PVM mokėtojas)
        inv_svs_sale_q = (
            Q(buyer_country_iso__in=EU_COUNTRIES_NO_LT)
            & ~Q(buyer_vat_code__isnull=True)
            & ~Q(buyer_vat_code="")
        )
        for r in inv_qs.filter(inv_svs_sale_q).exclude(seller_name="") \
                .values("seller_id", "seller_name", "seller_vat_code").annotate(cnt=Count("id")):
            key = _company_key(r["seller_name"], r["seller_vat_code"], r["seller_id"])
            if key and key in merged:
                merged[key]["svs_count"] += int(r["cnt"] or 0)

        items = [
            item for item in merged.values()
            if search in (item["display_name"] or "").lower()
            or search in (item["vat"] or "").lower()
            or search in (item["code"] or "").lower()
        ]
        items.sort(key=lambda x: (-x["svs_count"], -x["count"]))
        return Response(items[:30])


# ──────────────────────────────────────────────────────────────
# Generate
# ──────────────────────────────────────────────────────────────

class SVSReportGenerateView(APIView):
    """SVS ataskaitos generavimas (PVM101 + FR0564) su puslapiavimu."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        contractor_keys = request.data.get("contractor_keys", [])
        date_from = request.data.get("date_from")
        date_to = request.data.get("date_to")
        sources = request.data.get("sources", ["skaitmenizavimas", "israsymas"])
        offset = int(request.data.get("offset", 0))
        limit = int(request.data.get("limit", 25))

        if not contractor_keys:
            return Response({"error": "Nepasirinktas nė vienas kontrahentas"}, status=400)

        result = self._build_report(
            request.user, contractor_keys, date_from, date_to, sources,
        )

        total_count = len(result["entries"])
        page = result["entries"][offset:offset + limit]

        return Response({
            "pvm101_summary": result["pvm101_summary"],
            "fr0564_summary": result["fr0564_summary"],
            "grand_totals": result["grand_totals"],
            "total_count": total_count,
            "offset": offset,
            "entries": page,
        })

    @staticmethod
    def _build_report(user, contractor_keys, date_from, date_to, sources):
        contractor_keys_set = set(contractor_keys)
        raw_rows = []

        def _iter_docs(source_label, qs):
            for obj in qs.iterator():
                seller_key = _company_key(obj.seller_name, obj.seller_vat_code, obj.seller_id)
                buyer_key = _company_key(obj.buyer_name, obj.buyer_vat_code, obj.buyer_id)
                row = _classify_svs_doc(
                    source=source_label,
                    doc_id=obj.id,
                    seller_key=seller_key,
                    buyer_key=buyer_key,
                    seller_country_iso=obj.seller_country_iso,
                    buyer_country_iso=obj.buyer_country_iso,
                    buyer_vat_code=obj.buyer_vat_code,
                    buyer_name=obj.buyer_name,
                    seller_name=obj.seller_name,
                    vat_percent=obj.vat_percent,
                    vat_amount=obj.vat_amount,
                    amount_wo_vat=obj.amount_wo_vat,
                    amount_with_vat=obj.amount_with_vat,
                    traded_type=getattr(obj, "traded_type", None),
                    invoice_date=obj.invoice_date,
                    document_series=obj.document_series,
                    document_number=obj.document_number,
                    currency=obj.currency,
                    contractor_keys_set=contractor_keys_set,
                )
                if row:
                    raw_rows.append(row)

        # ── ScannedDocument ──
        if "skaitmenizavimas" in sources:
            sd_qs = ScannedDocument.objects.filter(
                user=user,
                status__in=["completed", "exported"],
                is_archive_container=False,
                invoice_date__isnull=False,
            )
            if date_from:
                sd_qs = sd_qs.filter(invoice_date__gte=date_from)
            if date_to:
                sd_qs = sd_qs.filter(invoice_date__lte=date_to)
            _iter_docs("skaitmenizavimas", sd_qs)

        # ── Invoice ──
        if "israsymas" in sources:
            inv_qs = Invoice.objects.filter(
                user=user,
                status__in=["issued", "sent", "partially_paid", "paid"],
                invoice_type__in=["saskaita", "pvm_saskaita"],
                invoice_date__isnull=False,
            )
            if date_from:
                inv_qs = inv_qs.filter(invoice_date__gte=date_from)
            if date_to:
                inv_qs = inv_qs.filter(invoice_date__lte=date_to)
            _iter_docs("israsymas", inv_qs)

        # ── Deduplikacija: israsymas prioritetas ──
        raw_rows.sort(key=lambda r: (0 if r["source"] == "israsymas" else 1))
        seen_keys = {}
        entries = []

        for row in raw_rows:
            dk = _svs_dedup_key(
                row["document_series"], row["document_number"],
                row["amount_wo_vat_eur"], row["invoice_date"],
            )
            is_duplicate = False
            if dk != ("", "", "", ""):
                if dk in seen_keys:
                    if row["source"] == "skaitmenizavimas" and seen_keys[dk] == "israsymas":
                        is_duplicate = True
                    elif row["source"] == "israsymas" and seen_keys[dk] == "skaitmenizavimas":
                        for prev in entries:
                            prev_dk = _svs_dedup_key(
                                prev["document_series"], prev["document_number"],
                                prev.get("_amount_raw"), prev.get("_date_raw"),
                            )
                            if prev_dk == dk and prev["source"] == "skaitmenizavimas":
                                prev["is_duplicate"] = True
                        is_duplicate = False
                        seen_keys[dk] = "israsymas"
                    else:
                        is_duplicate = True
                else:
                    seen_keys[dk] = row["source"]

            series = row["document_series"]
            number = row["document_number"]
            serija_nr = f"{series}-{number}" if series else number
            c_iso = row["counterparty_country_iso"]

            entries.append({
                "doc_id": row["doc_id"],
                "source": row["source"],
                "svs_code": row["svs_code"],
                "invoice_date": row["invoice_date"].strftime("%Y-%m-%d") if row["invoice_date"] else "",
                "serija_nr": serija_nr,
                "counterparty_name": row["counterparty_name"],
                "counterparty_country_iso": c_iso,
                "counterparty_country_name": EU_COUNTRY_NAMES_LT.get(c_iso, c_iso or ""),
                "counterparty_vat_code": row.get("counterparty_vat_code", ""),
                "amount_wo_vat": str(row["amount_wo_vat_eur"].quantize(Decimal("0.01"))),
                "is_duplicate": is_duplicate,
                # Internal fields dlia dedup
                "document_series": row["document_series"],
                "document_number": row["document_number"],
                "_amount_raw": row["amount_wo_vat_eur"],
                "_date_raw": row["invoice_date"],
            })

        entries.sort(key=lambda r: r["invoice_date"], reverse=True)

        # ── PVM101 suvestinė: pagal kodą ──
        pvm101_map = defaultdict(lambda: {"taxable": Decimal("0"), "doc_count": 0})
        for entry in entries:
            if entry["is_duplicate"]:
                continue
            code = entry["svs_code"]
            pvm101_map[code]["taxable"] += _safe_d(entry["amount_wo_vat"])
            pvm101_map[code]["doc_count"] += 1

        pvm101_summary = []
        grand_taxable = Decimal("0")
        grand_vat = Decimal("0")
        grand_doc_count = 0

        for code in sorted(pvm101_map.keys()):
            amounts = pvm101_map[code]
            taxable = amounts["taxable"].quantize(Decimal("0.01"))
            if code in ("140", "141"):
                vat = (taxable * Decimal("0.21")).quantize(Decimal("0.01"))
                vat_pct = "21"
            else:
                vat = Decimal("0")
                vat_pct = "0"

            pvm101_summary.append({
                "code": code,
                "label": CODE_LABELS.get(code, code),
                "vat_percent": vat_pct,
                "taxable_amount": str(taxable),
                "vat_amount": str(vat),
                "doc_count": amounts["doc_count"],
            })
            grand_taxable += taxable
            grand_vat += vat
            grand_doc_count += amounts["doc_count"]

        # ── FR0564 suvestinė: pagal kiekvieną pirkėją (tik kodas 043) ──
        fr0564_map = defaultdict(lambda: {
            "country_iso": "", "vat_code_raw": "",
            "buyer_name": "", "amount": Decimal("0"), "doc_count": 0,
        })
        for entry in entries:
            if entry["is_duplicate"] or entry["svs_code"] != "043":
                continue
            vat_raw = entry.get("counterparty_vat_code", "")
            c_iso = entry["counterparty_country_iso"]
            fr_key = (c_iso, vat_raw)
            fr0564_map[fr_key]["country_iso"] = c_iso
            fr0564_map[fr_key]["vat_code_raw"] = vat_raw
            fr0564_map[fr_key]["buyer_name"] = entry["counterparty_name"]
            fr0564_map[fr_key]["amount"] += _safe_d(entry["amount_wo_vat"])
            fr0564_map[fr_key]["doc_count"] += 1

        fr0564_summary = []
        for (c_iso, vat_raw), data in sorted(fr0564_map.items()):
            # FR0564 sumos apvalinamos iki sveikų skaičių (taisyklių 11 p.)
            amount_rounded = data["amount"].quantize(Decimal("1"))
            fr0564_summary.append({
                "country_iso": c_iso,
                "country_name": EU_COUNTRY_NAMES_LT.get(c_iso, c_iso),
                "vat_code": _strip_vat_prefix(vat_raw, c_iso),
                "vat_code_full": vat_raw,
                "buyer_name": data["buyer_name"],
                "services_amount": str(amount_rounded),
                "doc_count": data["doc_count"],
            })

        grand_totals = {
            "taxable_amount": str(grand_taxable.quantize(Decimal("0.01"))),
            "vat_amount": str(grand_vat.quantize(Decimal("0.01"))),
            "documents_count": grand_doc_count,
            "duplicates_count": sum(1 for e in entries if e["is_duplicate"]),
        }

        # Valome entries prieš grąžinant
        clean_entries = []
        for entry in entries:
            clean_entries.append({
                "doc_id": entry["doc_id"],
                "source": entry["source"],
                "svs_code": entry["svs_code"],
                "invoice_date": entry["invoice_date"],
                "serija_nr": entry["serija_nr"],
                "counterparty_name": entry["counterparty_name"],
                "counterparty_country_iso": entry["counterparty_country_iso"],
                "counterparty_country_name": entry["counterparty_country_name"],
                "counterparty_vat_code": entry.get("counterparty_vat_code", ""),
                "amount_wo_vat": entry["amount_wo_vat"],
                "is_duplicate": entry["is_duplicate"],
            })

        return {
            "pvm101_summary": pvm101_summary,
            "fr0564_summary": fr0564_summary,
            "grand_totals": grand_totals,
            "entries": clean_entries,
        }


# ──────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────

class SVSReportExportView(APIView):
    """SVS ataskaitos XLSX eksportas."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        contractor_keys = request.data.get("contractor_keys", [])
        date_from = request.data.get("date_from")
        date_to = request.data.get("date_to")
        sources = request.data.get("sources", ["skaitmenizavimas", "israsymas"])

        if not contractor_keys:
            return Response({"error": "Nepasirinktas nė vienas kontrahentas"}, status=400)

        result = SVSReportGenerateView._build_report(
            request.user, contractor_keys, date_from, date_to, sources,
        )

        wb = Workbook()
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def _hdr(ws, row, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.border = thin
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Sheet 1: PVM101 suvestinė ──
        ws1 = wb.active
        ws1.title = "PVM101"
        _hdr(ws1, 1, [
            "Kodas", "Aprašymas", "PVM tarifas, %",
            "Apmokestinamoji vertė (EUR)", "PVM suma (EUR)", "Dok.",
        ])

        for i, row in enumerate(result["pvm101_summary"], 2):
            ws1.cell(row=i, column=1, value=row["code"]).border = thin
            ws1.cell(row=i, column=2, value=row["label"]).border = thin
            c = ws1.cell(row=i, column=3, value=int(row["vat_percent"]))
            c.border = thin
            c = ws1.cell(row=i, column=4, value=float(row["taxable_amount"]))
            c.border = thin
            c.number_format = "#,##0.00"
            c = ws1.cell(row=i, column=5, value=float(row["vat_amount"]))
            c.border = thin
            c.number_format = "#,##0.00"
            ws1.cell(row=i, column=6, value=row["doc_count"]).border = thin

        tr = len(result["pvm101_summary"]) + 2
        c = ws1.cell(row=tr, column=1, value="Viso:")
        c.font = header_font
        c.fill = total_fill
        c.border = thin
        for col in (2, 3):
            c = ws1.cell(row=tr, column=col, value="")
            c.fill = total_fill
            c.border = thin
        c = ws1.cell(row=tr, column=4, value=float(result["grand_totals"]["taxable_amount"]))
        c.font = header_font
        c.fill = total_fill
        c.border = thin
        c.number_format = "#,##0.00"
        c = ws1.cell(row=tr, column=5, value=float(result["grand_totals"]["vat_amount"]))
        c.font = header_font
        c.fill = total_fill
        c.border = thin
        c.number_format = "#,##0.00"
        c = ws1.cell(row=tr, column=6, value=result["grand_totals"]["documents_count"])
        c.font = header_font
        c.fill = total_fill
        c.border = thin

        ws1.column_dimensions["A"].width = 10
        ws1.column_dimensions["B"].width = 55
        ws1.column_dimensions["C"].width = 14
        ws1.column_dimensions["D"].width = 26
        ws1.column_dimensions["E"].width = 18
        ws1.column_dimensions["F"].width = 8

        # ── Sheet 2: FR0564 (tik jei yra 043 pardavimų) ──
        if result["fr0564_summary"]:
            ws2 = wb.create_sheet("FR0564")
            _hdr(ws2, 1, [
                "Valstybės kodas", "PVM mokėtojo kodas (be prefikso)",
                "Pirkėjas", "Paslaugų vertė (EUR)",
            ])

            for i, row in enumerate(result["fr0564_summary"], 2):
                ws2.cell(row=i, column=1, value=row["country_iso"]).border = thin
                ws2.cell(row=i, column=2, value=row["vat_code"]).border = thin
                ws2.cell(row=i, column=3, value=row["buyer_name"]).border = thin
                c = ws2.cell(row=i, column=4, value=int(row["services_amount"]))
                c.border = thin
                c.number_format = "#,##0"

            ws2.column_dimensions["A"].width = 18
            ws2.column_dimensions["B"].width = 30
            ws2.column_dimensions["C"].width = 30
            ws2.column_dimensions["D"].width = 22

        # ── Sheet 3: Dokumentai ──
        ws3 = wb.create_sheet("Dokumentai")
        _hdr(ws3, 1, [
            "Eil. Nr.", "Kodas", "Data", "Serija ir numeris",
            "Kontrahentas", "Šalis", "PVM kodas",
            "Apmokestinamoji vertė (EUR)", "Šaltinis",
        ])

        row_num = 0
        for entry in result["entries"]:
            if entry["is_duplicate"]:
                continue
            row_num += 1
            r = row_num + 1
            ws3.cell(row=r, column=1, value=row_num).border = thin
            ws3.cell(row=r, column=2, value=entry["svs_code"]).border = thin
            ws3.cell(row=r, column=3, value=entry["invoice_date"]).border = thin
            ws3.cell(row=r, column=4, value=entry["serija_nr"]).border = thin
            ws3.cell(row=r, column=5, value=entry["counterparty_name"]).border = thin
            ws3.cell(row=r, column=6, value=entry.get("counterparty_country_name", "")).border = thin
            ws3.cell(row=r, column=7, value=entry.get("counterparty_vat_code", "")).border = thin
            c = ws3.cell(row=r, column=8, value=float(entry["amount_wo_vat"]))
            c.border = thin
            c.number_format = "#,##0.00"
            src = "Išrašymas" if entry["source"] == "israsymas" else "Skaitmenizavimas"
            ws3.cell(row=r, column=9, value=src).border = thin

        ws3.column_dimensions["A"].width = 8
        ws3.column_dimensions["B"].width = 8
        ws3.column_dimensions["C"].width = 14
        ws3.column_dimensions["D"].width = 20
        ws3.column_dimensions["E"].width = 30
        ws3.column_dimensions["F"].width = 16
        ws3.column_dimensions["G"].width = 22
        ws3.column_dimensions["H"].width = 26
        ws3.column_dimensions["I"].width = 18

        # ── Response ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="svs_ataskaita.xlsx"'
        return response

# ────────────────────────────────────────────────────────────
# END ─── SVS
# ────────────────────────────────────────────────────────────

# ─── Newsletter Campaign ───────────────────────────────────
class NewsletterSendView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request):
        """Создать кампанию и запустить рассылку."""
        from .serializers import NewsletterCampaignCreateSerializer
        from .models import NewsletterCampaign, NewsletterRecipient
        from .tasks import send_newsletter_task

        ser = NewsletterCampaignCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        d = ser.validated_data

        # Построить список получателей
        User = get_user_model()
        qs = User.objects.filter(is_active=True)

        exclude_ids = d.get("exclude_user_ids", [])
        if exclude_ids:
            qs = qs.exclude(id__in=exclude_ids)

        sources = d.get("registration_sources", [])
        if sources:
            source_q = Q()
            for src in sources:
                if src == "null":
                    source_q |= Q(registration_source__isnull=True) | Q(registration_source="")
                else:
                    source_q |= Q(registration_source=src)
            qs = qs.filter(source_q)

        users = list(qs.order_by("id").values_list("id", "email", named=True))

        if not users:
            return Response({"detail": "Nėra gavėjų pagal pasirinktus filtrus."}, status=400)

        # Создать кампанию
        campaign = NewsletterCampaign.objects.create(
            subject=d["subject"],
            body=d["body"],
            sender=request.user,
            exclude_user_ids=exclude_ids,
            registration_sources=sources,
            batch_size=d.get("batch_size") or 190,
            total_recipients=len(users),
        )

        # Создать получателей
        recipients = [
            NewsletterRecipient(campaign=campaign, user_id=u.id, email=u.email)
            for u in users
        ]
        NewsletterRecipient.objects.bulk_create(recipients, batch_size=500)

        # Запустить таск
        result = send_newsletter_task.delay(campaign_id=campaign.id)
        campaign.celery_task_id = result.id
        campaign.save(update_fields=["celery_task_id"])

        return Response({
            "campaign_id": campaign.id,
            "task_id": result.id,
            "total_recipients": len(users),
            "status": "sending",
        })

    def get(self, request):
        """Список кампаний или превью количества получателей."""
        from .models import NewsletterCampaign
        from .serializers import NewsletterCampaignSerializer

        # Если есть параметр preview — возвращаем count (для совместимости)
        if "sources" in request.query_params or "exclude_ids" in request.query_params:
            sources = request.query_params.getlist("sources")
            exclude_raw = request.query_params.get("exclude_ids", "")
            exclude_ids = [
                int(x.strip()) for x in exclude_raw.split(",")
                if x.strip().isdigit()
            ]
            User = get_user_model()
            qs = User.objects.filter(is_active=True)
            if exclude_ids:
                qs = qs.exclude(id__in=exclude_ids)
            if sources:
                source_q = Q()
                for src in sources:
                    if src == "null":
                        source_q |= Q(registration_source__isnull=True) | Q(registration_source="")
                    else:
                        source_q |= Q(registration_source=src)
                qs = qs.filter(source_q)
            return Response({"recipient_count": qs.count()})

        # Иначе — список кампаний
        campaigns = NewsletterCampaign.objects.all()[:20]
        serializer = NewsletterCampaignSerializer(campaigns, many=True)
        return Response({"campaigns": serializer.data})

    def put(self, request):
        """Тестовая отправка на фиксированный email."""
        from .serializers import NewsletterCampaignCreateSerializer

        ser = NewsletterCampaignCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        d = ser.validated_data
        try:
            msg = EmailMultiAlternatives(
                subject=f"[TEST] {d['subject'].strip()}",
                body=d["body"],
                from_email=formataddr(("Denis iš DokSkeno", settings.DEFAULT_FROM_EMAIL)),
                to=["orlov.projects@gmail.com"],
            )
            try:
                msg.tags = ["newsletter_test"]
                msg.metadata = {"event": "newsletter_test", "sender_user_id": request.user.id}
            except Exception:
                pass
            msg.send()
            return Response({"status": "sent", "to": "orlov.projects@gmail.com"})
        except Exception as e:
            logger.exception(f"[NEWSLETTER TEST ERROR] {e}")
            return Response({"detail": str(e)}, status=500)
# END ─── Newsletter Campaign ──────────────────────────────

# ────────────────────────────────────────────────────────────
# ─── Waybill scan ───
# ────────────────────────────────────────────────────────────

"""
waybill_views.py — API views для важтарашчей.
"""
import logging
import os
import uuid
from decimal import Decimal

from django.conf import settings
from django.db import transaction
from django.db.models import F, Q
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import (
    ScannedWaybill,
    WaybillUploadSession,
    WaybillChunkedUpload,
    CustomUser,
    WAYBILL_CREDIT_COST,
)
from .serializers import (
    WaybillSessionCreateSerializer,
    WaybillSessionStatusSerializer,
    WaybillSessionFinalizeSerializer,
    WaybillChunkedUploadInitSerializer,
    WaybillChunkedUploadStatusSerializer,
    ScannedWaybillListSerializer,
    ScannedWaybillDetailSerializer,
    ScannedWaybillUpdateSerializer,
)

logger = logging.getLogger("docscanner_app")


# ============================================================
# Upload Session
# ============================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_session_create(request):
    """Создать сессию загрузки важтарашчей."""
    ser = WaybillSessionCreateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)

    s = WaybillUploadSession.objects.create(
        user=request.user,
        client_total_files=ser.validated_data["client_total_files"],
        archive_formats=ser.validated_data.get("archive_formats", []),
        multi_doc=ser.validated_data.get("multi_doc", False),
        scan_type="detaliai",
    )

    return Response(
        {"id": str(s.id), "session_id": str(s.id), "stage": s.stage},
        status=status.HTTP_201_CREATED,
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def waybill_session_status(request, session_id):
    """Получить статус сессии для polling."""
    try:
        s = WaybillUploadSession.objects.get(id=session_id, user=request.user)
    except WaybillUploadSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

    return Response(WaybillSessionStatusSerializer(s).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_session_finalize(request, session_id):
    """
    Финализация: uploading → credit_check → queued/processing.
    Резервирует кредиты и запускает обработку.
    """
    with transaction.atomic():
        try:
            s = WaybillUploadSession.objects.select_for_update().get(
                id=session_id, user=request.user,
            )
        except WaybillUploadSession.DoesNotExist:
            return Response({"error": "Session not found"}, status=404)

        if s.stage != "uploading":
            return Response(
                {"error": f"Cannot finalize session in stage '{s.stage}'"},
                status=400,
            )

        u = CustomUser.objects.select_for_update().get(id=request.user.id)

        # Считаем expected items
        expected = s.expected_items or s.uploaded_files or s.client_total_files
        cost_per_doc = WAYBILL_CREDIT_COST
        total_cost = cost_per_doc * Decimal(expected)

        available = (u.credits or Decimal("0")) - (u.credits_reserved or Decimal("0"))

        if total_cost > available:
            s.stage = "blocked"
            s.error_message = f"Nepakanka kreditų: reikia {total_cost}, turima {available}"
            s.save(update_fields=["stage", "error_message", "updated_at"])
            return Response(
                {"error": s.error_message, "stage": "blocked"},
                status=402,
            )

        # Резервируем кредиты
        u.credits_reserved = (u.credits_reserved or Decimal("0")) + total_cost
        u.save(update_fields=["credits_reserved"])

        s.reserved_credits = total_cost
        s.reserved_items = expected

        # Если уже есть processing — в очередь
        has_processing = WaybillUploadSession.objects.filter(
            user=request.user, stage="processing",
        ).exclude(id=s.id).exists()

        if has_processing:
            s.stage = "queued"
        else:
            s.stage = "processing"
            s.started_at = timezone.now()

        s.save(update_fields=[
            "stage", "reserved_credits", "reserved_items",
            "started_at", "error_message", "updated_at",
        ])

    # Вне транзакции — запускаем обработку
    if s.stage == "processing":
        from .tasks import start_waybill_session_processing
        start_waybill_session_processing.delay(str(s.id))

    return Response({"stage": s.stage, "session_id": str(s.id)})


# ============================================================
# Chunked Upload
# ============================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_chunked_upload_init(request):
    """Инициализация chunked upload."""
    ser = WaybillChunkedUploadInitSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data

    try:
        session = WaybillUploadSession.objects.get(id=d["session_id"], user=request.user)
    except WaybillUploadSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

    if session.stage != "uploading":
        return Response({"error": "Session is not in uploading stage"}, status=400)

    upload = WaybillChunkedUpload.objects.create(
        user=request.user,
        session=session,
        filename=d["filename"],
        total_size=d["total_size"],
        chunk_size=d["chunk_size"],
        total_chunks=d["total_chunks"],
    )

    return Response(
        {"upload_id": str(upload.id), "status": upload.status},
        status=201,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_chunked_upload_chunk(request, upload_id):
    """Загрузить один chunk."""
    try:
        upload = WaybillChunkedUpload.objects.get(id=upload_id, user=request.user)
    except WaybillChunkedUpload.DoesNotExist:
        return Response({"error": "Upload not found"}, status=404)

    if upload.status != "uploading":
        return Response({"error": "Upload is not active"}, status=400)

    chunk_index = request.data.get("chunk_index")
    chunk_file = request.FILES.get("chunk")

    if chunk_index is None or chunk_file is None:
        return Response({"error": "chunk_index and chunk file required"}, status=400)

    chunk_index = int(chunk_index)

    # Путь для tmp файла
    if not upload.tmp_path:
        tmp_dir = os.path.join(settings.MEDIA_ROOT, "waybills_tmp", str(request.user.id))
        os.makedirs(tmp_dir, exist_ok=True)
        upload.tmp_path = os.path.join(tmp_dir, f"{upload.id}.part")

    # Записываем chunk
    mode = "ab" if os.path.exists(upload.tmp_path) else "wb"
    with open(upload.tmp_path, mode) as f:
        for c in chunk_file.chunks():
            f.write(c)

    received = upload.received or []
    if chunk_index not in received:
        received.append(chunk_index)
    upload.received = received
    upload.save(update_fields=["received", "tmp_path", "updated_at"])

    return Response({
        "upload_id": str(upload.id),
        "received": len(received),
        "total_chunks": upload.total_chunks,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_chunked_upload_complete(request, upload_id):
    """Завершить chunked upload — собрать файл и создать ScannedWaybill."""
    try:
        upload = WaybillChunkedUpload.objects.get(id=upload_id, user=request.user)
    except WaybillChunkedUpload.DoesNotExist:
        return Response({"error": "Upload not found"}, status=404)

    received = upload.received or []
    if len(received) < upload.total_chunks:
        return Response(
            {"error": f"Not all chunks received: {len(received)}/{upload.total_chunks}"},
            status=400,
        )

    if not upload.tmp_path or not os.path.exists(upload.tmp_path):
        return Response({"error": "Temporary file not found"}, status=400)

    try:
        session = upload.session

        # Создаём ScannedWaybill
        doc = ScannedWaybill(
            user=request.user,
            original_filename=upload.filename,
            status="pending",
            upload_session=session,
        )

        with open(upload.tmp_path, 'rb') as f:
            from django.core.files.base import ContentFile
            doc.file.save(upload.filename, ContentFile(f.read()), save=False)

        doc.uploaded_size_bytes = upload.total_size
        doc.save()

        # Обновляем счётчики сессии
        WaybillUploadSession.objects.filter(id=session.id).update(
            uploaded_files=F("uploaded_files") + 1,
            uploaded_bytes=F("uploaded_bytes") + upload.total_size,
            expected_items=F("expected_items") + 1,
        )

        upload.status = "complete"
        upload.save(update_fields=["status", "updated_at"])

        # Удаляем tmp
        try:
            os.remove(upload.tmp_path)
        except Exception:
            pass

        return Response({
            "doc_id": doc.id,
            "upload_id": str(upload.id),
            "status": "complete",
        })

    except Exception as e:
        logger.exception("[WAYBILL-UPLOAD] Complete failed: %s", e)
        upload.status = "failed"
        upload.error_message = str(e)
        upload.save(update_fields=["status", "error_message", "updated_at"])
        return Response({"error": str(e)}, status=500)


# ============================================================
# Waybill List
# ============================================================

class WaybillPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = "page_size"
    max_page_size = 100


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def waybill_list(request):
    """Список важтарашчей с фильтрацией и пагинацией."""
    qs = (
        ScannedWaybill.objects
        .filter(user=request.user, is_archive_container=False, is_multi_doc_container=False)
        .order_by("-uploaded_at")
    )

    # Фильтры
    status_filter = request.query_params.get("status")
    if status_filter:
        qs = qs.filter(status=status_filter)

    waybill_type = request.query_params.get("waybill_type")
    if waybill_type:
        qs = qs.filter(waybill_type=waybill_type)

    search = request.query_params.get("search")
    if search:
        qs = qs.filter(document_number__icontains=search)

    date_from = request.query_params.get("date_from")
    if date_from:
        qs = qs.filter(uploaded_at__date__gte=date_from)

    date_to = request.query_params.get("date_to")
    if date_to:
        qs = qs.filter(uploaded_at__date__lte=date_to)

    session_id = request.query_params.get("session_id")
    if session_id:
        qs = qs.filter(upload_session_id=session_id)

    paginator = WaybillPagination()
    page = paginator.paginate_queryset(qs, request)
    ser = ScannedWaybillListSerializer(page, many=True)
    return paginator.get_paginated_response(ser.data)


# ============================================================
# Waybill Detail / Update / Delete
# ============================================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def waybill_detail(request, pk):
    try:
        if request.user.is_superuser:
            doc = ScannedWaybill.objects.get(pk=pk)
        else:
            doc = ScannedWaybill.objects.get(pk=pk, user=request.user)
    except ScannedWaybill.DoesNotExist:
        return Response({"error": "Not found"}, status=404)
    return Response(ScannedWaybillDetailSerializer(doc).data)


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def waybill_update(request, pk):
    """Обновить поля важтарашчиса (ручная правка)."""
    try:
        doc = ScannedWaybill.objects.get(pk=pk, user=request.user)
    except ScannedWaybill.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    ser = ScannedWaybillUpdateSerializer(doc, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()

    return Response(ScannedWaybillDetailSerializer(doc).data)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def waybill_delete(request, pk):
    """Удалить один важтарашчис."""
    try:
        doc = ScannedWaybill.objects.get(pk=pk, user=request.user)
    except ScannedWaybill.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    # Удаляем файл
    if doc.file:
        try:
            doc.file.delete(save=False)
        except Exception:
            pass

    doc.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_bulk_delete(request):
    """Массовое удаление важтарашчей."""
    ids = request.data.get("ids", [])
    if not ids:
        return Response({"error": "No ids provided"}, status=400)

    docs = ScannedWaybill.objects.filter(pk__in=ids, user=request.user)

    # Удаляем файлы
    for doc in docs:
        if doc.file:
            try:
                doc.file.delete(save=False)
            except Exception:
                pass

    count = docs.count()
    docs.delete()

    return Response({"deleted": count})


"""
waybill_export_xls — экспорт важтарашчей в XLSX.
Добавить в views.py + URL: path("waybills/export-xls/", waybill_export_xls, name="waybill-export-xls"),
"""
import io
import logging
from datetime import datetime

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import ScannedWaybill

logger = logging.getLogger("docscanner_app")

# Polja dlia eksporta i ix zagolovki
EXPORT_COLUMNS = [
    ("document_number", "Numeris"),
    ("document_date", "Data"),
    ("airport", "Oro uostas"),
    ("payment_type", "Mokėjimo būdas"),
    ("delivery_receipt", "Važtaraštis kurui užpilti"),
    ("defuelling_receipt", "Važtaraštis kurui išpilti"),
    ("buyer_iata_code", "IATA kodas"),
    ("buyer_name", "Pavadinimas"),
    ("buyer_address", "Adresas"),
    ("buyer_vat_code", "PVM kodas"),
    ("buyer_remark_half_income", "Aviakompanija su > 1/2 pajamų"),
    ("buyer_remark_other", "Kita"),
    ("aircraft_type", "Orlaivio tipas"),
    ("flight_type", "Tipas"),
    ("outside_eu", "Už ES ribų"),
    ("flight_nature", "Skrydžio pobūdis"),
    ("time_departure", "Išvykimas"),
    ("time_arrival", "Atvykimas"),
    ("time_start", "Pradžia"),
    ("time_finish", "Pabaiga"),
    ("time_return", "Grįžimas"),
    ("from_city", "Iš (miestas)"),
    ("from_airport_code", "Iš (oro uostas)"),
    ("from_country_iso", "Iš (šalies kodas)"),
    ("to_city", "Į (miestas)"),
    ("to_airport_code", "Į (oro uostas)"),
    ("to_country_iso", "Į (šalies kodas)"),
    ("refueller_number", "Autocisternos Nr."),
    ("reading_before", "Prieš užpildymą"),
    ("reading_after", "Po užpildymo"),
    ("reading_difference", "Skirtumas"),
    ("company_representative", "Įmonės įgaliotas asmuo"),
    ("density_observed", "Tankis (faktinis)"),
    ("temperature_observed", "Temp. °C (faktinė)"),
    ("quantity_liters_observed", "Litrai (faktiniai)"),
    ("quantity_kg_observed", "Kilogramai (faktiniai)"),
    ("density_standard", "Tankis (+15°C)"),
    ("temperature_standard", "Temp. °C (+15°C)"),
    ("quantity_liters_standard", "Litrai (+15°C)"),
]

BOOLEAN_FIELDS = {
    "delivery_receipt",
    "defuelling_receipt",
    "outside_eu",
    "buyer_remark_half_income",
}


def _format_value(field_name, val):
    """Formatuoja reikšmę XLS celei.
    - Boolean laukai: None -> "Ne", True -> "Taip", False -> "Ne".
    - Kiti laukai: None -> "".
    """
    if field_name in BOOLEAN_FIELDS:
        if val is True:
            return "Taip"
        # val is False or None
        return "Ne"

    if val is None:
        return ""
    return val


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_export_xls(request):
    """Eksportuoti pasirinktus važtaraščius į XLSX failą."""
    ids = request.data.get("ids", [])
    if not ids:
        return Response({"error": "Nepateikti dokumentų ID"}, status=400)

    docs = ScannedWaybill.objects.filter(
        pk__in=ids,
        user=request.user,
        status__in=("completed", "exported"),
    ).order_by("document_date", "document_number")

    if not docs.exists():
        return Response({"error": "Nerasta eksportuojamų dokumentų"}, status=404)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return Response({"error": "openpyxl neįdiegtas serveryje"}, status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Važtaraščiai"

    # Stiliai
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Zagolovki
    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Danyje
    for row_idx, doc in enumerate(docs, start=2):
        for col_idx, (field, _) in enumerate(EXPORT_COLUMNS, start=1):
            val = getattr(doc, field, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=_format_value(field, val))
            cell.border = thin_border

    # Avto-shirina stolbcov
    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        max_len = len(label)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 30)

    # Zamorozka zagolovkov
    ws.freeze_panes = "A2"

    # Pomecijajem kak exported
    doc_ids = list(docs.values_list("id", flat=True))
    ScannedWaybill.objects.filter(id__in=doc_ids).update(status="exported")

    # Otdajem fail
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"vaztarasciai_{timestamp}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    logger.info(
        "[WAYBILL-EXPORT] User %s exported %d waybills to XLS",
        request.user.email, len(doc_ids),
    )

    return response


"""
Дополнительные views для важтарашчей — batch upload и active sessions.
Добавить в views.py рядом с остальными waybill views.
Добавить URLs.
"""


# ============================================================
# Batch upload (как SF upload_batch)
# ============================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_upload_batch(request, session_id):
    """Batch upload файлов через FormData — как SF upload_batch."""
    try:
        s = WaybillUploadSession.objects.get(id=session_id, user=request.user)
    except WaybillUploadSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

    if s.stage != "uploading":
        return Response({"error": f"Session stage is '{s.stage}', expected 'uploading'"}, status=400)

    files = request.FILES.getlist("files")
    if not files:
        return Response({"error": "No files provided"}, status=400)

    created_ids = []
    total_bytes = 0

    for f in files:
        doc = ScannedWaybill(
            user=request.user,
            original_filename=f.name,
            status="pending",
            upload_session=s,
            uploaded_size_bytes=f.size,
        )
        doc.file.save(f.name, f, save=False)
        doc.save()
        created_ids.append(doc.id)
        total_bytes += f.size

    # Обновляем счётчики сессии
    WaybillUploadSession.objects.filter(id=s.id).update(
        uploaded_files=F("uploaded_files") + len(created_ids),
        uploaded_bytes=F("uploaded_bytes") + total_bytes,
        expected_items=F("expected_items") + len(created_ids),
    )

    return Response({
        "uploaded": len(created_ids),
        "doc_ids": created_ids,
    })


# ============================================================
# Chunked upload (через session URL, как SF)
# ============================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_chunk_init(request, session_id):
    """Инициализация chunked upload через session URL."""
    try:
        session = WaybillUploadSession.objects.get(id=session_id, user=request.user)
    except WaybillUploadSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

    if session.stage != "uploading":
        return Response({"error": "Session is not in uploading stage"}, status=400)

    filename = request.data.get("filename")
    total_size = int(request.data.get("total_size", 0))
    chunk_size = int(request.data.get("chunk_size", 0))
    total_chunks = int(request.data.get("total_chunks", 0))

    if not filename or not total_size or not total_chunks:
        return Response({"error": "filename, total_size, total_chunks required"}, status=400)

    upload = WaybillChunkedUpload.objects.create(
        user=request.user,
        session=session,
        filename=filename,
        total_size=total_size,
        chunk_size=chunk_size,
        total_chunks=total_chunks,
    )

    return Response({"upload_id": str(upload.id)}, status=201)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def waybill_chunk_upload(request, session_id, upload_id, index):
    """Загрузить один chunk (raw bytes)."""
    try:
        upload = WaybillChunkedUpload.objects.get(
            id=upload_id, session_id=session_id, user=request.user,
        )
    except WaybillChunkedUpload.DoesNotExist:
        return Response({"error": "Upload not found"}, status=404)

    if upload.status != "uploading":
        return Response({"error": "Upload is not active"}, status=400)

    if not upload.tmp_path:
        tmp_dir = os.path.join(settings.MEDIA_ROOT, "waybills_tmp", str(request.user.id))
        os.makedirs(tmp_dir, exist_ok=True)
        upload.tmp_path = os.path.join(tmp_dir, f"{upload.id}.part")

    mode = "ab" if os.path.exists(upload.tmp_path) else "wb"
    with open(upload.tmp_path, mode) as f:
        f.write(request.body)

    received = upload.received or []
    if index not in received:
        received.append(index)
    upload.received = received
    upload.save(update_fields=["received", "tmp_path", "updated_at"])

    return Response({"received": len(received), "total_chunks": upload.total_chunks})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_chunk_complete(request, session_id, upload_id):
    """Завершить chunked upload — собрать файл и создать ScannedWaybill."""
    try:
        upload = WaybillChunkedUpload.objects.get(
            id=upload_id, session_id=session_id, user=request.user,
        )
    except WaybillChunkedUpload.DoesNotExist:
        return Response({"error": "Upload not found"}, status=404)

    received = upload.received or []
    if len(received) < upload.total_chunks:
        return Response({"error": f"Not all chunks: {len(received)}/{upload.total_chunks}"}, status=400)

    if not upload.tmp_path or not os.path.exists(upload.tmp_path):
        return Response({"error": "Temp file not found"}, status=400)

    try:
        session = upload.session
        doc = ScannedWaybill(
            user=request.user,
            original_filename=upload.filename,
            status="pending",
            upload_session=session,
            uploaded_size_bytes=upload.total_size,
        )

        from django.core.files.base import ContentFile
        with open(upload.tmp_path, 'rb') as f:
            doc.file.save(upload.filename, ContentFile(f.read()), save=False)
        doc.save()

        WaybillUploadSession.objects.filter(id=session.id).update(
            uploaded_files=F("uploaded_files") + 1,
            uploaded_bytes=F("uploaded_bytes") + upload.total_size,
            expected_items=F("expected_items") + 1,
        )

        upload.status = "complete"
        upload.save(update_fields=["status", "updated_at"])

        try:
            os.remove(upload.tmp_path)
        except Exception:
            pass

        return Response({"doc_id": doc.id, "status": "complete"})

    except Exception as e:
        logger.exception("[WAYBILL-UPLOAD] Chunk complete failed: %s", e)
        upload.status = "failed"
        upload.error_message = str(e)
        upload.save(update_fields=["status", "error_message", "updated_at"])
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def waybill_chunk_status(request, session_id, upload_id):
    """Статус chunked upload."""
    try:
        upload = WaybillChunkedUpload.objects.get(
            id=upload_id, session_id=session_id, user=request.user,
        )
    except WaybillChunkedUpload.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    return Response({
        "upload_id": str(upload.id),
        "status": upload.status,
        "received": len(upload.received or []),
        "total_chunks": upload.total_chunks,
    })


# ============================================================
# Active sessions (для ProcessingStatusBar)
# ============================================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def waybill_active_sessions(request):
    from datetime import timedelta

    # Активные сессии
    active = WaybillUploadSession.objects.filter(
        user=request.user,
        stage__in=["processing", "queued", "credit_check", "blocked"],
    ).order_by("-created_at")[:10]

    # "done" только за последние 30 секунд
    cutoff = timezone.now() - timedelta(seconds=30)
    done = WaybillUploadSession.objects.filter(
        user=request.user,
        stage="done",
        finished_at__gte=cutoff,
    ).order_by("-created_at")[:5]

    all_sessions = list(active) + list(done)

    result = []
    for s in all_sessions:
        result.append({
            "id": str(s.id),
            "stage": s.stage,
            "uploaded_files": s.uploaded_files,
            "expected_items": s.expected_items,
            "actual_items": s.actual_items,
            "processed_items": s.processed_items,
            "done_items": s.done_items,
            "failed_items": s.failed_items,
            "pending_archives": s.pending_archives,
            "error_message": s.error_message,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "started_at": s.started_at.isoformat() if s.started_at else None,
            "finished_at": s.finished_at.isoformat() if s.finished_at else None,
        })

    return Response({"sessions": result})


# ============================================================
# Retry / Cancel blocked session
# ============================================================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_retry_blocked(request, session_id):
    """Повторить blocked waybill session (после пополнения кредитов)."""
    with transaction.atomic():
        try:
            s = WaybillUploadSession.objects.select_for_update().get(
                id=session_id, user=request.user, stage="blocked",
            )
        except WaybillUploadSession.DoesNotExist:
            return Response({"error": "Session not found or not blocked"}, status=404)

        u = CustomUser.objects.select_for_update().get(id=request.user.id)

        expected = s.expected_items or s.uploaded_files
        total_cost = WAYBILL_CREDIT_COST * Decimal(expected)
        available = (u.credits or Decimal("0")) - (u.credits_reserved or Decimal("0"))

        if total_cost > available:
            return Response({
                "error": f"Nepakanka kreditų: reikia {total_cost}, turima {available}",
            }, status=402)

        u.credits_reserved = (u.credits_reserved or Decimal("0")) + total_cost
        u.save(update_fields=["credits_reserved"])

        s.reserved_credits = total_cost
        s.reserved_items = expected
        s.stage = "processing"
        s.started_at = timezone.now()
        s.error_message = ""
        s.save(update_fields=["stage", "reserved_credits", "reserved_items", "started_at", "error_message", "updated_at"])

    from .tasks import start_waybill_session_processing
    start_waybill_session_processing.delay(str(s.id))

    return Response({"stage": "processing"})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def waybill_cancel_blocked(request, session_id):
    """Отменить blocked waybill session."""
    with transaction.atomic():
        try:
            s = WaybillUploadSession.objects.select_for_update().get(
                id=session_id, user=request.user, stage="blocked",
            )
        except WaybillUploadSession.DoesNotExist:
            return Response({"error": "Session not found or not blocked"}, status=404)

        if s.reserved_credits > 0:
            u = CustomUser.objects.select_for_update().get(id=request.user.id)
            u.credits_reserved = max(
                (u.credits_reserved or Decimal("0")) - s.reserved_credits,
                Decimal("0"),
            )
            u.save(update_fields=["credits_reserved"])

        # Удаляем pending документы
        ScannedWaybill.objects.filter(upload_session=s, status="pending").delete()

        s.stage = "failed"
        s.finished_at = timezone.now()
        s.reserved_credits = Decimal("0")
        s.error_message = "Atšaukta vartotojo"
        s.save(update_fields=["stage", "finished_at", "reserved_credits", "error_message", "updated_at"])

    return Response({"stage": "failed"})

# ────────────────────────────────────────────────────────────
# END ─── Waybill scan ───
# ────────────────────────────────────────────────────────────


class CompanyProfileViewSet(viewsets.ModelViewSet):
    serializer_class = CompanyProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return CompanyProfile.objects.filter(user=self.request.user)

    def perform_destroy(self, instance):
        user = self.request.user
        count = user.company_profiles.count()
        if count <= 1:
            raise ValidationError(
                {"detail": "Negalima pašalinti paskutinės įmonės."}
            )

        was_active = user.active_company_profile_id == instance.id
        instance.delete()

        if was_active:
            new_active = user.company_profiles.first()
            user.active_company_profile = new_active
            user.save(update_fields=["active_company_profile"])

    @action(detail=True, methods=["post"], url_path="set-active")
    def set_active(self, request, pk=None):
        profile = self.get_object()
        user = request.user
        user.active_company_profile = profile

        update_fields = ["active_company_profile"]
        if profile.accounting_program:
            user.default_accounting_program = profile.accounting_program
            update_fields.append("default_accounting_program")

        user.save(update_fields=update_fields)
        return Response({"detail": "OK", "active_id": profile.id})





# ────────────────────────────────────────────────────────────
# ─── Pirkimai ───
# ────────────────────────────────────────────────────────────

from django.db.models import Count, Prefetch
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import LimitOffsetPagination

class PurchasePagination(LimitOffsetPagination):
    default_limit = 50
    max_limit = 100

class PurchaseViewSet(viewsets.ModelViewSet):
    serializer_class = PurchaseSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PurchasePagination

    def get_queryset(self):
        qs = (
            Purchase.objects
            .filter(user=self.request.user)
            .select_related("company_profile", "scanned_document")
            .annotate(line_items_count=Count("line_items"))
        )

        # Company filter
        company_id = self.request.query_params.get("company_profile")
        if company_id:
            qs = qs.filter(company_profile_id=company_id)
        else:
            active = self.request.user.active_company_profile_id
            if active:
                qs = qs.filter(company_profile_id=active)

        # Search
        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(document_number__icontains=search) |
                Q(document_series__icontains=search) |
                Q(seller_name__icontains=search) |
                Q(seller_id__icontains=search) |
                Q(seller_vat_code__icontains=search)
            )

        # Payment status
        payment = self.request.query_params.get("payment_status")
        if payment and payment != "visi":
            qs = qs.filter(payment_status=payment)

        # Status (including computed reikia_perziuros)
        status_filter = self.request.query_params.get("status_filter")
        if status_filter == "nauja":
            qs = qs.filter(status="new")
        elif status_filter == "reikia_perziuros":
            qs = qs.filter(Q(ready_for_export=False) | Q(math_validation_passed=False) | Q(kor_balanced=False))
        elif status_filter == "uzregistruota":
            qs = qs.filter(status="accounted")

        # Period
        period_from = self.request.query_params.get("period_from")
        period_to = self.request.query_params.get("period_to")
        if period_from:
            qs = qs.filter(
                Q(period__gte=period_from) | Q(period__isnull=True, invoice_date__gte=period_from)
            )
        if period_to:
            qs = qs.filter(
                Q(period__lte=period_to) | Q(period__isnull=True, invoice_date__lte=period_to)
            )

        return qs.order_by("-created_at")

    def perform_destroy(self, instance):
        scanned_doc = instance.scanned_document
        instance.delete()

        if scanned_doc:
            has_others = Purchase.objects.filter(
                scanned_document=scanned_doc
            ).exists()
            # Pozже добавить:
            # or Invoice.objects.filter(scanned_document=scanned_doc).exists()

            if not has_others:
                scanned_doc.perkelta_i_apskaita = False
                scanned_doc.perkelta_i_apskaita_at = None
                scanned_doc.perkelta_i_company_profile = None
                scanned_doc.save(update_fields=[
                    "perkelta_i_apskaita",
                    "perkelta_i_apskaita_at",
                    "perkelta_i_company_profile",
                ])

    @action(
        detail=True,
        methods=["patch"],
        url_path=r"line-items/(?P<line_id>[^/.]+)",
    )
    def update_line_item(self, request, pk=None, line_id=None):
        purchase = self.get_object()

        try:
            line = purchase.line_items.get(id=line_id)
        except PurchaseLine.DoesNotExist:
            return Response(
                {"detail": "Pirkimo eilutė nerasta."},
                status=status.HTTP_404_NOT_FOUND,
            )

        allowed_fields = {
            "debeto_saskaita",
            "kredito_saskaita",
            "pvm_saskaita",
        }

        data = {
            key: value
            for key, value in request.data.items()
            if key in allowed_fields
        }

        if not data:
            return Response(
                {"detail": "Nėra leidžiamų laukų atnaujinimui."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = PurchaseLineSerializer(
            line,
            data=data,
            partial=True,
            context=self.get_serializer_context(),
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # Пересчитать баланс
        purchase.kor_balanced = compute_kor_balanced(purchase)
        purchase.save(update_fields=["kor_balanced"])

        from .utils.journal_generators import sync_purchase_journal_entry
        try:
            sync_purchase_journal_entry(purchase)
        except Exception as e:
            logger.warning("[Purchase] DK sync failed for %s: %s", purchase.id, e)

        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="add-line-item")
    def add_line_item(self, request, pk=None):
        purchase = self.get_object()
        line = PurchaseLine.objects.create(purchase=purchase)
 
        # Пересчитать валидацию
        purchase.ready_for_export = check_required_fields_for_export(purchase)
        is_math_valid, _ = validate_document_math_for_export(purchase)
        purchase.math_validation_passed = is_math_valid
        purchase.kor_balanced = compute_kor_balanced(purchase)
        purchase.save(update_fields=["ready_for_export", "math_validation_passed", "kor_balanced"])

        from .utils.journal_generators import sync_purchase_journal_entry
        try:
            sync_purchase_journal_entry(purchase)
        except Exception as e:
            logger.warning("[Purchase] DK sync failed for %s: %s", purchase.id, e)

        serializer = PurchaseLineSerializer(line)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
 
    @action(
        detail=True,
        methods=["delete"],
        url_path=r"delete-line-item/(?P<line_id>[^/.]+)",
    )
    def delete_line_item(self, request, pk=None, line_id=None):
        purchase = self.get_object()
        try:
            line = purchase.line_items.get(id=line_id)
        except PurchaseLine.DoesNotExist:
            return Response(
                {"detail": "Line item not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
 
        line.delete()
 
        # Пересчитать валидацию
        purchase.ready_for_export = check_required_fields_for_export(purchase)
        is_math_valid, _ = validate_document_math_for_export(purchase)
        purchase.math_validation_passed = is_math_valid
        purchase.kor_balanced = compute_kor_balanced(purchase)
        purchase.save(update_fields=["ready_for_export", "math_validation_passed", "kor_balanced"])

        from .utils.journal_generators import sync_purchase_journal_entry
        try:
            sync_purchase_journal_entry(purchase)
        except Exception as e:
            logger.warning("[Purchase] DK sync failed for %s: %s", purchase.id, e)

        return Response(status=status.HTTP_204_NO_CONTENT)
 
    def perform_update(self, serializer):
        """Override чтобы пересчитывать валидацию при PATCH через ViewSet."""
        instance = serializer.save()
        changed = set(serializer.validated_data.keys())
 
        update_fields = []
        if changed & PURCHASE_REQUIRED_FIELDS:
            instance.ready_for_export = check_required_fields_for_export(instance)
            update_fields.append("ready_for_export")
        if changed & PURCHASE_MATH_FIELDS:
            is_valid, _ = validate_document_math_for_export(instance)
            instance.math_validation_passed = is_valid
            update_fields.append("math_validation_passed")
        if changed & (PURCHASE_MATH_FIELDS | {"debeto_saskaita", "kredito_saskaita", "pvm_saskaita"}):
            instance.kor_balanced = compute_kor_balanced(instance)
            update_fields.append("kor_balanced")
        if update_fields:
            instance.save(update_fields=update_fields)

        from .utils.journal_generators import sync_purchase_journal_entry
        try:
            sync_purchase_journal_entry(instance)
        except Exception as e:
            logger.warning("[Purchase] DK sync failed for %s: %s", instance.id, e)


def compute_kor_balanced(document):
    from decimal import Decimal

    tolerance = Decimal("0.02")

    def dec(value):
        if value in (None, ""):
            return Decimal("0")
        return Decimal(str(value))

    lines = list(document.line_items.all())

    is_invoice = getattr(document, "pirkimas_pardavimas", None) == "pardavimas" or hasattr(document, "invoice_type")
    is_purchase = getattr(document, "pirkimas_pardavimas", None) == "pirkimas" or hasattr(document, "is_credit_invoice") and not is_invoice

    # ═══════════════════════════════════════════════════════
    # PARDAVIMAI / Invoice
    # D 2410 = amount_with_vat
    # K 5000/5001 = amount_wo_vat
    # K 4492 = vat_amount
    # ═══════════════════════════════════════════════════════
    if is_invoice:
        # Debetas pardavimui yra document-level
        if not document.debeto_saskaita:
            return False

        d_total = Decimal("0")
        k_total = Decimal("0")

        amount_with_vat = dec(document.amount_with_vat)

        if amount_with_vat != 0:
            d_total += amount_with_vat

        if lines:
            for line in lines:
                subtotal = dec(line.subtotal)
                vat = dec(line.vat)

                kredito_saskaita = (
                    getattr(line, "kredito_saskaita", None)
                    or document.kredito_saskaita
                )

                pvm_saskaita = (
                    getattr(line, "pvm_saskaita", None)
                    or document.pvm_saskaita
                )

                if subtotal != 0 and not kredito_saskaita:
                    return False

                if vat != 0 and not pvm_saskaita:
                    return False

                k_total += subtotal + vat

        else:
            amount_wo_vat = dec(document.amount_wo_vat)
            vat_amount = dec(document.vat_amount)

            if amount_wo_vat != 0 and not document.kredito_saskaita:
                return False

            if vat_amount != 0 and not document.pvm_saskaita:
                return False

            k_total += amount_wo_vat + vat_amount

        return abs(d_total - k_total) <= tolerance

    # ═══════════════════════════════════════════════════════
    # PIRKIMAI / Purchase
    # D sąnaudos / prekės / turtas = subtotal
    # D 2441 = vat
    # K 4430 = total
    # ═══════════════════════════════════════════════════════

    # Kredito sąskaita pirkimui reikalinga visais atvejais
    if not document.kredito_saskaita:
        return False

    if lines:
        d_total = Decimal("0")
        k_total = Decimal("0")

        for line in lines:
            subtotal = dec(line.subtotal)
            vat = dec(line.vat)
            total = dec(line.total)

            debeto_saskaita = (
                getattr(line, "effective_debeto", None)
                or getattr(line, "debeto_saskaita", None)
                or document.debeto_saskaita
            )

            kredito_saskaita = (
                getattr(line, "kredito_saskaita", None)
                or document.kredito_saskaita
            )

            pvm_saskaita = (
                getattr(line, "pvm_saskaita", None)
                or document.pvm_saskaita
            )

            if subtotal != 0 and not debeto_saskaita:
                return False

            if total != 0 and not kredito_saskaita:
                return False

            if vat != 0 and not pvm_saskaita:
                return False

            d_total += subtotal + vat
            k_total += total

    else:
        amount_wo_vat = dec(document.amount_wo_vat)
        vat_amount = dec(document.vat_amount)
        amount_with_vat = dec(document.amount_with_vat)

        if amount_wo_vat != 0 and not document.debeto_saskaita:
            return False

        if amount_with_vat != 0 and not document.kredito_saskaita:
            return False

        if vat_amount != 0 and not document.pvm_saskaita:
            return False

        d_total = amount_wo_vat + vat_amount
        k_total = amount_with_vat

    return abs(d_total - k_total) <= tolerance


def _next_free_number_int(user, company_profile_id, invoice_type, prefix, start_from):
    """
    Возвращает первое свободное числовое значение номера для серии
    (prefix + invoice_type) в пределах профиля, начиная со start_from.
    Сравнение числовое (padding-agnostic): document_number хранится строкой
    с ведущими нулями, поэтому парсим в int.
    """
    from .models import Invoice

    taken_raw = (
        Invoice.objects
        .filter(
            user=user,
            company_profile_id=company_profile_id,
            document_series=prefix,
            invoice_type=invoice_type,
        )
        .exclude(status="cancelled")
        .values_list("document_number", flat=True)
    )

    taken = set()
    for val in taken_raw:
        try:
            taken.add(int(str(val).strip()))
        except (ValueError, TypeError):
            continue  # нечисловые номера игнорируем

    n = max(int(start_from), 1)
    while n in taken:
        n += 1
    return n


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def transfer_to_accounting(request):
    from decimal import Decimal, InvalidOperation
    from django.utils import timezone
    from .models import (
        CompanyProfile,
        ScannedDocument,
        Purchase,
        PurchaseLine,
        Invoice,
        InvoiceLineItem,
    )

    user = request.user
    doc_ids = request.data.get("document_ids", [])
    company_profile_id = request.data.get("company_profile_id")
    cp_key = request.data.get("cp_key", "")
    dry_run = request.data.get("dry_run", False)
    replace_document_company = request.data.get("replace_document_company") in (
        True,
        "true",
        "True",
        "1",
        1,
    )

    if not doc_ids:
        return Response(
            {"detail": "Nepateikti dokumentų ID."},
            status=400,
        )

    try:
        profile = CompanyProfile.objects.get(
            id=company_profile_id,
            user=user,
        )
    except CompanyProfile.DoesNotExist:
        return Response(
            {"detail": "Įmonės profilis nerastas."},
            status=404,
        )

    documents = (
        ScannedDocument.objects
        .filter(id__in=doc_ids, user=user)
        .prefetch_related("line_items")
    )

    # ═══════════════════════════════════════════════════════
    # Helpers: signs / amounts
    # ═══════════════════════════════════════════════════════

    def _dec(val, default=None):
        if val is None or val == "":
            return default

        try:
            return Decimal(str(val))
        except (InvalidOperation, TypeError, ValueError):
            return default

    def _has_amount(val):
        d = _dec(val)
        return d is not None and abs(d) > Decimal("0.001")

    def _signed_amount(val, is_credit):
        """
        Normalizuoja sumas:
        - įprastai SF visada teigiama
        - kreditinei SF visada neigiama
        """
        d = _dec(val)
        if d is None:
            return None

        if d == 0:
            return Decimal("0")

        return -abs(d) if is_credit else abs(d)

    def _signed_quantity(val, is_credit, default=1):
        """
        Normalizuoja kiekį:
        - įprastai SF kiekis teigiamas
        - kreditinei SF kiekis neigiamas
        """
        d = _dec(val, Decimal(str(default)))

        if d is None:
            return None

        if d == 0:
            return Decimal("0")

        return -abs(d) if is_credit else abs(d)

    def _positive_amount(val, default=None):
        """
        Kaina visada turi būti teigiama.
        """
        d = _dec(val, default)

        if d is None:
            return None

        return abs(d)

    def _line_price(price, subtotal, quantity):
        """
        Jeigu OCR davė price, imam abs(price).
        Jeigu price nėra, bandom paskaičiuoti iš subtotal / quantity.
        """
        price_dec = _dec(price)

        if price_dec is not None and price_dec != 0:
            return abs(price_dec)

        subtotal_abs = _positive_amount(subtotal)
        quantity_abs = _positive_amount(quantity)

        if subtotal_abs is not None and quantity_abs not in (None, Decimal("0")):
            return subtotal_abs / quantity_abs

        return subtotal_abs

    def _mark_doc_transferred(doc):
        doc.perkelta_i_apskaita = True
        doc.perkelta_i_apskaita_at = timezone.now()
        doc.perkelta_i_company_profile = profile
        doc.save(update_fields=[
            "perkelta_i_apskaita",
            "perkelta_i_apskaita_at",
            "perkelta_i_company_profile",
        ])

    # ═══════════════════════════════════════════════════════
    # Company detection helpers
    # ═══════════════════════════════════════════════════════

    profile_ids = set()

    if profile.company_code:
        profile_ids.add(profile.company_code.strip().upper())

    if profile.vat_code:
        profile_ids.add(profile.vat_code.strip().upper())

    profile_name_norm = (profile.name or "").strip().upper()

    if profile_name_norm:
        profile_ids.add(profile_name_norm)

    def _ids_from_doc(id_val, vat_val, name_val):
        s = set()

        if id_val:
            s.add(id_val.strip().upper())

        if vat_val:
            s.add(vat_val.strip().upper())

        if name_val:
            s.add(name_val.strip().upper())

        return s

    def _cp_key(id_val, vat_val, name_val):
        id_str = (str(id_val) if id_val else "").strip()

        if id_str:
            return f"id:{id_str}"

        norm_vat = (vat_val or "").strip().lower()

        if norm_vat:
            return norm_vat

        return (name_val or "").strip().lower()

    def _norm(value):
        return (str(value) if value else "").strip().upper()


    def _selected_role_for_doc(doc):
        """
        Grąžina, kuri dokumento pusė buvo pasirinkta kaip 'mano įmonė':
        buyer / seller / None.
        """
        if not cp_key:
            return None

        seller_cp = _cp_key(
            doc.seller_id,
            doc.seller_vat_code,
            doc.seller_name,
        )

        buyer_cp = _cp_key(
            doc.buyer_id,
            doc.buyer_vat_code,
            doc.buyer_name,
        )

        if cp_key == seller_cp:
            return "seller"

        if cp_key == buyer_cp:
            return "buyer"

        return None


    def _party_payload(doc, role):
        prefix = "seller" if role == "seller" else "buyer"

        return {
            "role": "Pardavėjas" if role == "seller" else "Pirkėjas",
            "name": getattr(doc, f"{prefix}_name", "") or "",
            "code": getattr(doc, f"{prefix}_id", "") or "",
            "vat": getattr(doc, f"{prefix}_vat_code", "") or "",
        }


    def _profile_payload():
        return {
            "name": profile.name or "",
            "code": profile.company_code or "",
            "vat": profile.vat_code or "",
        }


    def _clean_name(value):
        s = _norm(value)

        remove_parts = [
            "UAB",
            "AB",
            "MB",
            "VŠĮ",
            "VSI",
            "IĮ",
            "II",
            "INDIVIDUALI ĮMONĖ",
            "INDIVIDUALI IMONE",
        ]

        for part in remove_parts:
            s = s.replace(part, " ")

        s = s.replace('"', " ").replace("'", " ").replace(",", " ").replace(".", " ")

        return " ".join(s.split())


    def _party_matches_profile(party):
        profile_code = _norm(profile.company_code)
        party_code = _norm(party.get("code"))

        profile_vat = _norm(profile.vat_code)
        party_vat = _norm(party.get("vat"))

        # 1. Jeigu įmonės kodas sutampa — laikome match
        if profile_code and party_code and profile_code == party_code:
            return True

        # 2. Jeigu PVM kodas sutampa — laikome match
        if profile_vat and party_vat and profile_vat == party_vat:
            return True

        # 3. Jeigu abu kodai yra ir abu nesutampa — mismatch
        strong_checks_exist = False

        if profile_code and party_code:
            strong_checks_exist = True

        if profile_vat and party_vat:
            strong_checks_exist = True

        if strong_checks_exist:
            return False

        # 4. Jei kodų palyginti negalime, naudojame name fallback
        profile_name = _clean_name(profile.name)
        party_name = _clean_name(party.get("name"))

        if profile_name and party_name:
            return profile_name == party_name

        # 5. Jei nėra pakankamai duomenų — nerodome mismatch
        return True


    def _company_mismatch_for_docs(doc_list):
        mismatches = []

        for d in doc_list:
            role = _selected_role_for_doc(d)

            if not role:
                continue

            party = _party_payload(d, role)

            if not _party_matches_profile(party):
                mismatches.append({
                    "doc_id": d.id,
                    "selected_company": party,
                })

        if not mismatches:
            return None

        first = mismatches[0]

        return {
            "selected_company": first["selected_company"],
            "active_profile": _profile_payload(),
            "affected_count": len(mismatches),
        }


    def _profile_field(*names):
        for name in names:
            value = getattr(profile, name, None)

            if value not in (None, ""):
                return value

        return ""


    def _apply_profile_party_to_doc_in_memory(doc, role):
        """
        Pakeičia pasirinktą dokumento pusę tik šiame request'e.
        ScannedDocument DB įrašas nekeičiamas.
        Pakeisti duomenys naudojami tik kuriant Purchase / Invoice.
        """
        prefix = "seller" if role == "seller" else "buyer"

        model_fields = {f.name for f in doc._meta.fields}
        updates = {}

        updates[f"{prefix}_name"] = profile.name or ""
        updates[f"{prefix}_id"] = profile.company_code or ""
        updates[f"{prefix}_vat_code"] = profile.vat_code or ""
        updates[f"{prefix}_name_normalized"] = _norm(profile.name)

        optional_updates = {
            f"{prefix}_address": _profile_field("address", "company_address"),
            f"{prefix}_country": _profile_field("country", "company_country"),
            f"{prefix}_country_iso": _profile_field("country_iso", "company_country_iso"),
            f"{prefix}_iban": _profile_field("iban", "bank_iban", "bank_account"),
            f"{prefix}_is_person": False,
        }

        for field, value in optional_updates.items():
            if value not in (None, ""):
                updates[field] = value

        for field, value in updates.items():
            if field in model_fields:
                setattr(doc, field, value)

        return updates

    def detect_direction(doc):
        seller_cp = _cp_key(
            doc.seller_id,
            doc.seller_vat_code,
            doc.seller_name,
        )

        buyer_cp = _cp_key(
            doc.buyer_id,
            doc.buyer_vat_code,
            doc.buyer_name,
        )

        # Multi mode: kryptis nustatoma pagal pasirinktą įmonę dokumente
        if cp_key:
            if cp_key == seller_cp:
                return "pardavimas"

            if cp_key == buyer_cp:
                return "pirkimas"

        # Fallback pagal aktyvų profilį
        buyer_ids = _ids_from_doc(
            doc.buyer_id,
            doc.buyer_vat_code,
            doc.buyer_name,
        )

        seller_ids = _ids_from_doc(
            doc.seller_id,
            doc.seller_vat_code,
            doc.seller_name,
        )

        if profile_ids & buyer_ids:
            return "pirkimas"

        if profile_ids & seller_ids:
            return "pardavimas"

        return None

    # ═══════════════════════════════════════════════════════
    # Split docs
    # ═══════════════════════════════════════════════════════

    purchase_docs = []
    sale_docs = []
    skipped = []

    for doc in documents:
        direction = detect_direction(doc)

        if direction == "pirkimas":
            existing_purchase = Purchase.objects.filter(
                scanned_document=doc,
                company_profile=profile,
            ).first()

            if existing_purchase:
                if not doc.perkelta_i_apskaita:
                    _mark_doc_transferred(doc)

                skipped.append({
                    "id": doc.id,
                    "reason": "Jau perkeltas į pirkimus",
                })
                continue

            selected_role = _selected_role_for_doc(doc)

            if replace_document_company and not dry_run and selected_role:
                selected_party = _party_payload(doc, selected_role)

                if not _party_matches_profile(selected_party):
                    _apply_profile_party_to_doc_in_memory(doc, selected_role)

            purchase_docs.append(doc)

        elif direction == "pardavimas":
            existing_invoice = Invoice.objects.filter(
                scanned_document=doc,
                company_profile=profile,
            ).first()

            if existing_invoice:
                if not doc.perkelta_i_apskaita:
                    _mark_doc_transferred(doc)

                skipped.append({
                    "id": doc.id,
                    "reason": "Jau perkeltas į pardavimus",
                })
                continue

            selected_role = _selected_role_for_doc(doc)

            if replace_document_company and not dry_run and selected_role:
                selected_party = _party_payload(doc, selected_role)

                if not _party_matches_profile(selected_party):
                    _apply_profile_party_to_doc_in_memory(doc, selected_role)

            sale_docs.append(doc)

        else:
            skipped.append({
                "id": doc.id,
                "filename": doc.original_filename,
                "reason": "Nepavyko nustatyti krypties",
            })

    # ═══════════════════════════════════════════════════════
    # Dry run
    # ═══════════════════════════════════════════════════════

    company_mismatch = _company_mismatch_for_docs(
        purchase_docs + sale_docs
    )

    if dry_run:
        return Response({
            "company_name": profile.name,
            "purchase_count": len(purchase_docs),
            "sale_count": len(sale_docs),
            "skipped": skipped,
            "company_mismatch": company_mismatch,
        })

    # ═══════════════════════════════════════════════════════
    # Create purchases
    # ═══════════════════════════════════════════════════════

    created_purchases = []

    for doc in purchase_docs:
        is_credit = bool(doc.is_credit_invoice)

        doc_pvm_kodas = auto_select_pvm_code(
            pirkimas_pardavimas="pirkimas",
            buyer_country_iso=doc.buyer_country_iso,
            seller_country_iso=doc.seller_country_iso,
            preke_paslauga=doc.preke_paslauga,
            vat_percent=(
                float(doc.vat_percent)
                if doc.vat_percent is not None
                else None
            ),
            separate_vat=bool(doc.separate_vat),
            buyer_has_vat_code=bool(doc.buyer_vat_code),
            seller_has_vat_code=bool(doc.seller_vat_code),
            doc_96_str=bool(getattr(doc, "doc_96_str", False)),
        )

        purchase = Purchase.objects.create(
            user=user,
            company_profile=profile,
            scanned_document=doc,
            status="new",

            # Korespondencijos
            debeto_saskaita=doc.pirkimo_saskaita or "6312",
            kredito_saskaita="4430",
            pvm_saskaita="2441" if _has_amount(doc.vat_amount) else None,

            period=doc.invoice_date.replace(day=1) if doc.invoice_date else None,

            # Dokumento duomenys
            document_type=doc.document_type,
            is_credit_invoice=doc.is_credit_invoice,
            is_debit_invoice=doc.is_debit_invoice,
            document_series=doc.document_series,
            document_number=doc.document_number,
            invoice_date=doc.invoice_date,
            due_date=doc.due_date,
            operation_date=doc.operation_date,

            # Seller / tiekėjas
            seller_name=doc.seller_name,
            seller_id=doc.seller_id,
            seller_vat_code=doc.seller_vat_code,
            seller_address=doc.seller_address,
            seller_country=doc.seller_country,
            seller_country_iso=doc.seller_country_iso,
            seller_iban=doc.seller_iban,
            seller_is_person=doc.seller_is_person,
            seller_vat_val=getattr(doc, "seller_vat_val", None),

            # Buyer / mes
            buyer_name=doc.buyer_name,
            buyer_id=doc.buyer_id,
            buyer_vat_code=doc.buyer_vat_code,
            buyer_address=doc.buyer_address,
            buyer_country=doc.buyer_country,
            buyer_country_iso=doc.buyer_country_iso,
            buyer_iban=doc.buyer_iban,
            buyer_is_person=doc.buyer_is_person,

            # Sumos
            currency=doc.currency or "EUR",
            amount_wo_vat=_signed_amount(doc.amount_wo_vat, is_credit),
            vat_amount=_signed_amount(doc.vat_amount, is_credit),
            vat_percent=doc.vat_percent,
            amount_with_vat=_signed_amount(doc.amount_with_vat, is_credit),
            invoice_discount_with_vat=_signed_amount(
                doc.invoice_discount_with_vat,
                is_credit,
            ),
            invoice_discount_wo_vat=_signed_amount(
                doc.invoice_discount_wo_vat,
                is_credit,
            ),
            separate_vat=doc.separate_vat,
            doc_96_str=doc.doc_96_str,

            # iSAF / klasifikatoriai
            pirkimas_pardavimas="pirkimas",
            report_to_isaf=doc.report_to_isaf,
            document_type_code=doc.document_type_code or "",
            pvm_kodas=doc_pvm_kodas or "",

            # Prekė / sumiškai
            prekes_kodas=doc.prekes_kodas,
            prekes_pavadinimas=doc.prekes_pavadinimas,
            preke_paslauga=doc.preke_paslauga,

            # Meta
            scan_type=doc.scan_type,
            order_number=doc.order_number,
            paid_by_cash=doc.paid_by_cash,
            is_long_term_asset_candidate=doc.is_long_term_asset_candidate,
            suggested_asset_type=doc.suggested_asset_type or "",
        )

        purchase_lines = []
        source_lines = list(doc.line_items.all())

        if source_lines:
            for i, li in enumerate(source_lines):
                li_vat_pct = (
                    float(li.vat_percent)
                    if li.vat_percent is not None
                    else (
                        float(doc.vat_percent)
                        if doc.vat_percent is not None
                        else None
                    )
                )

                li_preke_paslauga = (
                    li.preke_paslauga
                    or doc.preke_paslauga
                )

                li_pvm_kodas = auto_select_pvm_code(
                    pirkimas_pardavimas="pirkimas",
                    buyer_country_iso=doc.buyer_country_iso,
                    seller_country_iso=doc.seller_country_iso,
                    preke_paslauga=li_preke_paslauga,
                    vat_percent=li_vat_pct,
                    separate_vat=False,
                    buyer_has_vat_code=bool(doc.buyer_vat_code),
                    seller_has_vat_code=bool(doc.seller_vat_code),
                    doc_96_str=bool(getattr(doc, "doc_96_str", False)),
                )

                raw_quantity = (
                    li.quantity
                    if li.quantity is not None
                    else 1
                )

                purchase_lines.append(
                    PurchaseLine(
                        purchase=purchase,

                        prekes_kodas=li.prekes_kodas or doc.prekes_kodas or "",
                        prekes_barkodas=li.prekes_barkodas or "",
                        prekes_pavadinimas=(
                            li.prekes_pavadinimas
                            or doc.prekes_pavadinimas
                            or "Prekės / paslaugos"
                        ),
                        preke_paslauga=li_preke_paslauga or "",

                        unit=li.unit or "vnt.",

                        # Kreditinei kiekis neigiamas, kaina teigiama
                        quantity=_signed_quantity(raw_quantity, is_credit),
                        price=_line_price(
                            li.price,
                            li.subtotal,
                            raw_quantity,
                        ),

                        subtotal=_signed_amount(li.subtotal, is_credit),
                        vat=_signed_amount(li.vat, is_credit),
                        vat_percent=(
                            li.vat_percent
                            if li.vat_percent is not None
                            else doc.vat_percent
                        ),
                        total=_signed_amount(li.total, is_credit),

                        discount_with_vat=_signed_amount(
                            li.discount_with_vat,
                            is_credit,
                        ),
                        discount_wo_vat=_signed_amount(
                            li.discount_wo_vat,
                            is_credit,
                        ),

                        pvm_kodas=li_pvm_kodas or "",
                        sort_order=i,
                    )
                )

        else:
            fallback_name = (
                doc.prekes_pavadinimas
                or doc.original_filename
                or "Prekės / paslaugos"
            )

            fallback_subtotal = _signed_amount(doc.amount_wo_vat, is_credit)
            fallback_vat = _signed_amount(doc.vat_amount, is_credit)
            fallback_total = _signed_amount(doc.amount_with_vat, is_credit)

            purchase_lines.append(
                PurchaseLine(
                    purchase=purchase,

                    prekes_kodas=doc.prekes_kodas or "",
                    prekes_barkodas="",
                    prekes_pavadinimas=fallback_name,
                    preke_paslauga=doc.preke_paslauga or "",

                    unit="vnt.",

                    # Sumiškai:
                    # įprasta SF:  1 x 100 = 100
                    # kreditinė:  -1 x 100 = -100
                    quantity=_signed_quantity(1, is_credit),
                    price=_positive_amount(doc.amount_wo_vat),

                    subtotal=fallback_subtotal,
                    vat=fallback_vat,
                    vat_percent=doc.vat_percent,
                    total=fallback_total,

                    discount_with_vat=_signed_amount(
                        doc.invoice_discount_with_vat,
                        is_credit,
                    ),
                    discount_wo_vat=_signed_amount(
                        doc.invoice_discount_wo_vat,
                        is_credit,
                    ),

                    pvm_kodas=doc_pvm_kodas or "",
                    sort_order=0,
                )
            )

        if purchase_lines:
            PurchaseLine.objects.bulk_create(purchase_lines)

        purchase.ready_for_export = check_required_fields_for_export(purchase)

        is_math_valid, _ = validate_document_math_for_export(purchase)
        purchase.math_validation_passed = is_math_valid

        purchase.kor_balanced = compute_kor_balanced(purchase)

        purchase.save(update_fields=[
            "ready_for_export",
            "math_validation_passed",
            "kor_balanced",
        ])

        try:
            from .utils.journal_generators import sync_purchase_journal_entry

            sync_purchase_journal_entry(purchase)
        except Exception as e:
            logger.warning(
                "[Transfer] sync_purchase_journal_entry failed for %s: %s",
                purchase.id,
                e,
            )

        try:
            from .utils.payment_invoice_matching import match_purchase_on_transfer

            alloc = match_purchase_on_transfer(purchase)

            if alloc:
                logger.info(
                    "[Transfer] Purchase %s auto-matched to bank txn, amount=%s",
                    purchase.id,
                    alloc.amount,
                )
        except Exception as e:
            logger.warning(
                "[Transfer] match_purchase_on_transfer failed: %s",
                e,
            )

        created_purchases.append(purchase.id)
        _mark_doc_transferred(doc)

    # ═══════════════════════════════════════════════════════
    # Create sales / pardavimai
    # ═══════════════════════════════════════════════════════

    created_sales = []

    for doc in sale_docs:
        if not doc.document_number:
            skipped.append({
                "id": doc.id,
                "filename": doc.original_filename,
                "reason": "Trūksta dokumento numerio",
            })
            continue

        dup_qs = (
            Invoice.objects
            .filter(
                user=user,
                company_profile=profile,
                document_series=doc.document_series or "",
                document_number=doc.document_number,
            )
            .exclude(status="cancelled")
        )

        if dup_qs.exists():
            _hint_type = "kreditine" if bool(doc.is_credit_invoice) else (
                "pvm_saskaita"
                if (_has_amount(doc.vat_amount) or doc.separate_vat)
                else "saskaita"
            )
            try:
                _start = int(str(doc.document_number).strip())
            except (ValueError, TypeError):
                _start = 1

            _free_int = _next_free_number_int(
                user, profile.id, _hint_type,
                doc.document_series or "", _start,
            )

            skipped.append({
                "id": doc.id,
                "filename": doc.original_filename,
                "reason": (
                    f"Dokumentas {doc.document_series or ''}-{doc.document_number} "
                    f"jau yra apskaitoje. Laisvas numeris šioje serijoje: "
                    f"{doc.document_series or ''}-{_free_int}"
                ),
            })
            continue

        with transaction.atomic():
            is_credit = bool(doc.is_credit_invoice)

            if is_credit:
                inv_type = "kreditine"
            elif _has_amount(doc.vat_amount):
                inv_type = "pvm_saskaita"
            elif doc.separate_vat:
                inv_type = "pvm_saskaita"
            else:
                inv_type = "saskaita"

            doc_pvm_kodas = auto_select_pvm_code(
                pirkimas_pardavimas="pardavimas",
                buyer_country_iso=doc.buyer_country_iso,
                seller_country_iso=doc.seller_country_iso,
                preke_paslauga=doc.preke_paslauga,
                vat_percent=(
                    float(doc.vat_percent)
                    if doc.vat_percent is not None
                    else None
                ),
                separate_vat=bool(doc.separate_vat),
                buyer_has_vat_code=bool(doc.buyer_vat_code),
                seller_has_vat_code=bool(doc.seller_vat_code),
                doc_96_str=bool(getattr(doc, "doc_96_str", False)),
            )

            # Pardavimo korespondencijos
            debit_account = "2410"  # Pirkėjų skolos

            credit_account = (
                getattr(doc, "pardavimo_saskaita", None)
                or "5001"
            )

            pvm_account = "4492" if _has_amount(doc.vat_amount) else None

            entry_date = doc.invoice_date
            period = entry_date.replace(day=1) if entry_date else None

            invoice = Invoice.objects.create(
                user=user,
                company_profile=profile,
                scanned_document=doc,
                invoice_type=inv_type,
                status="issued",

                # Numeracija
                document_series=doc.document_series or "",
                document_number=doc.document_number or "",

                # Datos
                invoice_date=doc.invoice_date,
                due_date=doc.due_date,
                operation_date=doc.operation_date,

                # Korespondencijos
                debeto_saskaita=debit_account,
                kredito_saskaita=credit_account,
                pvm_saskaita=pvm_account,
                period=period,

                # Seller / mes
                seller_name=doc.seller_name or "",
                seller_name_normalized=(doc.seller_name or "").strip().upper(),
                seller_id=doc.seller_id or "",
                seller_vat_code=doc.seller_vat_code or "",
                seller_address=doc.seller_address or "",
                seller_country=doc.seller_country or "",
                seller_country_iso=doc.seller_country_iso or "",
                seller_iban=doc.seller_iban or "",
                seller_is_person=doc.seller_is_person,

                # Buyer / pirkėjas
                buyer_name=doc.buyer_name or "",
                buyer_name_normalized=(doc.buyer_name or "").strip().upper(),
                buyer_id=doc.buyer_id or "",
                buyer_vat_code=doc.buyer_vat_code or "",
                buyer_address=doc.buyer_address or "",
                buyer_country=doc.buyer_country or "",
                buyer_country_iso=doc.buyer_country_iso or "",
                buyer_iban=doc.buyer_iban or "",
                buyer_is_person=doc.buyer_is_person,

                # Sumos
                currency=doc.currency or "EUR",
                pvm_tipas=(
                    "taikoma"
                    if inv_type in ("pvm_saskaita", "kreditine")
                    else "netaikoma"
                ),
                vat_percent=doc.vat_percent,

                amount_wo_vat=_signed_amount(doc.amount_wo_vat, is_credit),
                vat_amount=_signed_amount(doc.vat_amount, is_credit),
                amount_with_vat=_signed_amount(doc.amount_with_vat, is_credit),

                invoice_discount_with_vat=_signed_amount(
                    doc.invoice_discount_with_vat,
                    is_credit,
                ),
                invoice_discount_wo_vat=_signed_amount(
                    doc.invoice_discount_wo_vat,
                    is_credit,
                ),

                separate_vat=doc.separate_vat,
                doc_96_str=doc.doc_96_str,

                # iSAF
                pirkimas_pardavimas="pardavimas",
                report_to_isaf=doc.report_to_isaf,
                document_type_code=doc.document_type_code or "",
                document_type=doc.document_type or "",
                pvm_kodas=doc_pvm_kodas or "",

                # Prekė / sumiškai
                prekes_kodas=doc.prekes_kodas or "",
                prekes_pavadinimas=doc.prekes_pavadinimas or "",
                preke_paslauga=doc.preke_paslauga or "",

                # Meta
                public_link_enabled=False,
            )

            invoice_lines = []
            source_lines = list(doc.line_items.all())

            if source_lines:
                for i, li in enumerate(source_lines):
                    li_vat_pct = (
                        float(li.vat_percent)
                        if li.vat_percent is not None
                        else (
                            float(doc.vat_percent)
                            if doc.vat_percent is not None
                            else None
                        )
                    )

                    li_preke_paslauga = (
                        li.preke_paslauga
                        or doc.preke_paslauga
                    )

                    li_pvm_kodas = auto_select_pvm_code(
                        pirkimas_pardavimas="pardavimas",
                        buyer_country_iso=doc.buyer_country_iso,
                        seller_country_iso=doc.seller_country_iso,
                        preke_paslauga=li_preke_paslauga,
                        vat_percent=li_vat_pct,
                        separate_vat=False,
                        buyer_has_vat_code=bool(doc.buyer_vat_code),
                        seller_has_vat_code=bool(doc.seller_vat_code),
                        doc_96_str=bool(getattr(doc, "doc_96_str", False)),
                    )

                    raw_quantity = (
                        li.quantity
                        if li.quantity is not None
                        else 1
                    )

                    invoice_lines.append(
                        InvoiceLineItem(
                            invoice=invoice,

                            prekes_kodas=li.prekes_kodas or doc.prekes_kodas or "",
                            prekes_barkodas=li.prekes_barkodas or "",
                            prekes_pavadinimas=(
                                li.prekes_pavadinimas
                                or doc.prekes_pavadinimas
                                or "Prekės / paslaugos"
                            ),
                            preke_paslauga=li_preke_paslauga or "",

                            unit=li.unit or "vnt.",

                            # Kreditinei kiekis neigiamas, kaina teigiama
                            quantity=_signed_quantity(raw_quantity, is_credit),
                            price=_line_price(
                                li.price,
                                li.subtotal,
                                raw_quantity,
                            ),

                            subtotal=_signed_amount(li.subtotal, is_credit),
                            vat=_signed_amount(li.vat, is_credit),
                            vat_percent=(
                                li.vat_percent
                                if li.vat_percent is not None
                                else doc.vat_percent
                            ),
                            total=_signed_amount(li.total, is_credit),

                            discount_with_vat=_signed_amount(
                                li.discount_with_vat,
                                is_credit,
                            ),
                            discount_wo_vat=_signed_amount(
                                li.discount_wo_vat,
                                is_credit,
                            ),

                            pvm_kodas=li_pvm_kodas or "",

                            # Korespondencijos line-level
                            kredito_saskaita=credit_account,
                            pvm_saskaita=pvm_account,

                            sort_order=i,
                        )
                    )

            else:
                fallback_name = (
                    doc.prekes_pavadinimas
                    or doc.original_filename
                    or "Prekės / paslaugos"
                )

                fallback_subtotal = _signed_amount(doc.amount_wo_vat, is_credit)
                fallback_vat = _signed_amount(doc.vat_amount, is_credit)
                fallback_total = _signed_amount(doc.amount_with_vat, is_credit)

                invoice_lines.append(
                    InvoiceLineItem(
                        invoice=invoice,

                        prekes_kodas=doc.prekes_kodas or "",
                        prekes_barkodas="",
                        prekes_pavadinimas=fallback_name,
                        preke_paslauga=doc.preke_paslauga or "",

                        unit="vnt.",

                        # Sumiškai:
                        # įprasta SF:  1 x 100 = 100
                        # kreditinė:  -1 x 100 = -100
                        quantity=_signed_quantity(1, is_credit),
                        price=_positive_amount(doc.amount_wo_vat),

                        subtotal=fallback_subtotal,
                        vat=fallback_vat,
                        vat_percent=doc.vat_percent,
                        total=fallback_total,

                        discount_with_vat=_signed_amount(
                            doc.invoice_discount_with_vat,
                            is_credit,
                        ),
                        discount_wo_vat=_signed_amount(
                            doc.invoice_discount_wo_vat,
                            is_credit,
                        ),

                        pvm_kodas=doc_pvm_kodas or "",

                        kredito_saskaita=credit_account,
                        pvm_saskaita=pvm_account,

                        sort_order=0,
                    )
                )

            if invoice_lines:
                InvoiceLineItem.objects.bulk_create(invoice_lines)

            invoice.ready_for_export = check_required_fields_for_export(invoice)

            is_math_valid, _ = validate_document_math_for_export(invoice)
            invoice.math_validation_passed = is_math_valid

            invoice.kor_balanced = compute_kor_balanced(invoice)

            invoice.save(update_fields=[
                "ready_for_export",
                "math_validation_passed",
                "kor_balanced",
            ])

            try:
                from .utils.journal_generators import sync_invoice_journal_entry

                sync_invoice_journal_entry(invoice)
            except Exception as e:
                logger.warning(
                    "[Transfer] sync_invoice_journal_entry failed for %s: %s",
                    invoice.id,
                    e,
                )
            try:
                from .utils.payment_invoice_matching import match_invoice_on_transfer

                alloc = match_invoice_on_transfer(invoice)

                if alloc:
                    logger.info(
                        "[Transfer] Invoice %s auto-matched to bank txn, amount=%s",
                        invoice.id,
                        alloc.amount,
                    )
            except Exception as e:
                logger.warning(
                    "[Transfer] match_invoice_on_transfer failed: %s",
                    e,
                )

            # ── Вариант 1: подтянуть счётчик серии за перенесённым номером ──
            try:
                transferred_int = int(str(invoice.document_number).strip())
            except (ValueError, TypeError):
                transferred_int = None

            if transferred_int is not None:
                series_to_bump = InvoiceSeries.objects.select_for_update().filter(
                    user=user,
                    company_profile=profile,
                    prefix=invoice.document_series,
                    invoice_type=inv_type,
                    is_active=True,
                ).first()

                if series_to_bump and transferred_int >= series_to_bump.next_number:
                    series_to_bump.next_number = transferred_int + 1
                    series_to_bump.save(update_fields=["next_number"])

            created_sales.append(invoice.id)
            _mark_doc_transferred(doc)

    return Response({
        "created_purchases": created_purchases,
        "created_sales": created_sales,
        "skipped": skipped,
    })


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def invoice_update_kor(request, pk):
    """
    PATCH /api/invoicing/invoices/<id>/update-kor/
    Body: { line_item_ids: [1,2,3], kredito_saskaita: "5000" }
    """
    invoice = get_object_or_404(Invoice, pk=pk, user=request.user)

    line_ids = request.data.get("line_item_ids", [])
    new_code = (request.data.get("kredito_saskaita") or "").strip()

    if not new_code or not line_ids:
        return Response({"detail": "line_item_ids ir kredito_saskaita privalomi"}, status=400)

    updated = invoice.line_items.filter(id__in=line_ids).update(kredito_saskaita=new_code)

    # Document level — если все line items одинаковые
    codes = list(invoice.line_items.values_list("kredito_saskaita", flat=True).distinct())
    if len(codes) == 1 and codes[0]:
        invoice.kredito_saskaita = codes[0]
    invoice.kor_balanced = compute_kor_balanced(invoice)
    invoice.save(update_fields=["kredito_saskaita", "kor_balanced", "updated_at"])

    # Regenerate DK
    from .utils.journal_generators import sync_invoice_journal_entry
    try:
        sync_invoice_journal_entry(invoice)
    except Exception:
        pass

    return Response(InvoiceDetailSerializer(invoice, context={"request": request}).data)





# ── Allowed fields ──────────────────────────────────────────
 
PURCHASE_ALLOWED_DOC_FIELDS = {
    "invoice_date", "due_date", "operation_date",
    "document_series", "document_number",
    "amount_wo_vat", "vat_amount", "vat_percent", "amount_with_vat",
    "currency",
    "invoice_discount_wo_vat", "invoice_discount_with_vat",
    "seller_name", "seller_id", "seller_vat_code", "seller_iban",
    "debeto_saskaita", "kredito_saskaita", "pvm_saskaita", "pvm_kodas",
    "order_number", "prekes_kodas", "prekes_pavadinimas", "prekes_barkodas",
}
 
PURCHASE_ALLOWED_LINE_FIELDS = {
    "prekes_kodas", "prekes_pavadinimas", "prekes_barkodas",
    "unit", "quantity", "price", "subtotal", "vat", "vat_percent", "total",
    "debeto_saskaita", "kredito_saskaita", "pvm_saskaita", "pvm_kodas",
}
 
PURCHASE_REQUIRED_FIELDS = {
    "invoice_date", "document_number",
    "amount_wo_vat", "vat_amount", "amount_with_vat",
    "currency", "seller_name",
}
 
PURCHASE_MATH_FIELDS = {
    "amount_wo_vat", "vat_amount", "vat_percent", "amount_with_vat",
    "invoice_discount_wo_vat", "invoice_discount_with_vat",
}
 
PURCHASE_LINE_MATH_FIELDS = {
    "quantity", "price", "subtotal", "vat", "vat_percent", "total",
}
 
 
# ── Inline doc update ───────────────────────────────────────
 
class PurchaseInlineDocUpdateView(APIView):
    permission_classes = [IsAuthenticated]
 
    def patch(self, request, purchase_id):
        purchase = get_object_or_404(
            Purchase, pk=purchase_id, user=request.user,
        )
 
        field = request.data.get("field")
        value = request.data.get("value")
 
        if field not in PURCHASE_ALLOWED_DOC_FIELDS:
            return Response({"detail": "Field not allowed"}, status=400)
 
        if value in ("", None):
            value = None
 
        setattr(purchase, field, value)
        purchase.save(update_fields=[field])
 
        response_data = {
            "ok": True,
            "id": purchase.id,
            field: getattr(purchase, field),
        }
 
        try:
            KOR_BALANCE_FIELDS = {
                "amount_wo_vat", "vat_amount", "amount_with_vat",
                "debeto_saskaita", "kredito_saskaita", "pvm_saskaita",
            }

            if field in PURCHASE_REQUIRED_FIELDS:
                is_valid = check_required_fields_for_export(purchase)
                purchase.ready_for_export = is_valid
                response_data["ready_for_export"] = is_valid

            if field in PURCHASE_MATH_FIELDS:
                is_valid, _ = validate_document_math_for_export(purchase)
                purchase.math_validation_passed = is_valid
                response_data["math_validation_passed"] = is_valid

            if field in KOR_BALANCE_FIELDS:
                purchase.kor_balanced = compute_kor_balanced(purchase)
                response_data["kor_balanced"] = purchase.kor_balanced

            update_fields = []
            if field in PURCHASE_REQUIRED_FIELDS:
                update_fields.append("ready_for_export")
            if field in PURCHASE_MATH_FIELDS:
                update_fields.append("math_validation_passed")
            if field in KOR_BALANCE_FIELDS:
                update_fields.append("kor_balanced")
            if update_fields:
                purchase.save(update_fields=update_fields)
        except Exception as e:
            logger.error(f"Purchase validation error: {e}")

        from .utils.journal_generators import sync_purchase_journal_entry
        try:
            sync_purchase_journal_entry(purchase)
        except Exception as e:
            logger.warning("[Purchase] DK sync failed for %s: %s", purchase.id, e)

        return Response(response_data)


class PurchaseSearchView(APIView):
    """GET /api/purchases/search/?q=telia&limit=10 — для manual match в bank operations."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.db.models import Q

        q = request.query_params.get("q", "").strip()
        limit = min(int(request.query_params.get("limit", 10)), 50)

        qs = Purchase.objects.filter(
            user=request.user,
            payment_status__in=["unpaid", "partially_paid"],
        )
        if q:
            qs = qs.filter(
                Q(seller_name__icontains=q)
                | Q(document_number__icontains=q)
                | Q(document_series__icontains=q)
                | Q(seller_id__icontains=q)
                | Q(seller_iban__icontains=q)
                | Q(seller_name_normalized__icontains=q.upper())
            )
        qs = qs.order_by("-invoice_date")[:limit]

        results = [
            {
                "id": p.id,
                "document_series": p.document_series or "",
                "document_number": p.document_number or "",
                "seller_name": p.seller_name or "",
                "seller_id": p.seller_id or "",
                "amount_with_vat": p.amount_with_vat,
                "invoice_date": p.invoice_date,
                "payment_status": p.payment_status,
            }
            for p in qs
        ]
        return Response({"results": results})
 
 
# ── Inline line update ──────────────────────────────────────
 
class PurchaseInlineLineUpdateView(APIView):
    permission_classes = [IsAuthenticated]
 
    def patch(self, request, purchase_id, line_id):
        purchase = get_object_or_404(
            Purchase, pk=purchase_id, user=request.user,
        )
        line = get_object_or_404(
            PurchaseLine, pk=line_id, purchase=purchase,
        )
 
        field = request.data.get("field")
        value = request.data.get("value")
 
        if field not in PURCHASE_ALLOWED_LINE_FIELDS:
            return Response({"detail": "Field not allowed"}, status=400)
 
        if value in ("", None):
            value = None
 
        setattr(line, field, value)
        line.save(update_fields=[field])
 
        response_data = {
            "ok": True,
            "id": line.id,
            field: getattr(line, field),
        }
 
        try:
            LINE_KOR_BALANCE_FIELDS = {
                "subtotal", "vat", "total",
                "debeto_saskaita", "kredito_saskaita", "pvm_saskaita",
            }

            update_fields = []

            if field in PURCHASE_LINE_MATH_FIELDS:
                is_valid, _ = validate_document_math_for_export(purchase)
                purchase.math_validation_passed = is_valid
                response_data["math_validation_passed"] = is_valid
                update_fields.append("math_validation_passed")

            if field in {"subtotal", "total", "price", "quantity", "vat"}:
                is_valid = check_required_fields_for_export(purchase)
                purchase.ready_for_export = is_valid
                response_data["ready_for_export"] = is_valid
                update_fields.append("ready_for_export")

            if field in LINE_KOR_BALANCE_FIELDS:
                purchase.kor_balanced = compute_kor_balanced(purchase)
                response_data["kor_balanced"] = purchase.kor_balanced
                update_fields.append("kor_balanced")

            if update_fields:
                purchase.save(update_fields=update_fields)
        except Exception as e:
            logger.error(f"Purchase line validation error: {e}")

        from .utils.journal_generators import sync_purchase_journal_entry
        try:
            sync_purchase_journal_entry(purchase)
        except Exception as e:
            logger.warning("[Purchase] DK sync failed for %s: %s", purchase.id, e)

        return Response(response_data)
 
 
# ── Paginated line items ────────────────────────────────────
 
class PurchaseLineItemsListView(APIView):
    permission_classes = [IsAuthenticated]
 
    def get(self, request, purchase_id):
        purchase = get_object_or_404(
            Purchase, pk=purchase_id, user=request.user,
        )
 
        qs = purchase.line_items.all().order_by("sort_order", "id")
 
        limit = min(int(request.query_params.get("limit", 30)), 100)
        offset = max(int(request.query_params.get("offset", 0)), 0)
        count = qs.count()
 
        items = qs[offset: offset + limit]
        serializer = PurchaseLineSerializer(items, many=True)
 
        return Response({
            "count": count,
            "results": serializer.data,
        })
 
# ────────────────────────────────────────────────────────────
# END ─── Pirkimai ───
# ────────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════
# DK
# ═══════════════════════════════════════════════════════════
from django.db.models import Sum, Q, F, DecimalField, Count, Max, ExpressionWrapper, Value
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import date


def _get_active_profile(request):
    """Returns active CompanyProfile or None."""
    company_id = request.query_params.get("company_profile")
    if company_id:
        try:
            return CompanyProfile.objects.get(
                id=company_id, user=request.user
            )
        except CompanyProfile.DoesNotExist:
            return None
    return request.user.active_company_profile


def _parse_period(period_str):
    """'2026-07' → date(2026, 7, 1). Возвращает None если некорректно."""
    if not period_str:
        return None
    try:
        year, month = period_str.split("-")
        return date(int(year), int(month), 1)
    except (ValueError, AttributeError):
        return None


# ═══════════════════════════════════════════════════════════
# TAB 1: SKOLOS — кто кому должен
# ═══════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def apskaita_skolos(request):
    """
    Skolos pagal kontrahentus.

    GET /apskaita/skolos/?type=customer&limit=25&offset=0&search=abc&as_of=2026-07-31
    GET /apskaita/skolos/?type=supplier&limit=25&offset=0&search=abc&as_of=2026-07-31

    Возвращает только открытые долги:
    balance = amount_with_vat - paid_amount > 0
    """
    profile = _get_active_profile(request)
    if not profile:
        return Response({"detail": "Nepasirinktas įmonės profilis."}, status=400)

    debt_type = request.query_params.get("type", "customer")
    if debt_type not in ["customer", "supplier"]:
        return Response(
            {"detail": "Netinkamas skolos tipas. Naudokite customer arba supplier."},
            status=400,
        )

    search = request.query_params.get("search", "").strip()
    as_of = request.query_params.get("as_of")

    try:
        limit = int(request.query_params.get("limit", 25))
    except ValueError:
        limit = 25

    try:
        offset = int(request.query_params.get("offset", 0))
    except ValueError:
        offset = 0

    limit = max(1, min(limit, 100))
    offset = max(0, offset)

    as_of_date = None
    if as_of:
        try:
            year, month, day = as_of.split("-")
            as_of_date = date(int(year), int(month), int(day))
        except (ValueError, AttributeError):
            as_of_date = None

    money_field = DecimalField(max_digits=14, decimal_places=4)
    zero = Value(Decimal("0.00"), output_field=money_field)

    open_balance_expr = ExpressionWrapper(
        Coalesce(F("amount_with_vat"), zero) - Coalesce(F("paid_amount"), zero),
        output_field=money_field,
    )

    if debt_type == "customer":
        qs = Invoice.objects.filter(company_profile=profile)

        date_field = "invoice_date"
        name_field = "buyer_name"
        code_field = "buyer_id"
        counterparty_type = "pirkejas"

        if as_of_date:
            qs = qs.filter(invoice_date__lte=as_of_date)

        if search:
            qs = qs.filter(
                Q(buyer_name__icontains=search) |
                Q(buyer_id__icontains=search)
            )

    else:
        qs = Purchase.objects.filter(company_profile=profile)

        date_field = "invoice_date"
        name_field = "seller_name"
        code_field = "seller_id"
        counterparty_type = "tiekejas"

        if as_of_date:
            qs = qs.filter(invoice_date__lte=as_of_date)

        if search:
            qs = qs.filter(
                Q(seller_name__icontains=search) |
                Q(seller_id__icontains=search)
            )

    # Берём только открытые invoices/purchases
    qs = qs.annotate(open_balance=open_balance_expr).filter(
        open_balance__gt=Decimal("0.009")
    )

    grouped = (
        qs.values(name_field, code_field)
        .annotate(
            total_invoiced=Coalesce(Sum("amount_with_vat"), Decimal("0")),
            total_paid=Coalesce(Sum("paid_amount"), Decimal("0")),
            balance=Coalesce(Sum("open_balance"), Decimal("0")),
            invoice_count=Count("id"),
            newest_invoice_date=Max(date_field),
        )
    )

    rows = []
    total_balance = Decimal("0.00")

    for row in grouped:
        balance = row["balance"] or Decimal("0")
        if balance <= Decimal("0.009"):
            continue

        paid = row["total_paid"] or Decimal("0")
        payment_status = "partially_paid" if paid > 0 else "unpaid"

        newest_date = row["newest_invoice_date"]

        total_balance += balance

        rows.append({
            "counterparty_name": row.get(name_field) or "",
            "counterparty_code": row.get(code_field) or "",
            "counterparty_type": counterparty_type,
            "total_invoiced": str(row["total_invoiced"] or Decimal("0")),
            "total_paid": str(paid),
            "balance": str(balance),
            "invoice_count": row["invoice_count"] or 0,
            "newest_invoice_date": newest_date.isoformat() if newest_date else None,
            "payment_status": payment_status,
        })

    rows.sort(
        key=lambda x: x["newest_invoice_date"] or "",
        reverse=True,
    )

    total_count = len(rows)
    paginated = rows[offset:offset + limit]
    next_offset = offset + limit

    return Response({
        "type": debt_type,
        "results": paginated,
        "total_count": total_count,
        "limit": limit,
        "offset": offset,
        "has_more": next_offset < total_count,
        "next_offset": next_offset if next_offset < total_count else None,
        "summary": {
            "total_balance": str(total_balance),
        },
    })




def _get_document_number(obj):
    """
    Универсально достаём номер документа, потому что в разных моделях
    поле может называться invoice_number / document_number / number.
    """
    return (
        getattr(obj, "invoice_number", None)
        or getattr(obj, "document_number", None)
        or getattr(obj, "number", None)
        or f"#{obj.id}"
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def apskaita_skolos_invoices(request):
    """
    Возвращает открытые sąskaitos выбранного kontrahento.

    GET /apskaita/skolos/invoices/?type=customer&counterparty_code=123&counterparty_name=Client&as_of=2026-07-31
    GET /apskaita/skolos/invoices/?type=supplier&counterparty_code=123&counterparty_name=Telia&as_of=2026-07-31
    """
    profile = _get_active_profile(request)
    if not profile:
        return Response({"detail": "Nepasirinktas įmonės profilis."}, status=400)

    debt_type = request.query_params.get("type", "customer")
    if debt_type not in ["customer", "supplier"]:
        return Response(
            {"detail": "Netinkamas skolos tipas. Naudokite customer arba supplier."},
            status=400,
        )

    counterparty_code = request.query_params.get("counterparty_code", "").strip()
    counterparty_name = request.query_params.get("counterparty_name", "").strip()
    as_of = request.query_params.get("as_of")

    as_of_date = None
    if as_of:
        try:
            year, month, day = as_of.split("-")
            as_of_date = date(int(year), int(month), int(day))
        except (ValueError, AttributeError):
            as_of_date = None

    money_field = DecimalField(max_digits=14, decimal_places=4)
    zero = Value(Decimal("0.00"), output_field=money_field)

    open_balance_expr = ExpressionWrapper(
        Coalesce(F("amount_with_vat"), zero) - Coalesce(F("paid_amount"), zero),
        output_field=money_field,
    )

    if debt_type == "customer":
        qs = Invoice.objects.filter(company_profile=profile)
        source_type = "sale"

        if as_of_date:
            qs = qs.filter(invoice_date__lte=as_of_date)

        if counterparty_code:
            qs = qs.filter(buyer_id=counterparty_code)
        elif counterparty_name:
            qs = qs.filter(buyer_name=counterparty_name)
        else:
            return Response({"detail": "Trūksta kontrahento."}, status=400)

    else:
        qs = Purchase.objects.filter(company_profile=profile)
        source_type = "purchase"

        if as_of_date:
            qs = qs.filter(invoice_date__lte=as_of_date)

        if counterparty_code:
            qs = qs.filter(seller_id=counterparty_code)
        elif counterparty_name:
            qs = qs.filter(seller_name=counterparty_name)
        else:
            return Response({"detail": "Trūksta kontrahento."}, status=400)

    qs = (
        qs.annotate(open_balance=open_balance_expr)
        .filter(open_balance__gt=Decimal("0.009"))
        .order_by("-invoice_date", "-id")
    )

    results = []

    for obj in qs:
        amount = obj.amount_with_vat or Decimal("0")
        paid = obj.paid_amount or Decimal("0")
        balance = amount - paid

        if balance <= Decimal("0.009"):
            continue

        payment_status = "partially_paid" if paid > 0 else "unpaid"

        results.append({
            "id": obj.id,
            "source_type": source_type,
            "document_number": _get_document_number(obj),
            "invoice_date": obj.invoice_date.isoformat() if obj.invoice_date else None,
            "amount_with_vat": str(amount),
            "paid_amount": str(paid),
            "balance": str(balance),
            "payment_status": payment_status,
            "scanned_document_id": getattr(obj, "scanned_document_id", None),
        })

    return Response({
        "type": debt_type,
        "results": results,
    })


# ═══════════════════════════════════════════════════════════
# TAB 2: LIKUČIAI — остатки по sąskaitos
# ═══════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def apskaita_likuciai(request):
    """
    Возвращает остатки по всем sąskaitos с движениями в выбранном периоде.
    Pradinis likutis = все движения до начала периода.
    Debetas/Kreditas per periodą = только выбранный период.
    Galutinis likutis = pradinis + движение.
    """
    profile = _get_active_profile(request)
    if not profile:
        return Response({"detail": "Nepasirinktas įmonės profilis."}, status=400)

    period_str = request.query_params.get("period")
    period_start = _parse_period(period_str)
    if not period_start:
        # По умолчанию — текущий месяц
        today = date.today()
        period_start = date(today.year, today.month, 1)

    # Первый день следующего месяца
    if period_start.month == 12:
        period_end = date(period_start.year + 1, 1, 1)
    else:
        period_end = date(period_start.year, period_start.month + 1, 1)

    # ── Все sąskaitos с движениями (до конца периода) ──
    all_lines = JournalEntryLine.objects.filter(
        entry__company_profile=profile,
        entry__entry_date__lt=period_end,
        entry__status__in=[
            JournalEntry.STATUS_DRAFT,
            JournalEntry.STATUS_POSTED,
            JournalEntry.STATUS_NEEDS_REVIEW,
        ],
    )

    account_codes = (
        all_lines
        .values_list("account_code", flat=True)
        .distinct()
        .order_by("account_code")
    )

    result = []
    for code in account_codes:
        # Pradinis likutis: движения ДО начала периода
        opening_lines = all_lines.filter(
            account_code=code,
            entry__entry_date__lt=period_start,
        )
        opening_d = opening_lines.filter(side="D").aggregate(
            s=Coalesce(Sum("amount"), Decimal("0"))
        )["s"]
        opening_k = opening_lines.filter(side="K").aggregate(
            s=Coalesce(Sum("amount"), Decimal("0"))
        )["s"]
        opening_balance = opening_d - opening_k

        # Движения в периоде
        period_lines = all_lines.filter(
            account_code=code,
            entry__entry_date__gte=period_start,
            entry__entry_date__lt=period_end,
        )
        period_d = period_lines.filter(side="D").aggregate(
            s=Coalesce(Sum("amount"), Decimal("0"))
        )["s"]
        period_k = period_lines.filter(side="K").aggregate(
            s=Coalesce(Sum("amount"), Decimal("0"))
        )["s"]

        closing_balance = opening_balance + period_d - period_k

        # Пропускаем sąskaitos без движений и без остатков
        if (opening_balance == 0 and period_d == 0 and period_k == 0
                and closing_balance == 0):
            continue

        # Название
        first_line = all_lines.filter(account_code=code).first()
        account_name = first_line.account_name if first_line else ""

        result.append({
            "code": code,
            "name": account_name,
            "opening_balance": str(opening_balance),
            "opening_side": "D" if opening_balance >= 0 else "K",
            "period_debit": str(period_d),
            "period_credit": str(period_k),
            "closing_balance": str(abs(closing_balance)),
            "closing_side": "D" if closing_balance >= 0 else "K",
        })

    return Response({
        "period": period_start.strftime("%Y-%m"),
        "accounts": result,
    })


# ═══════════════════════════════════════════════════════════
# TAB 3: OPERACIJOS — DK įrašai (журнал операций)
# ═══════════════════════════════════════════════════════════

MANUAL_DK_DEBT_CODES = {"2080", "2410", "4430"}
MANUAL_DK_TOLERANCE = Decimal("0.01")
MANUAL_DK_MONEY = Decimal("0.01")


def _next_manual_dk_number(profile):
    """
    Возвращает следующий номер:
    RDK-00001, RDK-00002...
    """
    prefix = "RDK-"
    max_number = 0

    numbers = JournalEntry.objects.filter(
        company_profile=profile,
        source_type=JournalEntry.SOURCE_MANUAL,
        document_number__startswith=prefix,
    ).values_list("document_number", flat=True)

    for document_number in numbers:
        suffix = str(document_number or "")[len(prefix):]

        if suffix.isdigit():
            max_number = max(
                max_number,
                int(suffix),
            )

    return f"{prefix}{max_number + 1:05d}"


def _parse_manual_dk_payload(payload):
    """
    Tikrina ir normalizuoja rankinio DK įrašo duomenis.
    Aprašymas nėra privalomas.
    """
    from .utils.journal_generators import _get_account_name

    entry_date_raw = str(
        payload.get("entry_date") or ""
    ).strip()

    try:
        entry_date = date.fromisoformat(
            entry_date_raw
        )
    except (TypeError, ValueError):
        return None, "Nurodykite tinkamą datą."

    description = str(
        payload.get("description") or ""
    ).strip()[:255]

    counterparty_name = str(
        payload.get("counterparty_name") or ""
    ).strip()[:255]

    counterparty_code = str(
        payload.get("counterparty_code") or ""
    ).strip()[:50]

    counterparty_vat_code = str(
        payload.get("counterparty_vat_code") or ""
    ).strip()[:32]

    raw_lines = payload.get("lines")

    if not isinstance(raw_lines, list):
        return None, "DK eilutės turi būti sąrašas."

    if len(raw_lines) < 2:
        return None, "Reikia bent dviejų DK eilučių."

    parsed_lines = []
    total_debit = Decimal("0")
    total_credit = Decimal("0")
    has_debit = False
    has_credit = False

    for index, raw_line in enumerate(
        raw_lines,
        start=1,
    ):
        if not isinstance(raw_line, dict):
            return (
                None,
                f"Netinkama {index} DK eilutė.",
            )

        side = str(
            raw_line.get("side") or ""
        ).strip().upper()

        if side not in ("D", "K"):
            return (
                None,
                f"{index} eilutėje pasirinkite D arba K.",
            )

        account_code = str(
            raw_line.get("account_code") or ""
        ).strip()

        if not account_code:
            return (
                None,
                f"{index} eilutėje pasirinkite sąskaitą.",
            )

        if len(account_code) > 20:
            return (
                None,
                f"{index} eilutės sąskaitos kodas per ilgas.",
            )

        account_name = _get_account_name(
            account_code
        )

        if not account_name:
            return (
                None,
                f"Nežinoma sąskaita: {account_code}.",
            )

        raw_amount = str(
            raw_line.get("amount") or ""
        ).strip().replace(",", ".")

        try:
            amount = Decimal(
                raw_amount
            ).quantize(
                MANUAL_DK_MONEY,
                rounding=ROUND_HALF_UP,
            )
        except (
            InvalidOperation,
            TypeError,
            ValueError,
        ):
            return (
                None,
                f"{index} eilutėje nurodykite tinkamą sumą.",
            )

        if amount <= Decimal("0"):
            return (
                None,
                f"{index} eilutės suma turi būti didesnė už nulį.",
            )

        line_description = str(
            raw_line.get("description") or ""
        ).strip()[:255]

        parsed_lines.append({
            "side": side,
            "account_code": account_code,
            "account_name": account_name,
            "amount": amount,
            "description": line_description,
            "sort_order": index - 1,
        })

        if side == "D":
            total_debit += amount
            has_debit = True
        else:
            total_credit += amount
            has_credit = True

    requires_counterparty = any(
        line["account_code"] in MANUAL_DK_DEBT_CODES
        for line in parsed_lines
    )

    if requires_counterparty and not counterparty_name:
        return (
            None,
            "Pasirinkite arba įveskite kontrahentą, nes naudojama skolų sąskaita.",
        )

    if not has_debit:
        return None, "Reikia bent vienos debeto eilutės."

    if not has_credit:
        return None, "Reikia bent vienos kredito eilutės."

    difference = total_debit - total_credit

    if abs(difference) > MANUAL_DK_TOLERANCE:
        return (
            None,
            (
                "DK įrašas nesubalansuotas. "
                f"Skirtumas: {abs(difference):.2f} EUR."
            ),
        )

    return {
        "entry_date": entry_date,
        "period": entry_date.replace(day=1),
        "description": description,
        "counterparty_name": counterparty_name,
        "counterparty_code": counterparty_code,
        "counterparty_vat_code": counterparty_vat_code,
        "lines": parsed_lines,
        "total_debit": total_debit,
        "total_credit": total_credit,
    }, None


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def manual_dk_company_search(request):
    query = str(request.query_params.get("q") or "").strip()[:100]

    if len(query) < 2:
        return Response({"results": []})

    normalized_query = " ".join(query.upper().split())

    companies = (
        Company.objects
        .filter(
            Q(pavadinimas__icontains=query)
            | Q(normalized_pavadinimas__icontains=normalized_query)
            | Q(im_kodas__icontains=query)
            | Q(pvm_kodas__icontains=query)
        )
        .order_by("pavadinimas", "im_kodas")[:20]
    )

    return Response({
        "results": [
            {
                "id": company.id,
                "pavadinimas": company.pavadinimas or "",
                "im_kodas": company.im_kodas or "",
                "pvm_kodas": company.pvm_kodas or "",
                "adresas": company.adresas or "",
            }
            for company in companies
        ]
    })

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def manual_dk_collection(request):
    """
    GET:
      grąžina kitą RDK numerį.

    POST:
      sukuria iš karto užregistruotą rankinį DK įrašą.
    """
    from .utils.journal_generators import (
        finalize_journal_entry,
    )

    profile = _get_active_profile(request)

    if not profile:
        return Response(
            {"detail": "Nepasirinktas profilis."},
            status=400,
        )

    if request.method == "GET":
        return Response({
            "next_number":
                _next_manual_dk_number(profile),
        })

    parsed, error = _parse_manual_dk_payload(
        request.data
    )

    if error:
        return Response(
            {"detail": error},
            status=400,
        )

    with transaction.atomic():
        # Серилизуем выдачу RDK-номеров внутри профиля.
        type(profile).objects.select_for_update().get(
            pk=profile.pk
        )

        document_number = _next_manual_dk_number(
            profile
        )

        entry = JournalEntry.objects.create(
            user=request.user,
            company_profile=profile,
            source_type=JournalEntry.SOURCE_MANUAL,
            entry_date=parsed["entry_date"],
            period=parsed["period"],
            document_number=document_number,
            counterparty_name=parsed["counterparty_name"],
            counterparty_code=parsed["counterparty_code"],
            counterparty_vat_code=parsed["counterparty_vat_code"],
            description=parsed["description"],
            currency="EUR",
            status=JournalEntry.STATUS_POSTED,
        )

        JournalEntryLine.objects.bulk_create([
            JournalEntryLine(
                entry=entry,
                side=line["side"],
                account_code=line["account_code"],
                account_name=line["account_name"],
                amount=line["amount"],
                description=line["description"],
                sort_order=line["sort_order"],
            )
            for line in parsed["lines"]
        ])

        finalize_journal_entry(entry)

        entry.refresh_from_db()

    serializer = JournalEntrySerializer(
        entry,
        context={"request": request},
    )

    return Response(
        serializer.data,
        status=201,
    )


@api_view(["GET", "PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def manual_dk_detail(request, pk):
    """
    GET:
      rankinio DK įrašo duomenys.

    PUT:
      atnaujina visą rankinį DK įrašą.

    DELETE:
      ištrina rankinį DK įrašą.
    """
    from .utils.journal_generators import (
        finalize_journal_entry,
    )

    profile = _get_active_profile(request)

    if not profile:
        return Response(
            {"detail": "Nepasirinktas profilis."},
            status=400,
        )

    base_queryset = JournalEntry.objects.filter(
        pk=pk,
        company_profile=profile,
        source_type=JournalEntry.SOURCE_MANUAL,
    )

    if request.method == "GET":
        entry = (
            base_queryset
            .prefetch_related("lines")
            .first()
        )

        if not entry:
            return Response(
                {"detail": "Rankinis DK įrašas nerastas."},
                status=404,
            )

        serializer = JournalEntrySerializer(
            entry,
            context={"request": request},
        )

        return Response(serializer.data)

    if request.method == "DELETE":
        with transaction.atomic():
            entry = (
                base_queryset
                .select_for_update()
                .first()
            )

            if not entry:
                return Response(
                    {
                        "detail":
                        "Rankinis DK įrašas nerastas."
                    },
                    status=404,
                )

            entry.delete()

        return Response(status=204)

    parsed, error = _parse_manual_dk_payload(
        request.data
    )

    if error:
        return Response(
            {"detail": error},
            status=400,
        )

    with transaction.atomic():
        entry = (
            base_queryset
            .select_for_update()
            .first()
        )

        if not entry:
            return Response(
                {"detail": "Rankinis DK įrašas nerastas."},
                status=404,
            )

        entry.entry_date = parsed["entry_date"]
        entry.period = parsed["period"]
        entry.description = parsed["description"]
        entry.counterparty_name = parsed["counterparty_name"]
        entry.counterparty_code = parsed["counterparty_code"]
        entry.counterparty_vat_code = parsed["counterparty_vat_code"]
        entry.currency = "EUR"
        entry.status = JournalEntry.STATUS_POSTED

        entry.save(update_fields=[
            "entry_date",
            "period",
            "description",
            "counterparty_name",
            "counterparty_code",
            "counterparty_vat_code",
            "currency",
            "status",
            "updated_at",
        ])

        entry.lines.all().delete()

        JournalEntryLine.objects.bulk_create([
            JournalEntryLine(
                entry=entry,
                side=line["side"],
                account_code=line["account_code"],
                account_name=line["account_name"],
                amount=line["amount"],
                description=line["description"],
                sort_order=line["sort_order"],
            )
            for line in parsed["lines"]
        ])

        finalize_journal_entry(entry)

        entry.refresh_from_db()

    serializer = JournalEntrySerializer(
        entry,
        context={"request": request},
    )

    return Response(serializer.data)

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def patch_dk_line(request, pk):
    """
    PATCH /api/apskaita/dk-eilutes/<id>/

    Pakeičia leidžiamą DK sąskaitą.
    Pirkimo / pardavimo atveju taip pat atnaujina
    korespondenciją šaltinio dokumente.
    """
    from django.db import transaction
    from django.utils import timezone

    from .utils.journal_generators import (
        finalize_journal_entry,
        _get_account_name,
    )
    from .serializers import (
        LOCKED_CODES,
        LOCKED_PREFIXES,
    )

    profile = _get_active_profile(request)

    if not profile:
        return Response(
            {"detail": "Nepasirinktas profilis."},
            status=400,
        )

    new_code = (
        request.data.get("account_code") or ""
    ).strip()

    if not new_code:
        return Response(
            {"detail": "Sąskaitos kodas privalomas."},
            status=400,
        )

    with transaction.atomic():
        line = (
            JournalEntryLine.objects
            .select_for_update(of=("self",))
            .select_related(
                "entry",
                "entry__purchase",
                "entry__invoice",
            )
            .filter(
                entry__company_profile=profile,
                id=pk,
            )
            .first()
        )

        if not line:
            return Response(
                {"detail": "Nerasta."},
                status=404,
            )

        if line.entry.source_type == JournalEntry.SOURCE_MANUAL:
            return Response(
                {
                    "detail":
                    "Rankinis DK įrašas redaguojamas visas dialogo lange."
                },
                status=400,
            )

        old_code = (
            line.account_code or ""
        ).strip()

        # Paliekame dabartinį leidimų principą:
        # užrakintų eilučių redaguoti negalima.
        old_code_locked = (
            old_code in LOCKED_CODES
            or any(
                old_code.startswith(prefix)
                for prefix in LOCKED_PREFIXES
            )
        )

        if old_code_locked:
            return Response(
                {
                    "detail":
                    "Šios sąskaitos keisti negalima."
                },
                status=400,
            )

        # Neleidžiame redaguojamos eilutės pakeisti
        # į sisteminę / užrakintą sąskaitą.
        new_code_locked = (
            new_code in LOCKED_CODES
            or any(
                new_code.startswith(prefix)
                for prefix in LOCKED_PREFIXES
            )
        )

        if new_code_locked:
            return Response(
                {
                    "detail":
                    "Šios sąskaitos pasirinkti negalima."
                },
                status=400,
            )

        if new_code == old_code:
            return Response({
                "status": "unchanged",
                "account_code": old_code,
                "account_name": line.account_name,
                "is_user_modified":
                    line.is_user_modified,
            })

        entry = line.entry

        # ─────────────────────────────────────────────
        # PIRKIMAS
        # Redaguojama tik sąnaudų / turto DK eilutė.
        # ─────────────────────────────────────────────
        if entry.source_type == JournalEntry.SOURCE_PURCHASE:
            purchase = entry.purchase

            if not purchase:
                return Response(
                    {
                        "detail":
                        "Nerastas susietas pirkimo dokumentas."
                    },
                    status=409,
                )

            purchase_items = list(
                purchase.line_items
                .select_for_update()
                .all()
            )

            if purchase_items:
                changed_items = []

                for item in purchase_items:
                    effective_code = str(
                        item.effective_debeto or "6312"
                    ).strip()

                    if effective_code == old_code:
                        item.debeto_saskaita = new_code
                        changed_items.append(item)

                if not changed_items:
                    return Response(
                        {
                            "detail":
                            "Nepavyko rasti susietų "
                            "pirkimo eilučių."
                        },
                        status=409,
                    )

                purchase.line_items.model.objects.bulk_update(
                    changed_items,
                    ["debeto_saskaita"],
                )

                type(purchase).objects.filter(
                    pk=purchase.pk
                ).update(
                    updated_at=timezone.now(),
                )

            else:
                type(purchase).objects.filter(
                    pk=purchase.pk
                ).update(
                    debeto_saskaita=new_code,
                    updated_at=timezone.now(),
                )

        # ─────────────────────────────────────────────
        # PARDAVIMAS
        # Redaguojama tik pajamų DK eilutė.
        # ─────────────────────────────────────────────
        elif entry.source_type == JournalEntry.SOURCE_SALE:
            invoice = entry.invoice

            if not invoice:
                return Response(
                    {
                        "detail":
                        "Nerastas susietas pardavimo dokumentas."
                    },
                    status=409,
                )

            invoice_items = list(
                invoice.line_items
                .select_for_update()
                .all()
            )

            if invoice_items:
                changed_items = []

                for item in invoice_items:
                    supply_kind = str(
                        item.preke_paslauga or ""
                    ).strip().lower()

                    default_code = (
                        "5000"
                        if supply_kind in (
                            "preke",
                            "1",
                            "3",
                        )
                        else "5001"
                    )

                    effective_code = str(
                        item.kredito_saskaita
                        or default_code
                    ).strip()

                    if effective_code == old_code:
                        item.kredito_saskaita = new_code
                        changed_items.append(item)

                if not changed_items:
                    return Response(
                        {
                            "detail":
                            "Nepavyko rasti susietų "
                            "pardavimo eilučių."
                        },
                        status=409,
                    )

                invoice.line_items.model.objects.bulk_update(
                    changed_items,
                    ["kredito_saskaita"],
                )

                type(invoice).objects.filter(
                    pk=invoice.pk
                ).update(
                    updated_at=timezone.now(),
                )

            else:
                type(invoice).objects.filter(
                    pk=invoice.pk
                ).update(
                    kredito_saskaita=new_code,
                    updated_at=timezone.now(),
                )

        # Banko ir rankinio DK atveju šaltinio
        # dokumento korespondencijos neatnaujiname.

        new_account_name = _get_account_name(
            new_code
        )

        updated = (
            JournalEntryLine.objects
            .filter(pk=line.pk)
            .update(
                account_code=new_code,
                account_name=new_account_name,
                is_user_modified=True,
            )
        )

        if not updated:
            return Response(
                {
                    "detail":
                    "Įrašo nebėra – atnaujinkite puslapį."
                },
                status=409,
            )

        line.account_code = new_code
        line.account_name = new_account_name
        line.is_user_modified = True

        finalize_journal_entry(entry)

    return Response({
        "status": "updated",
        "account_code": new_code,
        "account_name": new_account_name,
        "is_user_modified": True,
    })

from .pagination import DkPagination


class OperacijosViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = JournalEntrySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = DkPagination

    def get_queryset(self):
        profile = _get_active_profile(self.request)
        if not profile:
            return JournalEntry.objects.none()

        qs = JournalEntry.objects.filter(
            company_profile=profile,
        ).select_related(
            "purchase", "purchase__scanned_document",
            "invoice", "invoice__scanned_document",
        ).prefetch_related("lines")

        # Filter by period
        period_str = self.request.query_params.get("period")
        date_from_str = self.request.query_params.get("date_from")
        date_to_str = self.request.query_params.get("date_to")

        period_start = _parse_period(period_str)
        if period_start:
            if period_start.month == 12:
                period_end = date(period_start.year + 1, 1, 1)
            else:
                period_end = date(period_start.year, period_start.month + 1, 1)
            qs = qs.filter(
                entry_date__gte=period_start,
                entry_date__lt=period_end,
            )
        else:
            if date_from_str:
                try:
                    qs = qs.filter(entry_date__gte=date.fromisoformat(date_from_str))
                except ValueError:
                    pass
            if date_to_str:
                try:
                    qs = qs.filter(entry_date__lte=date.fromisoformat(date_to_str))
                except ValueError:
                    pass

        # Filter by source_type
        source_type = self.request.query_params.get("source_type")
        if source_type:
            qs = qs.filter(source_type=source_type)

        # Filter by counterparty
        counterparty = self.request.query_params.get("counterparty", "").strip()
        if counterparty:
            qs = qs.filter(
                Q(counterparty_name__icontains=counterparty) |
                Q(counterparty_code__icontains=counterparty)
            )

        # Filter by account (показать только операции затрагивающие эту sąskaitą)
        account_code = self.request.query_params.get("account_code")
        if account_code:
            qs = qs.filter(lines__account_code=account_code).distinct()

        # Filter by status
        status_param = self.request.query_params.get("status")
        if status_param:
            qs = qs.filter(status=status_param)

        # Only problems (unbalanced/needs_review)
        only_problems = self.request.query_params.get("only_problems", "").lower() == "true"
        if only_problems:
            qs = qs.filter(status__in=[
                JournalEntry.STATUS_UNBALANCED,
                JournalEntry.STATUS_NEEDS_REVIEW,
            ])

        return qs.order_by("-entry_date", "-id")

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)


# ═══════════════════════════════════════════════════════════
# DASHBOARD CARDS — цифры для верха страницы Apskaita
# ═══════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def apskaita_summary_cards(request):
    """Возвращает цифры для карточек наверху страницы Apskaita."""
    profile = _get_active_profile(request)
    if not profile:
        return Response({"detail": "Nepasirinktas įmonės profilis."}, status=400)

    period_str = request.query_params.get("period")
    date_from_str = request.query_params.get("date_from")
    date_to_str = request.query_params.get("date_to")

    period_start = _parse_period(period_str)
    if period_start:
        if period_start.month == 12:
            period_end = date(period_start.year + 1, 1, 1)
        else:
            period_end = date(period_start.year, period_start.month + 1, 1)
    elif date_from_str or date_to_str:
        try:
            period_start = date.fromisoformat(date_from_str) if date_from_str else date(2000, 1, 1)
        except ValueError:
            period_start = date(2000, 1, 1)
        try:
            d = date.fromisoformat(date_to_str) if date_to_str else date.today()
            period_end = d + timedelta(days=1)
        except ValueError:
            period_end = date.today() + timedelta(days=1)
    else:
        today = date.today()
        period_start = date(today.year, today.month, 1)
        if today.month == 12:
            period_end = date(today.year + 1, 1, 1)
        else:
            period_end = date(today.year, today.month + 1, 1)

    # Customer debt (2410 balance)
    customer_debt = Decimal("0")
    for p in Purchase.objects.filter(company_profile=profile):
        customer_debt += (p.amount_with_vat or Decimal("0")) - (p.paid_amount or Decimal("0"))
    customer_debt = Decimal("0")
    supplier_debt = Decimal("0")

    for inv in Invoice.objects.filter(company_profile=profile):
        balance = (inv.amount_with_vat or Decimal("0")) - (inv.paid_amount or Decimal("0"))
        if balance > 0:
            customer_debt += balance

    for p in Purchase.objects.filter(company_profile=profile):
        balance = (p.amount_with_vat or Decimal("0")) - (p.paid_amount or Decimal("0"))
        if balance > 0:
            supplier_debt += balance

    # PVM per periodą
    period_lines = JournalEntryLine.objects.filter(
        entry__company_profile=profile,
        entry__entry_date__gte=period_start,
        entry__entry_date__lt=period_end,
    )
    vat_receivable = period_lines.filter(
        account_code="2441", side="D",
    ).aggregate(s=Coalesce(Sum("amount"), Decimal("0")))["s"]
    vat_payable = period_lines.filter(
        account_code="4492", side="K",
    ).aggregate(s=Coalesce(Sum("amount"), Decimal("0")))["s"]
    vat_net = vat_payable - vat_receivable

    # DK problems
    unbalanced_count = JournalEntry.objects.filter(
        company_profile=profile,
        status=JournalEntry.STATUS_UNBALANCED,
    ).count()

    return Response({
        "period": period_start.strftime("%Y-%m"),
        "customer_debt": str(customer_debt),
        "supplier_debt": str(supplier_debt),
        "vat_receivable": str(vat_receivable),
        "vat_payable": str(vat_payable),
        "vat_net": str(vat_net),
        "unbalanced_entries": unbalanced_count,
    })

# ═══════════════════════════════════════════════════════════
# END - DK
# ═══════════════════════════════════════════════════════════
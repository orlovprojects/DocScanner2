"""
Парсеры банковских выписок.

Каждый парсер возвращает list[dict] с унифицированными полями:
  transaction_date, value_date, doc_number, bank_operation_code,
  counterparty_name, counterparty_code, counterparty_account,
  payment_purpose, reference_number, amount (positive), currency, direction
"""

import csv
import io
import logging
import re
from abc import ABC, abstractmethod
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import BinaryIO, Optional, Union
from xml.etree import ElementTree as ET

logger = logging.getLogger("docscanner_app")


class BaseBankParser(ABC):
    bank_name: str = ""

    @abstractmethod
    def parse(self, file_content: Union[bytes, BinaryIO]) -> list[dict]:
        pass

    def _to_bytes(self, file_content) -> bytes:
        if isinstance(file_content, bytes):
            return file_content
        if hasattr(file_content, "read"):
            return file_content.read()
        return bytes(file_content)

    def _parse_date(self, date_str: str) -> Optional[date]:
        if not date_str:
            return None
        date_str = date_str.strip()
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%d.%m.%Y", "%d/%m/%Y", "%Y%m%d"):
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_amount(self, amount_str: str) -> Optional[Decimal]:
        if not amount_str:
            return None
        amount_str = amount_str.strip().replace("\xa0", "").replace(" ", "")
        if "," in amount_str and "." in amount_str:
            amount_str = amount_str.replace(".", "").replace(",", ".")
        elif "," in amount_str:
            amount_str = amount_str.replace(",", ".")
        try:
            return Decimal(amount_str)
        except InvalidOperation:
            return None

    def _detect_encoding(self, raw: bytes) -> str:
        for enc in ("utf-8-sig", "utf-8", "windows-1257", "iso-8859-13", "latin-1"):
            try:
                raw.decode(enc)
                return enc
            except (UnicodeDecodeError, LookupError):
                continue
        return "utf-8"

    def _extract_metadata(self, transactions: list[dict]) -> dict:
        dates = [t["transaction_date"] for t in transactions if t.get("transaction_date")]
        return {
            "period_from": min(dates) if dates else None,
            "period_to": max(dates) if dates else None,
        }

    def _detect_separator(self, text: str) -> str:
        """
        Detect CSV separator by checking which delimiter produces real header columns.

        Important:
        SEB CSV uses semicolon, but amounts/descriptions contain commas:
        562,78
        FACEBK...,fb.me/ads,IE

        So raw comma counting is unreliable.
        """
        lines = [line for line in text.splitlines()[:50] if line.strip()]
        if not lines:
            return ";"

        candidates = [";", ",", "\t"]

        header_keywords = [
            "data",
            "suma",
            "valiuta",
            "dok",
            "paskirtis",
            "gavėjo",
            "gavejo",
            "mokėtojo",
            "moketojo",
            "debetas",
            "kreditas",
            "sąskaita",
            "saskaita",
            "description",
            "amount",
            "currency",
            "started date",
            "completed date",
        ]

        scores = {}

        for sep in candidates:
            best_score = 0
            best_cols = 0
            best_hits = 0

            for line in lines:
                try:
                    parsed = next(csv.reader([line], delimiter=sep, quotechar='"'))
                except Exception:
                    continue

                cols = [c.strip().lower() for c in parsed if c.strip()]
                if len(cols) <= 1:
                    continue

                joined = " ".join(cols)
                hits = sum(1 for kw in header_keywords if kw in joined)

                # Header-like row with many real columns should win.
                # hits are more important than raw column count.
                score = hits * 100 + len(cols)

                if score > best_score:
                    best_score = score
                    best_cols = len(cols)
                    best_hits = hits

            scores[sep] = {
                "score": best_score,
                "cols": best_cols,
                "hits": best_hits,
            }

        best_sep = max(scores, key=lambda s: scores[s]["score"])

        if scores[best_sep]["score"] <= 1:
            best_sep = ";"

        logger.info(
            "[Parser] Separator detection scores: %s → chose %r",
            scores,
            best_sep,
        )

        return best_sep


# ────────────────────────────────────────────────────────────
# Swedbank CSV
# ────────────────────────────────────────────────────────────


class SwedbankCSVParser(BaseBankParser):
    bank_name = "swedbank"

    def parse(self, file_content) -> list[dict]:
        raw = self._to_bytes(file_content)
        encoding = self._detect_encoding(raw)
        text = raw.decode(encoding)

        logger.info("[SwedbankCSV] Encoding: %s, length: %d", encoding, len(text))

        delimiter = self._detect_separator(text)
        logger.info("[SwedbankCSV] Using delimiter: %s", repr(delimiter))

        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        logger.info("[SwedbankCSV] Total rows: %d", len(rows))

        if not rows:
            logger.warning("[SwedbankCSV] No rows found")
            return []

        for i, row in enumerate(rows[:3]):
            logger.info("[SwedbankCSV] Row %d (%d cols): %s", i, len(row), row[:5])

        header_idx = None
        for i, row in enumerate(rows):
            joined = ";".join(row).lower()
            if "operacijos data" in joined or "dok" in joined:
                header_idx = i
                break

        if header_idx is None:
            logger.warning("[SwedbankCSV] Could not find header row")
            return []

        logger.info("[SwedbankCSV] Header at row %d: %s", header_idx, rows[header_idx][:8])

        headers = [h.strip().lower() for h in rows[header_idx]]
        col_map = self._map_columns(headers)
        logger.info("[SwedbankCSV] Column mapping: %s", col_map)

        transactions = []
        skipped = 0
        for row in rows[header_idx + 1:]:
            if len(row) < 5:
                skipped += 1
                continue
            txn = self._parse_row(row, col_map)
            if txn and txn.get("amount") and txn.get("transaction_date"):
                transactions.append(txn)
            else:
                skipped += 1

        logger.info(
            "[SwedbankCSV] Result: %d transactions, %d skipped",
            len(transactions), skipped,
        )
        return transactions

    def _map_columns(self, headers):
        mapping = {}
        patterns = {
            "transaction_date": ["operacijos data", "data"],
            "value_date": ["knyg. data", "knygavimo data"],
            "doc_number": ["dok. nr", "dok nr"],
            "bank_operation_code": ["mokėjimo kodas", "mokejimo kodas", "banko žyma"],
            "counterparty_name": ["gavėjas/mokėtojas", "gavejas/moketojas"],
            "counterparty_account": ["gavėjo/mokėtojo sąskaita", "gavejos saskaita"],
            "counterparty_code": ["įmonės kodas", "imones kodas"],
            "amount": ["suma"],
            "direction_flag": ["d/k"],
            "payment_purpose": ["operacijos paskirtis", "paskirtis"],
            "currency": ["valiuta"],
        }
        for field, terms in patterns.items():
            for i, h in enumerate(headers):
                if any(t in h for t in terms):
                    mapping[field] = i
                    break
        return mapping

    def _parse_row(self, row, col_map):
        def get(f):
            idx = col_map.get(f)
            return row[idx].strip() if idx is not None and idx < len(row) else ""

        amount = self._parse_amount(get("amount"))
        if not amount:
            return None

        dk = get("direction_flag").upper()
        direction = "credit" if dk == "K" else "debit" if dk == "D" else ("credit" if amount > 0 else "debit")

        return {
            "transaction_date": self._parse_date(get("transaction_date")),
            "value_date": self._parse_date(get("value_date")),
            "doc_number": get("doc_number"),
            "bank_operation_code": get("bank_operation_code"),
            "counterparty_name": get("counterparty_name"),
            "counterparty_code": get("counterparty_code"),
            "counterparty_account": get("counterparty_account"),
            "payment_purpose": get("payment_purpose"),
            "reference_number": "",
            "amount": abs(amount),
            "currency": get("currency") or "EUR",
            "direction": direction,
        }


# ────────────────────────────────────────────────────────────
# ISO 20022 camt.053 XML (Swedbank, SEB, Luminor)
# ────────────────────────────────────────────────────────────


class ISO20022Parser(BaseBankParser):
    bank_name = "iso20022"

    def parse(self, file_content) -> list[dict]:
        raw = self._to_bytes(file_content)
        text = raw.decode(self._detect_encoding(raw))

        logger.info("[ISO20022] File length: %d", len(text))

        root = ET.fromstring(text)
        ns = self._detect_ns(root)
        logger.info("[ISO20022] Namespace: %s", ns)

        if not ns:
            logger.warning("[ISO20022] No namespace detected")
            return []

        transactions = []
        stmt_count = 0
        for stmt in root.iter(f"{{{ns}}}Stmt"):
            stmt_count += 1
            for ntry in stmt.iter(f"{{{ns}}}Ntry"):
                txn = self._parse_entry(ntry, ns)
                if txn:
                    transactions.append(txn)

        logger.info(
            "[ISO20022] Statements: %d, Transactions: %d",
            stmt_count, len(transactions),
        )
        return transactions

    def _detect_ns(self, root):
        tag = root.tag
        if "{" in tag:
            return tag.split("}")[0].lstrip("{")
        return None

    def _txt(self, el, path, default=""):
        node = el.find(path)
        return node.text.strip() if node is not None and node.text else default

    def _parse_entry(self, ntry, ns):
        amt_el = ntry.find(f"{{{ns}}}Amt")
        if amt_el is None:
            return None
        amount = self._parse_amount(amt_el.text)
        if not amount:
            return None

        cdi = self._txt(ntry, f"{{{ns}}}CdtDbtInd")
        direction = "credit" if cdi == "CRDT" else "debit"

        booking = self._txt(ntry, f"{{{ns}}}BookgDt/{{{ns}}}Dt")
        val_date = self._txt(ntry, f"{{{ns}}}ValDt/{{{ns}}}Dt")

        dtls = ntry.find(f".//{{{ns}}}NtryDtls/{{{ns}}}TxDtls")
        cp_name = cp_code = cp_acct = purpose = ref = doc = ""

        if dtls is not None:
            party_key = "Dbtr" if direction == "credit" else "Cdtr"
            acct_key = f"{party_key}Acct"

            party = dtls.find(f"{{{ns}}}RltdPties/{{{ns}}}{party_key}")
            if party is not None:
                cp_name = self._txt(party, f"{{{ns}}}Nm")
                org = party.find(f"{{{ns}}}Id/{{{ns}}}OrgId/{{{ns}}}Othr/{{{ns}}}Id")
                if org is not None and org.text:
                    cp_code = org.text.strip()

            cp_acct = self._txt(dtls, f"{{{ns}}}RltdPties/{{{ns}}}{acct_key}/{{{ns}}}Id/{{{ns}}}IBAN")

            ustrd = dtls.find(f"{{{ns}}}RmtInf/{{{ns}}}Ustrd")
            if ustrd is not None and ustrd.text:
                purpose = ustrd.text.strip()

            ref_el = dtls.find(f"{{{ns}}}Refs/{{{ns}}}EndToEndId")
            if ref_el is not None and ref_el.text and ref_el.text != "NOTPROVIDED":
                ref = ref_el.text.strip()

            doc_el = dtls.find(f"{{{ns}}}Refs/{{{ns}}}AcctSvcrRef")
            if doc_el is not None and doc_el.text:
                doc = doc_el.text.strip()

        return {
            "transaction_date": self._parse_date(booking),
            "value_date": self._parse_date(val_date),
            "doc_number": doc,
            "bank_operation_code": "",
            "counterparty_name": cp_name,
            "counterparty_code": cp_code,
            "counterparty_account": cp_acct,
            "payment_purpose": purpose,
            "reference_number": ref,
            "amount": abs(amount),
            "currency": amt_el.get("Ccy", "EUR"),
            "direction": direction,
        }


# ────────────────────────────────────────────────────────────
# SEB CSV
# ────────────────────────────────────────────────────────────


class SEBCSVParser(BaseBankParser):
    """
    SEB bank CSV parser.

    Line 1: title (SĄSKAITOS ... IŠRAŠAS)
    Line 2: headers
    Line 3+: data

    Headers:
      DOK NR.; DATA; VALIUTA; SUMA; MOKĖTOJO ARBA GAVĖJO PAVADINIMAS;
      MOKĖTOJO ARBA GAVĖJO IDENTIFIKACINIS KODAS; SĄSKAITA;
      KREDITO ĮSTAIGOS PAVADINIMAS; KREDITO ĮSTAIGOS SWIFT KODAS;
      MOKĖJIMO PASKIRTIS; TRANSAKCIJOS KODAS; DOKUMENTO DATA;
      TRANSAKCIJOS TIPAS; NUORODA; DEBETAS/KREDITAS;
      SUMA SĄSKAITOS VALIUTA; SĄSKAITOS NR; SĄSKAITOS VALIUTA
    """

    bank_name = "seb"

    def parse(self, file_content) -> list[dict]:
        raw = self._to_bytes(file_content)
        encoding = self._detect_encoding(raw)
        text = raw.decode(encoding)

        delimiter = self._detect_separator(text)
        logger.info("[SEBCSV] Detected delimiter: %s", repr(delimiter))
        reader = csv.reader(io.StringIO(text), delimiter=delimiter, quotechar='"')
        logger.info("[SEBCSV] Encoding: %s, length: %d", encoding, len(text))
        logger.info("[SEBCSV] Using delimiter: %s", repr(delimiter))
        rows = list(reader)
        if len(rows) < 3:
            return []

        header_idx = None
        for i, row in enumerate(rows):
            joined = ";".join(row).upper()
            if "DATA" in joined and "SUMA" in joined and "PASKIRTIS" in joined:
                header_idx = i
                break

        if header_idx is None:
            return []

        headers = [h.strip().upper() for h in rows[header_idx]]

        col = {}
        for i, h in enumerate(headers):
            if h == "DOK NR." or h == "DOK NR":
                col["doc_number"] = i
            elif h == "DATA":
                col["date"] = i
            elif h == "VALIUTA" and "currency" not in col:
                col["currency"] = i
            elif h == "SUMA" and "amount" not in col:
                col["amount"] = i
            elif "MOKĖTOJO ARBA GAVĖJO PAVADINIMAS" in h or "MOKETOJO ARBA GAVEJO PAVADINIMAS" in h:
                col["counterparty_name"] = i
            elif "IDENTIFIKACINIS KODAS" in h:
                col["counterparty_code"] = i
            elif h == "SĄSKAITA" or h == "SASKAITA":
                col["counterparty_account"] = i
            elif "MOKĖJIMO PASKIRTIS" in h or "MOKEJIMO PASKIRTIS" in h:
                col["purpose"] = i
            elif h == "DOKUMENTO DATA":
                col["value_date"] = i
            elif h == "NUORODA":
                col["reference"] = i
            elif "DEBETAS/KREDITAS" in h or h == "D/K":
                col["dk"] = i
            elif "TRANSAKCIJOS TIPAS" in h:
                col["tx_type"] = i
            elif "TRANSAKCIJOS KODAS" in h:
                col["tx_code"] = i
            elif h == "SĄSKAITOS NR" or h == "SASKAITOS NR":
                col["account_iban"] = i

        logger.info("[SEBCSV] Header at row %d: %s", header_idx, rows[header_idx][:8])
        logger.info("[SEBCSV] Column mapping: %s", col)

        transactions = []
        for row in rows[header_idx + 1:]:
            if len(row) < 5:
                continue

            def get(field):
                idx = col.get(field)
                if idx is not None and idx < len(row):
                    return row[idx].strip()
                return ""

            amount = self._parse_amount(get("amount"))
            if not amount:
                continue

            txn_date = self._parse_date(get("date"))
            if not txn_date:
                continue

            dk = get("dk").upper()
            direction = "credit" if dk == "C" or dk == "K" else "debit"

            transactions.append({
                "transaction_date": txn_date,
                "value_date": self._parse_date(get("value_date")),
                "doc_number": get("doc_number"),
                "bank_operation_code": get("tx_type"),
                "counterparty_name": get("counterparty_name"),
                "counterparty_code": get("counterparty_code"),
                "counterparty_account": get("counterparty_account"),
                "payment_purpose": get("purpose"),
                "reference_number": get("reference") or get("tx_code"),
                "amount": abs(amount),
                "currency": get("currency") or "EUR",
                "direction": direction,
            })

        logger.info("[SEBCSV] Result: %d transactions", len(transactions))
        return transactions


# ────────────────────────────────────────────────────────────
# Luminor CSV
# ────────────────────────────────────────────────────────────


class LuminorCSVParser(BaseBankParser):
    bank_name = "luminor"

    def parse(self, file_content) -> list[dict]:
        raw = self._to_bytes(file_content)
        encoding = self._detect_encoding(raw)
        text = raw.decode(encoding)

        logger.info("[LuminorCSV] Encoding: %s, length: %d", encoding, len(text))

        delimiter = self._detect_separator(text)
        logger.info("[LuminorCSV] Using delimiter: %s", repr(delimiter))

        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter, quotechar='"'))
        if not rows:
            return []

        header_idx = None
        for i, row in enumerate(rows):
            joined = " | ".join(c.strip().lower() for c in row)
            if "data" in joined and "suma" in joined and "c/d" in joined:
                header_idx = i
                break

        if header_idx is None:
            return []

        headers = [h.strip().lower() for h in rows[header_idx]]
        col_map = self._map_columns(headers)

        logger.info("[LuminorCSV] Header at row %d: %s", header_idx, rows[header_idx][:8])
        logger.info("[LuminorCSV] Column mapping: %s", col_map)

        transactions = []
        for row in rows[header_idx + 1:]:
            if not any(cell.strip() for cell in row):
                continue

            txn = self._parse_row(row, col_map)
            if txn and txn.get("amount") and txn.get("transaction_date"):
                transactions.append(txn)

        logger.info("[LuminorCSV] Result: %d transactions", len(transactions))
        return transactions

    def _map_columns(self, headers):
        mapping = {}

        patterns = {
            "transaction_type": [
                "operacijos/balanso tipas",
            ],
            "transaction_date": [
                "data",
            ],
            "transaction_time": [
                "laikas",
            ],
            "amount": [
                "suma",
            ],
            "equivalent": [
                "ekvivalentas",
            ],
            "direction_flag": [
                "c/d",
            ],
            "orig_amount": [
                "orig. suma",
                "orig suma",
            ],
            "orig_currency": [
                "orig. valiuta",
                "orig valiuta",
            ],
            "doc_number": [
                "operacijos dok. nr.",
                "operacijos dok. nr",
            ],
            "transaction_id": [
                "operacijos eilutė (identifikatorius)",
                "operacijos eilute (identifikatorius)",
                "transaction id",
            ],
            "beneficiary_customer_code": [
                "kliento kodas gavėjo informac. sistemoje",
                "kliento kodas gavejo informac. sistemoje",
            ],
            "payment_code": [
                "įmokos kodas",
                "imokos kodas",
            ],
            "payment_purpose": [
                "mokėjimo paskirtis",
                "mokejimo paskirtis",
                "structured details",
            ],
            "bic": [
                "kitos pusės bic",
                "kitos puses bic",
            ],
            "bank_name": [
                "kitos pusės kredito įstaigos pavadinimas",
                "kitos puses kredito istaigos pavadinimas",
            ],
            "counterparty_account": [
                "kitos pusės sąskaitos nr.",
                "kitos puses saskaitos nr.",
                "account number",
            ],
            "counterparty_name": [
                "kitos pusės pavadinimas",
                "kitos puses pavadinimas",
                "designation",
            ],
            "counterparty_code": [
                "kitos pusės asmens kodas/registracijos nr.",
                "kitos puses asmens kodas/registracijos nr.",
                "reg no",
            ],
            "payer_customer_code": [
                "kitos pusės kliento kodas mokėtojo informacinėje sistemoje",
                "kitos puses kliento kodas moketojo informacineje sistemoje",
            ],
            "ultimate_payer_account": [
                "pradinio mokėtojo sąskaitos nr.",
                "pradinio moketojo saskaitos nr.",
            ],
            "ultimate_payer_name": [
                "pradinio mokėtojo vardas ir pavardė/pavadinimas",
                "pradinio moketojo vardas ir pavarde/pavadinimas",
            ],
            "ultimate_payer_code": [
                "pradinio mokėtojo asmens kodas/registracijos nr.",
                "pradinio moketojo asmens kodas/registracijos nr.",
            ],
            "ultimate_beneficiary_account": [
                "galutinio gavėjo sąskaitos nr.",
                "galutinio gavejo saskaitos nr.",
            ],
            "ultimate_beneficiary_name": [
                "galutinio gavėjo vardas ir pavardė/pavadinimas",
                "galutinio gavejo vardas ir pavarde/pavadinimas",
            ],
            "ultimate_beneficiary_code": [
                "galutinio gavėjo asmens kodas/registracijos nr.",
                "galutinio gavejo asmens kodas/registracijos nr.",
            ],
        }

        for field, terms in patterns.items():
            for i, h in enumerate(headers):
                if any(term in h for term in terms):
                    mapping[field] = i
                    break

        return mapping

    def _parse_row(self, row, col_map):
        def get(field):
            idx = col_map.get(field)
            return row[idx].strip() if idx is not None and idx < len(row) else ""

        amount = self._parse_amount(get("amount"))
        if amount is None:
            return None

        cd = get("direction_flag").upper()
        if cd == "C":
            direction = "credit"
        elif cd == "D":
            direction = "debit"
        else:
            direction = "credit" if amount > 0 else "debit"

        payment_purpose = get("payment_purpose")
        payment_code = get("payment_code")
        bank_operation_code = get("transaction_type") or payment_code

        currency = get("orig_currency") or "EUR"

        return {
            "transaction_date": self._parse_date(get("transaction_date")),
            "value_date": self._parse_date(get("transaction_date")),
            "doc_number": get("doc_number"),
            "bank_operation_code": bank_operation_code,
            "counterparty_name": get("counterparty_name"),
            "counterparty_code": get("counterparty_code"),
            "counterparty_account": get("counterparty_account"),
            "payment_purpose": payment_purpose,
            "reference_number": get("transaction_id"),
            "amount": abs(amount),
            "currency": currency,
            "direction": direction,
        }


# ────────────────────────────────────────────────────────────
# Revolut CSV
# ────────────────────────────────────────────────────────────


class RevolutCSVParser(BaseBankParser):
    bank_name = "revolut"

    def _norm_key(self, key: str) -> str:
        """
        Normalize CSV header:
        - trims spaces
        - lowercases
        - turns dashes/underscores into spaces
        - collapses multiple spaces
        """
        key = (key or "").replace("\ufeff", "").strip().lower()

        # Different dash symbols + underscore → space
        key = re.sub(r"[\u2010\u2011\u2012\u2013\u2014\u2015\-_]+", " ", key)

        # Remove repeated spaces
        key = re.sub(r"\s+", " ", key).strip()

        return key

    def _normalize_row(self, row: dict) -> dict:
        """
        Convert row keys to normalized form.
        Example:
        'Started-Date' -> 'started date'
        'Started_Date' -> 'started date'
        'STARTED DATE ' -> 'started date'
        """
        normalized = {}

        for key, value in row.items():
            norm_key = self._norm_key(key)
            if norm_key:
                normalized[norm_key] = value

        return normalized

    def _get(self, row: dict, *names: str, default: str = "") -> str:
        """
        Get value by normalized header aliases.
        """
        for name in names:
            norm_name = self._norm_key(name)
            val = row.get(norm_name)
            if val not in (None, ""):
                return str(val).strip()

        return default

    def _revolut_ref(self, row: dict) -> str:
        """
        Revolut CSV neturi Transaction ID, todėl kuriame stabilų synthetic ref.
        Svarbiausia: Started Date su laiku + Completed Date + Balance.
        Tai leidžia atskirti 2 vienodus card payments tą pačią dieną.
        """
        tx_type = self._get(row, "Type")
        product = self._get(row, "Product")
        started = self._get(row, "Date started (UTC)", "Started Date", "Started")
        completed = self._get(row, "Date completed (UTC)", "Completed Date", "Completed")
        balance = self._get(row, "Balance", "Balance after")
        state = self._get(row, "State", "Status").upper()

        return f"REV|{tx_type}|{product}|{started}|{completed}|{balance}|{state}"[:255]

    def parse(self, file_content) -> list[dict]:
        raw = self._to_bytes(file_content)
        text = raw.decode(self._detect_encoding(raw))

        logger.info("[RevolutCSV] File length: %d", len(text))

        reader = csv.DictReader(io.StringIO(text))
        logger.info("[RevolutCSV] Headers: %s", reader.fieldnames)

        normalized_headers = [
            self._norm_key(h) for h in (reader.fieldnames or [])
        ]
        logger.info("[RevolutCSV] Normalized headers: %s", normalized_headers)

        transactions = []
        skipped = 0

        for raw_row in reader:
            row = self._normalize_row(raw_row)

            amt_str = self._get(row, "Amount")
            amount = self._parse_amount(amt_str)

            if amount is None:
                skipped += 1
                continue

            completed_date = self._get(row, "Date completed (UTC)", "Completed Date", "Completed")
            started_date = self._get(row, "Date started (UTC)", "Started Date", "Started")
            date_str = completed_date or started_date

            desc = self._get(
                row,
                "Description",
                "Merchant",
                "Counterparty",
            )

            reference_text = self._get(row, "Reference")
            purpose = " ".join(x for x in (desc, reference_text) if x).strip() or desc

            state = self._get(row, "State", "Status")

            if state.lower() in ("reverted", "failed", "declined"):
                skipped += 1
                continue

            txn_date = self._parse_date(date_str.split()[0] if date_str else "")
            if not txn_date:
                skipped += 1
                logger.warning(
                    "[RevolutCSV] Skipped row: could not parse date. date_str=%r row=%s",
                    date_str,
                    raw_row,
                )
                continue

            tx_type = self._get(row, "Type")
            bank_operation_code = " | ".join(
                x for x in [tx_type, state]
                if x
            )

            currency = self._get(row, "Payment currency", "Currency", "Orig currency", default="EUR")

            transactions.append({
                "transaction_date": txn_date,
                "value_date": None,
                "doc_number": "",
                "bank_operation_code": bank_operation_code,
                "counterparty_name": desc,
                "counterparty_code": "",
                "counterparty_account": "",
                "payment_purpose": purpose,
                "reference_number": self._get(row, "ID", "Transaction ID") or self._revolut_ref(row),
                "amount": abs(amount),
                "currency": currency or "EUR",
                "direction": "credit" if amount > 0 else "debit",
            })

        logger.info(
            "[RevolutCSV] Result: %d transactions, %d skipped",
            len(transactions),
            skipped,
        )

        return transactions


# ────────────────────────────────────────────────────────────
# PayPal XLSX (activity export)
# ────────────────────────────────────────────────────────────


class PayPalXLSXParser(BaseBankParser):
    """
    PayPal activity export (.xlsx).
    - Режем внутренние движения: reserve/hold/authorization/currency conversion.
    - Дедуп по Transaction ID (Memo-эхо схлопывается).
    - Направление по знаку Gross.
    - Комиссия (Fee, отрицательная) — отдельной операцией.
    - Валюта: EUR-нога General Currency Conversion (Reference Txn ID == Transaction ID
      платежа) даёт amount_eur; курс = foreign/eur.
    """

    bank_name = "paypal"

    SKIP_TYPE = re.compile(r"reserve|hold|authorization|currency conversion", re.I)

    HEADER_ALIASES = {
        "date": "transaction_date", "name": "name", "type": "type",
        "status": "status", "currency": "currency", "gross": "gross",
        "fee": "fee", "net": "net", "from email address": "from_email",
        "to email address": "to_email", "transaction id": "txn_id",
        "item title": "item_title", "invoice number": "invoice_number",
        "reference txn id": "ref_txn_id", "balance impact": "balance_impact",
        "subject": "subject", "note": "note",
    }

    def _dec(self, v):
        if v is None or v == "":
            return None
        if isinstance(v, (int, float, Decimal)):
            try:
                return Decimal(str(v))
            except (InvalidOperation, ValueError):
                return None
        s = str(v).strip().replace("\xa0", "").replace(" ", "")
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None

    def _to_date(self, v):
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        if isinstance(v, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y.%m.%d"):
                try:
                    return datetime.strptime(v.strip(), fmt).date()
                except ValueError:
                    continue
        return None

    def parse(self, file_content) -> list[dict]:
        from collections import defaultdict
        from openpyxl import load_workbook

        raw = self._to_bytes(file_content)
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active

        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return []

        col = {}
        for i, h in enumerate(header):
            key = self.HEADER_ALIASES.get(str(h or "").strip().lower())
            if key and key not in col:
                col[key] = i

        if "gross" not in col or "txn_id" not in col:
            logger.warning("[PayPalXLSX] Missing key columns: %s", list(col))
            return []

        def g(row, field):
            i = col.get(field)
            return row[i] if i is not None and i < len(row) else None

        allrows = list(it)

        # ── Pass 0: EUR-ноги конверсий по Reference Txn ID ──
        conv_eur = defaultdict(Decimal)
        for row in allrows:
            if "Currency Conversion" in str(g(row, "type") or "") \
                    and str(g(row, "currency") or "") == "EUR":
                ref = str(g(row, "ref_txn_id") or "").strip()
                amt = self._dec(g(row, "gross"))
                if ref and amt is not None:
                    conv_eur[ref] += abs(amt)

        # ── Pass 1: группировка реальных платежей по Transaction ID ──
        groups = {}
        synth = 0
        for row in allrows:
            if self.SKIP_TYPE.search(str(g(row, "type") or "")):
                continue
            if self._dec(g(row, "gross")) is None:
                continue
            tid = str(g(row, "txn_id") or "").strip()
            if not tid:
                synth += 1
                tid = "__NOID_%d" % synth
            groups.setdefault(tid, []).append(row)

        def first_nonempty(grp, field):
            for r in grp:
                v = g(r, field)
                if v not in (None, ""):
                    return str(v).strip()
            return ""

        Q2 = Decimal("0.01")
        Q6 = Decimal("0.000001")
        transactions = []
        for tid, grp in groups.items():
            rep = next(
                (r for r in grp if self._dec(g(r, "fee")) not in (None, Decimal("0"))),
                None,
            )
            if rep is None:
                rep = next(
                    (r for r in grp if str(g(r, "balance_impact") or "") in ("Credit", "Debit")),
                    None,
                )
            if rep is None:
                rep = grp[0]

            gross = self._dec(g(rep, "gross")) or Decimal("0")
            txn_date = self._to_date(g(rep, "transaction_date"))
            if not txn_date:
                continue

            direction = "credit" if gross > 0 else "debit"
            currency = first_nonempty(grp, "currency") or "EUR"
            amount = abs(gross)
            real_tid = "" if tid.startswith("__NOID_") else tid

            if currency == "EUR":
                amount_eur, rate, rate_date = amount, None, None
            else:
                eur = conv_eur.get(real_tid)
                if eur and eur > 0:
                    amount_eur = eur.quantize(Q2, ROUND_HALF_UP)
                    rate = (amount / eur).quantize(Q6, ROUND_HALF_UP)
                    rate_date = txn_date
                else:
                    amount_eur, rate, rate_date = None, None, None

            name = first_nonempty(grp, "name")
            cp_email = (
                first_nonempty(grp, "from_email")
                if direction == "credit"
                else first_nonempty(grp, "to_email")
            )

            transactions.append({
                "transaction_date": txn_date,
                "value_date": txn_date,
                "doc_number": first_nonempty(grp, "invoice_number"),
                "bank_operation_code": first_nonempty(grp, "type"),
                "counterparty_name": name or cp_email,
                "counterparty_code": "",
                "counterparty_account": "",
                "payment_purpose": first_nonempty(grp, "type"),
                "reference_number": real_tid,
                "amount": amount,
                "currency": currency,
                "amount_eur": amount_eur,
                "exchange_rate": rate,
                "exchange_rate_date": rate_date,
                "direction": direction,
            })

            fee = self._dec(g(rep, "fee"))
            if fee is not None and fee != 0:
                fee_amt = abs(fee)
                if currency == "EUR":
                    fee_eur = fee_amt
                elif rate:
                    fee_eur = (fee_amt / rate).quantize(Q2, ROUND_HALF_UP)
                else:
                    fee_eur = None
                transactions.append({
                    "transaction_date": txn_date,
                    "value_date": txn_date,
                    "doc_number": "",
                    "bank_operation_code": "PayPal Fee",
                    "counterparty_name": "PayPal",
                    "counterparty_code": "",
                    "counterparty_account": "",
                    "payment_purpose": "PayPal komisinis mokestis | %s" % real_tid,
                    "reference_number": (real_tid + "-FEE") if real_tid else "",
                    "amount": fee_amt,
                    "currency": currency,
                    "amount_eur": fee_eur,
                    "exchange_rate": rate,
                    "exchange_rate_date": rate_date,
                    "direction": "debit" if fee < 0 else "credit",
                })

        logger.info("[PayPalXLSX] %d ops emitted", len(transactions))
        return transactions


# ────────────────────────────────────────────────────────────
# Registry & Detection
# ────────────────────────────────────────────────────────────


PARSER_REGISTRY = {
    ("swedbank", "csv"): SwedbankCSVParser,
    ("swedbank", "xml"): ISO20022Parser,
    ("seb", "csv"): SEBCSVParser,
    ("seb", "xml"): ISO20022Parser,
    ("luminor", "csv"): LuminorCSVParser,
    ("luminor", "xml"): ISO20022Parser,
    ("siauliu", "csv"): SwedbankCSVParser,
    ("revolut", "csv"): RevolutCSVParser,
    ("paypal", "xlsx"): PayPalXLSXParser,
}


def get_parser(bank_name: str, file_format: str) -> BaseBankParser:
    key = (bank_name.lower(), file_format.lower())
    cls = PARSER_REGISTRY.get(key)
    if not cls:
        raise ValueError(f"No parser for {key}. Supported: {list(PARSER_REGISTRY.keys())}")
    return cls()

def _detect_bank_from_camt(content: bytes) -> Optional[str]:
    """
    camt.053 XML: banką nustatome pagal sąskaitos aptarnautojo BIC (<Svcr>),
    o ne pagal kontrahentų bankus operacijose.
    """
    try:
        root = None
        for enc in ("utf-8-sig", "utf-8", "windows-1257", "iso-8859-13"):
            try:
                root = ET.fromstring(content.decode(enc))
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if root is None:
            return None
    except ET.ParseError:
        return None

    tag = root.tag
    ns = tag.split("}")[0].lstrip("{") if "{" in tag else ""

    def q(t):
        return f"{{{ns}}}{t}" if ns else t

    svcr = root.find(f".//{q('Acct')}/{q('Svcr')}/{q('FinInstnId')}")
    bic = name = ""
    if svcr is not None:
        b = svcr.find(q("BIC"))
        if b is not None and b.text:
            bic = b.text.strip().upper()
        n = svcr.find(q("Nm"))
        if n is not None and n.text:
            name = n.text.strip().lower()

    bic_map = {"HABA": "swedbank", "CBVI": "seb", "AGBL": "luminor", "CBSB": "siauliu"}
    if bic and bic[:4] in bic_map:
        return bic_map[bic[:4]]

    if "swedbank" in name:
        return "swedbank"
    if "seb" in name:
        return "seb"
    if "luminor" in name:
        return "luminor"
    if "artea" in name or "šiaulių" in name or "siauli" in name:
        return "siauliu"

    return None

def _detect_paypal_xlsx(content: bytes) -> bool:
    """PayPal activity export (xlsx): узнаём по заголовкам (оба формата — старый без Balance Impact)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        header = next(ws.iter_rows(max_row=1, values_only=True), ())
        cols = {str(h or "").strip().lower() for h in header}
        return {"transaction id", "gross", "fee", "net"}.issubset(cols)
    except Exception:
        return False


def detect_bank_from_content(content: bytes) -> Optional[str]:
    # ── PayPal XLSX (бинарный zip, текстом не читается) ──
    if content[:2] == b"PK" and _detect_paypal_xlsx(content):
        return "paypal"

    text = ""

    for enc in ("utf-8-sig", "utf-8", "windows-1257"):
        try:
            text = content[:10000].decode(enc).lower()
            break
        except UnicodeDecodeError:
            continue

    compact = re.sub(r"\s+", " ", text)

    # ── ISO 20022 camt.053 XML (Swedbank/SEB/Luminor/Šiaulių) ──
    # Banką nustatome pagal aptarnautojo BIC (<Svcr>), ne pagal kontrahentus.
    if "camt.05" in compact or "<document" in compact:
        camt_bank = _detect_bank_from_camt(content)
        if camt_bank:
            return camt_bank

    # ── Revolut ─────────────────────────────────────────
    # Du eksporto formatai: senas ("Started Date"/"Completed Date")
    # ir Business ("Date started (UTC)"/"Date completed (UTC)").
    rev_started = "started date" in compact or "date started" in compact
    rev_completed = "completed date" in compact or "date completed" in compact
    if rev_started and rev_completed and "balance" in compact:
        return "revolut"

    # Business formato specifiniai stulpeliai
    if (
        "date completed (utc)" in compact
        or "spend program" in compact
        or ("orig currency" in compact and "payment currency" in compact)
    ):
        return "revolut"

    # ── SEB ─────────────────────────────────────────────
    # Svarbu: SEB tikriname prieš Swedbank.
    # SEB faile "Swedbank AB" gali būti tik kontrahento bankas.
    seb_header_signals = [
        "mokėtojo arba gavėjo pavadinimas",
        "moketojo arba gavejo pavadinimas",
        "debetas/kreditas",
        "sąskaitos nr",
        "saskaitos nr",
        "sąskaitos valiuta",
        "saskaitos valiuta",
        "transakcijos tipas",
        "transakcijos kodas",
    ]

    if (
        ("dok nr" in compact or "dok nr." in compact)
        and "data" in compact
        and "suma" in compact
        and sum(1 for s in seb_header_signals if s in compact) >= 2
    ):
        return "seb"

    # SEB IBAN bank code 7044 kaip fallback.
    if "7044" in compact and "debetas/kreditas" in compact:
        return "seb"

    # ── Luminor ─────────────────────────────────────────
    if "luminor" in compact or "dnb" in compact:
        return "luminor"

    # ── Artea (buvęs Šiaulių bankas) ───────────────────
    if "artea" in compact or "šiaulių" in compact or "siauliu" in compact:
        return "siauliu"

    # ── Swedbank ────────────────────────────────────────
    # Neužtenka tiesiog rasti žodį "swedbank",
    # nes SEB išraše jis gali būti kontrahento banko pavadinime.
    swedbank_header_signals = [
        "gavėjas/mokėtojas",
        "gavejas/moketojas",
        "operacijos paskirtis",
        "banko žyma",
        "d/k",
    ]

    if (
        ("swedbank" in compact or "habalt" in compact)
        and sum(1 for s in swedbank_header_signals if s in compact) >= 2
    ):
        return "swedbank"

    return None


def detect_format_from_content(content: bytes) -> str:
    start = content[:100].strip()
    if start.startswith(b"<?xml") or start.startswith(b"<Document"):
        return "xml"
    if content[:2] == b"PK":       # XLSX (OOXML zip)
        return "xlsx"
    return "csv"
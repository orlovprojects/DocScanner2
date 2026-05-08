from django.db import transaction
from ..models import ProductAutocomplete, ClientAutocomplete
from openpyxl import load_workbook
import io
from ..utils.client_autocomplete_upsert import normalize_name


def _get_xlsx_rows(file, required_fields):
    """
    Читает строки из XLSX и возвращает list[dict] с НОРМАЛИЗОВАННЫМИ ключами заголовков.
    Нормализация: trim, lower, удаление конечной '*' (pvz. 'imones_kodas*' -> 'imones_kodas').
    """
    file.seek(0)
    file_bytes = file.read()
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # 1) заголовки raw
    raw_header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    raw_headers = [str(c.value).strip() if c.value else "" for c in raw_header_cells]

    # 2) normalizacija
    def norm(h: str) -> str:
        h = (h or "").strip()
        if '(' in h:
            h = h[:h.index('(')]
        h = h.strip()                 
        if h.endswith("*"):
            h = h[:-1]
        return h.strip().lower()

    norm_headers = [norm(h) for h in raw_headers]

    # 3) validate required (jau be žvaigždučių)
    need = {f.strip().lower() for f in required_fields}
    have = set(norm_headers)
    if not need.issubset(have):
        missing = ", ".join(sorted(need - have))
        raise Exception(f"Nerasta {missing} stulpelio")

    # 4) rows -> dict su normalizuotais raktai
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        d = {}
        for i, val in enumerate(row[:len(norm_headers)]):
            key = norm_headers[i]
            d[key] = (str(val).strip() if val is not None else "")
        rows.append(d)

    return rows


def _norm_preke_paslauga(value) -> str:
    """
    Normalizuoja 'preke_paslauga_kodas' į '1' | '2' | '3' arba ''.
    Priima 1/2/3 (su tarpais, kableliais), bei žodinius sinonimus.
    """
    if value is None:
        return ""
    s = str(value).strip().lower()
    if not s:
        return ""

    # skaičiai (leidžiam '1', '1.0', ' 2 ', '3,0')
    s_num = s.replace(",", ".")
    try:
        n = int(float(s_num))
        if n in (1, 2, 3, 4):
            return str(n)
    except ValueError:
        pass

    # žodiniai variantai
    preke_syn = {"preke", "prekė", "prekes", "prekės"}
    paslauga_syn = {"paslauga", "paslaugos"}
    kodas_syn = {"kodas", "kodai"}

    if s in preke_syn:
        return "1"
    if s in paslauga_syn:
        return "2"
    if s in kodas_syn:
        return "3"

    return ""  # neatspažinta — paliekam tuščią


def import_products_from_xlsx(user, file):
    imported = 0
    updated = 0
    skipped_empty = 0
    skipped_duplicate = 0
    total = 0
    errors = []

    try:
        rows = _get_xlsx_rows(file, required_fields=["prekes_kodas", "prekes_pavadinimas"])
    except Exception as e:
        return {"error": str(e)}

    seen_codes = set()

    with transaction.atomic():
        for row_num, data in enumerate(rows, start=2):
            total += 1

            prekes_kodas = (data.get('prekes_kodas') or '').strip()
            prekes_pavadinimas = (data.get('prekes_pavadinimas') or '').strip()

            preke_paslauga = _norm_preke_paslauga(data.get('preke_paslauga_kodas'))

            missing = []
            if not prekes_pavadinimas:
                missing.append("prekes_pavadinimas")
            if not prekes_kodas:
                missing.append("prekes_kodas")
            if not preke_paslauga:
                missing.append("preke_paslauga_kodas")
            if missing:
                skipped_empty += 1
                errors.append(f"Eilutė {row_num}: trūksta {', '.join(missing)}")
                continue

            # ── Дубликат внутри файла ──
            if prekes_kodas in seen_codes:
                skipped_duplicate += 1
                errors.append(f"Eilutė {row_num}: kodas '{prekes_kodas}' kartojasi faile")
                continue
            seen_codes.add(prekes_kodas)

            preke_paslauga = _norm_preke_paslauga(data.get('preke_paslauga_kodas'))

            field_values = dict(
                prekes_pavadinimas=prekes_pavadinimas,
                prekes_barkodas=(data.get('prekes_barkodas') or '').strip(),
                preke_paslauga=preke_paslauga,
            )

            existing = ProductAutocomplete.objects.filter(
                user=user, prekes_kodas=prekes_kodas
            ).first()

            if existing:
                for attr, val in field_values.items():
                    setattr(existing, attr, val)
                existing.save()
                updated += 1
            else:
                ProductAutocomplete.objects.create(
                    user=user, prekes_kodas=prekes_kodas, **field_values
                )
                imported += 1

    return {
        "imported": imported,
        "updated": updated,
        "skipped_empty": skipped_empty,
        "skipped_duplicate": skipped_duplicate,
        "processed": total,
        "errors": errors,
    }


def _norm_fizinis_asmuo(value):
    """
    Grąžina True, False arba None (jei tuščia/neatpažinta).
    Priima: Taip/Ne, True/False, 1/0, Yes/No, T/N, F ir t.t.
    """
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if s in ("taip", "true", "1", "yes", "t"):
        return True
    if s in ("ne", "false", "0", "no", "n", "f"):
        return False
    return None  # neatpažinta reikšmė


def import_clients_from_xlsx(user, file):
    imported = 0
    updated = 0
    skipped_empty = 0
    skipped_duplicate = 0
    errors = []

    try:
        rows = _get_xlsx_rows(file, required_fields=["kodas", "pavadinimas"])
    except Exception as e:
        return {"error": str(e)}

    seen_codes = set()

    with transaction.atomic():
        for row_num, data in enumerate(rows, start=2):
            name = (data.get('pavadinimas') or '').strip()
            code = (data.get('kodas') or '').strip()
            pvm = (data.get('pvm_kodas') or '').strip()
            fizinis = _norm_fizinis_asmuo(data.get('fizinis_asmuo'))

            # ── Валидация обязательных полей ──
            missing = []
            if not name:
                missing.append("pavadinimas")
            if not code:
                missing.append("kodas")
            if fizinis is None:
                missing.append("fizinis_asmuo")

            country = (data.get('salies_kodas') or '').strip().upper()
            if len(country) != 2 or not country.isalpha():
                country = ""
            if not country:
                missing.append("salies_kodas")

            if missing:
                skipped_empty += 1
                errors.append(f"Eilutė {row_num}: trūksta {', '.join(missing)}")
                continue

            # ── Дубликат внутри файла ──
            if code in seen_codes:
                skipped_duplicate += 1
                errors.append(f"Eilutė {row_num}: kodas '{code}' kartojasi faile")
                continue
            seen_codes.add(code)

            field_values = dict(
                pavadinimas=name,
                is_person=fizinis,
                pvm_kodas=pvm,
                ibans=(data.get('iban') or '').strip(),
                address=(data.get('adresas') or '').strip(),
                country_iso=country,
                kodas_programoje=(data.get('kodas_programoje') or '').strip(),
                name_normalized=normalize_name(name),
            )

            # ── Upsert ──
            existing = None
            if code:
                existing = ClientAutocomplete.objects.filter(
                    user=user, imones_kodas=code, source="imported"
                ).first()
            if not existing and pvm:
                existing = ClientAutocomplete.objects.filter(
                    user=user, pvm_kodas=pvm, source="imported"
                ).first()

            try:
                if existing:
                    for attr, val in field_values.items():
                        setattr(existing, attr, val)
                    if code:
                        existing.imones_kodas = code
                    existing.save()
                    updated += 1
                else:
                    ClientAutocomplete.objects.create(
                        user=user,
                        imones_kodas=code,
                        source="imported",
                        doc_count=0,
                        **field_values,
                    )
                    imported += 1
            except Exception as e:
                errors.append(f"Eilutė {row_num}: {str(e)}")

    return {
        "imported": imported,
        "updated": updated,
        "skipped_empty": skipped_empty,
        "skipped_duplicate": skipped_duplicate,
        "processed": len(rows),
        "errors": errors,
    }
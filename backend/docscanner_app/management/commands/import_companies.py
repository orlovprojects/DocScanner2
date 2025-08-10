import os
import datetime
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from docscanner_app.models import Company

class Command(BaseCommand):
    help = "Импорт компаний из Excel-файла (XLSX) в базу данных"

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Путь к Excel-файлу')

    def handle(self, *args, **options):
        filepath = options['filepath']
        if not os.path.exists(filepath):
            self.stderr.write(self.style.ERROR(f"Файл не найден: {filepath}"))
            return

        batch_size = 10000
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {k: i for i, k in enumerate(header)}

        def parse_date(val):
            if not val:
                return None
            if isinstance(val, (datetime.datetime, datetime.date)):
                return val
            for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%d.%m.%Y"):
                try:
                    return datetime.datetime.strptime(val, fmt).date()
                except Exception:
                    continue
            return None

        objects = []
        total = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            obj = Company(
                im_kodas    = row[idx["ja_kodas"]],
                pavadinimas = row[idx["pavadinimas"]],
                ireg_data   = parse_date(row[idx["ireg_data"]]),
                isreg_data  = parse_date(row[idx["isreg_data"]]),
                tipas       = row[idx["tipo_kodas"]],
                pvm_kodas   = row[idx["pvm_kodas"]],
                pvm_ireg    = parse_date(row[idx["pvm_iregistruota"]]),
                pvm_isreg   = parse_date(row[idx["pvm_isregistruota"]])
            )
            objects.append(obj)
            if len(objects) >= batch_size:
                Company.objects.bulk_create(objects, ignore_conflicts=True)
                total += len(objects)
                self.stdout.write(f"Импортировано {total} записей...")
                objects = []
        if objects:
            Company.objects.bulk_create(objects, ignore_conflicts=True)
            total += len(objects)
            self.stdout.write(f"Импортировано {total} записей (финальная партия).")

        self.stdout.write(self.style.SUCCESS("Импорт завершён!"))
